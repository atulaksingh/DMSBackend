from rest_framework.response import Response
from django.shortcuts import get_object_or_404
from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework import status
from api.models import *
from api.serializers import *
from rest_framework_simplejwt.serializers import TokenObtainPairSerializer
from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework.permissions import IsAuthenticated, IsAdminUser
from django.contrib.auth.hashers import make_password
from decimal import Decimal, InvalidOperation
# for sending mails and generate token
from django.template.loader import render_to_string # used returns the resulting content as a string
from django.utils.http import urlsafe_base64_decode, urlsafe_base64_encode #  used to safely encode and decode data in a URL-friendly format
from .utils import TokenGenerator, generate_token
from django.utils.encoding import force_bytes, force_text, DjangoUnicodeDecodeError #  helps in managing string and byte conversions
from django.utils.encoding import force_str
from django.core.mail import EmailMessage # used to construct and send email messages
from django.conf import settings
from django.views.generic import View
from django.shortcuts import render
from rest_framework.generics import GenericAPIView
from rest_framework.mixins import CreateModelMixin
from rest_framework import generics
from openpyxl import load_workbook
from openpyxl import load_workbook
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from .models import Client, PF
from .serializers import PfSerializer
import pandas as pd
from django.http import JsonResponse
from django.db.models import Sum, F
from django.db.models.functions import Coalesce
from django.core.exceptions import ObjectDoesNotExist
from collections import defaultdict
from django.http import QueryDict
from decimal import Decimal, InvalidOperation
from django.db import transaction
import traceback
from rest_framework import status
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from rest_framework.views import APIView
from openpyxl import load_workbook
from rest_framework.decorators import api_view
from rest_framework.response import Response
from .models import PF  # Make sure this matches your model import
from rest_framework.decorators import api_view
from rest_framework.response import Response
from .models import PF
from django.db.models import Sum
from django.db import models
from rest_framework.decorators import api_view
from rest_framework.response import Response
from django.db.models import Sum
from .models import PF
from django.http import FileResponse, Http404
# from .permissions import IsSuperAdminOrOwnClient
from api.permission import IsSuperAdminOrOwnClient

#********************************************* safe_decimal**************************************************

# Helper function to safely convert to Decimal
def safe_decimal(value, default='0'):
    try:
        return Decimal(value)
    except (ValueError, InvalidOperation):
        return Decimal(default)

def safe_decimal(value, default=0):
    try:
        return Decimal(value)
    except (InvalidOperation, TypeError, ValueError):
        return Decimal(default)

def safe_decimal(value):
    try:
        return Decimal(value)
    except Exception:
        return Decimal(0)

# *******************************************Client View's***********************************************

# ****************************************Frontend create-client api*************************************
@api_view(['POST'])
def create_client(request):
    if request.method == 'POST':
        print("Received data:", request.data)  # Debugging output
        client_serializer = ClientSerializer(data=request.data)

        if client_serializer.is_valid():
            client = client_serializer.save()
            print('clien_serializer',client)
            # Access fileinfos data
            fileinfos_data = []
            for key in request.POST.keys():
                if key.startswith('fileinfos['):
                    # Parse the index from the key
                    index = key.split('[')[1].split(']')[0]
                    if index.isdigit():
                        fileinfo_str = request.POST.get(f'fileinfos[{index}]')
                        if fileinfo_str:
                            fileinfo_data = json.loads(fileinfo_str)
                            files = request.FILES.getlist(f'fileinfos[{index}].files[]')
                            fileinfos_data.append({
                                'fileinfo': fileinfo_data,
                                'files': files
                            })
            print('clien_serializer222',fileinfos_data)

            # Save FileInfo and File instances
            for entry in fileinfos_data:
                fileinfo = FileInfo.objects.create(client=client, **entry['fileinfo'])
                for file in entry['files']:
                    File.objects.create(fileinfo=fileinfo, files=file)

            # return Response(client_serializer.data, status=201)
            return Response({'message': 'Client created successfully','data': client_serializer.data},status=status.HTTP_201_CREATED)

        return Response({'message': 'data not created','error_message': client_serializer.errors},status=status.HTTP_400_BAD_REQUEST)

# *****************************Backend create-client api*************************************
# @api_view(['POST'])
# def create_client(request):
#     if request.method == 'POST':
#         print("Received data:", request.data)  # Debugging output
#         client_serializer = ClientSerializer(data=request.data)

#         if client_serializer.is_valid():
#             client = client_serializer.save()

#             # Access fileinfos data
#             fileinfos_data = []
#             index = 0
#             while f'fileinfos[{index}].login' in request.POST:
#                 fileinfo_data = {
#                     'login': request.POST.get(f'fileinfos[{index}].login'),
#                     'password': request.POST.get(f'fileinfos[{index}].password'),
#                     'document_type': request.POST.get(f'fileinfos[{index}].document_type'),
#                     'remark': request.POST.get(f'fileinfos[{index}].remark'),
#                 }

#                 files = request.FILES.getlist(f'fileinfos[{index}].files[]')

#                 fileinfos_data.append({
#                     'fileinfo': fileinfo_data,
#                     'files': files
#                 })

#                 index += 1

#             # Save FileInfo and File instances
#             for entry in fileinfos_data:
#                 fileinfo = FileInfo.objects.create(client=client, **entry['fileinfo'])
#                 for file in entry['files']:
#                     File.objects.create(fileinfo=fileinfo, files=file)

#             print('Data', client_serializer.data)

#             return Response(client_serializer.data, status=201)

#         return Response(client_serializer.errors, status=400)


# @api_view(['GET'])
# # @permission_classes([IsAuthenticated])
# def list_client(request):
    if request.method == 'GET':
        client = Client.objects.all()
        hsn = HSNCode.objects.all()
        product = Product.objects.all()
        product_description = ProductDescription.objects.all()

        serializers = ClientSerializer(client, many=True)
        serializers2 = HSNSerializer(hsn, many=True)
        serializers3 = ProductSerializer(product, many=True)
        serializers4 = ProductDescriptionSerializer(product_description, many=True)

        context = {
            'clients':serializers.data,
            'hsn':serializers2.data,
            'product':serializers3.data,
            'product_description':serializers4.data
        }

        return Response(context)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def list_client(request):
    user = request.user

    if user.role == 'superuser':
        clients = Client.objects.all()
        hsn = HSNCode.objects.all()
        product = Product.objects.all()
        product_description = ProductDescription.objects.all()

    elif user.role == 'clientuser':
        clients = Client.objects.filter(id=user.client_id)
        hsn = HSNCode.objects.all()
        product = Product.objects.all()
        product_description = ProductDescription.objects.all()
    
    elif user.role == 'customeruser':
        clients = Client.objects.filter(id=user.client_id)
        hsn = HSNCode.objects.all()
        product = Product.objects.all()
        product_description = ProductDescription.objects.all()
    else:
        return Response({'error': 'Unauthorized user role'}, status=403)

    hsn = HSNCode.objects.all()
    product = Product.objects.all()
    product_description = ProductDescription.objects.all()

    client_serializer = ClientSerializer(clients, many=True)
    hsn_serializer = HSNSerializer(hsn, many=True)
    product_serializer = ProductSerializer(product, many=True)
    desc_serializer = ProductDescriptionSerializer(product_description, many=True)

    return Response({
        'clients': client_serializer.data,
        'hsn': hsn_serializer.data,
        'product': product_serializer.data,
        'product_description': desc_serializer.data
    })


# @api_view(['GET'])
# @permission_classes([IsAuthenticated])
# def list_client(request):
#     user_email = request.user.email

#     try:
#         # Check if this email belongs to a ClientUser
#         client_user = ClientUser.objects.get(email=user_email)
#         # Filter client based on ClientUser's client
#         clients = Client.objects.filter(id=client_user.client.id)

#     except ClientUser.DoesNotExist:
#         # Maybe it's a DashboardUser or superuser? Show all
#         clients = Client.objects.all()

#     serializer = ClientSerializer(clients, many=True)
#     return Response(serializer.data, status=status.HTTP_200_OK)







# @api_view(['GET'])
# # @permission_classes([IsAuthenticated])
# def list_client(request):
    user = request.user

    if hasattr(user, 'dashboarduser') and user.dashboarduser.superadmin_user:
        # Superadmin → full access
        clients = Client.objects.all()
        hsns = HSNCode.objects.all()
        products = Product.objects.all()
        descriptions = ProductDescription.objects.all()
    elif hasattr(user, 'clientuser'):
        # Client user → only their client data
        client = user.clientuser.client
        clients = Client.objects.filter(id=client.id)
        hsns = HSNCode.objects.filter(client=client)
        products = Product.objects.filter(client=client)
        descriptions = ProductDescription.objects.filter(client=client)
    else:
        # Not allowed
        return Response({"detail": "Permission denied."}, status=403)

    # Serializing
    serializers = ClientSerializer(clients, many=True)
    serializers2 = HSNSerializer(hsns, many=True)
    serializers3 = ProductSerializer(products, many=True)
    serializers4 = ProductDescriptionSerializer(descriptions, many=True)

    context = {
        'clients': serializers.data,
        'hsn': serializers2.data,
        'product': serializers3.data,
        'product_description': serializers4.data
    }

    return Response(context)



# @api_view(['GET', 'POST'])
# def edit_client(request, pk):
#     try:
#         client = Client.objects.get(id=pk)
#     except Client.DoesNotExist:
#         return Response({"error_message": "Client not found"}, status=404)

#     # Handle GET request: Retrieve the client and pre-populate the frontend form
#     if request.method == 'GET':
#         client_serializer = ClientSerializer(client)
#         return Response(client_serializer.data, status=200)

#     # Handle POST request: Update the client and associated FileInfo and File models
#     elif request.method == 'POST':
#         # Update client fields
#         client_serializer = ClientSerializer(client, data=request.data)

#         if client_serializer.is_valid():
#             print('data', request.POST)
#             client_serializer.save()

#             # Access fileinfos data for updating
#             fileinfos_data = []
#             index = 0

#             # Loop through fileinfos in request.POST
#             while f'fileinfos[{index}].login' in request.POST:
#                 fileinfo_data = {
#                     'login': request.POST.get(f'fileinfos[{index}].login'),
#                     'password': request.POST.get(f'fileinfos[{index}].password'),
#                     'document_type': request.POST.get(f'fileinfos[{index}].document_type'),
#                     'remark': request.POST.get(f'fileinfos[{index}].remark'),
#                 }

#                 # Get files associated with this fileinfo
#                 files = request.FILES.getlist(f'fileinfos[{index}].files')
#                 # fileinfo_id = request.POST.get(f'fileinfos[{index}].id')  # Assuming you have a field to identify the FileInfo
#                 # if fileinfo_id:  # Update existing FileInfo if ID is provided
#                 #     try:
#                 #         fileinfo = FileInfo.objects.get(id=fileinfo_id, client=client)

#                 #         # Update existing FileInfo fields
#                 #         for attr, value in fileinfo_data.items():
#                 #             setattr(fileinfo, attr, value)
#                 #         fileinfo.save()


#                 #     except FileInfo.DoesNotExist:
#                 #         # If FileInfo doesn't exist, create a new one
#                 #         fileinfo = FileInfo.objects.create(client=client, **fileinfo_data)

#                 # else:
#                 #     # Create a new FileInfo if no ID was provided
#                 #     fileinfo = FileInfo.objects.create(client=client, **fileinfo_data)

#                 # Create new files associated with this FileInfo
#                 for file in files:
#                     File.objects.create(files=file)

#                 index += 1

#             # return Response(client_serializer.data, status=200)
#             return Response({'message': 'Client updated successfully','data': client_serializer.data},status=status.HTTP_200_OK)


#         return Response({'message': 'Fail to update client', 'error_message' : client_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

# @api_view(['GET', 'POST'])
# def edit_client(request, pk):
#     try:
#         client = Client.objects.get(id=pk)
#     except Client.DoesNotExist:
#         return Response({"error_message": "Client not found"}, status=404)

#     # Handle GET request: Retrieve the client and pre-populate the frontend form
#     if request.method == 'GET':
#         client_serializer = ClientSerializer(client)
#         return Response(client_serializer.data, status=200)

#     # Handle POST request: Update the client and associated FileInfo and File models
#     elif request.method == 'POST':
#         client_serializer = ClientSerializer(client, data=request.data)

#         if client_serializer.is_valid():
#             client_serializer.save()

#             # ✅ Step 1: Delete all existing FileInfo and File records for this client
#             FileInfo.objects.filter(client=client).delete()

#             index = 0
#             while f'fileinfos[{index}].login' in request.POST:
#                 fileinfo_data = {
#                     'login': request.POST.get(f'fileinfos[{index}].login'),
#                     'password': request.POST.get(f'fileinfos[{index}].password'),
#                     'document_type': request.POST.get(f'fileinfos[{index}].document_type'),
#                     'remark': request.POST.get(f'fileinfos[{index}].remark'),
#                 }

#                 files = request.FILES.getlist(f'fileinfos[{index}].files')

#                 # ✅ Step 2: Create new FileInfo
#                 # fileinfo = FileInfo.objects.filter(client=client, **fileinfo_data).first()
#                 fileinfo = FileInfo.objects.filter(client=client, **fileinfo_data).first()
#                 if not fileinfo:
#                     fileinfo = FileInfo.objects.create(client=client, **fileinfo_data)

                

#                 # ✅ Step 3: Create related File objects
#                 for file in files:
#                     File.objects.create(fileinfo=fileinfo, files=file)

#                 index += 1

#             return Response({
#                 'message': 'Client updated successfully',
#                 'data': client_serializer.data
#             }, status=200)

#         return Response({
#             'message': 'Fail to update client',
#             'error_message': client_serializer.errors
#         }, status=400)

# @api_view(['GET', 'POST'])
# def edit_client(request, pk):
#     try:
#         client = Client.objects.get(id=pk)
#     except Client.DoesNotExist:
#         return Response({"error_message": "Client not found"}, status=404)

#     # Handle GET request: Retrieve the client and pre-populate the frontend form
#     if request.method == 'GET':
#         client_serializer = ClientSerializer(client)
#         return Response(client_serializer.data, status=200)

#     # Handle POST request: Update the client and associated FileInfo and File models
#     elif request.method == 'POST':
#         try:
#             client = Client.objects.get(id=pk)
#         except Client.DoesNotExist:
#             return Response({"error_message": "Client not found"}, status=404)

#         client_serializer = ClientSerializer(client, data=request.data)
#         if client_serializer.is_valid():
#             client_serializer.save()

#             # Store IDs of fileinfos that were updated or newly created
#             received_fileinfo_ids = []

#             index = 0
#             while f'fileinfos[{index}].login' in request.POST:
#                 fileinfo_id = request.POST.get(f'fileinfos[{index}].id')
#                 fileinfo_data = {
#                     'login': request.POST.get(f'fileinfos[{index}].login'),
#                     'password': request.POST.get(f'fileinfos[{index}].password'),
#                     'document_type': request.POST.get(f'fileinfos[{index}].document_type'),
#                     'remark': request.POST.get(f'fileinfos[{index}].remark'),
#                 }
#                 files = request.FILES.getlist(f'fileinfos[{index}].files')

#                 if fileinfo_id:
#                     try:
#                         fileinfo = FileInfo.objects.get(id=int(fileinfo_id), client=client)
#                         print(f"✅ Updating FileInfo ID {fileinfo_id}")
#                         for attr, value in fileinfo_data.items():
#                             setattr(fileinfo, attr, value)
#                         fileinfo.save()

#                         # ✅ Only delete old files if new ones are uploaded
#                         if files:
#                             fileinfo.files.all().delete()

#                     except ObjectDoesNotExist:
#                         print(f"❌ FileInfo ID {fileinfo_id} not found — creating new one.")
#                         fileinfo = FileInfo.objects.create(client=client, **fileinfo_data)

#                 received_fileinfo_ids.append(fileinfo.id)

#                 for file in files:
#                     File.objects.create(fileinfo=fileinfo, files=file)

#                 index += 1

#             # Optional cleanup: delete any fileinfos not included in this update
#             FileInfo.objects.filter(client=client).exclude(id__in=received_fileinfo_ids).delete()

#             return Response({
#                 'message': 'Client updated successfully',
#                 'data': client_serializer.data
#             }, status=200)

#         return Response({
#             'message': 'Fail to update client',
#             'error_message': client_serializer.errors
#         }, status=400)

@api_view(['GET', 'POST'])
def edit_client(request, pk):
    try:
        client = Client.objects.get(id=pk)
    except Client.DoesNotExist:
        return Response({"error_message": "Client not found"}, status=404)

    if request.method == 'GET':
        client_serializer = ClientSerializer(client)
        return Response(client_serializer.data, status=200)

    elif request.method == 'POST':
        client_serializer = ClientSerializer(client, data=request.data)
        if client_serializer.is_valid():
            client_serializer.save()

            received_fileinfo_ids = []
            index = 0

            while f'fileinfos[{index}].login' in request.POST:
                fileinfo_id = request.POST.get(f'fileinfos[{index}].id', '').strip()
                fileinfo_data = {
                    'login': request.POST.get(f'fileinfos[{index}].login'),
                    'password': request.POST.get(f'fileinfos[{index}].password'),
                    'document_type': request.POST.get(f'fileinfos[{index}].document_type'),
                    'remark': request.POST.get(f'fileinfos[{index}].remark'),
                }

                files = request.FILES.getlist(f'fileinfos[{index}].files')
                fileinfo = None

                if fileinfo_id and fileinfo_id.isdigit():
                    try:
                        fileinfo = FileInfo.objects.get(id=int(fileinfo_id), client=client)
                        print(f"✅ Updating FileInfo ID {fileinfo_id}")
                        for attr, value in fileinfo_data.items():
                            setattr(fileinfo, attr, value)
                        fileinfo.save()

                        if files:  # only delete old files if new files are uploaded
                            fileinfo.files.all().delete()

                    except FileInfo.DoesNotExist:
                        print(f"❌ FileInfo ID {fileinfo_id} not found — creating new one.")
                        fileinfo = FileInfo.objects.create(client=client, **fileinfo_data)
                else:
                    print("➕ No ID provided — creating new FileInfo.")
                    fileinfo = FileInfo.objects.create(client=client, **fileinfo_data)

                received_fileinfo_ids.append(fileinfo.id)

                for file in files:
                    File.objects.create(fileinfo=fileinfo, files=file)

                index += 1

            # Optional cleanup: delete fileinfos that were removed from the frontend
            FileInfo.objects.filter(client=client).exclude(id__in=received_fileinfo_ids).delete()

            return Response({
                'message': 'Client updated successfully',
                'data': client_serializer.data
            }, status=200)

        return Response({
            'message': 'Failed to update client',
            'error_message': client_serializer.errors
        }, status=400)
@api_view(['GET'])
def single_fileinfo(request,pk,fileinfo_pk):
    client = Client.objects.get(id = pk)
    fileinfo = FileInfo.objects.get(id=fileinfo_pk, client=client)
    if request.method == 'GET':
        serializer = FileInfoSerializer(fileinfo)
        print(serializer)
        return Response(serializer.data)

@api_view(['DELETE'])
def delete_fileinfo(request,pk,fileinfo_pk):
    if request.method == 'DELETE':
        client = Client.objects.get(id=pk)
        fileinfo = FileInfo.objects.get(id=fileinfo_pk, client=client)
        fileinfo.delete()
        return Response ({'message':'Fileinfo Deleted'})
    return Response ({'message':'Fail to delete'}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
def delete_client(request,pk):
    if request.method == 'DELETE':
        client = Client.objects.get(id=pk)
        client.delete()
        return Response ({'message':'Company Deleted'})
    return Response ({'message':'Fail to delete'}, status=status.HTTP_400_BAD_REQUEST)


# ***********************************************Bank View's******************************************************

@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def create_bank(request,pk):
    client = Client.objects.get(id=pk)
    instance_data = request.data
    data = {key: value for key, value in instance_data.items()}

    account_number = request.data.get('account_number')
    if account_number == '':
        return Response({'error_message': 'Account number is required'}, status=status.HTTP_400_BAD_REQUEST)

    serializer = BankSerializer(data=data)
    # account_number = request.data.get('account_number')
    # if account_number == '':
    #     return Response({'error_message': 'Account number is required'}, status=status.HTTP_400_BAD_REQUEST)
    if serializer.is_valid(raise_exception=True):
        bank_instance = serializer.save(client=client)
        print(request.data)

        if request.FILES:
            files = dict((request.FILES).lists()).get('files',None)
            # files = request.FILES.getlist('files')
            if files:
                for file in files:
                    file_data = {
                        'bank' : bank_instance.pk,
                        'files' : file
                    }
                    file_serializer= FilesSerializer(data=file_data)
                    if file_serializer.is_valid(raise_exception=True):
                        file_serializer.save()

            return Response({'message':'Bank created successfully', 'data' : serializer.data},status=status.HTTP_201_CREATED)
    return Response({
        'message':'Fail to create bank', 
        'error_message' : serializer.errors
        },status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST', 'GET'])
@parser_classes([MultiPartParser, FormParser])
def edit_bank(request, pk, bank_pk):
    try:
        client = Client.objects.get(id=pk)
        bank = Bank.objects.get(id=bank_pk, client=client)
    except Client.DoesNotExist:
        return Response({'error_message': 'Client not found'}, status=status.HTTP_404_NOT_FOUND)
    except Bank.DoesNotExist:
        return Response({'error_message': 'Bank not found'}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'POST':
        # Update branch document fields
        print('Received raw data:', request.data)
        print('Received POST data:', request.POST)
        print('Received FILES data:', request.FILES)
        print('kjjkkj', request.data)
        bank_serializer = BankSerializer(instance=bank, data=request.data)
        if bank_serializer.is_valid():
            print('Updated data before save:', request.data)
            bank_serializer.save(client=client)
            print('Updated data after save:', request.data)

            # Handle file updates separately
            if request.FILES.getlist('files'):
                # Delete the old files if required
                Files.objects.filter(bank=bank).delete()

                # Add the new files
                files = request.FILES.getlist('files')
                for file in files:
                    file_data = {
                        'bank': bank.pk,
                        'files': file
                    }
                    file_serializer = FilesSerializer(data=file_data)
                    if file_serializer.is_valid():
                        file_serializer.save()
                    else:
                        return Response(file_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

            return Response({'message':'Bank updated successfully','data' : bank_serializer.data},status=status.HTTP_200_OK)
        else:
            return Response({'message': 'fail to update bank', 'error_message': bank_serializer.errors},status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'GET':
        serializer = BankSerializer(bank)
        return Response(serializer.data)

@api_view(['GET'])
def list_bank(request, pk):
    client = Client.objects.get(id=pk)
    off=OfficeLocation.objects.filter(branch__client=client)
    print('gh',off)
    # for i in off:
    #     print('fff',i)
    if request.method == 'GET':
        bank_list = Bank.objects.filter(client=client)
        serializers = BankSerializer(bank_list,many=True)
        print(serializers)
        return Response(serializers.data)

@api_view(['GET'])
def single_bank(request, pk, bank_pk):
    client = Client.objects.get(id=pk)
    bank = Bank.objects.get(id=bank_pk, client=client)
    if request.method == 'GET':
        serializer = BankSerializer(bank)
        print(serializer)
        return Response(serializer.data)

@api_view(['DELETE'])
def delete_bank(request,pk, bank_pk):
    client = get_object_or_404(Client, id=pk)
    bank = Bank.objects.get(id=bank_pk)
    if request.method == 'DELETE':
        bank.delete()
        return Response({'message':'Bank is Deleted'})
    return Response({'error_message':'Fail to Delete Bank'},status=status.HTTP_400_BAD_REQUEST)

# **********************************************Owners View's*******************************************

# @api_view(['POST', 'GET'])
# def create_owner(request, pk):
#     client = get_object_or_404(Client, id=pk)
    
#     if request.method == 'POST':
#         owner_serializer = OwnerSerializer(data=request.data)
#         if owner_serializer.is_valid():
#             # Calculate total shares assigned to the client so far
#             total_shares = Owner.objects.filter(client=client).aggregate(
#                 total_share=Coalesce(Sum(F('share')), 0)
#             )['total_share']
            
#             # Calculate remaining shares
#             remaining_shares = 100 - total_shares
            
#             # New share value from the current request
#             new_share = owner_serializer.validated_data['share']

#             if new_share > remaining_shares:
#                 # If entered shares are greater than the remaining shares, return an error
#                 return Response({
#                     'error_message': f'Cannot assign {new_share}%. Only {remaining_shares}% is left for assigning.'
#                 }, status=status.HTTP_400_BAD_REQUEST)

#             # if aadhar > 12:
#             #     return Response({'error_message': 'Aadhar number should be 12 digits'}, status=status.HTTP_400_BAD_REQUEST)
            
#             # Save the new owner
#             owner_serializer.save(client=client)
            
#             # Recalculate remaining shares after saving
#             total_shares += new_share
#             remaining_shares = 100 - total_shares

#             return Response({
#                 'message': f'Owner Created Successfully remaining shares are {remaining_shares}',
#                 'data': owner_serializer.data,
#                 'remaining_shares': remaining_shares,
#                 },status=status.HTTP_201_CREATED)

#         # Show errors if the data is not valid
#         return Response({
#                 'message': f'Fail to create owner {owner_serializer.errors}',
#                 'error_message': owner_serializer.errors,
#                 },status=status.HTTP_400_BAD_REQUEST)

#     elif request.method == 'GET':
#         # Calculate total shares and remaining shares
#         total_shares = Owner.objects.filter(client=client).aggregate(
#             total_share=Coalesce(Sum(F('share')), 0)
#         )['total_share']
#         remaining_shares = 100 - total_shares

#         return Response({
#             'total_shares': total_shares,
#             'remaining_shares': remaining_shares
#         }, status=status.HTTP_200_OK)

# @api_view(['POST', 'GET'])
# def create_owner(request, pk):
    client = get_object_or_404(Client, id=pk)
    
    if request.method == 'POST':

        pan = request.data.get('pan')
        if pan and len(pan) > 10 or len(pan) < 10:
            return Response({'error_message': 'PAN number should be at most 10 characters'}, status=status.HTTP_400_BAD_REQUEST)

        aadhar = request.data.get('aadhar')
        if aadhar and len(aadhar) > 12:
            return Response({'error_message': 'Aadhar number should be at most 12 characters'}, status=status.HTTP_400_BAD_REQUEST)

        mobile = request.data.get('mobile')
        if mobile and len(mobile) > 10 or len(mobile) < 10:
            return Response({'error_message': 'Mobile number should be at most 10 characters'}, status=status.HTTP_400_BAD_REQUEST)
        
        share = request.data.get('share')
        if share == '':
            return Response({'error_message': f'Share field is required'}, status=status.HTTP_400_BAD_REQUEST)

        user = request.data.get('username')
        if user == '':
            return Response({'error_message' : f'Username field is required'}, status=status.HTTP_400_BAD_REQUEST)

        owner_serializer = OwnerSerializer(data=request.data)
        if owner_serializer.is_valid():
            # Calculate total shares assigned to the client so far
            total_shares = Owner.objects.filter(client=client).aggregate(
                total_share=Coalesce(Sum(F('share')), 0)
            )['total_share']
            
            # Calculate remaining shares
            remaining_shares = 100 - total_shares
            
            # New share value from the current request
            new_share = owner_serializer.validated_data['share']

            if new_share > remaining_shares:
                # If entered shares are greater than the remaining shares, return an error
                return Response({
                    'error_message': f'Cannot assign {new_share}%. Only {remaining_shares}% is left for assigning.'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            if remaining_shares == 0:
                return Response({'error_message': f'Cannot Create Owner Cuz their is 0% shares are left'}, status=status.HTTP_400_BAD_REQUEST)

            # Save the new owner
            owner_serializer.save(client=client)
            
            # Recalculate remaining shares after saving
            total_shares += new_share
            remaining_shares = 100 - total_shares

            return Response({
                'message': f'Owner Created Successfully, remaining shares are {remaining_shares}%',
                'data': owner_serializer.data,
                'remaining_shares': remaining_shares,
            }, status=status.HTTP_201_CREATED)

        return Response({
            'message': 'Failed to create owner',
            'error_message': owner_serializer.errors,
        }, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'GET':
        # Calculate total shares and remaining shares
        total_shares = Owner.objects.filter(client=client).aggregate(
            total_share=Coalesce(Sum(F('share')), 0)
        )['total_share']
        remaining_shares = 100 - total_shares

        return Response({
            'total_shares': total_shares,
            'remaining_shares': remaining_shares
        }, status=status.HTTP_200_OK)

# @api_view(['POST', 'GET'])
# def create_owner(request, pk):
#     client = get_object_or_404(Client, id=pk)

#     if request.method == 'POST':
#         data = request.data
#         pan = request.data.get('pan')
#         if pan and len(pan) != 10:
#             return Response({'error_message': 'PAN number should be exactly 10 characters'}, status=status.HTTP_400_BAD_REQUEST)

#         aadhar = request.data.get('aadhar')
#         if aadhar and len(aadhar) > 12:
#             return Response({'error_message': 'Aadhar number should be at most 12 characters'}, status=status.HTTP_400_BAD_REQUEST)

#         mobile = request.data.get('mobile')
#         if mobile and len(mobile) != 10:
#             return Response({'error_message': 'Mobile number should be exactly 10 characters'}, status=status.HTTP_400_BAD_REQUEST)

#         share = request.data.get('share')
#         if share == '':
#             return Response({'error_message': 'Share field is required'}, status=status.HTTP_400_BAD_REQUEST)

#         user = request.data.get('username')
#         if user == '':
#             return Response({'error_message': 'Username field is required'}, status=status.HTTP_400_BAD_REQUEST)

#         owner_serializer = OwnerSerializer(data=request.data)

#         if not owner_serializer.is_valid():
#             return Response({
#                 'message': 'Failed to create owner',
#                 'error_message': owner_serializer.errors,
#             }, status=status.HTTP_400_BAD_REQUEST)

#         # Begin atomic transaction to prevent partial saves
#         with transaction.atomic():
#             total_shares = Owner.objects.filter(client=client).aggregate(
#                 total_share=Coalesce(Sum(F('share')), 0)
#             )['total_share']

#             remaining_shares = 100 - total_shares
#             new_share = owner_serializer.validated_data['share']

#             if remaining_shares == 0:
#                 return Response({'error_message': 'Cannot create owner because 0% shares are left'}, status=status.HTTP_400_BAD_REQUEST)

#             if new_share > remaining_shares:
#                 return Response({
#                     'error_message': f'Cannot assign {new_share}%. Only {remaining_shares}% is left for assigning.'
#                 }, status=status.HTTP_400_BAD_REQUEST)

#             # 💡 Generate password from owner name
#             # owner_name = owner_serializer.validated_data.get('owner_name', '')
#             # password = f"{owner_name[:4].lower()}@123"
#             # hashed_password = make_password(password)

#             # # 🔍 Print password in console
#             # print("Generated Password:", password)

#             owner_serializer.save(client=client)

#             # email = validated_data.get('email')
#             email = request.data.get('email')
#             # if not email:
#             #     return Response({'error_message': 'Email is required'}, status=status.HTTP_400_BAD_REQUEST)

#             if CustomUser.objects.filter(email=email).exists():
#                 return Response({'error_message': 'A user with this email already exists.'}, status=status.HTTP_400_BAD_REQUEST)

#             # name = request.data.get('owner_name')  
#             # password = f"{name[:4].lower()}@123"


#             # user = CustomUser.objects.create(
#             #     username=email,  # Username is same as email
#             #     email=email,
#             #     name=name,
#             #     password=make_password(password) if password else None,  # Secure the password
#             #     cus_user=True,
#             #     is_active=False,  # You can adjust this if needed
#             #     client=client,
#             # )
#             # email_subject = "Activate You Account"
#             # message = render_to_string(
#             #     "activate.html",
#             #     {
#             #         'user': user,
#             #         'domain': '127.0.0.1:8000',
#             #         'uid' : urlsafe_base64_encode(force_bytes(user.pk)),
#             #         'token' : generate_token.make_token(user),
#             #     }
#             # )
#             # # print(message)
#             # email_message = EmailMessage(email_subject,message,settings.EMAIL_HOST_USER,[data['email']])
#             # email_message.send()
#             # serializer = UserSerializerWithToken(user, many=False)

#             # Update remaining shares
#             remaining_shares -= new_share

#             return Response({
#                 'message': f'Owner Created Successfully, User Registered kindly activate ur account, remaining shares are {remaining_shares}%',
#                 'data': owner_serializer.data,
#                 # 'generated_password': password,
#                 'remaining_shares': remaining_shares,
#             }, status=status.HTTP_201_CREATED)

#     elif request.method == 'GET':
#         total_shares = Owner.objects.filter(client=client).aggregate(
#             total_share=Coalesce(Sum(F('share')), 0)
#         )['total_share']
#         remaining_shares = 100 - total_shares

#         return Response({
#             'total_shares': total_shares,
#             'remaining_shares': remaining_shares
#         }, status=status.HTTP_200_OK)


from django.db import transaction

# @api_view(['POST', 'GET'])
# def create_owner(request, pk):
    # client = get_object_or_404(Client, id=pk)

    # if request.method == 'POST':
    #     data = request.data
    #     pan = request.data.get('pan')
    #     if pan and len(pan) != 10:
    #         return Response({'error_message': 'PAN number should be exactly 10 characters'}, status=status.HTTP_400_BAD_REQUEST)

    #     aadhar = request.data.get('aadhar')
    #     if aadhar and len(aadhar) > 12:
    #         return Response({'error_message': 'Aadhar number should be at most 12 characters'}, status=status.HTTP_400_BAD_REQUEST)

    #     mobile = request.data.get('mobile')
    #     if mobile and len(mobile) != 10:
    #         return Response({'error_message': 'Mobile number should be exactly 10 characters'}, status=status.HTTP_400_BAD_REQUEST)

    #     share = request.data.get('share')
    #     if share == '':
    #         return Response({'error_message': 'Share field is required'}, status=status.HTTP_400_BAD_REQUEST)

    #     user = request.data.get('username')
    #     if user == '':
    #         return Response({'error_message': 'Username field is required'}, status=status.HTTP_400_BAD_REQUEST)

    #     owner_serializer = OwnerSerializer(data=request.data)

    #     if not owner_serializer.is_valid():
    #         return Response({
    #             # 'message': 'Failed to create owner',
    #             'error_message': owner_serializer.errors,
    #         }, status=status.HTTP_400_BAD_REQUEST)

    #     # Begin atomic transaction to prevent partial saves
    #     with transaction.atomic():
    #         total_shares = Owner.objects.filter(client=client).aggregate(
    #             total_share=Coalesce(Sum(F('share')), 0)
    #         )['total_share']

    #         remaining_shares = 100 - total_shares
    #         new_share = owner_serializer.validated_data['share']

    #         if remaining_shares == 0:
    #             return Response({'error_message': 'Cannot create owner because 0% shares are left'}, status=status.HTTP_400_BAD_REQUEST)

    #         if new_share > remaining_shares:
    #             return Response({
    #                 'error_message': f'Cannot assign {new_share}%. Only {remaining_shares}% is left for assigning.'
    #             }, status=status.HTTP_400_BAD_REQUEST)

    #         owner_serializer.save(client=client)

    #         # email = request.data.get('email')
    #         # if CustomUser.objects.filter(email=email).exists():
    #         #     return Response({'error_message': 'A user with this email already exists.'}, status=status.HTTP_400_BAD_REQUEST)
    #         remaining_shares -= new_share

    #         return Response({
    #             'message': f'Owner Created Successfully, User Registered kindly activate ur account, remaining shares are {remaining_shares}%',
    #             'data': owner_serializer.data,
    #             # 'generated_password': password,
    #             'remaining_shares': remaining_shares,
    #         }, status=status.HTTP_201_CREATED)
    #     # return Response(owner_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    # elif request.method == 'GET':
    #     total_shares = Owner.objects.filter(client=client).aggregate(
    #         total_share=Coalesce(Sum(F('share')), 0)
    #     )['total_share']
    #     remaining_shares = 100 - total_shares

    #     return Response({
    #         'total_shares': total_shares,
    #         'remaining_shares': remaining_shares
    #     }, status=status.HTTP_200_OK)

@api_view(['POST', 'GET'])
def create_owner(request, pk):
    client = get_object_or_404(Client, id=pk)

    if request.method == 'POST':
        data = request.data
        pan = request.data.get('pan')
        if pan and len(pan) != 10:
            return Response({'error_message': 'PAN number should be exactly 10 characters'}, status=status.HTTP_400_BAD_REQUEST)

        aadhar = request.data.get('aadhar')
        if aadhar and len(aadhar) > 12:
            return Response({'error_message': 'Aadhar number should be at most 12 characters'}, status=status.HTTP_400_BAD_REQUEST)

        mobile = request.data.get('mobile')
        if mobile and len(mobile) != 10:
            return Response({'error_message': 'Mobile number should be exactly 10 characters'}, status=status.HTTP_400_BAD_REQUEST)

        share = request.data.get('share')
        if share == '':
            return Response({'error_message': 'Share field is required'}, status=status.HTTP_400_BAD_REQUEST)

        user = request.data.get('username')
        if user == '':
            return Response({'error_message': 'Username field is required'}, status=status.HTTP_400_BAD_REQUEST)

        email = request.data.get('email')
        owner_serializer = OwnerSerializer(data=request.data)

        if not owner_serializer.is_valid():
            return Response({
                # 'message': 'Failed to create owner',
                'error_message': owner_serializer.errors,
            }, status=status.HTTP_400_BAD_REQUEST)

        # Begin atomic transaction to prevent partial saves
        with transaction.atomic():
            total_shares = Owner.objects.filter(client=client).aggregate(
                total_share=Coalesce(Sum(F('share')), 0)
            )['total_share']

            remaining_shares = 100 - total_shares
            new_share = owner_serializer.validated_data['share']

            if remaining_shares == 0:
                return Response({'error_message': 'Cannot create owner because 0% shares are left'}, status=status.HTTP_400_BAD_REQUEST)

            if new_share > remaining_shares:
                return Response({
                    'error_message': f'Cannot assign {new_share}%. Only {remaining_shares}% is left for assigning.'
                }, status=status.HTTP_400_BAD_REQUEST)

            # owner_serializer.save(client=client)
            user_password = owner_serializer.validated_data['user_password']
            owner = owner_serializer.save(client=client)
            clientuser = CommonUser.objects.create_user(
                username=owner.email,
                email=owner.email,
                name=owner.owner_name,
                password=user_password,  # from request
                role='clientuser',
                client=owner.client,
                is_active=True
            )
            email_subject = "Your Account is Registered"
            message = render_to_string("activate.html", {
                'user': clientuser,
                # 'domain': '127.0.0.1:8000',
                'password': user_password,
                'domain': 'admin.dms.zacoinfotech.com',
                'uid': urlsafe_base64_encode(force_bytes(clientuser.pk)),
                'token': generate_token.make_token(clientuser),
            })

            email_message = EmailMessage(email_subject, message, settings.EMAIL_HOST_USER, [email])
            email_message.send()

            # email = request.data.get('email')
            # if CustomUser.objects.filter(email=email).exists():
            #     return Response({'error_message': 'A user with this email already exists.'}, status=status.HTTP_400_BAD_REQUEST)
            remaining_shares -= new_share



            

            return Response({
                'message': f'Owner Created Successfully, User Registered kindly activate ur account, remaining shares are {remaining_shares}%',
                'data': owner_serializer.data,
                # 'generated_password': password,
                'remaining_shares': remaining_shares,
            }, status=status.HTTP_201_CREATED)
        # return Response(owner_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'GET':
        total_shares = Owner.objects.filter(client=client).aggregate(
            total_share=Coalesce(Sum(F('share')), 0)
        )['total_share']
        remaining_shares = 100 - total_shares

        return Response({
            'total_shares': total_shares,
            'remaining_shares': remaining_shares
        }, status=status.HTTP_200_OK)

@api_view(['POST','GET'])
def edit_owner(request, pk, owner_pk):
    client = get_object_or_404(Client, id= pk)
    owner = Owner.objects.get(id = owner_pk)
    if request.method == 'POST':
        
        owner_serializer = OwnerSerializer(data=request.data, instance = owner, partial=True)

        pan = request.data.get('pan')
        if pan and len(pan) > 10 or len(pan) < 10:
            return Response({'error_message': 'PAN number should be at most 10 characters'}, status=status.HTTP_400_BAD_REQUEST)

        aadhar = request.data.get('aadhar')
        if aadhar and (len(str(aadhar)) > 12 or len(str(aadhar)) < 12):
            return Response({'error_message': 'Aadhar number should be at most 12 characters'}, status=status.HTTP_400_BAD_REQUEST)

        mobile = request.data.get('mobile')
        if mobile and (len(str(mobile)) > 10 or len(str(mobile)) < 10):
            return Response({'error_message': 'Mobile number should be at most 10 characters'}, status=status.HTTP_400_BAD_REQUEST)
        
        share = request.data.get('share')
        if share == '':
            return Response({'error_message': f'Share field is required'}, status=status.HTTP_400_BAD_REQUEST)

        user = request.data.get('username')
        if user == '':
            return Response({'error_message' : f'Username field is required'}, status=status.HTTP_400_BAD_REQUEST)
    
        if owner_serializer.is_valid():
            # Calculate total shares assigned to the client so far
            total_shares = Owner.objects.filter(client=client).aggregate(
                total_share=Coalesce(Sum(F('share')), 0)
            )['total_share']
            
            # Calculate remaining shares
            remaining_shares = 100 - total_shares

            input_share = owner.share
            
            a = input_share + remaining_shares
            # New share value from the current request
            new_share = owner_serializer.validated_data['share']
            # new_share = 0

            if new_share > a:
                # If entered shares are greater than the remaining shares, return an error
                return Response({
                    'error_message': f'Cannot assign {new_share}%. Only {a}% is left for assigning.'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Save the new owner
            owner_serializer.save(client=client)

            remaining_shares = a - new_share
            return Response({'message':f'Owner Updated Successfully remaining shares are {remaining_shares}', 'remaining_shares': remaining_shares, 'status' : status.HTTP_200_OK})
            # return Response({'message':'Owner Updated Successfully', 'Status': status.HTTP_200_OK, 'remaining shares' : c})
        else:
            return Response({
                    'message': 'Fail to update owner',
                    'error_message': owner_serializer.errors,
                    },status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'GET':
        owner_serializer1 =OwnerSerializer(owner)
        return Response(owner_serializer1.data)

@api_view(['GET'])
def list_owner(request, pk):
    client = Client.objects.get(id = pk)
    owner_list = Owner.objects.filter(client=client)
    serializers = OwnerSerializer(owner_list, many=True)
    print(serializers)
    return Response(serializers.data)

@api_view(['GET'])
def single_owner(request, pk, owner_pk):
    client = Client.objects.get(id = pk)
    owner = Owner.objects.get(id = owner_pk)
    if request.method == 'GET':
        serializer = OwnerSerializer(owner)
        print(serializer)
        return Response(serializer.data)

# @api_view(['DELETE'])
# def delete_owner(request, pk, owner_pk):
#     client = get_object_or_404(Client, id=pk)
#     owner = Owner.objects.get(id = owner_pk)
#     try :
#         # storing the value of current owner shares in a variable
#         owner_share = owner.share
#         owner.delete()
#         # for loop for providing the remainig shares left
#         total_shares = sum([owner.share for owner in Owner.objects.all()])
#         remaining_shares = 100 - total_shares
#         return Response({'message': f'Owner is deleted.{owner_share}% share is added back. Avaliable shares: {remaining_shares}%'}, status=status.HTTP_200_OK)
#     except :
#         return Response({'error_message':'Owner not found'},status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
def delete_owner(request, pk, owner_pk):
    client = get_object_or_404(Client, id=pk)
    try:
        owner = Owner.objects.get(id=owner_pk, client=client)
        
        if not owner.is_active:
            return Response({'error_message': 'Owner already inactive.'}, status=status.HTTP_400_BAD_REQUEST)

        # Save current share and deactivate owner
        owner_share = owner.share
        owner.share = 0
        owner.is_active = False
        owner.save()

        # Calculate remaining shares
        total_shares = sum([owner.share for owner in Owner.objects.filter(is_active=True)])
        remaining_shares = 100 - total_shares

        return Response(
            {
                'message': f'Owner is disabled. {owner_share}% share is added back. Available shares: {remaining_shares}%',
            },
            status=status.HTTP_200_OK
        )
    except Owner.DoesNotExist:
        return Response({'error_message': 'Owner not found'}, status=status.HTTP_400_BAD_REQUEST)

# ******************************************User's Views*******************************************

# Client Login
# class MyTokenObtainPairSerializer(TokenObtainPairSerializer):
#     def validate(self, attrs):
#         data = super().validate(attrs)
#         serializer = UserSerializerWithToken(self.user).data
#         for k,v in serializer.items():
#              data[k]=v
#         return data

# class MyTokenObtainPairView(TokenObtainPairView):
#     serializer_class = MyTokenObtainPairSerializer

# User Profile
# @api_view(['GET'])
# @permission_classes([IsAuthenticated]) # the user should be valid
# def getUserProfile(request):
#     user = request.user # to get the specific user
#     serializer = UserSerializerWithToken(user, many=False)
#     return Response(serializer.data)

# Users List
# @api_view(['GET'])
# @permission_classes([IsAdminUser]) # the user should be an admin only
# def getUsers(request):
#     user = ClientUser.objects.all() # to get the list of all users
#     serializer = UserSerializerWithToken(user, many=True)
#     return Response(serializer.data)

# @api_view(['GET'])
# def single_clientuser(request, pk, user_pk): 
#     client = Client.objects.get(id = pk)
#     owner = ClientUser.objects.get(id = user_pk, client=client)
#     if request.method == 'GET':
#         serializer = UserSerializerWithToken(owner)
#         print(serializer)
#         return Response(serializer.data)

# @api_view(['POST'])
# def reset_clientuser_password(request, pk, user_pk):
#     client = Client.objects.get(id=pk)
#     user = ClientUser.objects.get(id=user_pk, client=client)
#     if request.method == 'POST':
#         data = request.data
#         passwords = request.data.get('password')
#         print("Previous (hashed) password:", passwords)
#         previous_password = request.data.get('previous_password')
#         new_password = request.data.get('new_password')
#         confirm_password = request.data.get('confirm_password')
#         if not user.check_password(previous_password):
#             return Response({'error_message': 'Incorrect previous password !!'}, status=status.HTTP_400_BAD_REQUEST)
#         if new_password != confirm_password:
#             return Response({'error_message': 'New password and confirm password do not match !!'}, status=status.HTTP_400_BAD_REQUEST)
#         user.set_password(new_password)  # Set the new password
#         user.save()
#         user_serializer = UserSerializerWithToken(data=data, instance=user, partial=True)
#         if user_serializer.is_valid():
#             user_serializer.save(client=client)
#             return Response({'message':'Client User Password Updated'})
#         return Response(user_serializer.errors,status=status.HTTP_400_BAD_REQUEST)

# @api_view(['POST'])
# def clientuser(request, pk):
#     client = get_object_or_404(Client, id=pk)
#     data = request.data
#     try:
#         if ClientUser.objects.filter(email=data['email']).exists():
#             return Response({'error_message': 'User Already Exists'}, status=status.HTTP_400_BAD_REQUEST)


#         name = data['name']
#         email = data['email']

#         # Generate password: first 4 letters of name (or less) + @123
#         name_part = name[:4].lower()
#         generated_password = f"{name_part}@123"

#         # Print password for testing
#         print("Generated password:", generated_password)
        
#         user = ClientUser.objects.create(
#             # first_name=data['first_name'],
#             # last_name=data['last_name'],
#             name = data['name'],
#             username=data['email'],
#             email=data['email'],
#             # password=make_password(data['password']),
#             password=make_password(generated_password),
#             is_active=True,
#             client=client
#         )

#         # Generate token for email sending
#         email_subject = "Activate Your Account"
#         message = render_to_string(
#             "activate.html",
#             {
#                 'user': user,
#                 'domain': '127.0.0.1:8000',
#                 'uid': urlsafe_base64_encode(force_bytes(user.pk)),
#                 'token': generate_token.make_token(user),
#             }
#         )

#         email_message = EmailMessage(email_subject, message, settings.EMAIL_HOST_USER, [data['email']])
#         email_message.send()

#         serializer = UserSerializerWithToken(user, many=False)
#         return Response({'Message': 'User Registered. Kindly activate your account.', 'Data': serializer.data})

#     except Exception as e:
#         return Response({'error_message': str(e)}, status=status.HTTP_400_BAD_REQUEST)

# # Email Activations
# class ActivateAccountView(View):
#     def get(self, request, uidb64, token):
#         try:
#             uid= force_text(urlsafe_base64_decode(uidb64))
#             user = CommonUser.objects.get(pk=uid)
#         except Exception as identifier:
#             user=None
#         if user is not None and generate_token.check_token(user,token):
#             user.is_active=True
#             user.save()
#             return render(request,"activatesuccess.html")
#         else:
#             return render(request,"activatefail.html")

# # Clientuser Update
# @api_view(['POST', 'GET'])
# def edit_clientuser(request, pk, user_pk):
    client = Client.objects.get(id=pk)
    user = ClientUser.objects.get(id = user_pk, client=client)
    # user_serializer = UserSerializerWithToken(data=request.data, instance=user, partial=True)
    if request.method == 'POST':
        print("Previous (hashed) password:", user.password)
        data = request.data.copy()
        if 'password' in data and data['password']:
            print("New password entered by user:", data['password'])  
            data['password'] = make_password(data['password'])      

        user_serializer = UserSerializerWithToken(data=data, instance=user, partial=True)
        if user_serializer.is_valid():
            user_serializer.save(client=client)
            return Response({'Message':'Client User Updated'})
        return Response(user_serializer.errors,status=status.HTTP_400_BAD_REQUEST)
    elif request.method == 'GET':
        user_ser = UserSerializerWithToken(user)
        return Response(user_ser.data)

# @api_view(['GET'])
# def list_users_by_client(request, pk):
#     client = get_object_or_404(Client, id=pk)
#     users = ClientUser.objects.filter(client=client)
#     serializer = UserSerializerWithToken(users, many=True)
#     return Response(serializer.data)

# ClientUser Delete
# @api_view(['DELETE'])
# def delete_clientuser(request,pk,user_pk):
#     client = Client.objects.get(id=pk)
#     user = ClientUser.objects.get(id = user_pk)
#     if request.method == 'DELETE':
#         user.delete()
#         return Response ({'message':'Client User is deleted'})
#     return Response ({'error_message':'Failed to delete Client User'},status=status.HTTP_400_BAD_REQUEST)


# ******************************************Dashboard User **************************************
# @api_view(['POST'])
# def dashboarduser(request):
#     data = request.data
#     email = data.get('email')
#     username = data.get('email')  # you're setting username as email
#     confirm_password = request.data.get('confirm_password')
#     password = request.data.get('password')

#     # ✅ Check if email or username already exists
#     if DashboardUser.objects.filter(email=email).exists():
#         return Response({'error_message': 'Email is already registered'}, status=status.HTTP_400_BAD_REQUEST)

#     if DashboardUser.objects.filter(username=username).exists():
#         return Response({'error_message': 'Username already exists'}, status=status.HTTP_400_BAD_REQUEST)
    
#     if password != data['confirm_password']:
#         return Response({'error_message': 'Password and ConfirmPassword do not match'}, status=status.HTTP_400_BAD_REQUEST)

#     try:
#         # ✅ Create user
#         dashboarduser = DashboardUser.objects.create(
#             first_name=data['first_name'],
#             last_name=data['last_name'],
#             username=username,
#             email=email,
#             password=make_password(data['password']),
#             # confirm_password=make_password(data['confirm_password']),
#             # superadmin_user=True,
#             is_active=False,
#             is_staff=True,         # ✅ allows login to admin
#             # is_superuser=True,
#         )

#         # ✅ Email activation token
#         email_subject = "Activate Your Account"
#         message = render_to_string("dashboarduseractivate.html", {
#             'user': dashboarduser,
#             'domain': '127.0.0.1:8000',
#             'uid': urlsafe_base64_encode(force_bytes(dashboarduser.pk)),
#             'token': generate_token.make_token(dashboarduser),
#         })

#         email_message = EmailMessage(email_subject, message, settings.EMAIL_HOST_USER, [email])
#         email_message.send()

#         serializer = SuperuserSerializerWithToken(dashboarduser, many=False)
#         return Response({'message': 'User Registered. Kindly activate your account.', 'data': serializer.data})
    
#     except Exception as e:
#         return Response({'error_message': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# @api_view(['POST', 'GET'])
# def edit_dashboardUser(request, user_pk):
#     # client = Client.objects.get(id=pk)
#     user = DashboardUser.objects.get(id=user_pk)
#     user_serializer = DashboardSerializerWithToken(data=request.data, instance=user)
#     if request.method == 'POST':
#         if user_serializer.is_valid():
#             user_serializer.save()
#             return Response({'message':'Dashboard User Updated','Data': user_serializer.data}, status=status.HTTP_200_OK)
#         return Response(user_serializer.errors,status=status.HTTP_400_BAD_REQUEST)
#     elif request.method == 'GET':
#         user_ser = DashboardSerializerWithToken(user)
#         return Response(user_ser.data)

# @api_view(['DELETE'])
# def delete_dashboarduser(request, user_pk):
#     user = DashboardUser.objects.get(id = user_pk)
#     if request.method == 'DELETE':
#         user.delete()
#         return Response({'message':'Dashboard User deleted'}, status=status.HTTP_200_OK)
#     return Response ({'error_message':'Failed to delete dashboard user'},status=status.HTTP_400_BAD_REQUEST)

# @api_view(['GET'])
# def get_dashboard_users(request):
#     users = DashboardUser.objects.all()
#     serializer = DashboardUserSerializer(users, many=True)
#     return Response(serializer.data)

# class DashboardUserActivateAccountView(APIView):
#     def get(self, request, uidb64, token):
#         print("✅ DashboardUserActivateAccountView called")  # Debug point
#         try:
#             uid = force_str(urlsafe_base64_decode(uidb64))
#             user = DashboardUser.objects.get(pk=uid)
#         except (TypeError, ValueError, OverflowError, DashboardUser.DoesNotExist):
#             user = None

#         if user is not None and generate_token.check_token(user, token):
#             user.is_active = True
#             user.save()
#             return render(request,"activatesuccess.html")
#         else:
#             return render(request,"activatefail.html")

# class MyTokenObtainPairSerializer(TokenObtainPairSerializer):
#     def validate(self, attrs):
#         try:
#             data = super().validate(attrs)
#             serializer = DashboardSerializerWithToken(self.user).data
#             for k, v in serializer.items(): 
#                 data[k] = v
#             return data
#         except Exception:
#             raise serializers.ValidationError({"error_message": "Invalid Credentials"})

# class MyTokenObtainPairViews(TokenObtainPairView):
#     # print("Received login data:", request.data)
#     serializer_class = MyTokenObtainPairSerializer

# class CommonLoginAPIView(APIView):
#     def post(self, request):
#         serializer = CommonTokenObtainPairSerializer(data=request.data)
#         if serializer.is_valid():
#             return Response(serializer.validated_data, status=status.HTTP_200_OK)
#         return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# @api_view(['GET'])
# @permission_classes([IsAuthenticated])
# def list_dashboard_users(request):
#     if not getattr(request.user, 'superadmin_user', False):
#         return Response({'detail': 'You are not authorized to view this.'}, status=403)

#     users = DashboardUser.objects.all()
#     serializer = DashboardUserSerializer(users, many=True)
#     return Response(serializer.data)




# ******************************************Commom Super User **************************************
@api_view(['POST'])
def create_common_superuser(request):
    data = request.data
    email = data.get('email')
    username = data.get('email')  # you're setting username as email
    confirm_password = request.data.get('confirm_password')
    password = request.data.get('password')

    # ✅ Check if email or username already exists
    if CommonUser.objects.filter(email=email).exists():
        return Response({'error_message': 'Email is already registered'}, status=status.HTTP_400_BAD_REQUEST)

    if CommonUser.objects.filter(username=username).exists():
        return Response({'error_message': 'Username already exists'}, status=status.HTTP_400_BAD_REQUEST)
    
    if password != data['confirm_password']:
        return Response({'error_message': 'Password and ConfirmPassword do not match'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        # ✅ Create user
        superuser = CommonUser.objects.create(
            # first_name=data['first_name'],
            # last_name=data['last_name'],
            name=data['name'],
            username=username,
            email=email,
            password=make_password(data['password']),
            role='superuser',
            # confirm_password=make_password(data['confirm_password']),
            # superadmin_user=True,
            is_active=False,
            is_staff=True,     
            is_superuser=True,
        )

        # ✅ Email activation token
        email_subject = "Activate Your Account"
        message = render_to_string("superuseractivate.html", {
            'user': superuser,
            # 'domain': '127.0.0.1:8000',
            'domain': 'admin.dms.zacoinfotech.com',
            'uid': urlsafe_base64_encode(force_bytes(superuser.pk)),
            'token': generate_token.make_token(superuser),
        })

        email_message = EmailMessage(email_subject, message, settings.EMAIL_HOST_USER, [email])
        email_message.send()

        serializer = SuperuserSerializerWithToken(superuser, many=False)
        return Response({'message': 'User Registered. Kindly activate your account.', 'data': serializer.data})
    
    except Exception as e:
        return Response({'error_message': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST', 'GET'])
def edit_common_superuser(request, user_pk):
    # client = Client.objects.get(id=pk)
    user = CommonUser.objects.get(id=user_pk)
    user_serializer = SuperuserSerializerWithToken(data=request.data, instance=user)
    if request.method == 'POST':
        if user_serializer.is_valid():
            user_serializer.save()
            return Response({'message':'Superuser User Updated','Data': user_serializer.data}, status=status.HTTP_200_OK)
        return Response(user_serializer.errors,status=status.HTTP_400_BAD_REQUEST)
    elif request.method == 'GET':
        user_ser = SuperuserSerializerWithToken(user)
        return Response(user_ser.data)

@api_view(['DELETE'])
def delete_common_superuser(request, user_pk):
    user = CommonUser.objects.get(id = user_pk)
    if request.method == 'DELETE':
        user.delete()
        return Response({'message':'Superuser User deleted'}, status=status.HTTP_200_OK)
    return Response ({'error_message':'Failed to delete superuser user'},status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
def get__superusers(request):
    users = CommonUser.objects.all()
    serializer = SuperuserSerializer(users, many=True)
    return Response(serializer.data)

class SuperuserActivateAccountView(APIView):
    def get(self, request, uidb64, token):
        print("✅ SuperuserActivateAccountView called")  # Debug point
        try:
            uid = force_str(urlsafe_base64_decode(uidb64))
            user = CommonUser.objects.get(pk=uid)
        except (TypeError, ValueError, OverflowError, CommonUser.DoesNotExist):
            user = None

        if user is not None and generate_token.check_token(user, token):
            user.is_active = True
            user.save()
            return render(request,"activatesuccess.html")
        else:
            return render(request,"activatefail.html")

class MyTokenObtainPairSerializer(TokenObtainPairSerializer):
    def validate(self, attrs):
        try:
            data = super().validate(attrs)
            serializer = SuperuserSerializerWithToken(self.user).data
            for k, v in serializer.items(): 
                data[k] = v
            return data
        except Exception:
            raise serializers.ValidationError({"error_message": "Invalid Credentials"})

class MyTokenObtainPairViews(TokenObtainPairView):
    # print("Received login data:", request.data)
    serializer_class = MyTokenObtainPairSerializer

class CommonLoginAPIView(APIView):
    def post(self, request):
        serializer = CommonTokenObtainPairSerializer(data=request.data)
        if serializer.is_valid():
            return Response(serializer.validated_data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def list_common_superuser(request):
    if not getattr(request.user, 'superadmin_user', False):
        return Response({'detail': 'You are not authorized to view this.'}, status=403)

    users = CommonUser.objects.all()
    serializer = SuperuserSerializer(users, many=True)
    return Response(serializer.data)

# ******************************************Commom Client User*************************************

@api_view(['POST'])
def create_common_clientuser(request, pk):
    data = request.data
    client = get_object_or_404(Client, id=pk)
    email = data.get('email')
    username = data.get('email')  # you're setting username as email
    # confirm_password = request.data.get('confirm_password')
    password = request.data.get('password')
    name = data.get('name')

    # ✅ Check if email or username already exists
    if CommonUser.objects.filter(email=email).exists():
        return Response({'error_message': 'Email is already registered'}, status=status.HTTP_400_BAD_REQUEST)

    if CommonUser.objects.filter(username=username).exists():
        return Response({'error_message': 'Username already exists'}, status=status.HTTP_400_BAD_REQUEST)

    # name_part = name[:4].lower()
    # generated_password = f"{name_part}@123"
    
    # if password != data['confirm_password']:
    #     return Response({'error_message': 'Password and ConfirmPassword do not match'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        # ✅ Create user
        clientuser = CommonUser.objects.create(
            # first_name=data['first_name'],
            # last_name=data['last_name'],
            name=data['name'],
            username=username,
            email=email,
            # password=make_password(generated_password),
            password=make_password(data['password']),
            role='clientuser',
            # confirm_password=make_password(data['confirm_password']),
            # superadmin_user=True,
            is_active=True,
            is_staff=False, 
            client=client    
            # is_superuser=True,
        )

        # ✅ Email activation token
        email_subject = "Your Account is Registered"
        message = render_to_string("activate.html", {
            'user': clientuser,
            # 'domain': '127.0.0.1:8000',
            'password': password,
            'domain': 'admin.dms.zacoinfotech.com',
            'uid': urlsafe_base64_encode(force_bytes(clientuser.pk)),
            'token': generate_token.make_token(clientuser),
        })

        email_message = EmailMessage(email_subject, message, settings.EMAIL_HOST_USER, [email])
        email_message.send()

        serializer = ClientuserSerializerWithToken(clientuser, many=False)
        return Response({'message': 'User Registered. Kindly activate your account.', 'data': serializer.data})
    
    except Exception as e:
        return Response({'error_message': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST', 'GET'])
def edit_common_clientuser(request,pk, user_pk):
    client = Client.objects.get(id=pk)
    user = CommonUser.objects.get(id=user_pk, client=client)
    user_serializer = ClientuserSerializerWithToken(data=request.data, instance=user, partial=True )
    if request.method == 'POST':
        if user_serializer.is_valid():
            user_serializer.save(client=client)
            return Response({'message':'Superuser User Updated','Data': user_serializer.data}, status=status.HTTP_200_OK)
        return Response(user_serializer.errors,status=status.HTTP_400_BAD_REQUEST)
    elif request.method == 'GET':
        user_ser = ClientuserSerializerWithToken(user)
        return Response(user_ser.data)

# Email Activations
class ClientuserActivateAccountView(View):
    def get(self, request, uidb64, token):
        try:
            uid= force_text(urlsafe_base64_decode(uidb64))
            user = CommonUser.objects.get(pk=uid)
        except Exception as identifier:
            user=None
        if user is not None and generate_token.check_token(user,token):
            user.is_active=True
            user.save()
            return render(request,"activatesuccess.html")
        else:
            return render(request,"activatefail.html")

@api_view(['DELETE'])
def delete_common_clientuser(request,pk, user_pk):
    client = Client.objects.get(id=pk)
    user = CommonUser.objects.get(id = user_pk, client=client)
    if request.method == 'DELETE':
        user.delete()
        return Response({'message':'Clientuser User deleted'}, status=status.HTTP_200_OK)
    return Response ({'error_message':'Failed to delete clientser user'},status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
def list_users_by_client(request, pk):
    client = get_object_or_404(Client, id=pk)
    users = CommonUser.objects.filter(client=client)
    serializer = ClientuserSerializer(users, many=True)
    return Response(serializer.data)

@api_view(['GET'])
def single_common_clientuser(request, pk, user_pk): 
    client = Client.objects.get(id = pk)
    user = CommonUser.objects.get(id = user_pk, client=client)
    if request.method == 'GET':
        serializer = ClientuserSerializerWithToken(user)
        print(serializer)
        return Response(serializer.data)

@api_view(['POST'])
def reset_clientuser_password(request, pk, user_pk):
    client = Client.objects.get(id=pk)
    user = CommonUser.objects.get(id=user_pk, client=client)
    if request.method == 'POST':
        data = request.data
        passwords = request.data.get('password')
        print("Previous (hashed) password:", passwords)
        previous_password = request.data.get('previous_password')
        new_password = request.data.get('new_password')
        confirm_password = request.data.get('confirm_password')
        if not user.check_password(previous_password):
            return Response({'error_message': 'Incorrect previous password !!'}, status=status.HTTP_400_BAD_REQUEST)
        if new_password != confirm_password:
            return Response({'error_message': 'New password and confirm password do not match !!'}, status=status.HTTP_400_BAD_REQUEST)
        user.set_password(new_password)  # Set the new password
        user.save()
        user_serializer = ClientuserSerializerWithToken(data=data, instance=user, partial=True)
        if user_serializer.is_valid():
            user_serializer.save(client=client)
            return Response({'message':'Client User Password Updated'})
        return Response(user_serializer.errors,status=status.HTTP_400_BAD_REQUEST)

# @api_view(['GET', 'POST'])
# def forget_password(request, user_pk):
#     data = request.data
#     user = CommonUser.objects.get(id=user_pk)    
#     # return Response(serializer.data)
#     if request.method == 'POST':
#         username = data.get('username')
#         if CommonUser.objects.filter(username=username):
#             email_subject = "Activate Your Account"
#             message = 'jhjhj'
#             email_message = EmailMessage(email_subject, message, settings.EMAIL_HOST_USER, [username])
#             email_message.send()
#             serializer = UserSerializer(user)
#             return Response({'message' : 'Kindly Check ur mail'})
#         else:
#             return Response ({'error_message' : "Given username or email doesn't exist"}, status=status.HTTP_400_BAD_REQUEST )
#     elif request.method == 'GET':
#         ser = UserSerializer(user)
#         return Response(ser.data)

from django.utils.http import urlsafe_base64_encode
from django.utils.encoding import force_bytes
from django.template.loader import render_to_string
from django.contrib.auth.tokens import PasswordResetTokenGenerator

token_generator = PasswordResetTokenGenerator()

@api_view(['POST'])
def forget_password(request):
    username = request.data.get('username')

    try:
        user = CommonUser.objects.get(username=username)
    except CommonUser.DoesNotExist:
        return Response({'error_message': "Given username doesn't exist"}, status=status.HTTP_400_BAD_REQUEST)

    # Generate token and UID
    uid = urlsafe_base64_encode(force_bytes(user.pk))
    token = token_generator.make_token(user)

    # Email content
    email_subject = "Reset Your Password"
    message = render_to_string("passwordreset.html", {
        'user': user,
        # 'domain': 'admin.dms.zacoinfotech.com', https://dms-frontend-new.vercel.app/forgetpassword
        # http://localhost:5173/reset-password/MzM/cu85ad-af76c1ed92e7af05f1df7d9f508c1b9d/
        # http://127.0.0.1:8000/reset-password/MzM/cu85en-236a306cd2999dd9db4ec85c0d0cf2c1/
        # 'domain': 'localhost:5173',
        'domain': 'dms-frontend-new.vercel.app',
        'uid': uid,
        'token': token,
    })

    email = EmailMessage(email_subject, message, settings.EMAIL_HOST_USER, [user.email])
    email.send()

    return Response({'message': 'Kindly check your email to reset your password.'}, status=status.HTTP_200_OK)

@api_view(['POST'])
def reset_password(request, uidb64, token):
    try:
        uid = urlsafe_base64_decode(uidb64).decode()
        user = CommonUser.objects.get(pk=uid)
    except (TypeError, ValueError, OverflowError, CommonUser.DoesNotExist):
        return Response({'error_message': 'Invalid link'}, status=status.HTTP_400_BAD_REQUEST)

    if not token_generator.check_token(user, token):
        return Response({'error_message': 'Invalid or expired token'}, status=status.HTTP_400_BAD_REQUEST)

    password = request.data.get('password')
    confirm_password = request.data.get('confirm_password')

    if password != confirm_password:
        return Response({'error_message': 'Passwords do not match'}, status=status.HTTP_400_BAD_REQUEST)

    user.set_password(password)
    user.save()

    return Response({'message': 'Password has been reset successfully.'}, status=status.HTTP_200_OK)


# ******************************************Commom Customer User*************************************

@api_view(['POST'])
def create_common_customeruser(request, pk):
    data = request.data
    client = get_object_or_404(Client, id=pk)
    email = data.get('email')
    username = data.get('email')  # you're setting username as email
    # confirm_password = request.data.get('confirm_password')
    password = request.data.get('password')
    name = data.get('name')

    # ✅ Check if email or username already exists
    if CommonUser.objects.filter(email=email).exists():
        return Response({'error_message': 'Email is already registered'}, status=status.HTTP_400_BAD_REQUEST)

    if CommonUser.objects.filter(username=username).exists():
        return Response({'error_message': 'Username already exists'}, status=status.HTTP_400_BAD_REQUEST)

    # name_part = name[:4].lower()
    # generated_password = f"{name_part}@123"
    
    # if password != data['confirm_password']:
    #     return Response({'error_message': 'Password and ConfirmPassword do not match'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        # ✅ Create user
        customeruser = CommonUser.objects.create(
            # first_name=data['first_name'],
            # last_name=data['last_name'],
            name=data['name'],
            username=username,
            email=email,
            # password=make_password(generated_password),
            password=make_password(data['password']),
            role='customeruser',
            # confirm_password=make_password(data['confirm_password']),
            # superadmin_user=True,
            is_active=True,
            is_staff=False, 
            client=client    
            # is_superuser=True,
        )

        # ✅ Email activation token
        email_subject = "Your Account is Registered"
        message = render_to_string("activate.html", {
            'user': customeruser,
            # 'domain': '127.0.0.1:8000',
            'password': password,
            'domain': 'admin.dms.zacoinfotech.com',
            'uid': urlsafe_base64_encode(force_bytes(customeruser.pk)),
            'token': generate_token.make_token(customeruser),
        })

        email_message = EmailMessage(email_subject, message, settings.EMAIL_HOST_USER, [email])
        email_message.send()

        serializer = CustomeruserSerializerWithToken(customeruser, many=False)
        return Response({'message': 'User Registered. Kindly Check your mail for password.', 'data': serializer.data})
    
    except Exception as e:
        return Response({'error_message': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST', 'GET'])
def edit_common_customeruser(request,pk, user_pk):
    client = Client.objects.get(id=pk)
    user = CommonUser.objects.get(id=user_pk, client=client)
    user_serializer = CustomeruserSerializerWithToken(data=request.data, instance=user, partial=True )
    if request.method == 'POST':
        if user_serializer.is_valid():
            user_serializer.save(client=client)
            return Response({'message':'Customeruser User Updated','Data': user_serializer.data}, status=status.HTTP_200_OK)
        return Response(user_serializer.errors,status=status.HTTP_400_BAD_REQUEST)
    elif request.method == 'GET':
        user_ser = CustomeruserSerializerWithToken(user)
        return Response(user_ser.data)

# Email Activations
class CustomeruserActivateAccountView(View):
    def get(self, request, uidb64, token):
        try:
            uid= force_text(urlsafe_base64_decode(uidb64))
            user = CommonUser.objects.get(pk=uid)
        except Exception as identifier:
            user=None
        if user is not None and generate_token.check_token(user,token):
            user.is_active=True
            user.save()
            return render(request,"activatesuccess.html")
        else:
            return render(request,"activatefail.html")

@api_view(['DELETE'])
def delete_common_customeruser(request,pk, user_pk):
    client = Client.objects.get(id=pk)
    user = CommonUser.objects.get(id = user_pk, client=client)
    if request.method == 'DELETE':
        user.delete()
        return Response({'message':'Customer User deleted'}, status=status.HTTP_200_OK)
    return Response ({'error_message':'Failed to delete customer user'},status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
def list_users_by_role(request, pk):
    client = get_object_or_404(Client, id=pk)
    users = CommonUser.objects.filter(client=client, role='customeruser')
    serializer = CustomeruserSerializer(users, many=True)
    return Response(serializer.data)

@api_view(['GET'])
def single_common_customeruser(request, pk, user_pk): 
    client = Client.objects.get(id = pk)
    user = CommonUser.objects.get(id = user_pk, client=client)
    if request.method == 'GET':
        serializer = CustomeruserSerializerWithToken(user)
        print(serializer)
        return Response(serializer.data)

# ******************************************Company Document **************************************

# @api_view(['POST'])
# @parser_classes([MultiPartParser, FormParser])
# def create_companydoc(request, pk):
#     client = Client.objects.get(id=pk)
#     instance_data = request.data
#     data = {key: value for key, value in instance_data.items()}

#     # Save the CompanyDocument
#     serializer = FileInfoSerializer(data=data)
#     if serializer.is_valid(raise_exception=True):
#         fileinfo = serializer.save(client=client)

#         # Handle file uploads correctly (expecting multiple files)
#         files = request.FILES.getlist('files')  # This gets a list of files
#         if files:
#             file_list = [{'fileinfo': fileinfo.pk, 'files': file} for file in files]
#             file_serializer = FileSerializer(data=file_list, many=True)  # Pass the list of file data
#             if file_serializer.is_valid(raise_exception=True):
#                 file_serializer.save()

#         # Success response
#         return Response(
#             {
#                 'message': 'Company Document created successfully.',
#                 'data': serializer.data,
#             },
#             status=status.HTTP_201_CREATED,
#         )

#     # Error response if serializer fails
#     return Response(
#         {
#             'message': 'Failed to create Company Document.',
#             'error_message': serializer.errors,
#         },
#         status=status.HTTP_400_BAD_REQUEST,
#     )

@api_view(['POST'])
def create_companydoc(request, pk):
    if request.method == 'POST':
        client_id = request.data.get('client_id')  # Expecting the client_id in the request

        try:
            client = Client.objects.get(id=pk)
        except Client.DoesNotExist:
            return Response({'message': 'Client not found'}, status=status.HTTP_404_NOT_FOUND)

        fileinfo_serializer = FileInfoSerializer(data=request.data)
        
        if fileinfo_serializer.is_valid():
            fileinfo = fileinfo_serializer.save(client=client)  # Associate with the client
            
            files = request.FILES.getlist('files')
            if files:
                for file in files:
                    file_data = {
                        'fileinfo': fileinfo.pk,
                        'files': file
                    }
                    file_serializer = FileSerializer(data=file_data)
                    if file_serializer.is_valid():
                        file_serializer.save()
            # Return the created fileinfo details
            return Response({
                'message': 'FileInfo and associated files created successfully',
                'data': fileinfo_serializer.data
            }, status=status.HTTP_201_CREATED)
        
        return Response({
            'message': 'Failed to create FileInfo',
            'error_message': fileinfo_serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST', 'GET'])
@parser_classes([MultiPartParser, FormParser])
def edit_companydoc(request, pk, file_pk):
    try:
        client = Client.objects.get(id=pk)
        fileinfo = FileInfo.objects.get(id=file_pk, client=client)
    except Client.DoesNotExist:
        return Response({'error_message': 'Client not found'}, status=status.HTTP_404_NOT_FOUND)
    except CompanyDocument.DoesNotExist:
        return Response({'error_message': 'Company Document not found'}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'POST':
        # Update branch document fields
        fileinfo_serializer = FileInfoSerializer(instance=fileinfo, data=request.data)
        print('before',request.data)
        if fileinfo_serializer.is_valid():
            fileinfo_serializer.save(client=client)
            print('after',request.data)

            # Handle file updates separately
            if request.FILES.getlist('files'):
                # Delete the old files if required
                File.objects.filter(fileinfo=fileinfo).delete()

                # Add the new files
                files = request.FILES.getlist('files')
                for file in files:
                    file_data = {
                        'fileinfo': fileinfo.pk,
                        'files': file
                    }
                    file_serializer = FileSerializer(data=file_data)
                    if file_serializer.is_valid():
                        file_serializer.save()
                    else:
                        return Response(file_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

            return Response({
                'message': 'Company Document updated',
                'data': fileinfo_serializer.data,
                'status' : status.HTTP_200_OK
            })
        else:
            return Response({
                'message': 'Fail to update Company Document',
                'error_message': fileinfo_serializer.errors,
                },status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'GET':
        fileinfo_serializer = FileInfoSerializer(fileinfo)
        return Response(fileinfo_serializer.data)       

@api_view(['GET'])
def list_companydoc(request,pk):
    client = Client.objects.get(id=pk)
    doc_list = FileInfo.objects.filter(client=client)
    serializer = FileInfoSerializer(doc_list,many=True)
    print(serializer)
    return Response(serializer.data)

@api_view(['DELETE'])
def delete_companydoc(request,pk, file_pk):
    client = Client.objects.get(id=pk)
    doc = FileInfo.objects.get(id=file_pk, client=client)
    if request.method == 'DELETE':
        doc.delete()
        return Response({'message':"Document Deleted"})
    return Response({'error_message':"Failed to delete document"},status=status.HTTP_400_BAD_REQUEST)

# ************************************** Branch View's ********************************************

@api_view(['POST'])
def create_branch(request, pk):
    client = Client.objects.get(id=pk)
    if request.method == 'POST':
        branch_serializer = BranchSerailizer(data=request.data)
        if branch_serializer.is_valid():
            branch_serializer.save(client=client)
            return Response(
                {
                    'message': 'Branch created',
                    'data': branch_serializer.data,
                },
                status=status.HTTP_201_CREATED  # Correct placement of status
            )
        else:
            return Response(
                {
                    'message': 'Fail to create branch',
                    'error_message': branch_serializer.errors,
                },
                status=status.HTTP_400_BAD_REQUEST  # Correct placement of status
            )

@api_view(['POST', 'GET'])
def edit_branch(request,pk,branch_pk):
    client = Client.objects.get(id=pk)
    branch = Branch.objects.get(id = branch_pk)
    if request.method == 'POST':
        branch_serializer = BranchSerailizer(instance=branch, data=request.data)
        if branch_serializer.is_valid():
            branch_serializer.save(client=client)
            return Response({
                'message': 'Branch updated',
                'data': branch_serializer.data,
                },status=status.HTTP_201_CREATED)
        return Response({
                'message': 'Fail to update branch',
                'error_message': branch_serializer.errors,
                },status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'GET':
        branch_ser = BranchSerailizer(branch)
        return Response (branch_ser.data)

@api_view(['GET'])
def list_branch(request,pk):
    client = Client.objects.get(id=pk)
    if request.method == 'GET':
        branch_list = Branch.objects.filter(client=client)
        branch_serializer = BranchSerailizer(branch_list, many=True)
        print(branch_serializer)
        return Response(branch_serializer.data)

@api_view(['DELETE'])
def delete_branch(request,pk,branch_pk):
    client = Client.objects.get(id=pk)
    branch = Branch.objects.get(id=branch_pk)
    if request.method == 'DELETE':
        branch.delete()
        return Response ({'message':'Branch deleted'})
    return Response ({'error_message':'Fail to delete branch'},status=status.HTTP_400_BAD_REQUEST)

# *************************************Office Location**********************************************

@api_view(['POST'])
def create_officelocation(request,branch_pk):
    branch = Branch.objects.get(id=branch_pk)
    if request.method == 'POST':
        officeLocation_serializer = OfficeLocationSerializer(data=request.data)
        if officeLocation_serializer.is_valid():
            officeLocation_serializer.save(branch=branch)
            return Response({
                'message': 'Office Loaction created',
                'data': officeLocation_serializer.data,
                },status=status.HTTP_201_CREATED)
        return Response({
                'message': 'Fail to create Office Location',
                'error_message': officeLocation_serializer.errors,
                },status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST', 'GET'])
def edit_officelocation(request,branch_pk,office_pk):
    branch = Branch.objects.get(id=branch_pk)
    office = OfficeLocation.objects.get(id=office_pk)
    officeLocation_serializer = OfficeLocationSerializer(instance=office, data=request.data)
    if request.method == 'POST':
        if officeLocation_serializer.is_valid():
            officeLocation_serializer.save(branch=branch)
            return Response({
                'message': 'Office Loaction updated',
                'data': officeLocation_serializer.data,
                'status' : status.HTTP_200_OK
            })
        return Response({
                'message': 'Fail to updated Office Location',
                'error_message': officeLocation_serializer.errors,
                },status=status.HTTP_400_BAD_REQUEST)
    elif request.method == 'GET':
        office_ser = OfficeLocationSerializer(office)
        return Response (office_ser.data)

@api_view(['GET'])
def single_officelocation(request, branch_pk, office_pk):
    branch = Branch.objects.get(id = branch_pk)
    owner = OfficeLocation.objects.get(id = office_pk, branch=branch)
    if request.method == 'GET':
        serializer = OfficeLocationSerializer(owner)
        print(serializer)
        return Response(serializer.data)

@api_view(['GET'])
def list_officelocation(request,pk, branch_pk):
    client = Client.objects.get(id=pk)
    branch = Branch.objects.get(id = branch_pk, client=client)
    if request.method == 'GET':
        list_officelocation= OfficeLocation.objects.filter(branch=branch)
        officeLocation_serializer = OfficeLocationSerializer(list_officelocation, many=True)
        print(officeLocation_serializer)
        return Response(officeLocation_serializer.data)

@api_view(['DELETE'])
def delete_officelocation(request,pk, branch_pk, office_pk):
    client = Client.objects.get(id=pk)
    branch = Branch.objects.get(id=branch_pk, client=client)
    office = OfficeLocation.objects.get(id = office_pk, branch = branch)
    if request.method == 'DELETE':
        office.delete()
        return Response({'message':'Office Location deleted'})
    return Response ({'error_message':'Failed to delete office location'}, status=status.HTTP_400_BAD_REQUEST)

# *************************************Customer Or Vendor **************************************

@api_view(['POST'])
def create_customer(request, pk):
    client = Client.objects.get(id=pk)
    if request.method == 'POST':
        customer_serializer = CustomerVendorSerializer(data=request.data)
        pan = request.data.get('pan')
        if pan and len(pan) > 10 or len(pan) < 10:
            return Response({'error_message': 'PAN number should be at most 10 characters'}, status=status.HTTP_400_BAD_REQUEST)

        mobile = request.data.get('contact')
        if mobile and (len(str(mobile)) > 10 or len(str(mobile)) < 10):
            return Response({'error_message': 'Mobile number should be at most 10 characters'}, status=status.HTTP_400_BAD_REQUEST)

        email = request.data.get('email')
        # if email and not re.match(r"[^@]+@[^@]+\.[^@]+", email) or email == '':
        if email == '':
            return Response({'error_message': 'Email field is required'}, status=status.HTTP_400_BAD_REQUEST)

        if customer_serializer.is_valid():
            customer_serializer.save(client=client)
            return Response({
                'message': 'Customer or Vendor created',
                'data': customer_serializer.data,
                },status=status.HTTP_201_CREATED)
        return Response({
                'message': 'Fail to create Customer or Vendor',
                'error_message': customer_serializer.errors,
                },status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST', 'GET'])
def edit_customer(request, pk, customer_pk):
    client = Client.objects.get(id=pk)
    customer = Customer.objects.get(id=customer_pk)
    customer_serializer = CustomerVendorSerializer(instance=customer, data=request.data)
    if request.method == 'POST':

        pan = request.data.get('pan')
        if pan and len(pan) > 10 or len(pan) < 10:
            return Response({'error_message': 'PAN number should be at most 10 characters'}, status=status.HTTP_400_BAD_REQUEST)

        mobile = request.data.get('contact')
        if mobile and (len(str(mobile)) > 10 or len(str(mobile)) < 10):
            return Response({'error_message': 'Mobile number should be at most 10 characters'}, status=status.HTTP_400_BAD_REQUEST)

        email = request.data.get('email')
        # if email and not re.match(r"[^@]+@[^@]+\.[^@]+", email) or email == '':
        if email == '':
            return Response({'error_message': 'Email field is required'}, status=status.HTTP_400_BAD_REQUEST)

        if customer_serializer.is_valid():
            customer_serializer.save(client=client)
            return Response({
                'message': 'Customer or Vendor updated',
                'data': customer_serializer.data,
                'status' : status.HTTP_200_OK
            })
        return Response({
                'message': 'Fail to update Customer or Vendor',
                'error_message': customer_serializer.errors,
                },status=status.HTTP_400_BAD_REQUEST)
    elif request.method == 'GET':
        customer_ser = CustomerVendorSerializer(customer)
        return Response(customer_ser.data)

@api_view(['GET'])
def list_customer(request, pk):
    client = Client.objects.get(id=pk)
    if request.method == 'GET':
        list_customer = Customer.objects.filter(client=client)
        customer_serializer = CustomerVendorSerializer(list_customer, many=True)
        print(customer_serializer)
        return Response(customer_serializer.data)

@api_view(['GET'])
def single_customer(request, pk, customer_pk):
    client = Client.objects.get(id = pk)
    owner = Customer.objects.get(id = customer_pk, client=client)
    if request.method == 'GET':
        serializer = CustomerVendorSerializer(owner)
        print(serializer)
        return Response(serializer.data)

@api_view(['DELETE'])
def delete_customer(request,pk, customer_pk):
    client = Client.objects.get(id=pk)
    customer = Customer.objects.get(id=customer_pk)
    if request.method == 'DELETE':
        customer.delete()
        return Response({'message':'Customer or Vendor deleted'})
    return Response({'error_message':'Fail to delete Customer or Vendor'},status=status.HTTP_400_BAD_REQUEST)

# **********************************************Branch Document*********************************************
# @api_view(['POST'])
# @parser_classes([MultiPartParser, FormParser])
# def create_branchdoc(request,branch_pk):
#     branch = Branch.objects.get(id=branch_pk)
#     instance_data = request.data
#     data = {key: value for key, value in instance_data.items()}

#     serializer = BranchDocSerailizer(data=data)
#     if serializer.is_valid(raise_exception=True):
#         doc_instance = serializer.save(branch=branch)
#         print(request.data)

#         if request.FILES:
#             files = dict((request.FILES).lists()).get('files',None)
#             # files = request.FILES.getlist('files')
#             if files:
#                 for file in files:
#                     file_data = {
#                         'branch_doc' : doc_instance.pk,
#                         'files' : file
#                     }
#                     file_serializer= FilesSerializer(data=file_data)
#                     if file_serializer.is_valid(raise_exception=True):
#                         file_serializer.save()
#                     else:
#                         return Response(file_serializer.errors, status=status.HTTP_400_BAD_REQUEST)
#         return Response({
#             'message': 'Branch Document created',
#             'data': serializer.data,
#             },status=status.HTTP_201_CREATED)
#     else:    
#         return Response({
#                 'message': 'Fail to create Branch Document',
#                 'error_message': serializer.errors,
#                 },status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def create_branchdoc(request, branch_pk):
    branch = Branch.objects.get(id=branch_pk)
    instance_data = request.data
    data = {key: value for key, value in instance_data.items()}

    serializer = BranchDocSerailizer(data=data)
    if serializer.is_valid(raise_exception=True):
        doc_instance = serializer.save(branch=branch)
        print(request.data)

        if request.FILES:
            files = dict((request.FILES).lists()).get('files', None)
            # files = request.FILES.getlist('files')
            if files:
                for file in files:
                    file_data = {
                        'branch_doc': doc_instance.pk,
                        'files': file
                    }
                    file_serializer = FilesSerializer(data=file_data)
                    if file_serializer.is_valid(raise_exception=True):
                        file_serializer.save()
                    else:
                        return Response(file_serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        return Response({


            
            'message': 'Branch Document created',
            'data': serializer.data,
        }, status=status.HTTP_201_CREATED)
    else:    
        return Response({
            'message': 'Failed to create Branch Document',
            'error_message': serializer.errors,
        }, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST', 'GET'])
@parser_classes([MultiPartParser, FormParser])
def edit_branchdoc(request, branch_pk, branchdoc_pk):
    try:
        branch = Branch.objects.get(id=branch_pk)
        branchdoc = BranchDocument.objects.get(id=branchdoc_pk, branch=branch)
    except Branch.DoesNotExist:
        return Response({'error_message': 'Branch not found'}, status=status.HTTP_404_NOT_FOUND)
    except BranchDocument.DoesNotExist:
        return Response({'error_message': 'Branch Document not found'}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'POST':
        # Update branch document fields
        branchdoc_serializer = BranchDocSerailizer(instance=branchdoc, data=request.data)
        if branchdoc_serializer.is_valid():
            branchdoc_serializer.save(branch=branch)

            # Handle file updates separately
            if request.FILES.getlist('files'):
                # Delete the old files if required
                Files.objects.filter(branch_doc=branchdoc).delete()

                # Add the new files
                files = request.FILES.getlist('files')
                for file in files:
                    file_data = {
                        'branch_doc': branchdoc.pk,
                        'files': file
                    }
                    file_serializer = FilesSerializer(data=file_data)
                    if file_serializer.is_valid():
                        file_serializer.save()
                    else:
                        return Response(file_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

            return Response({
                'message': 'Branch Document updated',
                'data': branchdoc_serializer.data,
                'status' : status.HTTP_200_OK
            })
        else:
            return Response({
                'message': 'Fail to update Branch Document',
                'error_message': branchdoc_serializer.errors,
                },status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'GET':
        branchdoc_serializer = BranchDocSerailizer(branchdoc)
        return Response(branchdoc_serializer.data)

@api_view(['GET'])
def list_branchdoc(request, branch_pk):
    branch = Branch.objects.get(id = branch_pk)
    # branchdoc = BranchDocument.objects.get(id = branchdoc_pk, branch=branch)
    if request.method == 'GET':
        list_branchdoc = BranchDocument.objects.filter(branch=branch)
        branchdoc_serializer = BranchDocSerailizer(list_branchdoc, many=True)
        print(branchdoc_serializer)
        return Response(branchdoc_serializer.data)

@api_view(['GET'])
def single_branchdoc(request, branch_pk, branchdoc_pk):
    branch = Branch.objects.get(id=branch_pk)
    doc = BranchDocument.objects.get(id = branchdoc_pk, branch=branch)
    if request.method == 'GET':
        ser = BranchDocSerailizer(doc)
        print(ser)
        return Response(ser.data)

@api_view(['DELETE'])
def delete_branchdoc(request,branch_pk, branchdoc_pk ):
    branch = Branch.objects.get(id = branch_pk)
    branchdoc = BranchDocument.objects.get(id = branchdoc_pk, branch=branch)
    if request.method == 'DELETE':
        branchdoc.delete()
        return Response({'message':'Branch Document Delete'})
    return Response({'error_message':'Fail to delete branch document'} ,status=status.HTTP_400_BAD_REQUEST)

# **********************************************# Income Tax Document******************************************
@api_view(['POST'])
def create_incometaxdoc(request,pk):
    client = Client.objects.get(id = pk)
    if request.method == 'POST':
        income_serializer = IncomeTaxDocumentSerializer(data=request.data)
        if income_serializer.is_valid():
            income_serializer.save(client=client)
            return Response({
                'message': 'Income Tax Documents created',
                'data': income_serializer.data,
                },status=status.HTTP_201_CREATED)
        return Response({
                'message': 'Fail to create Income Tax Documents',
                'error_message': income_serializer.errors,
                },status=status.HTTP_400_BAD_REQUEST)
@api_view(['POST','GET'])

def edit_incometaxdoc(request, pk, income_pk):
    client = Client.objects.get(id=pk)
    income = IncomeTaxDocument.objects.get(id = income_pk, client=client)
    income_serializer = IncomeTaxDocumentSerializer(instance=income, data=request.data)
    if request.method == 'POST':
        if income_serializer.is_valid():
            income_serializer.save(income=income)
            return Response({
                'message': 'Income Tax Documents updated',
                'data': income_serializer.data,
                'status' : status.HTTP_200_OK
            })
        return Response({
                'message': 'Fail to update Income Tax Documents',
                'error_message': income_serializer.errors,
                },status=status.HTTP_400_BAD_REQUEST)
    elif request.method == 'GET':
        income_ser = IncomeTaxDocumentSerializer(income)
        return Response(income_ser.data)

@api_view(['GET'])
def list_incometaxdoc(request,pk):
    client = Client.objects.get(id=pk)
    if request.method == 'GET':
        list_income = IncomeTaxDocument.objects.filter(client=client)
        income_serializer = IncomeTaxDocumentSerializer(list_income, many=True)
        print(income_serializer)
        return Response(income_serializer.data)

@api_view(['DELETE'])
def delete_incometaxdoc(request,pk,income_pk):
    client = Client.objects.get(id=pk)
    income = IncomeTaxDocument.objects.get(id=income_pk)
    if request.method == 'DELETE':
        income.delete()
        return Response({'message':'Income tax document deleted'})
    return Response ({'error_message':'Fail to delete Income tax document'}, status=status.HTTP_400_BAD_REQUEST)

#*****************************************************PF*******************************************************
# @api_view(['POST'])
# def create_pf(request,pk):
#     client = Client.objects.get(id = pk)
#     if request.method == 'POST':
#         pf_serializer = PfSerializer(data=request.data)
#         if pf_serializer.is_valid():
#             pf_serializer.save(client=client)
#             return Response({
#                 'message': 'PF created',
#                 'data': pf_serializer.data,
#                 },status=status.HTTP_201_CREATED)
#         return Response({
#                 'error_message': 'Fail to create PF',
#                 'error_message': pf_serializer.errors,
#                 },status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
def create_pf(request, pk):
    client = Client.objects.get(id=pk)
    
    if request.method == 'POST':
        employee_code = request.data.get('employee_code')
        uan = request.data.get('uan')
        pf_number = request.data.get('pf_number')
        month = request.data.get('month')

        # Check if entry already exists for this combination
        existing_pf = PF.objects.filter(
            client=client,
            employee_code=employee_code,
            uan=uan,
            pf_number=pf_number,
            month=month
        ).first()

        if existing_pf:
            return Response({
                'error_message': f'Data for employee {employee_code} for {month} already exists.'
            }, status=status.HTTP_400_BAD_REQUEST)

        pf_serializer = PfSerializer(data=request.data)
        if pf_serializer.is_valid():
            pf_serializer.save(client=client)
            return Response({
                'message': 'PF created',
                'data': pf_serializer.data,
            }, status=status.HTTP_201_CREATED)
        
        return Response({
            'error_message': 'Fail to create PF',
            'errors': pf_serializer.errors,
        }, status=status.HTTP_400_BAD_REQUEST)

# class ExcelImportView(APIView):
#     parser_classes = [MultiPartParser]

#     def post(self, request, pk, *args, **kwargs):
#         # Get the client using the pk from the URL
#         try:
#             client = Client.objects.get(pk=pk)
#         except Client.DoesNotExist:
#             return Response({"error_message": "Client not found"}, status=404)

#         file = request.FILES['file']

#         # Load the workbook and select the active worksheet
#         wb = load_workbook(file)
#         ws = wb.active

#         pf_entries = []

#         # Define the fields in a list to optimize the entry creation
#         fields = [
#             'employee_code', 'employee_name', 'uan', 'pf_number', 'pf_deducted',
#             'date_of_joining', 'status', 'month', 'gross_ctc', 'basic_pay',
#             'hra', 'statutory_bonus', 'special_allowance', 'pf', 'gratuity',
#             'total_gross_salary', 'number_of_days_in_month', 'present_days',
#             'lwp', 'leave_adjustment', 'gender', 'basic_pay_monthly',
#             'hra_monthly', 'statutory_bonus_monthly', 'special_allowance_monthly',
#             'total_gross_salary_monthly', 'provident_fund', 'professional_tax',
#             'advance', 'esic_employee', 'tds', 'total_deduction', 'net_pay',
#             'advance_esic_employer_cont'
#         ]

#         # Iterate through the rows in the Excel file and create or update PF objects
#         for row in ws.iter_rows(min_row=2):  # Skip the header row
#             # Create a dictionary for the current row
#             data = {field: row[i].value for i, field in enumerate(fields)}

#             employee_code = data['employee_code']
#             month = data['month']

#             # Skip rows without essential fields
#             if not employee_code or not month:
#                 continue

#             # Check if an entry with the same employee_code and month exists
#             instance = PF.objects.filter(employee_code=employee_code, month=month).first()

#             if instance:
#                 # Update existing entry
#                 for field, value in data.items():
#                     setattr(instance, field, value)
#                 instance.save()
#             else:
#                 # Create new PF entry and associate it with the client
#                 pf_entry = PF(client=client, **data)
#                 pf_entry.save()
#                 pf_entries.append(pf_entry)

#         return Response({"status": "success", "data": PfSerializer(pf_entries, many=True).data},status=status.HTTP_201_CREATED)

# class ExcelImportView(APIView):
#     parser_classes = [MultiPartParser]

#     def post(self, request, pk, *args, **kwargs):
#         try:
#             client = Client.objects.get(pk=pk)
#         except Client.DoesNotExist:
#             return Response({"error_message": "Client not found"}, status=status.HTTP_404_NOT_FOUND)

#         file = request.FILES.get('file')
#         if not file:
#             return Response({"error_message": "No file provided"}, status=status.HTTP_400_BAD_REQUEST)

#         wb = load_workbook(file)
#         ws = wb.active

#         fields = [
#             'employee_code', 'employee_name', 'uan', 'pf_number', 'pf_deducted',
#             'date_of_joining', 'status', 'month', 'gross_ctc', 'basic_pay',
#             'hra', 'statutory_bonus', 'special_allowance', 'pf', 'gratuity',
#             'total_gross_salary', 'number_of_days_in_month', 'present_days',
#             'lwp', 'leave_adjustment', 'gender', 'basic_pay_monthly',
#             'hra_monthly', 'statutory_bonus_monthly', 'special_allowance_monthly',
#             'total_gross_salary_monthly', 'provident_fund', 'professional_tax',
#             'advance', 'esic_employee', 'tds', 'total_deduction', 'net_pay',
#             'advance_esic_employer_cont'
#         ]

#         for row in ws.iter_rows(min_row=2):  # Skip the header row
#             data = {field: row[i].value for i, field in enumerate(fields)}

#             employee_code = data.get('employee_code')
#             month = data.get('month')

#             if not employee_code or not month:
#                 continue  # Skip invalid rows

#             # Check if employee_code already exists for the given month
#             if PF.objects.filter(employee_code=employee_code, month=month).exists():
#                 return Response(
#                     {"error_message": f"Employee {employee_code} already exists for month {month}"},
#                     status=status.HTTP_400_BAD_REQUEST
#                 )

#             # Create new PF entry
#             PF.objects.create(client=client, **data)

#         return Response({"status": "success", "message": "All entries uploaded successfully"}, status=status.HTTP_201_CREATED)

# @api_view(['POST','GET'])
# def edit_pf(request, pk, pf_pk):
#     client = Client.objects.get(id=pk)
#     # income = IncomeTaxDocument.objects.get(id = pf_pk, client=client)
#     pf = PF.objects.get(id = pf_pk, client=client)
#     pf_serializer = PfSerializer(instance=pf, data=request.data)
#     if request.method == 'POST':
#         if  pf_serializer.is_valid():
#             pf_serializer.save()
#             return Response({
#                 'message': 'PF updated',
#                 'data': pf_serializer.data,
#                 'status' : status.HTTP_200_OK
#             })
#         return Response({
#                 'message': 'Fail to update PF',
#                 'error_message': pf_serializer.errors,
#                 },status=status.HTTP_400_BAD_REQUEST)
#     elif request.method == 'GET':
#         pf_ser = PfSerializer(pf)
#         return Response(pf_ser.data)

import re

# class ExcelImportView(APIView):
#     parser_classes = [MultiPartParser]

#     def post(self, request, pk, *args, **kwargs):
#         try:
#             client = Client.objects.get(pk=pk)
#         except Client.DoesNotExist:
#             return Response({"error_message": "Client not found"}, status=status.HTTP_404_NOT_FOUND)

#         file = request.FILES.get('file')
#         if not file:
#             return Response({"error_message": "No file provided"}, status=status.HTTP_400_BAD_REQUEST)

#         wb = load_workbook(file)
#         df = pd.read_excel(file)
#         ws = wb.active

#         fields = [
#             'employee_code', 'employee_name', 'uan', 'pf_number', 'pf_deducted',
#             'date_of_joining', 'status', 'month', 'gross_ctc', 'basic_pay',
#             'hra', 'statutory_bonus', 'special_allowance', 'pf', 'gratuity',
#             'total_gross_salary', 'number_of_days_in_month', 'present_days',
#             'lwp', 'leave_adjustment', 'gender', 'basic_pay_monthly',
#             'hra_monthly', 'statutory_bonus_monthly', 'special_allowance_monthly',
#             'total_gross_salary_monthly', 'provident_fund', 'professional_tax',
#             'advance', 'esic_employee', 'tds', 'total_deduction', 'net_pay',
#             'advance_esic_employer_cont'
#         ]

        
#         skipped_rows = []
#         # month_pattern = re.compile(
#         #     r'^(jan|feb|mar|apr|many|jun|jul|aug|sep|oct|nov|dec'
#         #     r' january|february|march|april|may|june|july|august|september|october|november|december)\s+20\d{2}$',
#         #     re.IGNORECASE
#         # )

#         # errors = []

#         # for index, row in df.iterrows():
#         #     row_errors = {}
#         #     month_value = row.get('Month')

#         #     # emp_code = row.get('Employee_code')
#         #     emp_code = str(row.get('Employee Code', f'Row {index + 2}')).strip()


#         #     if not isinstance(month_value, str) or not month_pattern.match(month_value.strip()):
#         #         row_errors['Month'] = 'Invalid month format. Use "August 2024" or "Aug 2024".'

            
#         #     if row_errors:
#         #         errors.append({'employee_code':emp_code, 'errors': row_errors})

#         # if errors:
#         #     return Response({'error_message': errors}, status=status.HTTP_400_BAD_REQUEST)


#         for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):  # row index starts at 2
#             data = {field: row[i].value for i, field in enumerate(fields)}
#             employee_code = data.get('employee_code')
#             uan = data.get('uan')
#             pf_number = data.get('pf_number')
#             month = data.get('month')

#             if not (employee_code and uan and pf_number and month):
#                 skipped_rows.append(idx)
#                 continue  # Skip incomplete rows

#             if PF.objects.filter(
#                 employee_code=employee_code,
#                 uan=uan,
#                 pf_number=pf_number,
#                 month=month,
#                 client=client
#             ).exists():
#                 skipped_rows.append(idx)
#                 continue  # Skip duplicates

#             PF.objects.create(client=client, **data)

#         message = "Upload completed."
#         if skipped_rows:
#             message += f" Skipped rows due to duplication or missing data: {skipped_rows}"

#         return Response({"status": "success", "message": message}, status=status.HTTP_201_CREATED)

class PFExcelUploaadView(APIView):
    parser_classes = [MultiPartParser]
    def post (self, request, pk, *args, **kwargs):
        try:
            client = Client.objects.get(pk=pk)
        except Client.DoesNotExist:
            return Response({"error_message" : "Client not found"}, status=status.HTTP_404_NOT_FOUND)
        
        file = request.FILES.get('file')
        if not file:
            return Response({"error_message": "No file provided"}, status=status.HTTP_400_BAD_REQUEST)

        df = pd.read_excel(file)
        # df.columns = [col.strip().lower().replace(" ", "_") for col in df.columns]
        df.columns = [re.sub(r'[^\w]+', '_', col.strip().lower()).strip('_') for col in df.columns]

        required_fields = [
            'employee_code', 'employee_name', 'uan', 'pf_number', 'pf_deducted',
            'date_of_joining', 'status', 'month', 'gross_ctc', 'basic_pay',
            'hra', 'statutory_bonus', 'special_allowance', 'pf', 'gratuity',
            'total_gross_salary', 'number_of_days_in_month', 'present_days',
            'lwp', 'leave_adjustment', 'gender', 'basic_pay_monthly',
            'hra_monthly', 'statutory_bonus_monthly', 'special_allowance_monthly',
            'total_gross_salary_monthly', 'provident_fund', 'professional_tax',
            'advance', 'esic_employee', 'tds', 'total_deduction', 'net_pay',
            'advance_esic_employer_cont'
        ]

        errors = []
        success = []

        for field in required_fields:
            if field not in df.columns:
                return Response({'error': f'Missing required field: {field}'}, status=status.HTTP_400_BAD_REQUEST)
        
        for index, row in df.iterrows():
            data_dict = row.to_dict()
            exist = PF.objects.filter(
                client=client, 
                employee_code = row['employee_code'],
                uan = row['uan'],
                pf_number = row['pf_number'],
                month = row['month']
            ).exists()

            month = row.get('month')
            if not re.match(r"^(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec|january|february|march|april|june|july|august|september|october|november|december)\s\d{4}$",str(month), re.IGNORECASE):
                # errors.append(f"{row['employee_code']} month field must be in August 2024 or Aug 2024 format")
                return Response({'error_message':f'Employee Code: {row["employee_code"]} month field must be in August 2024 or aug 2024 formate'}, status=status.HTTP_400_BAD_REQUEST)

            # Get the raw value from Excel and strip leading/trailing whitespace
            statu = str(row.get('status', "")).strip().lower()

            # Validate the cleaned value
            if statu not in ['active', 'inactive']:
                return Response(
                    {'error_message': f'Employee Code: {row["employee_code"]} status field must be "active" or "inactive"'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            gen = str(row.get('gender', "")).strip().lower()

            if gen not in ['male', 'female']:
                return Response(
                    {'error_message': f'Employee Code: {row["employee_code"]} gender field must be "male" or "female"'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            if exist:
                errors.append(f"\n Employee Code- {row['employee_code']} data already exists.")
                continue

            data_dict['status'] = statu
            data_dict['gender'] = gen
            print('gggg',data_dict['gender'])
            
            PF.objects.create(client=client, **data_dict)      
            success.append(f"\n {row['employee_code']} Employee") 

        return Response({
            "success_message": success,
            "error_message": errors},
            status=status.HTTP_207_MULTI_STATUS)


@api_view(['GET'])
def pf_field_totals(request, pk):
    try:
        pf_entries = PF.objects.filter(client_id=pk)
        if not pf_entries.exists():
            return Response({"message": "No PF data found for this client."}, status=404)

        totals = {}
        first_entry = pf_entries.first()

        for field in first_entry._meta.get_fields():
            field_name = field.name

            # Skip ID, foreign keys, reverse relations, etc.
            if field.auto_created or field.is_relation or field_name in ['id', 'client']:
                continue

            try:
                # Check if we can sum the field
                total = pf_entries.aggregate(total=Sum(field_name))['total']
                if total is not None:
                    totals[field_name] = total
                else:
                    totals[field_name] = '-'
            except Exception:
                totals[field_name] = '-'

        return Response(totals)
    
    except Exception as e:
        return Response({'error_message': str(e)}, status=500)

@api_view(['GET'])
def get_pf_totals(request, pk):
    pf_records = PF.objects.filter(client_id=pk)

    # Fields where we return "-"
    text_fields = [
        'employee_code', 'employee_name', 'uan', 'pf_number', 
        'pf_deducted', 'date_of_joining', 'status', 'gender','month',
    ]

    # Use model to get all field names
    all_fields = [field.name for field in PF._meta.fields if field.name != 'id' and field.name != 'client']

    result = {}

    for field in all_fields:
        if field in text_fields:
            result[field] = "-"
        else:
            if pf_records.exists():
                total = pf_records.aggregate(total=Sum(field)).get("total")
                result[field] = total if total is not None else 0
            else:
                result[field] = 0

    return Response(result)

@api_view(['POST', 'GET'])
def edit_pf(request, pk, pf_pk):
    try:
        client = Client.objects.get(id=pk)
        pf = PF.objects.get(id=pf_pk, client=client)
    except (Client.DoesNotExist, PF.DoesNotExist):
        return Response({"error_message": "Client or PF record not found"}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'POST':
        new_data = request.data
        # Check for duplicates only if those fields are being changed
        employee_code = new_data.get('employee_code')
        uan = new_data.get('uan')
        pf_number = new_data.get('pf_number')
        month = new_data.get('month')

        if employee_code and uan and pf_number and month:
            exists = PF.objects.filter(
                employee_code=employee_code,
                uan=uan,
                pf_number=pf_number,
                month=month,
                client=client
            ).exclude(id=pf.id).exists()

            if exists:
                return Response(
                    {"error_message": f"PF data for employee {employee_code} already exists for month {month}"},
                    status=status.HTTP_400_BAD_REQUEST
                )

        pf_serializer = PfSerializer(instance=pf, data=request.data)
        if pf_serializer.is_valid():
            pf_serializer.save()
            return Response({
                'message': 'PF updated',
                'data': pf_serializer.data,
                'status': status.HTTP_200_OK
            })
        return Response({
            'message': 'Fail to update PF',
            'error_message': pf_serializer.errors,
        }, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'GET':
        pf_ser = PfSerializer(pf)
        return Response(pf_ser.data)

@api_view(['GET'])
def list_pf(request,pk):
    client = Client.objects.get(id=pk)
    if request.method == 'GET':
        list_pf = PF.objects.filter(client=client)
        pf_serializer = PfSerializer(list_pf, many=True)
        print(pf_serializer)
        return Response(pf_serializer.data)

@api_view(['DELETE'])
def delete_pf(request,pk,pf_pk):
    client = Client.objects.get(id=pk)
    pf = PF.objects.get(id=pf_pk)
    if request.method == 'DELETE':
        pf.delete()
        return Response({'message':'Pf deleted'})
    return Response ({'error_message':'Fail to delete Pf'}, status=status.HTTP_400_BAD_REQUEST)

# ***********************************************Tax Audit***************************************************
@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def create_taxaudit(request,pk):
    client = Client.objects.get(id=pk)
    instance_data = request.data
    data = {key: value for key, value in instance_data.items()}

    serializer = TaxAuditSerializer(data=data)
    if serializer.is_valid(raise_exception=True):
        tax_instance = serializer.save(client=client)
        print(request.data)

        if request.FILES:
            files = dict((request.FILES).lists()).get('files',None)
            # files = request.FILES.getlist('files')
            if files:
                for file in files:
                    file_data = {
                        'tax_audit' : tax_instance.pk,
                        'files' : file
                    }
                    file_serializer= FilesSerializer(data=file_data)
                    if file_serializer.is_valid(raise_exception=True):
                        file_serializer.save()

            return Response({
                'message': 'Tax Audit created',
                'data': serializer.data,
                },status=status.HTTP_201_CREATED)
        return Response({
                'message': 'Fail to create Tax Audit',
                'error_message': serializer.errors,
                },status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST', 'GET'])
@parser_classes([MultiPartParser, FormParser])
def edit_taxaudit(request, pk, taxaudit_pk):
    try:
        client = Client.objects.get(id=pk)
        taxaudit = TaxAudit.objects.get(id=taxaudit_pk, client=client)
    except Client.DoesNotExist:
        return Response({'error_message': 'Client not found'}, status=status.HTTP_404_NOT_FOUND)
    except TaxAudit.DoesNotExist:
        return Response({'error_message': 'Tax Aduit not found'}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'POST':
        # Update branch document fields
        taxaudit_serializer = TaxAuditSerializer(instance=taxaudit, data=request.data)
        if taxaudit_serializer.is_valid():
            taxaudit_serializer.save(client=client)

            # Handle file updates separately
            if request.FILES.getlist('files'):
                # Delete the old files if required
                Files.objects.filter(tax_audit=taxaudit).delete()

                # Add the new files
                files = request.FILES.getlist('files')
                for file in files:
                    file_data = {
                        'tax_audit': taxaudit.pk,
                        'files': file
                    }
                    file_serializer = FilesSerializer(data=file_data)
                    if file_serializer.is_valid():
                        file_serializer.save()
                    else:
                        return Response(file_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

            return Response({
                'message': 'Tax Audit updated',
                'data': taxaudit_serializer.data,
                'status' : status.HTTP_200_OK
            })
        else:
            return Response({
                'message': 'Fail to update Tax Audit',
                'error_message': taxaudit_serializer.errors,
                },status=status.HTTP_400_BAD_REQUEST)
    elif request.method == 'GET':
        taxaudit_serializer = TaxAuditSerializer(taxaudit)
        return Response(taxaudit_serializer.data)

@api_view(['GET'])
def list_taxaudit(request, pk):
    client = Client.objects.get(id = pk)
    if request.method == 'GET':
        list_taxaudit = TaxAudit.objects.filter(client=client)
        taxaudit_serializer = TaxAuditSerializer(list_taxaudit, many=True)
        print(taxaudit_serializer)
        return Response(taxaudit_serializer.data)

@api_view(['GET'])
def single_taxaudit(request, pk, taxaudit_pk):
    client = Client.objects.get(id=pk)
    taxaudit = TaxAudit.objects.get(id = taxaudit_pk, client=client)
    if request.method == 'GET':
        ser = TaxAuditSerializer(taxaudit)
        print(ser)
        return Response(ser.data)

@api_view(['DELETE'])
def delete_taxaudit(request, pk, taxaudit_pk ):
    client = Client.objects.get(id = pk)
    taxaudit = TaxAudit.objects.get(id = taxaudit_pk, client=client)
    if request.method == 'DELETE':
        taxaudit.delete()
        return Response({'message':'Tax Audit Delete'})
    return Response({'error_message':'Fail to delete Tax Audit'} ,status=status.HTTP_400_BAD_REQUEST)

# ******************************************************AIR****************************************************
# @api_view(['POST'])
# @parser_classes([MultiPartParser, FormParser])
# def create_air(request,pk):
#     client = Client.objects.get(id=pk)
#     instance_data = request.data
#     data = {key: value for key, value in instance_data.items()}

#     serializer = AIRSerializer(data=data)
#     if serializer.is_valid(raise_exception=True):
#         air_instance = serializer.save(client=client)
#         print(request.data)

#         if request.FILES:
#             files = dict((request.FILES).lists()).get('files',None)
#             # files = request.FILES.getlist('files')
#             if files:
#                 for file in files:
#                     file_data = {
#                         'air' : air_instance.pk,
#                         'files' : file
#                     }
#                     file_serializer= FilesSerializer(data=file_data)
#                     if file_serializer.is_valid(raise_exception=True):
#                         file_serializer.save()

#             return Response({
#                 'message': 'AIR created',
#                 'data': serializer.data,
#                 },status=status.HTTP_201_CREATED)
#         return Response({
#                 'message': 'Fail to create AIR',
#                 'error_message': serializer.errors,
#                 },status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def create_air(request, pk):
    client = Client.objects.get(id=pk)
    instance_data = request.data
    data = {key: value for key, value in instance_data.items()}

    # Check if files are present
    files = dict(request.FILES.lists()).get('files', None)
    if not files:
        return Response(
            {'error_message': 'You must upload at least one file.'},
            status=status.HTTP_400_BAD_REQUEST
        )

    serializer = AIRSerializer(data=data)
    if serializer.is_valid(raise_exception=True):
        air_instance = serializer.save(client=client)

        for file in files:
            file_data = {
                'air': air_instance.pk,
                'files': file
            }
            file_serializer = FilesSerializer(data=file_data)
            if file_serializer.is_valid(raise_exception=True):
                file_serializer.save()

        return Response({
            'message': 'AIR created',
            'data': serializer.data,
        }, status=status.HTTP_201_CREATED)

    return Response({
        'message': 'Failed to create AIR',
        'error_message': serializer.errors,
    }, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST', 'GET'])
@parser_classes([MultiPartParser, FormParser])
def edit_air(request, pk, air_pk):
    try:
        client = Client.objects.get(id=pk)
        air = AIR.objects.get(id=air_pk, client=client)
    except Client.DoesNotExist:
        return Response({'error_message': 'Client not found'}, status=status.HTTP_404_NOT_FOUND)
    except AIR.DoesNotExist:
        return Response({'error_message': 'AIR not found'}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'POST':
        # Update branch document fields
        air_serializer = AIRSerializer(instance=air, data=request.data)
        if air_serializer.is_valid():
            air_serializer.save(client=client)

            # Handle file updates separately
            if request.FILES.getlist('files'):
                # Delete the old files if required
                Files.objects.filter(air=air).delete()

                # Add the new files
                files = request.FILES.getlist('files')
                for file in files:
                    file_data = {
                        'air': air.pk,
                        'files': file
                    }
                    file_serializer = FilesSerializer(data=file_data)
                    if file_serializer.is_valid():
                        file_serializer.save()
                    else:
                        return Response(file_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

            return Response({
                'message': 'AIR updated',
                'data': air_serializer.data,
                'status' : status.HTTP_200_OK
            })
        else:
            return Response({
                'message': 'Fail to update AIR',
                'error_message': air_serializer.errors,
                },status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'GET':
        air_serializer = AIRSerializer(air)
        return Response(air_serializer.data)

@api_view(['GET'])
def list_air(request, pk):
    client = Client.objects.get(id = pk)
    if request.method == 'GET':
        list_air = AIR.objects.filter(client=client)
        air_serializer = AIRSerializer(list_air, many=True)
        print(air_serializer)
        return Response(air_serializer.data)

@api_view(['GET'])
def single_air(request, pk, air_pk):
    client = Client.objects.get(id=pk)
    air = AIR.objects.get(id = air_pk, client=client)
    if request.method == 'GET':
        ser = AIRSerializer(air)
        print(ser)
        return Response(ser.data)

@api_view(['DELETE'])
def delete_air(request, pk, air_pk ):
    client = Client.objects.get(id = pk)
    air = AIR.objects.get(id = air_pk, client=client)
    if request.method == 'DELETE':
        air.delete()
        return Response({'message':'AIR Delete'})
    return Response({'error_message':'Fail to delete AIR'} ,status=status.HTTP_400_BAD_REQUEST)

# ******************************************************SFT****************************************************
@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def create_sft(request,pk):
    client = Client.objects.get(id=pk)
    instance_data = request.data
    data = {key: value for key, value in instance_data.items()}

    serializer = SFTSerializer(data=data)
    if serializer.is_valid(raise_exception=True):
        sft_instance = serializer.save(client=client)
        print(request.data)

        if request.FILES:
            files = dict((request.FILES).lists()).get('files',None)
            # files = request.FILES.getlist('files')
            if files:
                for file in files:
                    file_data = {
                        'sft' : sft_instance.pk,
                        'files' : file
                    }
                    file_serializer= FilesSerializer(data=file_data)
                    if file_serializer.is_valid(raise_exception=True):
                        file_serializer.save()

            return Response({
                'message': 'SFT created',
                'data': serializer.data,
                },status=status.HTTP_201_CREATED)
        return Response({
                'message': 'Fail to create SFT',
                'error_message': serializer.errors,
                },status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST', 'GET'])
@parser_classes([MultiPartParser, FormParser])
def edit_sft(request, pk, sft_pk):
    try:
        client = Client.objects.get(id=pk)
        sft = SFT.objects.get(id=sft_pk, client=client)
    except Client.DoesNotExist:
        return Response({'error_message': 'Client not found'}, status=status.HTTP_404_NOT_FOUND)
    except SFT.DoesNotExist:
        return Response({'error_message': 'SFT not found'}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'POST':
        # Update branch document fields
        sft_serializer = SFTSerializer(instance=sft, data=request.data)
        if sft_serializer.is_valid():
            sft_serializer.save(client=client)

            # Handle file updates separately
            if request.FILES.getlist('files'):
                # Delete the old files if required
                Files.objects.filter(sft=sft).delete()

                # Add the new files
                files = request.FILES.getlist('files')
                for file in files:
                    file_data = {
                        'sft': sft.pk,
                        'files': file
                    }
                    file_serializer = FilesSerializer(data=file_data)
                    if file_serializer.is_valid():
                        file_serializer.save()
                    else:
                        return Response(file_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

            return Response({
                'message': 'SFT updated',
                'data': sft_serializer.data,
                'status' : status.HTTP_200_OK
            })
        else:
            return Response({
                'message': 'Fail to update SFT',
                'error_message': sft_serializer.errors,
                },status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'GET':
        sft_serializer = SFTSerializer(sft)
        return Response(sft_serializer.data)

@api_view(['GET'])
def list_sft(request, pk):
    client = Client.objects.get(id = pk)
    if request.method == 'GET':
        list_sft = SFT.objects.filter(client=client)
        sft_serializer = SFTSerializer(list_sft, many=True)
        print(sft_serializer)
        return Response(sft_serializer.data)

@api_view(['GET'])
def single_sft(request, pk, sft_pk):
    client = Client.objects.get(id=pk)
    sft = SFT.objects.get(id = sft_pk, client=client)
    if request.method == 'GET':
        ser = SFTSerializer(sft)
        print(ser)
        return Response(ser.data)

@api_view(['DELETE'])
def delete_sft(request, pk, sft_pk ):
    client = Client.objects.get(id = pk)
    sft = SFT.objects.get(id = sft_pk, client=client)
    if request.method == 'DELETE':
        sft.delete()
        return Response({'message':'SFT Delete'})
    return Response({'error_message':'Fail to delete SFT'} ,status=status.HTTP_400_BAD_REQUEST)

# **************************************************OTHERS***************************************************
@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def create_others(request, pk):
    client = Client.objects.get(id=pk)
    instance_data = request.data
    data = {key: value for key, value in instance_data.items()}

    # Check if files are present
    files = dict(request.FILES.lists()).get('files', None)
    if not files:
        return Response(
            {'error_message': 'You must upload at least one file.'},
            status=status.HTTP_400_BAD_REQUEST
        )

    serializer = OthersSerializer(data=data)
    if serializer.is_valid(raise_exception=True):
        others_instance = serializer.save(client=client)

        for file in files:
            file_data = {
                'others': others_instance.pk,
                'files': file
            }
            file_serializer = FilesSerializer(data=file_data)
            if file_serializer.is_valid(raise_exception=True):
                file_serializer.save()

        return Response({
            'message': 'Others created',
            'data': serializer.data,
        }, status=status.HTTP_201_CREATED)

    return Response({
        'message': 'Failed to create Others',
        'error_message': serializer.errors,
    }, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST', 'GET'])
@parser_classes([MultiPartParser, FormParser])
def edit_others(request, pk, others_pk):
    try:
        client = Client.objects.get(id=pk)
        others = Others.objects.get(id=others_pk, client=client)
    except Client.DoesNotExist:
        return Response({'error_message': 'Client not found'}, status=status.HTTP_404_NOT_FOUND)
    except Others.DoesNotExist:
        return Response({'error_message': 'Others not found'}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'POST':
        # Update branch document fields
        others_serializer = OthersSerializer(instance=others, data=request.data)
        if others_serializer.is_valid():
            others_serializer.save(client=client)

            # Handle file updates separately
            if request.FILES.getlist('files'):
                # Delete the old files if required
                Files.objects.filter(others=others).delete()

                # Add the new files
                files = request.FILES.getlist('files')
                for file in files:
                    file_data = {
                        'others': others.pk,
                        'files': file
                    }
                    file_serializer = FilesSerializer(data=file_data)
                    if file_serializer.is_valid():
                        file_serializer.save()
                    else:
                        return Response(file_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

            return Response({
                'message': 'Others updated',
                'data': others_serializer.data,
                'status' : status.HTTP_200_OK
            })
        else:
            return Response({
                'message': 'Fail to update Others',
                'error_message': others_serializer.errors,
                },status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'GET':
        others_serializer = OthersSerializer(others)
        return Response(others_serializer.data)

@api_view(['GET'])
def list_others(request, pk):
    client = Client.objects.get(id = pk)
    if request.method == 'GET':
        list_others = Others.objects.filter(client=client)
        others_serializer = OthersSerializer(list_others, many=True)
        print(others_serializer)
        return Response(others_serializer.data)

@api_view(['GET'])
def single_others(request, pk, others_pk):
    client = Client.objects.get(id=pk)
    others = Others.objects.get(id = others_pk, client=client)
    if request.method == 'GET':
        ser = OthersSerializer(others)
        print(ser)
        return Response(ser.data)

@api_view(['DELETE'])
def delete_others(request, pk, others_pk ):
    client = Client.objects.get(id = pk)
    others = Others.objects.get(id = others_pk, client=client)
    if request.method == 'DELETE':
        others.delete()
        return Response({'message':'Others Delete'})
    return Response({'error_message':'Fail to delete Others'} ,status=status.HTTP_400_BAD_REQUEST)


# *************************************************TDS Payment***********************************************
@api_view(['POST'])
def create_tdspayment(request, pk):
    client = Client.objects.get(id=pk)
    if request.method == 'POST':
        serializer = TDSPaymentSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(client=client)
            return Response({
                'message': 'TDS Payment created',
                'data': serializer.data,
                },status=status.HTTP_201_CREATED)
        return Response({
                'message': 'Fail to create TDS Payment',
                'error_message': serializer.errors,
                },status=status.HTTP_400_BAD_REQUEST)

class ExcelImportViewtds(APIView):
    parser_classes = [MultiPartParser]

    def post(self, request, pk, *args, **kwargs):
        try:
            client = Client.objects.get(pk=pk)
        except Client.DoesNotExist:
            return Response({"error_message": "Client not found"}, status=status.HTTP_404_NOT_FOUND)

        file = request.FILES.get('file')
        if not file:
            return Response({"error_message": "No file provided"}, status=status.HTTP_400_BAD_REQUEST)

        wb = load_workbook(file)
        ws = wb.active

        fields = [
            'client_name', 'date', 'PAN', 'amount', 
            'cgst', 'sgst', 'igst', 'total_amt', 
            'tds_rate', 'tds_section', 'tds_amount', 
            'net_amount', 'tds_payment_date', 'tds_challan_no' 
        ]

        for row in ws.iter_rows(min_row=2):  # Skip the header row
            data = {field: row[i].value for i, field in enumerate(fields)}

            # employee_code = data.get('employee_code')
            # month = data.get('month')

            # if not employee_code or not month:
            #     continue  # Skip invalid rows

            # # Check if employee_code already exists for the given month
            # if PF.objects.filter(employee_code=employee_code, month=month).exists():
            #     return Response(
            #         {"error_message": f"Employee {employee_code} already exists for month {month}"},
            #         status=status.HTTP_400_BAD_REQUEST
            #     )

            # Create new PF entry
            TDSPayment.objects.create(client=client, **data)

        return Response({"status": "success", "message": "All entries uploaded successfully"}, status=status.HTTP_201_CREATED)

@api_view(['GET', 'POST'])
def edit_tdspayment(request, pk, tdspayment_pk):
    client = Client.objects.get(id=pk)
    payment = TDSPayment.objects.get(id=tdspayment_pk, client=client)
    serializer = TDSPaymentSerializer(instance=payment, data=request.data)
    if request.method == 'GET':
        tds_serializer = TDSPaymentSerializer(payment)
        print(tds_serializer.data)
        return Response (tds_serializer.data)
    elif request.method == 'POST':
        if serializer.is_valid():
            serializer.save(client=client)
            return Response({
                'message': 'TDS Payment updated',
                'data': serializer.data,
                'status' : status.HTTP_200_OK
            })
        return Response({
                'message': 'Fail to update TDS Payment',
                'error_message': serializer.errors,
                },status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
def list_tdspayment(request, pk):
    client = Client.objects.get(id=pk)
    if request.method == 'GET':
        list_tdspayment = TDSPayment.objects.filter(client=client)
        tds_ser = TDSPaymentSerializer(list_tdspayment, many = True)
        return Response(tds_ser.data)

@api_view(['GET'])
def single_tdspayment(request, pk, tdspayment_pk):
    client = Client.objects.get(id=pk)
    tds = TDSPayment.objects.get(id=tdspayment_pk)
    if request.method == 'GET':
        ser = TDSPaymentSerializer(tds)
        return Response(ser.data)

@api_view(['DELETE'])
def delete_tdspayment(request, pk, tdspayment_pk):
    client = Client.objects.get(id=pk)
    tds = TDSPayment.objects.get(id=tdspayment_pk)
    if request.method == 'DELETE':
        tds.delete()
        return Response({'message':'TDS Payment delete'})
    return Response (status=status.HTTP_400_BAD_REQUEST)

# *************************************************TDS Return**************************************************
@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def create_tds(request,pk):
    client = Client.objects.get(id=pk)
    instance_data = request.data
    data = {key: value for key, value in instance_data.items()}

    serializer = TDSReturnSerializer(data=data)
    if serializer.is_valid(raise_exception=True):
        tds_instance = serializer.save(client=client)
        print(request.data)

        if request.FILES:
            files = dict((request.FILES).lists()).get('files',None)
            # files = request.FILES.getlist('files')
            if files:
                for file in files:
                    file_data = {
                        'tds' : tds_instance.pk,
                        'files' : file
                    }
                    file_serializer= FilesSerializer(data=file_data)
                    if file_serializer.is_valid(raise_exception=True):
                        file_serializer.save()

            return Response({
                'message': 'TDS Return created',
                'data': serializer.data,
                },status=status.HTTP_201_CREATED)
        return Response({
                'message': 'Fail to create TDS Return',
                'error_message': serializer.errors,
                },status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST', 'GET'])
@parser_classes([MultiPartParser, FormParser])
def edit_tds(request, pk, tds_pk):
    try:
        client = Client.objects.get(id=pk)
        tds = TDSReturn.objects.get(id=tds_pk, client=client)
    except Client.DoesNotExist:
        return Response({'error_message': 'Client not found'}, status=status.HTTP_404_NOT_FOUND)
    except TDSReturn.DoesNotExist:
        return Response({'error_message': 'TDS Return not found'}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'POST':
        # Update branch document fields
        tds_serializer = TDSReturnSerializer(instance=tds, data=request.data)
        if tds_serializer.is_valid():
            tds_serializer.save(client=client)

            # Handle file updates separately
            if request.FILES.getlist('files'):
                # Delete the old files if required
                Files.objects.filter(tds=tds).delete()

                # Add the new files
                files = request.FILES.getlist('files')
                for file in files:
                    file_data = {
                        'tds': tds.pk,
                        'files': file
                    }
                    file_serializer = FilesSerializer(data=file_data)
                    if file_serializer.is_valid():
                        file_serializer.save()
                    else:
                        return Response(file_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

            return Response({
                'message': 'TDS Return updated',
                'data': tds_serializer.data,
                'status' : status.HTTP_200_OK
            })
        else:
            return Response({
                'message': 'Fail to update TDS Return',
                'error_message': tds_serializer.errors,
                },status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'GET':
        tdsreturn_serializer = TDSReturnSerializer(tds)
        return Response(tdsreturn_serializer.data)

@api_view(['GET'])
def list_tds(request, pk):
    client = Client.objects.get(id = pk)
    if request.method == 'GET':
        list_tds = TDSReturn.objects.filter(client=client)
        tds_serializer = TDSReturnSerializer(list_tds, many=True)
        # print(sft_serializer)
        return Response(tds_serializer.data)

@api_view(['GET'])
def single_tds(request, pk, tds_pk):
    client = Client.objects.get(id=pk)
    tds = TDSReturn.objects.get(id = tds_pk, client=client)
    if request.method == 'GET':
        ser = TDSReturnSerializer(tds)
        print(ser)
        return Response(ser.data)

@api_view(['DELETE'])
def delete_tds(request, pk, tds_pk ):
    client = Client.objects.get(id = pk)
    tds = TDSReturn.objects.get(id = tds_pk, client=client)
    if request.method == 'DELETE':
        tds.delete()
        return Response({'message':'TDS Return Delete'})
    return Response({'error_message':'Fail to delete TDS Return'} ,status=status.HTTP_400_BAD_REQUEST) 
    

# *************************************************TDS Section**************************************************

# @api_view(['POST'])
# def create_tdssection(request, pk):
#     client = Client.objects.get(id=pk)
#     if request.method == 'POST':    
#         serializer = TDSSectionSerializer(data=request.data)
#         if serializer.is_valid():
#             serializer.save(client=client)
#             return Response({
#                 'message': 'TDS Section created',
#                 'data': serializer.data,
#                 },status=status.HTTP_201_CREATED)
#         return Response({
#                 'message': 'Fail to create TDS Section',
#                 'error_message': serializer.errors,
#                 },status=status.HTTP_400_BAD_REQUEST)

# class ExcelImportViewtdssection(APIView):
#     parser_classes = [MultiPartParser]

#     def post(self, request, pk, *args, **kwargs):
#         try:
#             client = Client.objects.get(pk=pk)
#         except Client.DoesNotExist:
#             return Response({"error_message": "Client not found"}, status=status.HTTP_404_NOT_FOUND)

#         file = request.FILES.get('file')
#         if not file:
#             return Response({"error_message": "No file provided"}, status=status.HTTP_400_BAD_REQUEST)

#         wb = load_workbook(file)
#         ws = wb.active

#         fields = [
#             'name'
#         ]

#         for row in ws.iter_rows(min_row=2):  # Skip the header row
#             data = {field: row[i].value for i, field in enumerate(fields)}

#             # employee_code = data.get('employee_code')
#             # month = data.get('month')

#             # if not employee_code or not month:
#             #     continue  # Skip invalid rows

#             # # Check if employee_code already exists for the given month
#             # if PF.objects.filter(employee_code=employee_code, month=month).exists():
#             #     return Response(
#             #         {"error_message": f"Employee {employee_code} already exists for month {month}"},
#             #         status=status.HTTP_400_BAD_REQUEST
#             #     )

#             # Create new PF entry
#             TDSSection.objects.create(client=client, **data)

#         return Response({"status": "success", "message": "All entries uploaded successfully"}, status=status.HTTP_201_CREATED)

# @api_view(['POST', 'GET'])
# def edit_tdssection(request, pk, tdssection_pk):
#     client = Client.objects.get(id=pk)
#     section = TDSSection.objects.get(id=tdssection_pk, client=client)
#     serializer = TDSSectionSerializer(instance=section, data=request.data)
#     if request.method == 'GET':
#         tds_serializer = TDSSectionSerializer(section)
#         print(tds_serializer.data)
#         return Response (tds_serializer.data)
#     elif request.method == 'POST':
#         if serializer.is_valid():
#             serializer.save(client=client)  
#             return Response({
#                 'message': 'TDS Section updated',
#                 'data': serializer.data,
#                 'status' : status.HTTP_200_OK
#             })
#         return Response({
#                 'message': 'Fail to update TDS Section',
#                 'error_message': serializer.errors,
#                 },status=status.HTTP_400_BAD_REQUEST)

# @api_view(['GET'])
# def list_tdssection(request, pk):
#     client = Client.objects.get(id=pk)
#     if request.method == 'GET':
#         list_tdssection = TDSSection.objects.filter(client=client)
#         tds_ser = TDSSectionSerializer(list_tdssection, many = True)
#         return Response(tds_ser.data)
        
# @api_view(['GET'])
# def single_tdssection(request, pk, tdssection_pk):
#     client = Client.objects.get(id=pk)
#     tds = TDSSection.objects.get(id=tdssection_pk)
#     if request.method == 'GET':
#         ser = TDSSectionSerializer(tds)
#         return Response(ser.data)

# @api_view(['DELETE'])
# def delete_tdssection(request, pk, tdssection_pk):
#     client = Client.objects.get(id=pk)
#     tds = TDSSection.objects.get(id=tdssection_pk)
#     if request.method == 'DELETE':
#         tds.delete()
#         return Response({'message':'TDS Section delete'}, status=status.HTTP_200_OK)
#     return Response({'error_message':'Fail to delete TDS Section'} ,status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
def create_tdssection(request):
    # client = Client.objects.get(id=pk)
    if request.method == 'POST':    
        serializer = TDSSectionSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({
                'message': 'TDS Section created',
                'data': serializer.data,
                },status=status.HTTP_201_CREATED)
        return Response({
                'message': 'Fail to create TDS Section',
                'error_message': serializer.errors,
                },status=status.HTTP_400_BAD_REQUEST)

class ExcelImportViewtdssection(APIView):
    parser_classes = [MultiPartParser]

    def post(self, request, *args, **kwargs):
        # try:
        #     # client = Client.objects.get(pk=pk)
        # except Client.DoesNotExist:
        #     return Response({"error_message": "Client not found"}, status=status.HTTP_404_NOT_FOUND)

        file = request.FILES.get('file')
        if not file:
            return Response({"error_message": "No file provided"}, status=status.HTTP_400_BAD_REQUEST)

        wb = load_workbook(file)
        ws = wb.active

        fields = [
            'name'
        ]

        for row in ws.iter_rows(min_row=2):  # Skip the header row
            data = {field: row[i].value for i, field in enumerate(fields)}

            # employee_code = data.get('employee_code')
            # month = data.get('month')

            # if not employee_code or not month:
            #     continue  # Skip invalid rows

            # # Check if employee_code already exists for the given month
            # if PF.objects.filter(employee_code=employee_code, month=month).exists():
            #     return Response(
            #         {"error_message": f"Employee {employee_code} already exists for month {month}"},
            #         status=status.HTTP_400_BAD_REQUEST
            #     )

            # Create new PF entry
            TDSSection.objects.create(**data)

        return Response({"status": "success", "message": "All entries uploaded successfully"}, status=status.HTTP_201_CREATED)

@api_view(['POST', 'GET'])
def edit_tdssection(request, tdssection_pk):
    # client = Client.objects.get(id=pk)
    section = TDSSection.objects.get(id=tdssection_pk)
    serializer = TDSSectionSerializer(instance=section, data=request.data)
    if request.method == 'GET':
        tds_serializer = TDSSectionSerializer(section)
        print(tds_serializer.data)
        return Response (tds_serializer.data)
    elif request.method == 'POST':
        if serializer.is_valid():
            serializer.save()  
            return Response({
                'message': 'TDS Section updated',
                'data': serializer.data,
                'status' : status.HTTP_200_OK
            })
        return Response({
                'message': 'Fail to update TDS Section',
                'error_message': serializer.errors,
                },status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
def list_tdssection(request):
    # client = Client.objects.get(id=pk)
    if request.method == 'GET':
        list_tdssection = TDSSection.objects.filter()
        tds_ser = TDSSectionSerializer(list_tdssection, many = True)
        return Response(tds_ser.data)
        
@api_view(['GET'])
def single_tdssection(request, tdssection_pk):
    # client = Client.objects.get(id=pk)
    tds = TDSSection.objects.get(id=tdssection_pk)
    if request.method == 'GET':
        ser = TDSSectionSerializer(tds)
        return Response(ser.data)

@api_view(['GET'])
def get_create_tdssection(request):
    # try:
    #     client = Client.objects.get(id=pk)
    # except Client.DoesNotExist:
    #     return Response({'error_message':'Client not Found.'}, status=404)

    if request.method == 'GET':
        context = {
            'tds_section':[],
        }
        tds_section = TDSSection.objects.all()
        tds_section_serializer = TDSSectionSerializer(tds_section, many=True)
        context.update({
            'tds_section': tds_section_serializer.data,
        })
        return Response(context)

@api_view(['DELETE'])
def delete_tdssection(request, tdssection_pk):
    # client = Client.objects.get(id=pk)
    tds = TDSSection.objects.get(id=tdssection_pk)
    if request.method == 'DELETE':
        tds.delete()
        return Response({'message':'TDS Section delete'}, status=status.HTTP_200_OK)
    return Response({'error_message':'Fail to delete TDS Section'} ,status=status.HTTP_400_BAD_REQUEST)

# ***********************************************Sales*******************************************

@api_view(['GET', 'POST'])
def create_sales_get(request, pk):
    try:
        client = Client.objects.get(id=pk)
    except Client.DoesNotExist:
        return Response({"error_message": "Client not found."}, status=404)

    if request.method == 'GET':
        # Initialize response context with default values
        context = {
            'serializer_customer': [],
            'serializer': [],
            'product_serializer': [],
            'message': None,
            'hsn': None,
            'location': None
        }

        # Retrieve query parameters
        received_value = request.GET.get('newValue')
        product_id = request.GET.get('productID')
        print('daadasdasdsa',received_value)
        # Process productID if it exists
        if product_id:
            try:
                hsn_cc = Product.objects.get(id=product_id).hsn
                context.update({
                    "message": "Product HSN found",
                    "hsn": HSNSerializer(hsn_cc).data
                })
            except (ValueError, Product.DoesNotExist):
                return Response({"error_message": "Invalid Product ID."}, status=400)

        # Process received_value if it exists
        if received_value:
            try:
                received_value = int(received_value)
                location = OfficeLocation.objects.get(id=received_value)
                branch_gst = location.branch.gst_no
                print("branch val",branch_gst)
                context.update({
                    "message": "Location found",
                    "location": OfficeLocationSerializer(location).data,
                    "branch_gst": branch_gst
                })
            except (ValueError, OfficeLocation.DoesNotExist):
                return Response({"error_message": "Invalid location ID."}, status=400)

        # Add data only if productID and received_value were not provided
        if not product_id and not received_value:
            off = OfficeLocation.objects.filter(branch__client=client)
            customer = Customer.objects.filter(client=client, customer=True)
            product = Product.objects.all()
            branch = Branch.objects.filter(client=client)

            serializer_customer = CustomerVendorSerializer(customer, many=True)
            serializer = OfficeLocationSerializer(off, many=True)
            product_serializer = ProductSerializer(product, many=True)
            branch_serializer = BranchSerailizer(branch, many=True)

            context.update({
                'serializer_customer': serializer_customer.data,
                'serializer': serializer.data,
                'product_serializer': product_serializer.data,
                'branch_serializer': branch_serializer.data
            })

        return Response(context)

@api_view(['POST'])
def create_sale(request, pk):
    client = Client.objects.get(id=pk)

    # Check if 'attach_e_way_bill' is in the request files
    if 'attach_invoice' in request.FILES:
        # Iterate through each file in the 'attach_e_way_bill' field
        for e_way_bill in request.FILES.getlist('attach_invoice'):
            # Prepare data for each file
            sale_data = {
                'attach_invoice': e_way_bill,  # The file being uploaded
                'client': client.id,  # Associate the file with the client
            }

            # Initialize the serializer for each file
            serializer = SalesSerializer2(data=sale_data)

            # Check if the serializer is valid
            if serializer.is_valid():
                # Save each file as a separate object
                serializer.save()
            else:
                return Response({'error_message': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

        # If all files are processed, return success response
        return Response({'message': 'Sales E-way bill(s) uploaded successfully.'}, status=status.HTTP_201_CREATED)
    else:
        return Response({'error_message': 'No files uploaded in the request.'}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
def get_sales_invoice_data(request, client_pk, invoice_pk):
    """
    GET request to fetch sales invoice data along with related product summaries and client info.
    """
    # Fetch the sales invoice with related client information
    sales_invoice = SalesInvoice.objects.filter(client_id=client_pk, id=invoice_pk).first()

    if not sales_invoice:
        return Response({"error_message": "Sales Invoice not found."}, status=status.HTTP_404_NOT_FOUND)

    # Serialize the sales invoice
    sales_invoice_data = SalesSerializer3(sales_invoice).data

    # Prepare product summaries data
    product_summaries = sales_invoice.product_summaries.all()
    product_summary_data = [
        {
            "id": summary.id,
            "hsnCode": summary.hsn.hsn_code,
            "gstRate": summary.hsn.gst_rate,
            "product": summary.product.product_name,
            "description": summary.prod_description.description,
            "unit": summary.prod_description.unit,
            "rate": summary.prod_description.rate,
            "product_amount": summary.prod_description.product_amount,
            "cgst": summary.prod_description.cgst,
            "sgst": summary.prod_description.sgst,
            "igst": summary.prod_description.igst,
        }
        for summary in product_summaries
    ]

    # Helper function to safely get attributes from related objects
    def get_safe_attr(obj, attr, default=None):
        return getattr(obj, attr, default) if obj else default

    # Prepare client location data (with safe attribute access)
    client_location = sales_invoice.client_Location
    client_location_data = {
        "id": get_safe_attr(client_location, 'id'),
        "location": get_safe_attr(client_location, 'location'),
        "contact": get_safe_attr(client_location, 'contact'),
        "address": get_safe_attr(client_location, 'address'),
        "city": get_safe_attr(client_location, 'city'),
        "state": get_safe_attr(client_location, 'state'),
        "country": get_safe_attr(client_location, 'country'),
    }

    # Prepare customer data (with safe attribute access)
    customer = sales_invoice.customer
    customer_data = {
        "id": get_safe_attr(customer, 'id'),
        "name": get_safe_attr(customer, 'name'),
        "gst_no": get_safe_attr(customer, 'gst_no'),
        "pan": get_safe_attr(customer, 'pan'),
        "customer_address": get_safe_attr(customer, 'address'),
        "customer": get_safe_attr(customer, 'customer'),
        "vendor": get_safe_attr(customer, 'vendor'),
        "email": get_safe_attr(customer, 'email'),     #nnnnn
        "contact": get_safe_attr(customer, 'contact'),  #nnnnn
    }

    # Prepare final response data
    response_data = {
        "sales_invoice": sales_invoice_data,
        "product_summaries": product_summary_data,
        "client_location": client_location_data,
        "customer": customer_data,
    }

    return Response(response_data, status=status.HTTP_200_OK)

@api_view(['GET', 'PUT'])
def update_sales_invoice(request, client_pk, invoice_pk):

    try:
        # Handle GET request
        if request.method == 'GET':
            sales_invoice = SalesInvoice.objects.filter(client_id=client_pk, id=invoice_pk).first()
            if not sales_invoice:
                return Response({"error_message": "Sales Invoice not found."}, status=status.HTTP_404_NOT_FOUND)

            sales_invoice_data = SalesSerializer3(sales_invoice).data
            
            print('sales',sales_invoice_data)

            product_summaries = sales_invoice.product_summaries.all()
            product_summary_data = [
                {
                    "id": summary.id,
                    "hsnCode": summary.hsn.hsn_code,
                    "gstRate": summary.hsn.gst_rate,
                    "product": summary.product.product_name,
                    "description": summary.prod_description.description,
                    "unit": summary.prod_description.unit,
                    "rate": summary.prod_description.rate,
                    "product_amount": summary.prod_description.product_amount,
                    "cgst": summary.prod_description.cgst,
                    "sgst": summary.prod_description.sgst,
                    "igst": summary.prod_description.igst,
                }
                for summary in product_summaries
            ]

            if sales_invoice.po_date:
                formatted_po = sales_invoice.po_date.strftime("%d-%m-%Y")  # Directly format date
            else:
                formatted_po = None

            if sales_invoice.invoice_date:
                formatted_invoice_date = sales_invoice.invoice_date.strftime("%d-%m-%Y")  # Directly format date
            else:
                formatted_invoice_date = None

            response_data = {
                "sales_invoice": {
                    **sales_invoice_data,
                    "po_date" : formatted_po,
                    "invoice_date" : formatted_invoice_date
                    
                    },
                "product_summaries": product_summary_data,
                "client_location": {
                    "id": sales_invoice.client_Location.id if sales_invoice.client_Location else None,
                    "location": sales_invoice.client_Location.location if sales_invoice.client_Location else None,
                    "contact": sales_invoice.client_Location.contact if sales_invoice.client_Location else None,
                    "address": sales_invoice.client_Location.address if sales_invoice.client_Location else None,
                    "city": sales_invoice.client_Location.city if sales_invoice.client_Location else None,
                    "state": sales_invoice.client_Location.state if sales_invoice.client_Location else None,
                    "country": sales_invoice.client_Location.country if sales_invoice.client_Location else None,
                    "branchID": sales_invoice.client_Location.branch.id if sales_invoice.client_Location else None,
                },
                "customer": {
                    "id": sales_invoice.customer.id if sales_invoice.customer else None,
                    "name": sales_invoice.customer.name if sales_invoice.customer else None,
                    "gst_no": sales_invoice.customer.gst_no if sales_invoice.customer else None,
                    "pan": sales_invoice.customer.pan if sales_invoice.customer else None,
                    "customer_address": sales_invoice.customer.address if sales_invoice.customer else None,
                    "customer": sales_invoice.customer.customer if sales_invoice.customer else None,
                    "vendor": sales_invoice.customer.vendor if sales_invoice.customer else None,
                    "email": sales_invoice.customer.email if sales_invoice.customer else None,  #nnnnn
                    "contact": sales_invoice.customer.contact if sales_invoice.customer else None,   #nnnnn
                },
            }
            print('bbbbbbbbbbbbbbbbbbbbbbb',response_data["sales_invoice"])
            return Response(response_data, status=status.HTTP_200_OK)

        # Handle PUT request
        elif request.method == 'PUT':
            print('youuuuuuuuu',request.FILES)
            sales_invoice = SalesInvoice.objects.filter(client_id=client_pk, id=invoice_pk).first()
            if not sales_invoice:
                return Response({"error_message": "Sales Invoice not found."}, status=status.HTTP_404_NOT_FOUND)

            # Extract data from the request
            # form_data = request.data.get('formData', {})
            # vendor_data = request.data.get('vendorData', {})
            # rows = request.data.get('rows', [])
            payload = request.data

            # Container to hold rows
            rows_data = defaultdict(dict)

            # Iterate over keys dynamically
            for key, value in payload.items():
                if key.startswith("rows["):  # Check if the key corresponds to rows
                    # Extract the row index and field name
                    row_index = key.split('[')[1].split(']')[0]  # Get the index (e.g., '0', '1', '2')
                    field_name = key.split('[')[2].split(']')[0]  # Get the field name (e.g., 'product')

                    # Store the value in the corresponding row and field
                    rows_data[int(row_index)][field_name] = value

            # Convert defaultdict to a regular list for easier use
            rows = [rows_data[index] for index in sorted(rows_data.keys())]

            # Output the rows data
            print(rows)

            invoice_data = request.data.get('invoiceData', [{}])[0]  # Extract the first item
            invoice_file = request.data.get('invoice_file')
            print('invoice_file',invoice_data)
            form_data = {
                "offLocID": request.data.get("formData[offLocID]"),
                "location": request.data.get("formData[location]"),
                "contact": request.data.get("formData[contact]"),
                "address": request.data.get("formData[address]"),
                "city": request.data.get("formData[city]"),
                "state": request.data.get("formData[state]"),
                "country": request.data.get("formData[country]"),
                "branchID": request.data.get("formData[branchID]"),
            }
            vendor_data = {
                "name": request.data.get("vendorData[name]"),
                "gst_no": request.data.get("vendorData[gst_no]"),
                "pan": request.data.get("vendorData[pan]"),
                "customer_address": request.data.get("vendorData[customer_address]"),
                "email": request.data.get("vendorData[email]"),  #nnnnn
                "contact": request.data.get("vendorData[contact]"),  #nnnnn
                "customer": request.data.get("vendorData[customer]").lower() == "true" if request.data.get("vendorData[customer]") else None,
                "vendor": request.data.get("vendorData[vendor]").lower() == "true" if request.data.get("vendorData[vendor]") else None,
            }

            # Access invoice_data in a similar way
            invoice_data = {
                "invoice_no": request.data.get("invoiceData[0][invoice_no]"),
                "invoice_date": request.data.get("invoiceData[0][invoice_date]"),
                "po_no": request.data.get("invoiceData[0][po_no]"),
                "po_date": request.data.get("invoiceData[0][po_date]"),
                # "month": request.data.get("invoiceData[0][month]"),
                "invoice_type": request.data.get("invoiceData[0][invoice_type]"),
                "entry_type": request.data.get("invoiceData[0][entry_type]"),
                "taxable_amount": request.data.get("invoiceData[0][taxable_amount]"),
                "totalall_gst": request.data.get("invoiceData[0][totalall_gst]"),
                "total_invoice_value": request.data.get("invoiceData[0][total_invoice_value]"),
                "tds_tcs_rate": request.data.get("invoiceData[0][tds_tcs_rate]"),
                "tcs": request.data.get("invoiceData[0][tcs]"),
                "tds": request.data.get("invoiceData[0][tds]"),
                "amount_receivable": request.data.get("invoiceData[0][amount_receivable]"),
                "attach_invoice": request.data.get("invoiceData[0][attach_invoice]"),
                "attach_e_way_bill": request.data.get("invoiceData[0][attach_e_way_bill]"),
            }

            if invoice_data["invoice_date"]:
                try:
                    invoice_data["invoice_date"] = datetime.strptime(invoice_data["invoice_date"], "%d-%m-%Y").date()
                except ValueError:
                    return Response({"error_message": "Invalid Invoice date format, expected DD-MM-YYYY."}, status=status.HTTP_400_BAD_REQUEST)
            print("Final invoice date:", invoice_data["invoice_date"])

            if invoice_data["po_date"]:
                try:
                    invoice_data["po_date"] = datetime.strptime(invoice_data["po_date"], "%d-%m-%Y").date()
                except ValueError:
                    return Response({"error_message": "Invalid PO date format, expected DD-MM-YYYY."},status=status.HTTP_400_BAD_REQUEST)


            attach_invoice = request.FILES.get("invoiceData[0][attach_invoice]")
            attach_e_way_bill = request.FILES.get("invoiceData[0][attach_e_way_bill]")
                        # Update the sales_invoice instance fields
            if attach_invoice:
                sales_invoice.attach_invoice = attach_invoice

            if attach_e_way_bill:
                sales_invoice.attach_e_way_bill = attach_e_way_bill

            for field, value in invoice_data.items():
                if field not in ['attach_invoice', 'attach_e_way_bill']:  # Skip file fields
                    if hasattr(sales_invoice, field):
                        setattr(
                            sales_invoice,
                            field,
                            safe_decimal(value) if field in [
                                'taxable_amount', 'totalall_gst', 'total_invoice_value',
                                'tds_tcs_rate', 'tds', 'tcs', 'amount_receivable'
                            ] else value
                        )


            attach_invoice = invoice_data.get('attach_invoice')
            print('uoiuoiuiuiuiuio',attach_invoice)
             # Log the flattened data
            # print("Flattened form_data:", form_data)
            print("Flattened invoice_data:", invoice_data)

            print('request payload',request.data)
            # print('form_data',form_data)
            # Update invoice file
            if invoice_file:
                sales_invoice.invoice_file = invoice_file
                sales_invoice.save()
                return Response({"message": "Invoice file uploaded successfully."}, status=status.HTTP_200_OK)

            if vendor_data:
                # Check for 'customer_address' in vendor data and map it to 'address'
                if 'customer_address' in vendor_data:
                    vendor_data['address'] = vendor_data.pop('customer_address')  # Replace 'customer_address' with 'address'

                vendor_id = request.data.get("vendorData[vendorID]")  # Retrieve vendorID if provided

                if vendor_id:
                    vendor_obj = Customer.objects.filter(id=vendor_id).first()
                    if vendor_obj:
                        if vendor_obj.gst_no == vendor_data.get("gst_no"):
                            vendor_serializer = CustomerVendorSerializer(vendor_obj, data=vendor_data, partial=True)
                            if vendor_serializer.is_valid():
                                vendor_serializer.save()
                            else:
                                return Response({"vendor_errors": vendor_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
                        else:
                            # Create a new vendor if gst_no is changed
                            vendor_serializer = CustomerVendorSerializer(data=vendor_data)
                            if vendor_serializer.is_valid():
                                vendor_obj = vendor_serializer.save(client=sales_invoice.client)
                            else:
                                return Response({"vendor_errors": vendor_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
                    else:
                        return Response({"error_message": f"Vendor with ID {vendor_id} not found."}, status=status.HTTP_404_NOT_FOUND)
                else:
                    existing_vendor = Customer.objects.filter(client=sales_invoice.client, gst_no=vendor_data.get("gst_no")).first()
                    if existing_vendor:
                        vendor_obj = existing_vendor
                        vendor_serializer = CustomerVendorSerializer(vendor_obj, data=vendor_data, partial=True)
                        if vendor_serializer.is_valid():
                            vendor_serializer.save()
                        else:
                            return Response({"vendor_errors": vendor_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
                    else:
                        vendor_serializer = CustomerVendorSerializer(data=vendor_data)
                        if vendor_serializer.is_valid():
                            vendor_obj = vendor_serializer.save(client=sales_invoice.client)
                        else:
                            return Response({"vendor_errors": vendor_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
                
                # Assign vendor_obj to expenses.vendor if defined
                if vendor_obj:
                    sales_invoice.customer = vendor_obj


            # Process rows for product summaries
            product_summaries = []
            # processed_product_names = set()
            for row in rows:
                hsn_code = row.get('hsnCode')
                gst_rate = safe_decimal(row.get('gstRate', '0'))
                product_name = row.get('product')
                description_text = row.get('description')
                unit_value = safe_decimal(row.get('unit', '0'))
                rate_value = safe_decimal(row.get('rate', '0'))
                amount = safe_decimal(row.get('product_amount', '0'))
                cgst = safe_decimal(row.get('cgst', '0'))
                sgst = safe_decimal(row.get('sgst', '0'))
                igst = safe_decimal(row.get('igst', '0'))

                # Update or create HSNCode
                hsn_code_obj, _ = HSNCode.objects.update_or_create(
                    hsn_code=hsn_code,
                    defaults={'gst_rate': gst_rate}
                )

                existing_product = Product.objects.filter(product_name=product_name).exclude(hsn=hsn_code_obj).first()
                if existing_product:
                    return Response(
                        {"error_message": f"Product name '{product_name}' already exists with a different HSN code."},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                # Update or create Product
                product_obj, _ = Product.objects.update_or_create(
                    product_name=product_name,
                    hsn=hsn_code_obj,
                    defaults={'unit_of_measure': unit_value}
                )

                # Update or create ProductDescription
                product_description_obj, _ = ProductDescription.objects.update_or_create(
                    product=product_obj,
                    description=description_text,
                    defaults={
                        'unit': unit_value,
                        'rate': rate_value,
                        'product_amount': amount,
                        'cgst': cgst,
                        'sgst': sgst,
                        'igst': igst
                    }      
                )

                # Update or create ProductSummary
                product_summary, _ = ProductSummary.objects.update_or_create(
                    hsn=hsn_code_obj,
                    product=product_obj,
                    prod_description=product_description_obj
                )    
                product_summaries.append(product_summary)

            sales_invoice.product_summaries.set(product_summaries)
            # sales_invoice.save()
            location_data = form_data.get('location')  # Location name entered by user
            location_id = form_data.get('offLocID')   # Existing location ID, if provided
            branch_id = form_data.get('branchID')     # Branch ID selected for new location

            if location_id:  # Update existing location
                # Fetch the existing location
                location_obj = OfficeLocation.objects.filter(id=location_id).first()
                if not location_obj:
                    return Response({"error_message": "Office Location not found."}, status=status.HTTP_404_NOT_FOUND)

                # Update the location details
                location_obj.location = location_data
                location_obj.contact = form_data.get('contact')
                location_obj.address = form_data.get('address')
                location_obj.city = form_data.get('city')
                location_obj.state = form_data.get('state')
                location_obj.country = form_data.get('country')
                location_obj.save()

            else:  # Create a new location
                # Validate branch selection
                if not branch_id:
                    return Response({"error_message": "Branch ID is required for creating a new location."}, status=status.HTTP_400_BAD_REQUEST)

                branch_instance = Branch.objects.filter(id=branch_id, client_id=sales_invoice.client.id).first()
                if not branch_instance:
                    return Response({"error_message": f"Branch with ID {branch_id} not found or doesn't belong to the client."},
                                    status=status.HTTP_404_NOT_FOUND)

                # Create the new location
                location_obj, _ = OfficeLocation.objects.get_or_create(
                    location=location_data,
                    contact=form_data.get('contact'),
                    address=form_data.get('address'),
                    city=form_data.get('city'),
                    state=form_data.get('state'),
                    country=form_data.get('country'),
                    branch=branch_instance  # Associate with the selected branch
                )

            # Associate the updated or newly created location with the sales invoice
            sales_invoice.client_Location = location_obj
            sales_invoice.save()

            response_data = {
                'message': 'Sales Invoice updated successfully.',
                'sales_invoice_data': SalesSerializer(sales_invoice).data,
                'product_summaries': [{'id': summary.id, 'product_name': summary.product.product_name} for summary in product_summaries]
            }
            return Response(response_data, status=status.HTTP_200_OK)

    except Exception as e:
        error_details = traceback.format_exc()
        # print({"Error in update_sales_invoice" : str(e)})
        return Response({"error_message":str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
def create_sales_invoice2(request, client_pk):
    try:
        with transaction.atomic():  # Start a transaction
            payload = request.data
            print('payload', payload)

            # Extract rows dynamically
            rows_data = defaultdict(dict)
            for key, value in payload.items():
                if key.startswith("rows["):  # Check if the key corresponds to rows
                    row_index = key.split('[')[1].split(']')[0]
                    field_name = key.split('[')[2].split(']')[0]
                    rows_data[int(row_index)][field_name] = value
            rows = [rows_data[index] for index in sorted(rows_data.keys())]
            print('Rows:', rows)

            # Extract form data, vendor data, and invoice data
            form_data = {
                "offLocID": payload.get("formData[offLocID]"),
                "location": payload.get("formData[location]"),
                "contact": payload.get("formData[contact]"),
                "address": payload.get("formData[address]"),
                "city": payload.get("formData[city]"),
                "state": payload.get("formData[state]"),
                "country": payload.get("formData[country]"),
                "branchID": payload.get("formData[branchID]"),
            }
            vendor_data = {
                "name": payload.get("vendorData[name]"),
                "gst_no": payload.get("vendorData[gst_no]"),
                "pan": payload.get("vendorData[pan]"),
                "address": payload.get("vendorData[address]"),
                "email": payload.get("vendorData[email]"), #nnnnn
                "contact": payload.get("vendorData[contact]"), #nnnnn
                "customer": payload.get("vendorData[customer]", "").lower() == "true",
                "vendor": payload.get("vendorData[vendor]", "").lower() == "true",
            }
            invoice_data = {
                "invoice_no": payload.get("invoiceData[0][invoice_no]"),
                "invoice_date": payload.get("invoiceData[0][invoice_date]"),
                "po_no": payload.get("invoiceData[0][po_no]"),
                "po_date": payload.get("invoiceData[0][po_date]"),
                # "month": payload.get("invoiceData[0][month]"),
                "invoice_type": payload.get("invoiceData[0][invoice_type]"),
                "entry_type": payload.get("invoiceData[0][entry_type]"),
                "taxable_amount": payload.get("invoiceData[0][taxable_amount]"),
                "totalall_gst": payload.get("invoiceData[0][totalall_gst]"),
                "total_invoice_value": payload.get("invoiceData[0][total_invoice_value]"),
                "tds_tcs_rate": payload.get("invoiceData[0][tds_tcs_rate]"),
                "tcs": payload.get("invoiceData[0][tcs]"),
                "tds": payload.get("invoiceData[0][tds]"),
                "amount_receivable": payload.get("invoiceData[0][amount_receivable]"),
            }
            if invoice_data["invoice_date"]:
                try:
                    invoice_data["invoice_date"] = datetime.strptime(invoice_data["invoice_date"], "%d-%m-%Y").date()
                except ValueError:
                    return Response({"error_message": "Invalid date format, expected DD-MM-YYYY."}, status=status.HTTP_400_BAD_REQUEST)
            print("Final invoice date:", invoice_data["invoice_date"])

            if invoice_data["po_date"]:
                try:
                    invoice_data["po_date"] = datetime.strptime(invoice_data["po_date"], "%d-%m-%Y").date()
                except ValueError:
                    return Response({"error_message": "Invalid PO date format, expected DD-MM-YYYY."},status=status.HTTP_400_BAD_REQUEST)

            print("Final month:", invoice_data["po_date"])


            attach_invoice = request.FILES.get("invoiceData[0][attach_invoice]")
            attach_e_way_bill = request.FILES.get("invoiceData[0][attach_e_way_bill]")

            # Handle Vendor creation or update
            vendor_obj = None
            if vendor_data.get("gst_no"):
                existing_vendor = Customer.objects.filter(client_id=client_pk, gst_no=vendor_data["gst_no"]).first()
                if existing_vendor:
                    vendor_serializer = CustomerVendorSerializer(existing_vendor, data=vendor_data, partial=True)
                    if vendor_serializer.is_valid():
                        vendor_obj = vendor_serializer.save()
                    else:
                        return Response({"vendor_errors": vendor_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
                else:
                    vendor_serializer = CustomerVendorSerializer(data=vendor_data)
                    if vendor_serializer.is_valid():
                        vendor_obj = vendor_serializer.save(client_id=client_pk)
                    else:
                        return Response({"vendor_errors": vendor_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

            # Validate all rows and handle Product Summaries
            product_summaries = []
            for row in rows:
                hsn_code = row.get('hsnCode')
                gst_rate = safe_decimal(row.get('gstRate', '0'))
                product_name = row.get('product')
                product_id = row.get('product_id')
                description_text = row.get('description', '')
                unit_value = safe_decimal(row.get('unit', '0'))
                rate_value = safe_decimal(row.get('rate', '0'))
                amount = safe_decimal(row.get('product_amount', '0'))
                cgst = safe_decimal(row.get('cgst', '0'))
                sgst = safe_decimal(row.get('sgst', '0'))
                igst = safe_decimal(row.get('igst', '0'))

                hsn_code_obj, _ = HSNCode.objects.get_or_create(
                    hsn_code=hsn_code, defaults={'gst_rate': gst_rate}
                )

                existing_product = Product.objects.filter(product_name=product_name).exclude(hsn=hsn_code_obj).first()
                if existing_product:
                    return Response(
                        {"error_message": f"Product name '{product_name}' already exists with a different HSN code."},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                if product_id:
                    product_obj = Product.objects.filter(id=product_id).first()
                    if not product_obj:
                        return Response({"error_message": f"Product with ID {product_id} not found."}, status=status.HTTP_404_NOT_FOUND)
                else:
                    product_obj, _ = Product.objects.get_or_create(
                        product_name=product_name, hsn=hsn_code_obj, defaults={'unit_of_measure': unit_value}
                    )

                product_description_obj, _ = ProductDescription.objects.get_or_create(
                    product=product_obj,
                    description=description_text,
                    defaults={
                        'unit': unit_value,
                        'rate': rate_value,
                        'product_amount': amount,
                        'cgst': cgst,
                        'sgst': sgst,
                        'igst': igst,
                    }
                )

                product_summary = ProductSummary(
                    hsn=hsn_code_obj,
                    product=product_obj,
                    prod_description=product_description_obj
                )
                product_summaries.append(product_summary)

            # Handle Office Location
            location_obj = None
            if form_data["offLocID"]:
                location_obj = OfficeLocation.objects.filter(id=form_data["offLocID"]).first()
                if not location_obj:
                    return Response({"error_message": "Office Location not found."}, status=status.HTTP_404_NOT_FOUND)
            else:
                branch_instance = Branch.objects.filter(id=form_data["branchID"], client_id=client_pk).first()
                if not branch_instance:
                    return Response({"error_message": f"Branch with ID {form_data['branchID']} not found or doesn't belong to the client."},
                                    status=status.HTTP_404_NOT_FOUND)
                location_obj, _ = OfficeLocation.objects.get_or_create(
                    location=form_data.get("location"),
                    contact=form_data.get("contact"),
                    address=form_data.get("address"),
                    city=form_data.get("city"),
                    state=form_data.get("state"),
                    country=form_data.get("country"),
                    branch=branch_instance
                )

            # Finally, create the Sales Invoice
            sales_invoice = SalesInvoice.objects.create(
                client_id=client_pk,
                customer=vendor_obj,
                attach_invoice=attach_invoice,
                attach_e_way_bill=attach_e_way_bill,
                client_Location=location_obj,
                **invoice_data
            )

            for product_summary in product_summaries:
                product_summary.save()
                sales_invoice.product_summaries.add(product_summary)

            print('888888888888',payload)

            return Response({"message": "Sales Invoice created successfully.", "invoice_id": sales_invoice.id}, status=status.HTTP_200_OK)

    except Exception as e:
        error_details = traceback.format_exc()
        return Response({"error_message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
   
@api_view(['DELETE'])
def delete_sales_invoice(request, client_pk, pk):
    """
    Deletes a SalesInvoice by its primary key (ID) along with its associated product summaries.
    """
    try:
        # Retrieve the Client instance
        client = Client.objects.filter(id=client_pk).first()

        if not client:
            return Response({"error_message": "Client not found."}, status=status.HTTP_404_NOT_FOUND)

        # Retrieve the SalesInvoice instance
        sales_invoice = SalesInvoice.objects.filter(id=pk, client=client).first()

        if not sales_invoice:
            return Response({"error_message": "Sales Invoice not found."}, status=status.HTTP_404_NOT_FOUND)

        # Handle deletion of associated ProductSummary instances
        product_summaries = sales_invoice.product_summaries.all()
        for product_summary in product_summaries:
            product_summary.delete()

        # Delete the SalesInvoice instance
        sales_invoice.delete()

        return Response({"message": "Sales Invoice deleted successfully."}, status=status.HTTP_200_OK)

    except Exception as e:
        error_details = traceback.format_exc()
        # print({"Error in update_sales_invoice" : str(e)})
        return Response({"error_message":str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET', 'PATCH'])
def sales_invoice_detail_view(request, client_pk, invoice_pk):
    try:
        # Fetch the sales invoice object
        sales_invoice = SalesInvoice.objects.get(client=client_pk, pk=invoice_pk)
    except SalesInvoice.DoesNotExist:
        return Response({"error_message": "Sales invoice not found."}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        # Serialize and return the sales invoice details
        serializer = SalesSerializerList(sales_invoice)
        return Response(serializer.data, status=status.HTTP_200_OK)

    elif request.method == 'PATCH':
        # Update only specific fields of the sales invoice
        data = request.data

        # Update fields from the request payload
        sales_invoice.invoice_no = data.get('invoice_no', sales_invoice.invoice_no)
        sales_invoice.invoice_date = data.get('invoice_date', sales_invoice.invoice_date)
        sales_invoice.po_no = data.get('po_no', sales_invoice.po_no)
        sales_invoice.po_date = data.get('po_date', sales_invoice.po_date)
        # sales_invoice.month = data.get('month', sales_invoice.month)
        sales_invoice.invoice_type = data.get('invoice_type', sales_invoice.invoice_type)
        sales_invoice.entry_type = data.get('entry_type', sales_invoice.entry_type)
        sales_invoice.taxable_amount = safe_decimal(data.get('taxable_amount', sales_invoice.taxable_amount))
        sales_invoice.total_gst = safe_decimal(data.get('totalall_gst', sales_invoice.total_gst))
        sales_invoice.total_invoice_value = safe_decimal(data.get('total_invoice_value', sales_invoice.total_invoice_value))
        sales_invoice.tds_tcs_rate = safe_decimal(data.get('tds_tcs_rate', sales_invoice.tds_tcs_rate))
        sales_invoice.tds = safe_decimal(data.get('tds', sales_invoice.tds))
        sales_invoice.tcs = safe_decimal(data.get('tcs', sales_invoice.tcs))
        sales_invoice.amount_receivable = safe_decimal(data.get('amount_receivable', sales_invoice.amount_receivable))

        # Update product summaries if provided
        rows = data.get('rows', [])
        product_summaries = []
        for row in rows:
            hsn_code = row.get('hsnCode')
            gst_rate = safe_decimal(row.get('gstRate', '0'))
            product_name = row.get('product')
            description_text = row.get('description')
            unit_value = safe_decimal(row.get('unit', '0'))
            rate_value = safe_decimal(row.get('rate', '0'))
            amount = safe_decimal(row.get('product_amount', '0'))
            cgst = safe_decimal(row.get('cgst', '0'))
            sgst = safe_decimal(row.get('sgst', '0'))
            igst = safe_decimal(row.get('igst', '0'))

            # Handle HSNCode
            hsn_code_obj, created = HSNCode.objects.get_or_create(
                hsn_code=hsn_code,
                defaults={'gst_rate': gst_rate}
            )
            if not created and hsn_code_obj.gst_rate != gst_rate:
                hsn_code_obj.gst_rate = gst_rate
                hsn_code_obj.save()

            # Handle Product
            product_obj = Product.objects.filter(product_name=product_name, hsn=hsn_code_obj).first()
            if not product_obj:
                product_obj = Product.objects.create(
                    product_name=product_name,
                    hsn=hsn_code_obj,
                    unit_of_measure=unit_value
                )

            # Handle ProductDescription
            product_description_obj, created = ProductDescription.objects.get_or_create(
                product=product_obj,
                description=description_text,
                defaults={
                    'unit': unit_value,
                    'rate': rate_value,
                    'product_amount': amount,
                    'cgst': cgst,
                    'sgst': sgst,
                    'igst': igst
                }
            )
            if not created:
                # Update fields in ProductDescription
                product_description_obj.unit = unit_value
                product_description_obj.rate = rate_value
                product_description_obj.product_amount = amount
                product_description_obj.cgst = cgst
                product_description_obj.sgst = sgst
                product_description_obj.igst = igst
                product_description_obj.save()

            # Create ProductSummary
            product_summary = ProductSummary.objects.create(
                hsn=hsn_code_obj,
                product=product_obj,
                prod_description=product_description_obj
            )
            product_summaries.append(product_summary)

        # Save updates to the sales invoice
        if product_summaries:
            sales_invoice.product_summaries.set(product_summaries)
        sales_invoice.save()

        # Return the updated object
        updated_serializer = SalesSerializer(sales_invoice)
        return Response(updated_serializer.data, status=status.HTTP_200_OK)


# **************************************************Purchase***************************************************
@api_view(['GET','POST'])
def create_purchase_get(request, pk):
    try:
        client = Client.objects.get(id=pk)
    except Client.DoesNotExist:
        return Response({'error_message':'Client not Found.'}, status=404)

    if request.method == 'GET':
        context = {
            'serializer_customer':[],
            'serializer': [],
            'product_serializer':[],
            'message':None,
            'hsn':None,
            'location':None,
        }
        received_value = request.GET.get('newValue')
        product_Id = request.GET.get('productID')
        if product_Id:
            try:
                hsn_cc = Product.objects.get(id=product_Id).hsn
                context.update({
                    'message':'Product HSN found',
                    'hsn':HSNSerializer(hsn_cc).data
                })
            except (ValueError, Product.DoesNotExist):
                return Response({'error_message':'Invalid Product ID'}, status=400)

        if received_value:
            try:
                received_value = int(received_value)
                location = OfficeLocation.objects.get(id=received_value)
                branch_gst = location.branch.gst_no
                print('branch val', branch_gst)
                context.update({
                    "message":"Location Found",
                    "location" : OfficeLocationSerializer(location).data,
                    "branch_gst": branch_gst
                })
            except (ValueError, OfficeLocation.DoesNotExist):
                return Response({'error_message':'Invalid location ID.'}, status=400)

        if not product_Id and not received_value:
            off = OfficeLocation.objects.filter(branch__client=client)
            customer = Customer.objects.filter(client=client, vendor=True)
            product = Product.objects.all()
            branch = Branch.objects.filter(client=client)

            serializer_customer = CustomerVendorSerializer(customer, many=True)
            serializers = OfficeLocationSerializer(off, many=True)
            product_serializer = ProductSerializer(product, many=True)
            branch_serializer = BranchSerailizer(branch, many= True)

            context.update({
                'serializer_customer' : serializer_customer.data,
                'serializer' : serializers.data,
                'product_serializer' : product_serializer.data,
                'branch_serializer' : branch_serializer.data
            })
        return Response(context)
    
@api_view(['POST'])
def create_purchase(request, pk):
    client = Client.objects.get(id=pk)

    # Check if 'attach_e_way_bill' is in the request files
    if 'attach_invoice' in request.FILES:
        # Iterate through each file in the 'attach_invoice' field
        for e_way_bill in request.FILES.getlist('attach_invoice'):
            # Prepare data for each file
            purchase_data = {
                'attach_invoice': e_way_bill,  # The file being uploaded
                'client': client.id,  # Associate the file with the client
            }

            # Initialize the serializer for each file
            serializer = PurchaseSerializer2(data=purchase_data)

            # Check if the serializer is valid
            if serializer.is_valid():
                # Save each file as a separate object
                serializer.save()
            else:
                return Response({'error_message': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

        # If all files are processed, return success response
        return Response({'message': 'Purchase E-way bill(s) uploaded successfully.'}, status=status.HTTP_201_CREATED)
    else:
        return Response({'error_message': 'No files uploaded in the request.'}, status=status.HTTP_400_BAD_REQUEST)
    
@api_view(['GET'])
def get_purchase_invoice_data(request, client_pk, invoice_pk):
    purchase_invoice = PurchaseInvoice.objects.filter(client_id=client_pk, id=invoice_pk).first()
    
    if not purchase_invoice:
        return Response({'error_message':'Purchase Invoice not found'}, status=status.HTTP_400_BAD_REQUEST)
    
    purchase_invoice_data = PurchaseSerializer3(purchase_invoice).data
    
    product_summaries = purchase_invoice.product_summaries.all()
    product_summary_data = [
        {
            "id": summary.id,
            "hsnCode": summary.hsn.hsn_code,
            "gstRate": summary.hsn.gst_rate,
            "product": summary.product.product_name,
            "description": summary.prod_description.description,
            "unit": summary.prod_description.unit,
            "rate": summary.prod_description.rate,
            "product_amount": summary.prod_description.product_amount,
            "cgst": summary.prod_description.cgst,
            "sgst": summary.prod_description.sgst,
            "igst": summary.prod_description.igst,
        }
        for summary in product_summaries
    ]
    
    def get_safe_attr(obj, attr, default=None):
        return getattr(obj, attr, default) if obj else default
    
    client_location = purchase_invoice.client_Location
    client_location_data = {
        'id' : get_safe_attr(client_location, 'id'),
        'location' : get_safe_attr(client_location, 'location'),
        'contact' : get_safe_attr(client_location,'contact'),
        'address' : get_safe_attr(client_location, 'address'),
        'city' : get_safe_attr(client_location,'city'),
        'state' : get_safe_attr(client_location, 'state'),
        'country' : get_safe_attr(client_location, 'country'),
        
    }
    vendor = purchase_invoice.vendor
    vendor_data ={
        "id" : get_safe_attr(vendor, 'id'),
        "name": get_safe_attr(vendor, 'name'),
        "gst_no": get_safe_attr(vendor, 'gst_no'),
        "pan": get_safe_attr(vendor, 'pan'),
        "vendor_address": get_safe_attr(vendor, 'address'),
        "customer": get_safe_attr(vendor, 'customer'),
        "vendor": get_safe_attr(vendor, 'vendor'),
        "email": get_safe_attr(vendor, 'email'),
        "contact": get_safe_attr(vendor, 'contact'),
    }
    response_data = {
        'purchase_invoice' : purchase_invoice_data,
        'product_summaries' : product_summary_data,
        'client_location' : client_location_data,
        'vendor' : vendor_data,
    }
    return Response(response_data, status=status.HTTP_200_OK)
    
@api_view(['GET','PUT'])
def update_purchase_invoice(request, client_pk, invoice_pk):
    try:
        print('my payload',request.data)
        if request.method == 'GET':
            purchase_invoice = PurchaseInvoice.objects.filter(client_id=client_pk, id= invoice_pk).first()
            if not purchase_invoice:
                return Response({"error_message": "Purchase Invoice not found."}, status=status.HTTP_404_NOT_FOUND)
            purchase_invoice_data = PurchaseSerializer3(purchase_invoice).data
            product_summaries = purchase_invoice.product_summaries.all()
            product_summary_data =[
                {
                    "id": summary.id,
                    "hsnCode": summary.hsn.hsn_code,
                    "gstRate": summary.hsn.gst_rate,
                    "product": summary.product.product_name,
                    "description": summary.prod_description.description,
                    "unit": summary.prod_description.unit,
                    "rate": summary.prod_description.rate,
                    "product_amount": summary.prod_description.product_amount,
                    "cgst": summary.prod_description.cgst,
                    "sgst": summary.prod_description.sgst,
                    "igst": summary.prod_description.igst,
                }
                for summary in product_summaries
            ]

            if purchase_invoice.month:
                formatted_month = purchase_invoice.month.strftime("%d-%m-%Y")  # Directly format date
            else:
                formatted_month = None

            if purchase_invoice.invoice_date:
                formatted_invoice_date = purchase_invoice.invoice_date.strftime("%d-%m-%Y")  # Directly format date
            else:
                formatted_invoice_date = None

            if purchase_invoice.utilise_month:
                formatted_utilise = purchase_invoice.utilise_month.strftime("%d-%m-%Y")
            else:
                formatted_utilise = None

            
            response_data ={
                # "purchase_invoice" : purchase_invoice_data,
                "purchase_invoice": {
                    **purchase_invoice_data,
                    "month": formatted_month,  # Ensure month is sent in dd/mm/yyyy format
                    "invoice_date" : formatted_invoice_date,
                    "utilise_month" : formatted_utilise,
                },
                "product_summaries" : product_summary_data,
                "client_location": {
                    "id": purchase_invoice.client_Location.id if purchase_invoice.client_Location else None,
                    "location": purchase_invoice.client_Location.location if purchase_invoice.client_Location else None,
                    "contact": purchase_invoice.client_Location.contact if purchase_invoice.client_Location else None,
                    "address": purchase_invoice.client_Location.address if purchase_invoice.client_Location else None,
                    "city": purchase_invoice.client_Location.city if purchase_invoice.client_Location else None,
                    "state": purchase_invoice.client_Location.state if purchase_invoice.client_Location else None,
                    "country": purchase_invoice.client_Location.country if purchase_invoice.client_Location else None,
                    "branchID": purchase_invoice.client_Location.branch.id if purchase_invoice.client_Location else None,
                },
                "vendor": {
                    "id": purchase_invoice.vendor.id if purchase_invoice.vendor else None,
                    "name": purchase_invoice.vendor.name if purchase_invoice.vendor else None,
                    "gst_no": purchase_invoice.vendor.gst_no if purchase_invoice.vendor else None,
                    "pan": purchase_invoice.vendor.pan if purchase_invoice.vendor else None,
                    "vendor_address": purchase_invoice.vendor.address if purchase_invoice.vendor else None,
                    "customer": purchase_invoice.vendor.customer if purchase_invoice.vendor else None,
                    "vendor": purchase_invoice.vendor.vendor if purchase_invoice.vendor else None,
                    "email": purchase_invoice.vendor.email if purchase_invoice.vendor else None,
                    "contact": purchase_invoice.vendor.contact if purchase_invoice.vendor else None,
                },
            }
            #  print('bbbbbbbbbbbbbbbbbbbbbbb',response_data["sales_invoice"])
            return Response(response_data, status=status.HTTP_200_OK)
        elif request.method == 'PUT':
            purchase_invoice = PurchaseInvoice.objects.filter(client_id=client_pk, id= invoice_pk).first()
            if not purchase_invoice:
                    return Response({"error_message":"Purchase Invoice not found"}, status=status.HTTP_404_NOT_FOUND)
                
            payload = request.data
            # ser = PurchaseSerializer(data=payload)
            # if ser.is_valid()
            rows_data = defaultdict(dict)
            for key, value in payload.items():
                if key.startswith("rows["):  # Check if the key corresponds to rows
                    # Extract the row index and field name
                    row_index = key.split('[')[1].split(']')[0]  # Get the index (e.g., '0', '1', '2')
                    field_name = key.split('[')[2].split(']')[0]  # Get the field name (e.g., 'product')

                    # Store the value in the corresponding row and field
                    rows_data[int(row_index)][field_name] = value
                    
            rows = [rows_data[index] for index in sorted(rows_data.keys())]
            print(rows)
            invoice_data = request.data.get('invoiceData', [{}])[0]  # Extract the first item
            invoice_file = request.data.get('invoice_file')
            print('invoice_file',invoice_data)
            form_data = {
                "offLocID": request.data.get("formData[offLocID]"),
                "location": request.data.get("formData[location]"),
                "contact": request.data.get("formData[contact]"),
                "address": request.data.get("formData[address]"),
                "city": request.data.get("formData[city]"),
                "state": request.data.get("formData[state]"),
                "country": request.data.get("formData[country]"),
                "branchID": request.data.get("formData[branchID]"),
            }
            vendor_data = {
                "name": request.data.get("vendorData[name]"),
                "gst_no": request.data.get("vendorData[gst_no]"),
                "pan": request.data.get("vendorData[pan]"),
                "vendor_address": request.data.get("vendorData[vendor_address]"),  #nnnnnnnnn
                "email": request.data.get("vendorData[email]"),  #nnnnnnnnn
                "contact": request.data.get("vendorData[contact]"),  #nnnnnnnnn
                "customer": request.data.get("vendorData[customer]").lower() == "true" if request.data.get("vendorData[customer]") else None,
                "vendor": request.data.get("vendorData[vendor]").lower() == "true" if request.data.get("vendorData[vendor]") else None,
            }
            invoice_data = {
                "invoice_no": request.data.get("invoiceData[0][invoice_no]"),
                "invoice_date": request.data.get("invoiceData[0][invoice_date]"),
                "month": request.data.get("invoiceData[0][month]"),
                "invoice_type": request.data.get("invoiceData[0][invoice_type]"),
                "entry_type": request.data.get("invoiceData[0][entry_type]"),
                "taxable_amount": request.data.get("invoiceData[0][taxable_amount]"),
                "totalall_gst": request.data.get("invoiceData[0][totalall_gst]"),
                "total_invoice_value": request.data.get("invoiceData[0][total_invoice_value]"),
                "tds_tcs_rate": request.data.get("invoiceData[0][tds_tcs_rate]"),
                "tcs": request.data.get("invoiceData[0][tcs]"),
                "tds": request.data.get("invoiceData[0][tds]"),
                "amount_receivable": request.data.get("invoiceData[0][amount_receivable]"),
                "attach_invoice": request.data.get("invoiceData[0][attach_invoice]"),
                "attach_e_way_bill": request.data.get("invoiceData[0][attach_e_way_bill]"),
                "utilise_edit": payload.get("invoiceData[0][utilise_edit]", "").lower() == "true", #nnnn
                "utilise_month": payload.get("invoiceData[0][utilise_month]"), #nnnnnn

            }

            if invoice_data["invoice_date"]:
                try:
                    invoice_data["invoice_date"] = datetime.strptime(invoice_data["invoice_date"], "%d-%m-%Y").date()
                except ValueError:
                    return Response({"error_message": "Invalid Invoice date format, expected DD-MM-YYYY."}, status=status.HTTP_400_BAD_REQUEST)
            print("Final invoice date:", invoice_data["invoice_date"])

            if invoice_data["month"]:
                try:
                    invoice_data["month"] = datetime.strptime(invoice_data["month"], "%d-%m-%Y").date()
                except ValueError:
                    return Response({"error_message": "Invalid Month date format, expected DD-MM-YYYY."},status=status.HTTP_400_BAD_REQUEST)

            if invoice_data["utilise_month"]:
                try:
                    invoice_data["utilise_month"] = datetime.strptime(invoice_data["utilise_month"], "%d-%m-%Y").date()
                except ValueError:
                    return Response({"error_message": "Invalid Utilise Month date format, expected DD-MM-YYYY."},status=status.HTTP_400_BAD_REQUEST)

            print("Final month:", invoice_data["month"])
            attach_invoice = request.FILES.get("invoiceData[0][attach_invoice]")
            attach_e_way_bill = request.FILES.get("invoiceData[0][attach_e_way_bill]")
            data = request.data
            ser = PurchaseSerializer(data=data)
            if ser.is_valid():
                if attach_invoice:
                    purchase_invoice.attach_invoice = attach_invoice
                    
                if attach_e_way_bill:
                    purchase_invoice.attach_e_way_bill = attach_e_way_bill
                
                for field, value in invoice_data.items():
                    if field not in ['attach_invoice','attach_e_way_bill']:
                        if hasattr(purchase_invoice, field):
                            setattr(
                                purchase_invoice,
                                field,
                                safe_decimal(value) if field in [
                                    'taxable_amount', 'totalall_gst', 'total_invoice_value',
                                    'tds_tcs_rate', 'tds', 'tcs', 'amount_receivable'
                                ] else value
                            )
                
                attach_invoice = invoice_data.get('attach_invoice')
                print('attach_invoice',attach_invoice)
                # Log the flattened data
                # print("Flattened form_data:", form_data)
                print("Flattened invoice_data:", invoice_data)
                print('request payload', request.data)
                if invoice_file:
                    purchase_invoice.invoice_file = invoice_file
                    purchase_invoice.save()
                    return Response ({'message':'Invoice file uploaded successfully '},status=status.HTTP_400_BAD_REQUEST)
                
                if vendor_data:
                    if 'vendor_address' in vendor_data:
                        vendor_data['address'] = vendor_data.pop('vendor_address')
                        
                    vendor_id = request.data.get("vendorData[vendorID]")

                    if vendor_id:
                        vendor_obj = Customer.objects.filter(id=vendor_id).first()
                        if vendor_obj:
                            if vendor_obj.gst_no == vendor_data.get("gst_no"):
                                vendor_serializer = CustomerVendorSerializer(vendor_obj, data=vendor_data, partial=True)
                                if vendor_serializer.is_valid():
                                    vendor_serializer.save()
                                else:
                                    return Response({"vendor_errors": vendor_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
                            else:
                                # Create a new vendor if gst_no is changed
                                vendor_serializer = CustomerVendorSerializer(data=vendor_data)
                                if vendor_serializer.is_valid():
                                    vendor_obj = vendor_serializer.save(client=purchase_invoice.client)
                                else:
                                    return Response({"vendor_errors": vendor_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
                        else:
                            return Response({"error_message": f"Vendor with ID {vendor_id} not found."}, status=status.HTTP_404_NOT_FOUND)
                    else:
                        existing_vendor = Customer.objects.filter(client=purchase_invoice.client, gst_no=vendor_data.get("gst_no")).first()
                        if existing_vendor:
                            vendor_obj = existing_vendor
                            vendor_serializer = CustomerVendorSerializer(vendor_obj, data=vendor_data, partial=True)
                            if vendor_serializer.is_valid():
                                vendor_serializer.save()
                            else:
                                return Response({"vendor_errors": vendor_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
                        else:
                            vendor_serializer = CustomerVendorSerializer(data=vendor_data)
                            if vendor_serializer.is_valid():
                                vendor_obj = vendor_serializer.save(client=purchase_invoice.client)
                            else:
                                return Response({"vendor_errors": vendor_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
                    
                    # Assign vendor_obj to expenses.vendor if defined
                    if vendor_obj:
                        purchase_invoice.vendor = vendor_obj
       
                product_summaries = []
                # processed_product_names = set()
                for row in rows:
                    hsn_code = row.get('hsnCode')
                    gst_rate = safe_decimal(row.get('gstRate', '0'))
                    product_name = row.get('product')
                    description_text = row.get('description')
                    unit_value = safe_decimal(row.get('unit', '0'))
                    rate_value = safe_decimal(row.get('rate', '0'))
                    amount = safe_decimal(row.get('product_amount', '0'))
                    cgst = safe_decimal(row.get('cgst', '0'))
                    sgst = safe_decimal(row.get('sgst', '0'))
                    igst = safe_decimal(row.get('igst', '0'))


                    # Update or create HSNCode
                    hsn_code_obj, _ = HSNCode.objects.update_or_create(
                        hsn_code=hsn_code,
                        defaults={'gst_rate': gst_rate}
                    )

                    existing_product = Product.objects.filter(product_name=product_name).exclude(hsn=hsn_code_obj).first()
                    if existing_product:
                        return Response(
                            {"error_message": f"Product name '{product_name}' already exists with a different HSN code."},
                            status=status.HTTP_400_BAD_REQUEST
                        )

                    # Update or create Product
                    product_obj, _ = Product.objects.update_or_create(
                        product_name=product_name,
                        hsn=hsn_code_obj,
                        defaults={'unit_of_measure': unit_value}
                    )

                    # Update or create ProductDescription
                    product_description_obj, _ = ProductDescription.objects.update_or_create(
                        product=product_obj,
                        description=description_text,
                        defaults={
                            'unit': unit_value,
                            'rate': rate_value,
                            'product_amount': amount,
                            'cgst': cgst,
                            'sgst': sgst,
                            'igst': igst
                        }
                    )
                    
                    product_summary, _ = ProductSummaryPurchase.objects.update_or_create(
                        hsn=hsn_code_obj,
                        product=product_obj,
                        prod_description=product_description_obj
                    )
                    product_summaries.append(product_summary)
                    
                    if invoice_data:
                        for field, value in invoice_data.items():
                            if field != 'client':
                                setattr(
                                    purchase_invoice,
                                    field,
                                    safe_decimal(value) if field in [
                                        'taxable_amount', 'totalall_gst', 'total_invoice_value',
                                        'tds_tcs_rate', 'tds', 'tcs', 'amount_receivable'
                                    ] else value
                                )
                purchase_invoice.product_summaries.set(product_summaries)
                # purchase_invoice.save()
                
                location_data = form_data.get('location')  # Location name entered by user
                location_id = form_data.get('offLocID')   # Existing location ID, if provided
                branch_id = form_data.get('branchID')     # Branch ID selected for new location
                
                if location_id:
                    location_obj = OfficeLocation.objects.filter(id=location_id).first()
                    if not location_obj:
                        return Response({'error_message':'Office Location not found. '}, status=status.HTTP_400_BAD_REQUEST)
                    location_obj.location = location_data
                    location_obj.contact = form_data.get('contact')
                    location_obj.address = form_data.get('address')
                    location_obj.city = form_data.get('city')
                    location_obj.state = form_data.get('state')
                    location_obj.country = form_data.get('country')
                    location_obj.save()
                else:
                    if not branch_id:
                        return Response({"error_message": "Branch ID is required for creating a new location."}, status=status.HTTP_400_BAD_REQUEST)
                    branch_instance = Branch.objects.filter(id=branch_id, client_id=purchase_invoice.client.id).first()
                    if not branch_instance:
                        return Response({"error_message": f"Branch with ID {branch_id} not found or doesn't belong to the client."},
                                        status=status.HTTP_404_NOT_FOUND)
                        
                    location_obj, _ = OfficeLocation.objects.get_or_create(
                        location = location_data,
                        contact = form_data.get('contact'),
                        address = form_data.get('address'),
                        city = form_data.get('city'),
                        state = form_data.get('state'),
                        country = form_data.get('country'),
                        branch = branch_instance
                    )
                    
                purchase_invoice.client_Location = location_obj
                print('my payload',request.data)
                purchase_invoice.save()
                
                
                response_data = {
                    'message' : 'Purchase Invoice Updated successfully. ',
                    'purchase_invoice_data' : PurchaseSerializer(purchase_invoice).data,
                    'product_summaries' : [{'id': summary.id, 'product_name': summary.product.product_name} for summary in product_summaries]
                }
                return Response(response_data, status=status.HTTP_200_OK)
            else:
                return Response({'error_message': ser.errors}, status=status.HTTP_400_BAD_REQUEST)

    except Exception as e:
        error_details = traceback.format_exc()
        # print({"Error in update_sales_invoice" : str(e)})
        return Response({"error_message":str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
def create_purchase_invoice2(request, client_pk):
    try:
        with transaction.atomic():
            data = request.data
            payload = request.data
            # print('payload',payload)
            # print('aaaaaaaaaaaaaa',invoice_data)

            rows_data = defaultdict(dict)
            for key, value in payload.items():
                if key.startswith("rows["):
                    row_index = key.split('[')[1].split(']')[0]
                    field_name = key.split('[')[2].split(']')[0]
                    rows_data[int(row_index)][field_name] = value
            rows = [rows_data[index] for index in sorted(rows_data.keys())]
            # print('aaaaaaaaaaaaaa',invoice_data)
            print('aaaaaaaaaaaaaacccc',request.data)
            
            form_data = {
                "offLocID" : payload.get("formData[offLocID]"),
                "location" : payload.get("formData[location]"),
                "contact" :  payload.get("formData[contact]"),
                "address" : payload.get("formData[address]"),
                "city" : payload.get("formData[city]"),
                "state" : payload.get("formData[state]"),
                "country" : payload.get("formData[country]"),
                "branchID" : payload.get("formData[branchID]"),
            }
            vendor_data = {
                "name" : payload.get("vendorData[name]"),
                "gst_no" : payload.get("vendorData[gst_no]"),
                "pan" : payload.get("vendorData[pan]"),
                "address" : payload.get("vendorData[vendor_address]"), #nnnnnnnnn
                "email" : payload.get("vendorData[email]"),  #nnnnnnnnn
                "contact" : payload.get("vendorData[contact]"),  #nnnnnnnnn
                "customer" : payload.get("vendorData[customer]", "").lower() == "true",
                "vendor" : payload.get("vendorData[vendor]", "").lower() == "true",
            }
            invoice_data = {
                "invoice_no" : payload.get("invoiceData[0][invoice_no]"),
                "invoice_date" : payload.get("invoiceData[0][invoice_date]"),
                # "invoice_date": payload.get("invoiceData[0][invoice_date]") or payload.get("selectedInvoiceDate"),
                "month" : payload.get("invoiceData[0][month]"),
                "invoice_type" : payload.get("invoiceData[0][invoice_type]"),
                "entry_type" : payload.get("invoiceData[0][entry_type]"),
                "taxable_amount" : payload.get("invoiceData[0][taxable_amount]"),
                "totalall_gst" : payload.get("invoiceData[0][totalall_gst]"),
                "total_invoice_value" : payload.get("invoiceData[0][total_invoice_value]"),
                "tds_tcs_rate": payload.get("invoiceData[0][tds_tcs_rate]"),
                "tcs": payload.get("invoiceData[0][tcs]"),
                "tds": payload.get("invoiceData[0][tds]"),
                "amount_receivable": payload.get("invoiceData[0][amount_receivable]"),
                # "utilise_edit": payload.get("invoiceData[0][utilise_edit]", "").lower() == "true", #nnnn
                # "utilise_month": payload.get("invoiceData[0][utilise_month]").lower(), #nnnnnn
                
            }
            if invoice_data["invoice_date"]:
                try:
                    invoice_data["invoice_date"] = datetime.strptime(invoice_data["invoice_date"], "%d-%m-%Y").date()
                except ValueError:
                    return Response({"error_message": "Invalid date format, expected DD-MM-YYYY."}, status=status.HTTP_400_BAD_REQUEST)
            print("Final invoice date:", invoice_data["invoice_date"])

            if invoice_data["month"]:
                try:
                    invoice_data["month"] = datetime.strptime(invoice_data["month"], "%d-%m-%Y").date()
                except ValueError:
                    return Response({"error_message": "Invalid Month date format, expected DD-MM-YYYY."},status=status.HTTP_400_BAD_REQUEST)

            print("Final month:", invoice_data["month"])

            attach_invoice = request.FILES.get("invoiceData[0][attach_invoice]")
            attach_e_way_bill = request.FILES.get("invoiceData[0][attach_e_way_bill]")
            print('aaaaaaaaaaaaaad',invoice_data)

            vendor_obj = None
            if vendor_data.get("gst_no"):
                existing_vendor = Customer.objects.filter(client_id=client_pk, gst_no=vendor_data["gst_no"]).first()
                if existing_vendor:
                    vendor_serializer = CustomerVendorSerializer(existing_vendor, data=vendor_data, partial=True)
                    if vendor_serializer.is_valid():
                        vendor_obj = vendor_serializer.save(status=status.HTTP_200_OK)
                    else:
                        return Response({"vendor_errors": vendor_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
                else:
                    vendor_serializer = CustomerVendorSerializer(data=vendor_data)
                    if vendor_serializer.is_valid():
                        vendor_obj = vendor_serializer.save(client_id=client_pk)
                    else:
                        return Response({"vendor_errors": vendor_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

            product_summaries = []  # To store created product summaries
            for row in rows:
                hsn_code = row.get('hsnCode')
                gst_rate = safe_decimal(row.get('gstRate', '0'))
                product_name = row.get('product')
                product_id = row.get('product_id')  # Assuming the frontend sends this if selecting an existing product
                description_text = row.get('description', '')
                unit_value = safe_decimal(row.get('unit', '0'))
                rate_value = safe_decimal(row.get('rate', '0'))
                amount = safe_decimal(row.get('product_amount', '0'))
                cgst = safe_decimal(row.get('cgst', '0'))
                sgst = safe_decimal(row.get('sgst', '0'))
                igst = safe_decimal(row.get('igst', '0'))
 
                hsn_code_obj, _ = HSNCode.objects.get_or_create(
                    hsn_code=hsn_code, defaults={'gst_rate': gst_rate}
                )

                existing_product = Product.objects.filter(product_name=product_name).exclude(hsn=hsn_code_obj).first()
                if existing_product:
                    return Response(
                        {"error_message": f"Product name '{product_name}' already exists with a different HSN code."},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                # Handle Product (existing or new)
                if product_id:
                    # Use existing product
                    product_obj = Product.objects.filter(id=product_id).first()
                    if not product_obj:
                        return Response({"error_message": f"Product with ID {product_id} not found."}, status=status.HTTP_404_NOT_FOUND)
                else:
                    # Create new product
                    product_obj, _ = Product.objects.get_or_create(
                        product_name=product_name, hsn=hsn_code_obj, defaults={'unit_of_measure': unit_value}
                    )

                # Handle ProductDescription
                product_description_obj, _ = ProductDescription.objects.get_or_create(
                    product=product_obj,
                    description=description_text,
                    defaults={
                        'unit': unit_value,
                        'rate': rate_value,
                        'product_amount': amount,
                        'cgst': cgst,
                        'sgst': sgst,
                        'igst': igst,
                    }
                )

                # Create ProductSummary
                # print('aaaaaaaaaaaaaad',invoice_data)
                product_summary = ProductSummaryPurchase(
                    hsn=hsn_code_obj,
                    product=product_obj,
                    prod_description=product_description_obj
                )
                product_summaries.append(product_summary)

            location_obj = None
            if form_data["offLocID"] :
                location_obj = OfficeLocation.objects.filter(id=form_data["offLocID"]).first()
                if not location_obj:
                    return Response({"error_message":"Office Location not found"}, status=status.HTTP_400_BAD_REQUEST)
            else:
                branch_instance = Branch.objects.filter(id=form_data["branchID"], client_id=client_pk).first()
                if not branch_instance:
                    return Response({"error_message": f"Branch with ID {form_data['branchID']} not found or doesn't belong to the client."},
                                    status=status.HTTP_404_NOT_FOUND)
                location_obj, _ = OfficeLocation.objects.get_or_create(
                    location = form_data.get("location"),
                    contact = form_data.get("contact"),
                    address = form_data.get("address"),
                    city = form_data.get("city"),
                    state = form_data.get("state"),
                    country = form_data.get("country"),
                    branch = branch_instance

                )
        
            purchase_invoice= PurchaseInvoice.objects.create(
                client_id = client_pk,
                vendor = vendor_obj,
                attach_invoice = attach_invoice,
                attach_e_way_bill = attach_e_way_bill,
                client_Location = location_obj,
                **invoice_data
            )

            for product_summary in product_summaries:
                product_summary.save()
                purchase_invoice.product_summaries.add(product_summary)  # Add the product summary to the invoice

            print('fffffffffffffffff',payload)
            return Response({"message": "Purchase Invoice created successfully.", "invoice_id": purchase_invoice.id}, status=status.HTTP_201_CREATED)
            

    except Exception as e:
        error_details = traceback.format_exc()
        # print({"Error in update_sales_invoice" : str(e)})
        return Response({"error_message":str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET','PATCH'])
def purchase_invoice_detail_view(request, client_pk, invoice_pk):
    try:
        purchase_invoice = PurchaseInvoice.objects.get(client=client_pk, pk=invoice_pk)
    except PurchaseInvoice.DoesNotExist:
        return Response({'error_message':'purchase invoice not found'}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        serializers = PurchaseSerializerList(purchase_invoice)
        return Response(serializers.data, status=status.HTTP_200_OK)

    elif request.method == 'PATCH':
        data = request.data
        purchase_invoice.invoice_no = data.get('invoice_no',purchase_invoice.invoice_no)
        purchase_invoice.invoice_date = data.get('invoice_date', purchase_invoice.invoice_date)
        purchase_invoice.month = data.get('invoice_month',purchase_invoice.month)
        purchase_invoice.invoice_type = data.get('invoice_type', purchase_invoice.invoice_type)
        purchase_invoice.entry_type= data.get('entry_type',purchase_invoice.entry_type)
        purchase_invoice.taxable_amount = safe_decimal(data.get('taxable_amount', purchase_invoice.taxable_amount))
        purchase_invoice.totalall_gst = safe_decimal(data.get('totalall_gst', purchase_invoice.totalall_gst))
        purchase_invoice.total_invoice_value = safe_decimal(data.get('total_invoice_value', purchase_invoice.total_invoice_value))
        purchase_invoice.tds_tcs_rate = safe_decimal(data.get('tds_tcs_rate', purchase_invoice.tds_tcs_rate))
        purchase_invoice.tds = safe_decimal(data.get('tds',purchase_invoice.tds))
        purchase_invoice.tcs = safe_decimal(data.get('tcs', purchase_invoice.tcs))
        purchase_invoice.amount_receivable = safe_decimal(data.get('amount_receivable', purchase_invoice.amount_receivable))
        purchase_invoice.utilise_edit = data.get('utilise_edit', purchase_invoice.utilise_edit)
        purchase_invoice.utilise_month = data.get('utilise_month', purchase_invoice.utilise_month)


        rows = data.get('rows',[])
        product_summaries = []
        for row in rows:
            hsn_code = row.get('hsnCode')
            gst_rate = safe_decimal(row.get('gstRate','0'))
            product_name = row.get('product')
            description_text = row.get('description')
            unit_value = safe_decimal(row.get('unit','0'))
            rate_value = safe_decimal(row.get('rate','0'))
            amount = safe_decimal(row.get('product_amount','0'))
            cgst = safe_decimal(row.get('cgst','0'))
            sgst = safe_decimal(row.get('sgst','0'))
            igst = safe_decimal(row.get('igst','0'))

            hsn_code_obj, created = HSNCode.objects.get_or_create(
                hsn_code = hsn_code,
                defaults={'gst_rate': gst_rate}
            )
            if not created and hsn_code_obj.gst_rate != gst_rate:
                hsn_code_obj.gst_rate = gst_rate
                hsn_code_obj.save()

            product_obj = Product.objects.filter(product_name=product_name, hsn=hsn_code_obj).first()
            if not product_obj:
                product_obj = Product.objects.create(
                    product_name = product_name,
                    hsn = hsn_code_obj,
                    unit_of_measure = unit_value
                )

            product_description_obj, created = ProductDescription.objects.get_or_create(
                product = product_obj,
                description = description_text,
                defaults={
                    'unit': unit_value,
                    'rate': rate_value,
                    'product_amount' : amount,
                    'cgst' : cgst,
                    'sgst' : sgst,
                    'igst' : igst,
                }
            )
            if not created:
                product_description_obj.unit = unit_value
                product_description_obj.rate = rate_value
                product_description_obj.product_amount = amount
                product_description_obj.cgst = cgst
                product_description_obj.sgst = sgst
                product_description_obj.igst = igst
                product_description_obj.save()
            product_summary = ProductSummaryPurchase.objects.create(
                hsn= hsn_code_obj,
                product = product_obj,
                prod_description = product_description_obj
            )
            product_summaries.append(product_summary)

        if product_summaries:
            purchase_invoice.product_summaries.set(product_summaries)
        purchase_invoice.save()

        updated_serializer = PurchaseSerializer(purchase_invoice)
        return Response(updated_serializer.data, status=status.HTTP_200_OK)
    
@api_view(['DELETE'])
def delete_purchase_invoice(request, client_pk, pk):
    """
    Deletes a PurchaseInvoice by its primary key (ID) along with its associated product summaries.
    """
    try:
        # Retrieve the Client instance
        client = Client.objects.filter(id=client_pk).first()

        if not client:
            return Response({"error_message": "Client not found."}, status=status.HTTP_404_NOT_FOUND)

        # Retrieve the SalesInvoice instance
        purchase_invoice = PurchaseInvoice.objects.filter(id=pk, client=client).first()

        if not purchase_invoice:
            return Response({"error_message": "Purchase Invoice not found."}, status=status.HTTP_404_NOT_FOUND)

        # Handle deletion of associated ProductSummary instances
        product_summaries = purchase_invoice.product_summaries.all()
        for product_summary in product_summaries:
            product_summary.delete()

        # Delete the SalesInvoice instance
        purchase_invoice.delete()

        return Response({"message": "Purchase Invoice deleted successfully."}, status=status.HTTP_200_OK)

    # except Exception as e:
    #     return Response({"error_message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    except Exception as e:
        error_details = traceback.format_exc()
        # print({"Error in update_sales_invoice" : str(e)})
        return Response({"error_message":str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# ******************************************************Debit Note******************************************************
@api_view(['GET', 'POST'])
def create_debit_note_get(request, pk):
    try:
        client = Client.objects.get(id=pk)

    except Client.DoesNotExist:
        return Response({"error_message": "Client not found."}, status=404)

    if request.method == 'GET':
        # Initialize response context with default values
        context = {
            'serializer_customer': [],
            'serializer': [],
            'product_serializer': [],
            'message': None,
            'hsn': None,
            'location': None
        }

        # Retrieve query parameters
        received_value = request.GET.get('newValue')
        product_id = request.GET.get('productID')
        print('daadasdasdsa',received_value)
        # Process productID if it exists
        if product_id:
            try:
                hsn_cc = Product.objects.get(id=product_id).hsn
                context.update({
                    "message": "Product HSN found",
                    "hsn": HSNSerializer(hsn_cc).data
                })
            except (ValueError, Product.DoesNotExist):
                return Response({"error_message": "Invalid Product ID."}, status=400)

        # Process received_value if it exists
        if received_value:
            try:
                received_value = int(received_value)
                location = OfficeLocation.objects.get(id=received_value)
                branch_gst = location.branch.gst_no
                print("branch val",branch_gst)
                context.update({
                    "message": "Location found",
                    "location": OfficeLocationSerializer(location).data,
                    "branch_gst": branch_gst
                })
            except (ValueError, OfficeLocation.DoesNotExist):
                return Response({"error_message": "Invalid location ID."}, status=400)

        # Add data only if productID and received_value were not provided
        if not product_id and not received_value:
            off = OfficeLocation.objects.filter(branch__client=client)
            customer = Customer.objects.filter(client=client, customer=True)
            product = Product.objects.all()
            branch = Branch.objects.filter(client=client)

            serializer_customer = CustomerVendorSerializer(customer, many=True)
            serializer = OfficeLocationSerializer(off, many=True)
            product_serializer = ProductSerializer(product, many=True)
            branch_serializer = BranchSerailizer(branch, many=True)

            context.update({
                'serializer_customer': serializer_customer.data,
                'serializer': serializer.data,
                'product_serializer': product_serializer.data,
                'branch_serializer': branch_serializer.data
            })

        return Response(context)

@api_view(['POST'])
def create_debit_note(request, client_pk, invoice_pk):
    client = Client.objects.get(id=client_pk)
    sales = SalesInvoice.objects.get(id=invoice_pk, client=client)

    # Check if 'attach_e_way_bill' is in the request files
    if 'attach_invoice' in request.FILES:
        # Iterate through each file in the 'attach_e_way_bill' field
        for e_way_bill in request.FILES.getlist('attach_invoice'):
            # Prepare data for each file
            debitnote_data = {
                'attach_invoice': e_way_bill,  # The file being uploaded
                'client': client.id,  # Associate the file with the client
                'sales_invoice' : sales.id,
            }

            # Initialize the serializer for each file
            serializer = DebitNoteSerializer2(data=debitnote_data)

            # Check if the serializer is valid
            if serializer.is_valid():
                # Save each file as a separate object
                serializer.save()
            else:
                return Response({'error_message': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

        # If all files are processed, return success response
        return Response({'message': 'Debit Note E-way bill(s) uploaded successfully.'}, status=status.HTTP_200_OK)
    else:
        return Response({'error_message': 'No files uploaded in the request.'}, status=status.HTTP_400_BAD_REQUEST)
    
@api_view(['GET'])
def get_debit_note_data(request, client_pk, invoice_pk, debit_pk):
    """
    GET request to fetch sales invoice data along with related product summaries and client info.
    """
    # Fetch the sales invoice with related client information
    debit_note = DebitNote.objects.filter(client_id=client_pk, sales_invoice_id=invoice_pk, id=debit_pk).first()

    if not debit_note:
        return Response({"error_message": "Debit Note not found."}, status=status.HTTP_404_NOT_FOUND)

    # Serialize the sales invoice
    debit_note_data = DebitNoteSerializer3(debit_note).data

    # Prepare product summaries data
    product_summaries = debit_note.product_summaries.all()
    product_summary_data = [
        {
            "id": summary.id,
            "hsnCode": summary.hsn.hsn_code,
            "gstRate": summary.hsn.gst_rate,
            "product": summary.product.product_name,
            "description": summary.prod_description.description,
            "unit": summary.prod_description.unit,
            "rate": summary.prod_description.rate,
            "product_amount": summary.prod_description.product_amount,
            "cgst": summary.prod_description.cgst,
            "sgst": summary.prod_description.sgst,
            "igst": summary.prod_description.igst,
        }
        for summary in product_summaries
    ]

    # Helper function to safely get attributes from related objects
    def get_safe_attr(obj, attr, default=None):
        return getattr(obj, attr, default) if obj else default

    # Prepare client location data (with safe attribute access)
    client_location = debit_note.client_Location
    client_location_data = {
        "id": get_safe_attr(client_location, 'id'),
        "location": get_safe_attr(client_location, 'location'),
        "contact": get_safe_attr(client_location, 'contact'),
        "address": get_safe_attr(client_location, 'address'),
        "city": get_safe_attr(client_location, 'city'),
        "state": get_safe_attr(client_location, 'state'),
        "country": get_safe_attr(client_location, 'country'),
    }

    # Prepare customer data (with safe attribute access)
    customer = debit_note.customer
    customer_data = {
        "id": get_safe_attr(customer, 'id'),
        "name": get_safe_attr(customer, 'name'),
        "gst_no": get_safe_attr(customer, 'gst_no'),
        "pan": get_safe_attr(customer, 'pan'),
        "customer_address": get_safe_attr(customer, 'address'),
        "customer": get_safe_attr(customer, 'customer'),
        "vendor": get_safe_attr(customer, 'vendor'),
        "email": get_safe_attr(customer, 'email'),
        "contact": get_safe_attr(customer, 'contact'),
    }

    # Prepare final response data
    response_data = {
        "debit_note": debit_note_data,
        "product_summaries": product_summary_data,
        "client_location": client_location_data,
        "customer": customer_data,
    }

    return Response(response_data, status=status.HTTP_200_OK)

@api_view(['GET', 'POST'])
def update_debit_note(request, client_pk, invoice_pk):
    try:
        # Handle GET request
        if request.method == 'GET':
            debit_note = SalesInvoice.objects.filter(client_id=client_pk, id=invoice_pk) \
                .select_related('client_Location', 'customer') \
                .prefetch_related('product_summaries__hsn', 'product_summaries__prod_description') \
                .first()

            if not debit_note:
                return Response({"error_message": "Debit Note not found."}, status=status.HTTP_404_NOT_FOUND)

            debit_notes = DebitNote.objects.filter(sales_invoice=debit_note)
            product_unit_sums = defaultdict(int)

            for dn in debit_notes:
                for product_summary in dn.product_summaries.all():
                    product_name = product_summary.product_name()
                    product_unit = product_summary.unit() or 0  # Agar unit None ho toh 0 lelo
                    product_unit_sums[product_name] += product_unit

            remaining_units = {}
            for product in debit_note.product_summaries.all():
                product_name = product.product_name()
                si_unit = product.unit() or 0  # Sales Invoice ke unit
                dn_unit_sum = product_unit_sums[product_name]  # Debit Notes ka unit sum
                remaining_units[product_name] = si_unit - dn_unit_sum  # Remaining units

            print('ooooooooo',remaining_units)

            debit_notes_details = []
            for debit_note in debit_notes:
                product_summaries = debit_note.product_summaries.all()
                print(f"Found {len(product_summaries)} ProductSummaries for DebitNote {debit_note.id}")

                for product_summary in product_summaries:
                    product_details = {
                        'debit_note_id': debit_note.id,
                        'product_name': product_summary.product_name(),
                        'unit': product_summary.unit(),
                    }
                    debit_notes_details.append(product_details)

            dd = debit_notes_details
            print('Debit Notes',dd)

            debit_note_data = SalesSerializer3(debit_note).data

            product_summaries = debit_note.product_summaries.all()
            product_summary_data = [
                {
                    "id": summary.id,
                    "hsnCode": summary.hsn.hsn_code,
                    "gstRate": summary.hsn.gst_rate,
                    "product": summary.product.product_name,
                    "description": summary.prod_description.description,
                    "rate": summary.prod_description.rate,
                    "product_amount": summary.prod_description.product_amount,
                    "unit": remaining_units.get(summary.product_name(), summary.prod_description.unit),
                    "cgst": summary.prod_description.cgst,
                    "sgst": summary.prod_description.sgst,
                    "igst": summary.prod_description.igst,
                }
                for summary in product_summaries
            ]

            response_data = {
                "debit_note": debit_note_data,
                "product_summaries": product_summary_data,
                "remaining_units": remaining_units,
                "client_location": {
                    "id": debit_note.client_Location.id if debit_note.client_Location else None,
                    "location": debit_note.client_Location.location if debit_note.client_Location else None,
                    "contact": debit_note.client_Location.contact if debit_note.client_Location else None,
                    "address": debit_note.client_Location.address if debit_note.client_Location else None,
                    "city": debit_note.client_Location.city if debit_note.client_Location else None,
                    "state": debit_note.client_Location.state if debit_note.client_Location else None,
                    "country": debit_note.client_Location.country if debit_note.client_Location else None,
                    "branchID": debit_note.client_Location.branch.id if debit_note.client_Location else None,
                },
                "customer": {
                    "id": debit_note.customer.id if debit_note.customer else None,
                    "name": debit_note.customer.name if debit_note.customer else None,
                    "gst_no": debit_note.customer.gst_no if debit_note.customer else None,
                    "pan": debit_note.customer.pan if debit_note.customer else None,
                    "customer_address": debit_note.customer.address if debit_note.customer else None,
                    "customer": debit_note.customer.customer if debit_note.customer else None,
                    "vendor": debit_note.customer.vendor if debit_note.customer else None,
                    "email": debit_note.customer.email if debit_note.customer else None,
                    "contact": debit_note.customer.contact if debit_note.customer else None,
                },
            }
            return Response(response_data, status=status.HTTP_200_OK)

        # Handle PUT request
        elif request.method == 'POST':
            try:
                with transaction.atomic(): 
                    payload = request.data
                    data = request.data

                    # Fetch the Client
                    client = Client.objects.filter(pk=client_pk).first()
                    if not client:
                        return Response({"error_message": "Client not found."}, status=status.HTTP_404_NOT_FOUND)

                    # Fetch the Sales Invoice
                    sales_invoice = SalesInvoice.objects.filter(client_id=client_pk, id=invoice_pk) \
                        .select_related('client_Location', 'customer') \
                        .prefetch_related('product_summaries__hsn', 'product_summaries__prod_description') \
                        .first()

                    if not sales_invoice:
                        return Response({"error_message": "Sales Invoice not found or does not belong to the client."},
                                        status=status.HTTP_404_NOT_FOUND)

                    # Calculate remaining units
                    debit_notes = DebitNote.objects.filter(sales_invoice=sales_invoice)
                    product_unit_sums = defaultdict(int)

                    for dn in debit_notes:
                        for product_summary in dn.product_summaries.all():
                            product_name = product_summary.product_name()
                            product_unit = product_summary.unit() or 0
                            product_unit_sums[product_name] += product_unit

                    remaining_units = {}
                    for product in sales_invoice.product_summaries.all():
                        product_name = product.product_name()
                        si_unit = product.unit() or 0
                        dn_unit_sum = product_unit_sums[product_name]
                        remaining_units[product_name] = si_unit - dn_unit_sum

                    print('gggggggggg',remaining_units)
                    
                    # Validations..........
                    if all(unit == 0 for unit in remaining_units.values()):
                        return Response(
                            {"message": "Cannot create a new Debit Note. All product remaining units are 0."},
                            status=status.HTTP_400_BAD_REQUEST
                        )
                    
                    for product_name, unit in remaining_units.items():
                        if unit is None:
                            return Response({"message" : f"Remaining units for product '{product_name}' is missing or undefined."}, status=status.HTTP_400_BAD_REQUEST)
                        if unit < 0:
                            return Response({"message" : f"Remaining units for prooooooooooooduct '{product_name}' cannot be negative. Current value: {unit}"}, status=status.HTTP_400_BAD_REQUEST)

                    # Extract rows dynamically
                    rows_data = defaultdict(dict)
                    for key, value in payload.items():
                        if key.startswith("rows["):
                            row_index = key.split('[')[1].split(']')[0]
                            field_name = key.split('[')[2].split(']')[0]
                            rows_data[int(row_index)][field_name] = value
                    rows = [rows_data[index] for index in sorted(rows_data.keys())]

                    # Extract form data, vendor data, and invoice data
                    form_data = {
                        "offLocID": payload.get("formData[offLocID]"),
                        "location": payload.get("formData[location]"),
                        "contact": payload.get("formData[contact]"),
                        "address": payload.get("formData[address]"),
                        "city": payload.get("formData[city]"),
                        "state": payload.get("formData[state]"),
                        "country": payload.get("formData[country]"),
                        "branchID": payload.get("formData[branchID]"),
                    }
                    vendor_data = {
                        "name": payload.get("vendorData[name]"),
                        "gst_no": payload.get("vendorData[gst_no]"),
                        "pan": payload.get("vendorData[pan]"),
                        "customer_address": payload.get("vendorData[customer_address]"),
                        "email": payload.get("vendorData[email]"),
                        "contact": payload.get("vendorData[contact]"),
                        "customer": payload.get("vendorData[customer]", "").lower() == "true",
                        "vendor": payload.get("vendorData[vendor]", "").lower() == "true",
                    }
                    invoice_data = {
                        "invoice_no": payload.get("invoiceData[0][invoice_no]"),
                        "invoice_date": payload.get("invoiceData[0][invoice_date]"),
                        # "month": payload.get("invoiceData[0][month]"),
                        "po_date": payload.get("invoiceData[0][po_date]"),
                        "po_no": payload.get("invoiceData[0][po_no]"),
                        "invoice_type": payload.get("invoiceData[0][invoice_type]"),
                        "entry_type": payload.get("invoiceData[0][entry_type]"),
                        "taxable_amount": payload.get("invoiceData[0][taxable_amount]"),
                        "totalall_gst": payload.get("invoiceData[0][totalall_gst]"),
                        "total_invoice_value": payload.get("invoiceData[0][total_invoice_value]"),
                        "tds_tcs_rate": payload.get("invoiceData[0][tds_tcs_rate]"),
                        "tcs": payload.get("invoiceData[0][tcs]"),
                        "tds": payload.get("invoiceData[0][tds]"),
                        "amount_receivable": payload.get("invoiceData[0][amount_receivable]"),
                    }

                    if invoice_data["invoice_date"]:
                        try:
                            invoice_data["invoice_date"] = datetime.strptime(invoice_data["invoice_date"], "%d-%m-%Y").date()
                        except ValueError:
                            return Response({"error_message": "Invalid Invoice date format, expected DD-MM-YYYY."}, status=status.HTTP_400_BAD_REQUEST)
                    print("Final invoice date:", invoice_data["invoice_date"])

                    if invoice_data["po_date"]:
                        try:
                            invoice_data["po_date"] = datetime.strptime(invoice_data["po_date"], "%d-%m-%Y").date()
                        except ValueError:
                            return Response({"error_message": "Invalid PO date format, expected DD-MM-YYYY."},status=status.HTTP_400_BAD_REQUEST)

                    # if invoice_data["utilise_month"]:
                    #     try:
                    #         invoice_data["utilise_month"] = datetime.strptime(invoice_data["utilise_month"], "%d-%m-%Y").date()
                    #     except ValueError:
                    #         return Response({"error_message": "Invalid Utilise Month date format, expected DD-MM-YYYY."},status=status.HTTP_400_BAD_REQUEST)

            # print("Final month:", invoice_data["month"])
                    attach_invoice = request.FILES.get("invoiceData[0][attach_invoice]")
                    attach_e_way_bill = request.FILES.get("invoiceData[0][attach_e_way_bill]")

                    # Handle Vendor creation or update
                    vendor_obj = None
                    if vendor_data.get("gst_no"):
                        existing_vendor = Customer.objects.filter(client=client, gst_no=vendor_data["gst_no"]).first()
                        if existing_vendor:
                            vendor_serializer = CustomerVendorSerializer(existing_vendor, data=vendor_data, partial=True)
                            if vendor_serializer.is_valid():
                                vendor_obj = vendor_serializer.save()
                            else:
                                return Response({"vendor_errors": vendor_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
                        else:
                            vendor_serializer = CustomerVendorSerializer(data=vendor_data)
                            if vendor_serializer.is_valid():
                                vendor_obj = vendor_serializer.save(client=client)
                            else:
                                return Response({"vendor_errors": vendor_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

                    if any(unit is None for unit in remaining_units.values()):
                        return Response({"message" : f"Remaining units for product {product} is missing."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
                    else:
                        debit_note_serializer = DebitNoteSerializer(data=data)

                        if debit_note_serializer.is_valid():
                            location_obj = None
                            if form_data["offLocID"]:
                                location_obj = OfficeLocation.objects.filter(id=form_data["offLocID"]).first()
                                if not location_obj:
                                    return Response({"error_message": "Office Location not found."}, status=status.HTTP_404_NOT_FOUND)
                            else:
                                branch_instance = Branch.objects.filter(id=form_data["branchID"], client=client).first()
                                if not branch_instance:
                                    return Response({"error_message": f"Branch with ID {form_data['branchID']} not found or doesn't belong to the client."},
                                                    status=status.HTTP_404_NOT_FOUND)
                                location_obj, _ = OfficeLocation.objects.get_or_create(
                                    location=form_data.get("location"),
                                    contact=form_data.get("contact"),
                                    address=form_data.get("address"),
                                    city=form_data.get("city"),
                                    state=form_data.get("state"),
                                    country=form_data.get("country"),
                                    branch=branch_instance
                                )

                            for row in rows:
                                product_name = row.get('product')
                                unit_value = safe_decimal(row.get('unit', '0'))
                                
                                # Skip rows with unit value as 0 or negative
                                if unit_value <= 0:
                                    continue  # Skip processing this row
                                
                                # Validate remaining units for the product
                                if product_name in remaining_units:
                                    if remaining_units[product_name] < unit_value:
                                        return Response(
                                            {
                                                "message": f"Not enough units remaining for the product '{product_name}'. "
                                                        f"Available: {remaining_units[product_name]}, Requested: {unit_value}."
                                            },
                                            status=status.HTTP_400_BAD_REQUEST
                                        )
                                    
                                    # Deduct the used units from the remaining_units
                                    remaining_units[product_name] -= unit_value
                                else:
                                    return Response(
                                        {"message": f"Product '{product_name}' not found in remaining units."},
                                        status=status.HTTP_400_BAD_REQUEST
                                    )
                
                        # if not debit_note_serializer.is_valid():
                            debit_note = DebitNote.objects.create(
                            client=client,
                            sales_invoice=sales_invoice,
                            client_Location=location_obj,
                            customer=vendor_obj,
                            attach_invoice=attach_invoice,
                            attach_e_way_bill=attach_e_way_bill,
                            **invoice_data
                            )
                
                            product_summaries = []
                            for row in rows:
                                product_name = row.get('product')
                                unit_value = safe_decimal(row.get('unit', '0'))
                                hsn_code = row.get('hsnCode')
                                gst_rate = safe_decimal(row.get('gstRate', '0'))
                                product_id = row.get('product_id')
                                description_text = row.get('description', '')
                                rate_value = safe_decimal(row.get('rate', '0'))
                                amount = safe_decimal(row.get('product_amount', '0'))
                                cgst = safe_decimal(row.get('cgst', '0'))
                                sgst = safe_decimal(row.get('sgst', '0'))
                                igst = safe_decimal(row.get('igst', '0'))

                                hsn_code_obj, _ = HSNCode.objects.get_or_create(
                                    hsn_code=hsn_code, defaults={'gst_rate': gst_rate}
                                )

                                if product_id:
                                    product_obj = Product.objects.filter(id=product_id).first()
                                    if not product_obj:
                                        return Response({"message": f"Product with ID {product_id} not found."}, status=status.HTTP_404_NOT_FOUND)
                                else:
                                    product_obj, _ = Product.objects.get_or_create(
                                        product_name=product_name, hsn=hsn_code_obj, defaults={'unit_of_measure': unit_value}
                                    )
                                    product_description_obj = ProductDescription.objects.create(
                                        product=product_obj,
                                        description=description_text,
                                        unit=unit_value,
                                        rate=rate_value,
                                        product_amount=amount,
                                        cgst=cgst,
                                        sgst=sgst,
                                        igst=igst,
                                    )
                                    l = product_description_obj.unit
                                    product_summary = ProductSummaryDebitNote.objects.create(
                                        hsn=hsn_code_obj,
                                        product=product_obj,
                                        prod_description=product_description_obj
                                    )
                                    product_summaries.append(product_summary)
                                    debit_note.product_summaries.add(product_summary)
                                    
                                debit_note.save()
                            return Response({"message": "Debit Note created successfully."}, status=status.HTTP_200_OK)
            except Exception as e:
                return Response({"error": str(e)}, status=400)

    # except Exception as e:
    #     return Response({"error_message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    except Exception as e:
        error_details = traceback.format_exc()
        # print({"Error in update_sales_invoice" : str(e)})
        return Response({"error_message":str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET', 'PATCH'])
def debit_note_detail_view(request, client_pk, invoice_pk, debit_pk):
    try:
        # Fetch the sales invoice object
        debit_note = DebitNote.objects.get(client=client_pk, sales_invoice=invoice_pk, pk=debit_pk)
    except SalesInvoice.DoesNotExist:
        return Response({"error_message": "Sales invoice not found."}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        # Serialize and return the sales invoice details
        serializer = DebitNoteSerializerList(debit_note)
        return Response(serializer.data, status=status.HTTP_200_OK)

    elif request.method == 'PATCH':
        # Update only specific fields of the sales invoice
        data = request.data

        # Update fields from the request payload
        debit_note.invoice_no = data.get('invoice_no', debit_note.invoice_no)
        debit_note.invoice_date = data.get('invoice_date', debit_note.invoice_date)
        debit_note.po_no = data.get('po_no', debit_note.po_no)
        debit_note.po_date = data.get('po_date', debit_note.po_date)
        debit_note.invoice_type = data.get('invoice_type', debit_note.invoice_type)
        debit_note.entry_type = data.get('entry_type', debit_note.entry_type)
        debit_note.taxable_amount = safe_decimal(data.get('taxable_amount', debit_note.taxable_amount))
        debit_note.totalall_gst = safe_decimal(data.get('totalall_gst', debit_note.totalall_gst))
        debit_note.total_invoice_value = safe_decimal(data.get('total_invoice_value', debit_note.total_invoice_value))
        debit_note.tds_tcs_rate = safe_decimal(data.get('tds_tcs_rate', debit_note.tds_tcs_rate))
        debit_note.tds = safe_decimal(data.get('tds', debit_note.tds))
        debit_note.tcs = safe_decimal(data.get('tcs', debit_note.tcs))
        debit_note.amount_receivable = safe_decimal(data.get('amount_receivable', debit_note.amount_receivable))

        # Update product summaries if provided
        rows = data.get('rows', [])
        product_summaries = []
        for row in rows:
            hsn_code = row.get('hsnCode')
            gst_rate = safe_decimal(row.get('gstRate', '0'))
            product_name = row.get('product')
            description_text = row.get('description')
            unit_value = safe_decimal(row.get('unit', '0'))
            rate_value = safe_decimal(row.get('rate', '0'))
            amount = safe_decimal(row.get('product_amount', '0'))
            cgst = safe_decimal(row.get('cgst', '0'))
            sgst = safe_decimal(row.get('sgst', '0'))
            igst = safe_decimal(row.get('igst', '0'))

            # Handle HSNCode
            hsn_code_obj, created = HSNCode.objects.get_or_create(
                hsn_code=hsn_code,
                defaults={'gst_rate': gst_rate}
            )
            if not created and hsn_code_obj.gst_rate != gst_rate:
                hsn_code_obj.gst_rate = gst_rate
                hsn_code_obj.save()

            # Handle Product
            product_obj = Product.objects.filter(product_name=product_name, hsn=hsn_code_obj).first()
            if not product_obj:
                product_obj = Product.objects.create(
                    product_name=product_name,
                    hsn=hsn_code_obj,
                    unit_of_measure=unit_value
                )

            # Handle ProductDescription
            product_description_obj, created = ProductDescription.objects.get_or_create(
                product=product_obj,
                description=description_text,
                defaults={
                    'unit': unit_value,
                    'rate': rate_value,
                    'product_amount': amount,
                    'cgst': cgst,
                    'sgst': sgst,
                    'igst': igst
                }
            )
            if not created:
                # Update fields in ProductDescription
                product_description_obj.unit = unit_value
                product_description_obj.rate = rate_value
                product_description_obj.product_amount = amount
                product_description_obj.cgst = cgst
                product_description_obj.sgst = sgst
                product_description_obj.igst = igst
                product_description_obj.save()

            # Create ProductSummary
            product_summary = ProductSummaryDebitNote.objects.create(
                hsn=hsn_code_obj,
                product=product_obj,
                prod_description=product_description_obj
            )
            product_summaries.append(product_summary)

        # Save updates to the sales invoice
        if product_summaries:
            debit_note.product_summaries.set(product_summaries)
        debit_note.save()

        # Return the updated object
        updated_serializer = DebitNoteSerializer(debit_note)
        return Response(updated_serializer.data, status=status.HTTP_200_OK)
       
@api_view(['DELETE'])
def delete_debit_note(request, client_pk, invoice_pk, pk):
    """
    Deletes a SalesInvoice by its primary key (ID) along with its associated product summaries.
    """
    try:
        # Retrieve the Client instance
        client = Client.objects.filter(id=client_pk).first()

        if not client:
            return Response({"error_message": "Client not found."}, status=status.HTTP_404_NOT_FOUND)
        
        sales = SalesInvoice.objects.filter(pk=invoice_pk, client=client).first()
        if not sales:
            return Response({"error_message": "Sales Invoice not found or does not belong to the client."},
                            status=status.HTTP_404_NOT_FOUND)

        # Retrieve the SalesInvoice instance
        debit_note = DebitNote.objects.filter(id=pk, sales_invoice=sales ,client=client).first()

        if not debit_note:
            return Response({"error_message": "Debit Note not found."}, status=status.HTTP_404_NOT_FOUND)

        # Handle deletion of associated ProductSummary instances
        product_summaries = debit_note.product_summaries.all()
        for product_summary in product_summaries:
            product_summary.delete()

        # Delete the SalesInvoice instance
        debit_note.delete()

        return Response({"message": "Debit Note deleted successfully."}, status=status.HTTP_200_OK)

    # except Exception as e:
    #     return Response({"error_message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    except Exception as e:
        error_details = traceback.format_exc()
        # print({"Error in update_sales_invoice" : str(e)})
        return Response({"error_message":str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
def create_debit_note2(request, client_pk, invoice_pk):
    try:
        payload = request.data
        files = request.FILES

        # Fetch the Client
        client = Client.objects.filter(pk=client_pk).first()
        if not client:
            return Response({"error_message": "Client not found."}, status=status.HTTP_404_NOT_FOUND)

        # Fetch the Sales Invoice
        sales_invoice = SalesInvoice.objects.filter(pk=invoice_pk, client=client).first()
        if not sales_invoice:
            return Response({"error_message": "Sales Invoice not found or does not belong to the client."},
                            status=status.HTTP_404_NOT_FOUND)
            
        product_summaries_sales = sales_invoice.product_summaries.all()

        # Get the total units of the sales invoice
        total_units = [summary.prod_description.unit for summary in product_summaries_sales]  # Assuming `unit_value` is the total units in the SalesInvoice

        # Extract rows dynamically
        rows_data = defaultdict(dict)
        for key, value in payload.items():
            if key.startswith("rows["):
                row_index = key.split('[')[1].split(']')[0]
                field_name = key.split('[')[2].split(']')[0]
                rows_data[int(row_index)][field_name] = value
        rows = [rows_data[index] for index in sorted(rows_data.keys())]

        # Extract form data, vendor data, and invoice data
        form_data = {
            "offLocID": payload.get("formData[offLocID]"),
            "location": payload.get("formData[location]"),
            "contact": payload.get("formData[contact]"),
            "address": payload.get("formData[address]"),
            "city": payload.get("formData[city]"),
            "state": payload.get("formData[state]"),
            "country": payload.get("formData[country]"),
            "branchID": payload.get("formData[branchID]"),
        }
        vendor_data = {
            "name": payload.get("vendorData[name]"),
            "gst_no": payload.get("vendorData[gst_no]"),
            "pan": payload.get("vendorData[pan]"),
            "customer_address": payload.get("vendorData[customer_address]"),
            "customer": payload.get("vendorData[customer]", "").lower() == "true",
            "vendor": payload.get("vendorData[vendor]", "").lower() == "true",
            "email": payload.get("vendorData[email]"),
            "contact": payload.get("vendorData[contact]"),
        }
        invoice_data = {
            "invoice_no": payload.get("invoiceData[0][invoice_no]"),
            "invoice_date": payload.get("invoiceData[0][invoice_date]"),
            "po_no": payload.get("invoiceData[0][po_no]"),
            "po_date": payload.get("invoiceData[0][po_date]"),
            "invoice_type": payload.get("invoiceData[0][invoice_type]"),
            "entry_type": payload.get("invoiceData[0][entry_type]"),
            "taxable_amount": payload.get("invoiceData[0][taxable_amount]"),
            "totalall_gst": payload.get("invoiceData[0][totalall_gst]"),
            "total_invoice_value": payload.get("invoiceData[0][total_invoice_value]"),
            "tds_tcs_rate": payload.get("invoiceData[0][tds_tcs_rate]"),
            "tcs": payload.get("invoiceData[0][tcs]"),
            "tds": payload.get("invoiceData[0][tds]"),
            "amount_receivable": payload.get("invoiceData[0][amount_receivable]"),
        }
        attach_invoice = files.get("invoiceData[0][attach_invoice]")
        attach_e_way_bill = files.get("invoiceData[0][attach_e_way_bill]")

        # Handle Office Location creation or selection
        location_obj = None
        if form_data["offLocID"]:
            location_obj = OfficeLocation.objects.filter(id=form_data["offLocID"]).first()
            if not location_obj:
                return Response({"error_message": "Office Location not found."}, status=status.HTTP_404_NOT_FOUND)
        else:
            branch_instance = Branch.objects.filter(id=form_data["branchID"], client=client).first()
            if not branch_instance:
                return Response({"error_message": f"Branch with ID {form_data['branchID']} not found or doesn't belong to the client."},
                                status=status.HTTP_404_NOT_FOUND)
            location_obj = OfficeLocation.objects.create(
                location=form_data.get("location"),
                contact=form_data.get("contact"),
                address=form_data.get("address"),
                city=form_data.get("city"),
                state=form_data.get("state"),
                country=form_data.get("country"),
                branch=branch_instance
            )

        # Handle Vendor creation or update
        vendor_obj = None
        if vendor_data.get("gst_no"):
            existing_vendor = Customer.objects.filter(client=client, gst_no=vendor_data["gst_no"]).first()
            if existing_vendor:
                vendor_serializer = CustomerVendorSerializer(existing_vendor, data=vendor_data, partial=True)
                if vendor_serializer.is_valid():
                    vendor_obj = vendor_serializer.save()
                else:
                    return Response({"vendor_errors": vendor_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
            else:
                vendor_serializer = CustomerVendorSerializer(data=vendor_data)
                if vendor_serializer.is_valid():
                    vendor_obj = vendor_serializer.save(client=client)
                else:
                    return Response({"vendor_errors": vendor_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

        # Create the Debit Note
        total_units_remaining = total_units
        created_debit_notes = []  # List to store created debit notes

        for row in rows:
            # If the total_units_remaining becomes 0 or negative, stop creating more debit notes
            if total_units_remaining <= 0:
                return Response({"error_message": "No units left to create debit notes."}, status=status.HTTP_400_BAD_REQUEST)

            # Calculate the units for the current debit note
            unit_value = safe_decimal(row.get('unit', '0'))
            if unit_value <= total_units_remaining:
                total_units_remaining -= unit_value
            else:
                unit_value = total_units_remaining
                total_units_remaining = 0

            # Create the Debit Note for the current row
            debit_note = DebitNote.objects.create(
                client=client,
                sales_invoice=sales_invoice,
                attach_invoice=attach_invoice,
                attach_e_way_bill=attach_e_way_bill,
                unit_value=unit_value,
                **{
                    "invoice_no": payload.get("invoiceData[0][invoice_no]"),
                    "invoice_date": payload.get("invoiceData[0][invoice_date]"),
                    # "month": payload.get("invoiceData[0][month]"),
                    "po_no": payload.get("invoiceData[0][po_no]"),
                    "po_date": payload.get("invoiceData[0][po_date]"),
                    "invoice_type": payload.get("invoiceData[0][invoice_type]"),
                    "entry_type": payload.get("invoiceData[0][entry_type]"),
                    "taxable_amount": payload.get("invoiceData[0][taxable_amount]"),
                    "totalall_gst": payload.get("invoiceData[0][totalall_gst]"),
                    "total_invoice_value": payload.get("invoiceData[0][total_invoice_value]"),
                    "tds_tcs_rate": payload.get("invoiceData[0][tds_tcs_rate]"),
                    "tcs": payload.get("invoiceData[0][tcs]"),
                    "tds": payload.get("invoiceData[0][tds]"),
                    "amount_receivable": payload.get("invoiceData[0][amount_receivable]"),
                }
            )
            created_debit_notes.append(debit_note.id)

            # Add product summaries to the debit note
            for row in rows:
                product_summary = create_product_summary(row, debit_note)
                debit_note.product_summaries.add(product_summary)

        return Response({
            "message": f"Debit Notes created successfully. Created Debit Notes: {created_debit_notes}",
            "invoice_id": debit_note.id
        }, status=status.HTTP_200_OK)

    # except Exception as e:
    #     return Response({"error_message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    except Exception as e:
        error_details = traceback.format_exc()
        # print({"Error in update_sales_invoice" : str(e)})
        return Response({"error_message":str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# ******************************************************Credit Note****************************************************   

@api_view(['GET', 'POST'])
def create_credit_note_get(request, pk):
    try:
        client = Client.objects.get(id=pk)

    except Client.DoesNotExist:
        return Response({"error_message": "Client not found."}, status=404)

    if request.method == 'GET':
        # Initialize response context with default values
        context = {
            'serializer_customer': [],
            'serializer': [],
            'product_serializer': [],
            'message': None,
            'hsn': None,
            'location': None
        }

        # Retrieve query parameters
        received_value = request.GET.get('newValue')
        product_id = request.GET.get('productID')
        print('daadasdasdsa',received_value)
        # Process productID if it exists
        if product_id:
            try:
                hsn_cc = Product.objects.get(id=product_id).hsn
                context.update({
                    "message": "Product HSN found",
                    "hsn": HSNSerializer(hsn_cc).data
                })
            except (ValueError, Product.DoesNotExist):
                return Response({"error_message": "Invalid Product ID."}, status=400)

        # Process received_value if it exists
        if received_value:
            try:
                received_value = int(received_value)
                location = OfficeLocation.objects.get(id=received_value)
                branch_gst = location.branch.gst_no
                print("branch val",branch_gst)
                context.update({
                    "message": "Location found",
                    "location": OfficeLocationSerializer(location).data,
                    "branch_gst": branch_gst
                })
            except (ValueError, OfficeLocation.DoesNotExist):
                return Response({"error_message": "Invalid location ID."}, status=400)

        # Add data only if productID and received_value were not provided
        if not product_id and not received_value:
            off = OfficeLocation.objects.filter(branch__client=client)
            customer = Customer.objects.filter(client=client, customer=True)
            product = Product.objects.all()
            branch = Branch.objects.filter(client=client)

            serializer_customer = CustomerVendorSerializer(customer, many=True)
            serializer = OfficeLocationSerializer(off, many=True)
            product_serializer = ProductSerializer(product, many=True)
            branch_serializer = BranchSerailizer(branch, many=True)

            context.update({
                'serializer_customer': serializer_customer.data,
                'serializer': serializer.data,
                'product_serializer': product_serializer.data,
                'branch_serializer': branch_serializer.data
            })

        return Response(context)
    
@api_view(['POST'])
def create_credit_note2(request, client_pk, invoice_pk):
    try:
        payload = request.data
        print('payload', payload)

        # Fetch the Client
        client = Client.objects.filter(pk=client_pk).first()
        if not client:
            return Response({"error_message": "Client not found."}, status=status.HTTP_404_NOT_FOUND)

        # Fetch the Purchase Invoice
        purchase_invoice = PurchaseInvoice.objects.filter(pk=invoice_pk, client=client).first()
        if not purchase_invoice:
            return Response({"error_message": "Purchase Invoice not found or does not belong to the client."},
                            status=status.HTTP_404_NOT_FOUND)

        # Extract rows dynamically
        rows_data = defaultdict(dict)
        for key, value in payload.items():
            if key.startswith("rows["):  # Check if the key corresponds to rows
                row_index = key.split('[')[1].split(']')[0]
                field_name = key.split('[')[2].split(']')[0]
                rows_data[int(row_index)][field_name] = value
        rows = [rows_data[index] for index in sorted(rows_data.keys())]
        print('rows', rows)

        # Extract form data, vendor data, and invoice data
        form_data = {
            "offLocID": payload.get("formData[offLocID]"),
            "location": payload.get("formData[location]"),
            "contact": payload.get("formData[contact]"),
            "address": payload.get("formData[address]"),
            "city": payload.get("formData[city]"),
            "state": payload.get("formData[state]"),
            "country": payload.get("formData[country]"),
            "branchID": payload.get("formData[branchID]"),
        }
        vendor_data = {
            "name": payload.get("vendorData[name]"),
            "gst_no": payload.get("vendorData[gst_no]"),
            "pan": payload.get("vendorData[pan]"),
            "customer_address": payload.get("vendorData[customer_address]"),
            "customer": payload.get("vendorData[customer]", "").lower() == "true",
            "vendor": payload.get("vendorData[vendor]", "").lower() == "true",
            "email": payload.get("vendorData[email]"),
            "contact": payload.get("vendorData[contact]"),
        }
        invoice_data = {
            "invoice_no": payload.get("invoiceData[0][invoice_no]"),
            "invoice_date": payload.get("invoiceData[0][invoice_date]"),
            "month": payload.get("invoiceData[0][month]"),
            "invoice_type": payload.get("invoiceData[0][invoice_type]"),
            "entry_type": payload.get("invoiceData[0][entry_type]"),
            "taxable_amount": payload.get("invoiceData[0][taxable_amount]"),
            "totalall_gst": payload.get("invoiceData[0][totalall_gst]"),
            "total_invoice_value": payload.get("invoiceData[0][total_invoice_value]"),
            "tds_tcs_rate": payload.get("invoiceData[0][tds_tcs_rate]"),
            "tcs": payload.get("invoiceData[0][tcs]"),
            "tds": payload.get("invoiceData[0][tds]"),
            "amount_receivable": payload.get("invoiceData[0][amount_receivable]"),
            "utilise_edit": payload.get("invoiceData[0][utilise_edit]", "").lower() == "true", #nnnn
            "utilise_month": payload.get("invoiceData[0][utilise_month]").lower(), #nnnnnn
            
        }
        attach_invoice = request.FILES.get("invoiceData[0][attach_invoice]")
        attach_e_way_bill = request.FILES.get("invoiceData[0][attach_e_way_bill]")

        # Handle Office Location creation or selection
        location_obj = None
        if form_data["offLocID"]:
            location_obj = OfficeLocation.objects.filter(id=form_data["offLocID"]).first()
            if not location_obj:
                return Response({"error_message": "Office Location not found."}, status=status.HTTP_404_NOT_FOUND)
        else:
            branch_instance = Branch.objects.filter(id=form_data["branchID"], client=client).first()
            if not branch_instance:
                return Response({"error_message": f"Branch with ID {form_data['branchID']} not found or doesn't belong to the client."},
                                status=status.HTTP_404_NOT_FOUND)
            location_obj = OfficeLocation.objects.create(
                location=form_data.get("location"),
                contact=form_data.get("contact"),
                address=form_data.get("address"),
                city=form_data.get("city"),
                state=form_data.get("state"),
                country=form_data.get("country"),
                branch=branch_instance
            )

        # Handle Vendor creation or update
        vendor_obj = None
        if vendor_data.get("gst_no"):
            existing_vendor = Customer.objects.filter(client=client, gst_no=vendor_data["gst_no"]).first()
            if existing_vendor:
                vendor_serializer = CustomerVendorSerializer(existing_vendor, data=vendor_data, partial=True)
                if vendor_serializer.is_valid():
                    vendor_obj = vendor_serializer.save()
                else:
                    return Response({"vendor_errors": vendor_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
            else:
                vendor_serializer = CustomerVendorSerializer(data=vendor_data)
                if vendor_serializer.is_valid():
                    vendor_obj = vendor_serializer.save(client=client)
                else:
                    return Response({"vendor_errors": vendor_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

        # Create the Debit Note
        credit_note = CreditNote.objects.create(
            client=client,
            purchase_invoice=purchase_invoice,
            client_Location=location_obj,
            vendor=vendor_obj,
            attach_invoice=attach_invoice,
            attach_e_way_bill=attach_e_way_bill,
            **invoice_data
        )
        product_summaries = []  # To store created product summaries
        for row in rows:
            hsn_code = row.get('hsnCode')
            gst_rate = safe_decimal(row.get('gstRate', '0'))
            product_name = row.get('product')
            product_id = row.get('product_id')  # Assuming the frontend sends this if selecting an existing product
            description_text = row.get('description', '')
            unit_value = safe_decimal(row.get('unit', '0'))
            rate_value = safe_decimal(row.get('rate', '0'))
            amount = safe_decimal(row.get('product_amount', '0'))
            cgst = safe_decimal(row.get('cgst', '0'))
            sgst = safe_decimal(row.get('sgst', '0'))
            igst = safe_decimal(row.get('igst', '0'))

            # Handle HSNCode
            hsn_code_obj, _ = HSNCode.objects.get_or_create(
                hsn_code=hsn_code, defaults={'gst_rate': gst_rate}
            )

            # Handle Product (existing or new)
            if product_id:
                # Use existing product
                product_obj = Product.objects.filter(id=product_id).first()
                if not product_obj:
                    return Response({"error_message": f"Product with ID {product_id} not found."}, status=status.HTTP_404_NOT_FOUND)
            else:
                # Create new product
                product_obj, _ = Product.objects.get_or_create(
                    product_name=product_name, hsn=hsn_code_obj, defaults={'unit_of_measure': unit_value}
                )

            # Handle ProductDescription
            product_description_obj, _ = ProductDescription.objects.get_or_create(
                product=product_obj,
                description=description_text,
                defaults={
                    'unit': unit_value,
                    'rate': rate_value,
                    'product_amount': amount,
                    'cgst': cgst,
                    'sgst': sgst,
                    'igst': igst,
                }
            )

            # Create ProductSummary
            product_summary = ProductSummaryCreditNote.objects.create(
                hsn=hsn_code_obj,
                product=product_obj,
                prod_description=product_description_obj
            )
            product_summaries.append(product_summary)

            # Link ProductSummary to the SalesInvoice
            credit_note.product_summaries.add(product_summary)  # Add the product summary to the invoice


        # Handle Product Summaries (same logic as before)

        return Response({"message": "Credit Note created successfully.", "invoice_id": credit_note.id}, status=status.HTTP_200_OK)

    # except Exception as e:
    #     return Response({"error_message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    except Exception as e:
        error_details = traceback.format_exc()
        # print({"Error in update_sales_invoice" : str(e)})
        return Response({"error_message":str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
  
@api_view(['POST'])
def create_credit_note(request, client_pk, invoice_pk):
    client = Client.objects.get(id=client_pk)
    purchase = PurchaseInvoice.objects.get(id=invoice_pk, client=client)

    # Check if 'attach_e_way_bill' is in the request files
    if 'attach_invoice' in request.FILES:
        # Iterate through each file in the 'attach_e_way_bill' field
        for e_way_bill in request.FILES.getlist('attach_invoice'):
            # Prepare data for each file
            creditnote_data = {
                'attach_invoice': e_way_bill,  # The file being uploaded
                'client': client.id,  # Associate the file with the client
                'purchase_invoice' : purchase.id,
            }

            # Initialize the serializer for each file
            serializer = CreditNoteSerializer2(data=creditnote_data)

            # Check if the serializer is valid
            if serializer.is_valid():
                # Save each file as a separate object
                serializer.save()
            else:
                return Response({'error_message': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

        # If all files are processed, return success response
        return Response({'message': 'Credit Note E-way bill(s) uploa ded successfully.'}, status=status.HTTP_200_OK)
    else:
        return Response({'error_message': 'No files uploaded in the request.'}, status=status.HTTP_400_BAD_REQUEST)
    
@api_view(['GET', 'POST'])
def update_credit_note(request, client_pk, invoice_pk):

    try:
        # Handle GET request
        if request.method == 'GET':
            credit_note = PurchaseInvoice.objects.filter(client_id=client_pk, id=invoice_pk )\
                .select_related('client_Location', 'vendor') \
                .prefetch_related('product_summaries__hsn', 'product_summaries__prod_description') \
                .first()
            if not credit_note:
                return Response({"error_message": "Credit Note not found."}, status=status.HTTP_404_NOT_FOUND)

            debit_notes = CreditNote.objects.filter(purchase_invoice=credit_note)
            product_unit_sums = defaultdict(int)

            for dn in debit_notes:
                for product_summary in dn.product_summaries.all():
                    product_name = product_summary.product_name()
                    product_unit = product_summary.unit() or 0  # Agar unit None ho toh 0 lelo
                    product_unit_sums[product_name] += product_unit

            remaining_units = {}
            for product in credit_note.product_summaries.all():
                product_name = product.product_name()
                si_unit = product.unit() or 0  # Sales Invoice ke unit
                dn_unit_sum = product_unit_sums[product_name]  # Debit Notes ka unit sum
                remaining_units[product_name] = si_unit - dn_unit_sum  # Remaining units

            print('ooooooooo',remaining_units)

            credit_notes_details = []
            for debit_note in debit_notes:
                product_summaries = debit_note.product_summaries.all()
                print(f"Found {len(product_summaries)} ProductSummaries for DebitNote {debit_note.id}")

                for product_summary in product_summaries:
                    product_details = {
                        'debit_note_id': debit_note.id,
                        'product_name': product_summary.product_name(),
                        'unit': product_summary.unit(),
                    }
                    credit_notes_details.append(product_details)

            dd = credit_notes_details
            print('Credit Notes',dd)


            credit_note_data = PurchaseSerializer3(credit_note).data
            
            print('creditnote',credit_note_data)

            product_summaries = credit_note.product_summaries.all()
            product_summary_data = [
                {
                    "id": summary.id,
                    "hsnCode": summary.hsn.hsn_code,
                    "gstRate": summary.hsn.gst_rate,
                    "product": summary.product.product_name,
                    "description": summary.prod_description.description,
                    # "unit": summary.prod_description.unit,
                    "unit": remaining_units.get(summary.product_name(), summary.prod_description.unit),
                    "rate": summary.prod_description.rate,
                    "product_amount": summary.prod_description.product_amount,
                    "cgst": summary.prod_description.cgst,
                    "sgst": summary.prod_description.sgst,
                    "igst": summary.prod_description.igst,
                }
                for summary in product_summaries
            ]

            response_data = {
                "credit_note": credit_note_data,
                "product_summaries": product_summary_data,
                "client_location": {
                    "id": credit_note.client_Location.id if credit_note.client_Location else None,
                    "location": credit_note.client_Location.location if credit_note.client_Location else None,
                    "contact": credit_note.client_Location.contact if credit_note.client_Location else None,
                    "address": credit_note.client_Location.address if credit_note.client_Location else None,
                    "city": credit_note.client_Location.city if credit_note.client_Location else None,
                    "state": credit_note.client_Location.state if credit_note.client_Location else None,
                    "country": credit_note.client_Location.country if credit_note.client_Location else None,
                    "branchID": credit_note.client_Location.branch.id if credit_note.client_Location else None,
                },
                "vendor": {
                    "id": credit_note.vendor.id if credit_note.vendor else None,
                    "name": credit_note.vendor.name if credit_note.vendor else None,
                    "gst_no": credit_note.vendor.gst_no if credit_note.vendor else None,
                    "pan": credit_note.vendor.pan if credit_note.vendor else None,
                    "customer_address": credit_note.vendor.address if credit_note.vendor else None,
                    "customer": credit_note.vendor.customer if credit_note.vendor else None,
                    "vendor": credit_note.vendor.vendor if credit_note.vendor else None,
                    "email": credit_note.vendor.email if credit_note.vendor else None,
                    "contact": credit_note.vendor.contact if credit_note.vendor else None,
                },
                # print('bbbbbbbbbbbbbbbbbbbbbbb',response_data["vendor"])
            }
            print('bbbbbbbbbbbbbbbbbbbbbbb',response_data["vendor"])
            return Response(response_data, status=status.HTTP_200_OK)

        # Handle PUT request
        elif request.method == 'POST':
            try:
                with transaction.atomic(): 
                    payload = request.data
                    data = request.data
                    # print('payload', payload)

                    # Fetch the Client
                    client = Client.objects.filter(pk=client_pk).first()
                    if not client:
                        return Response({"message": "Client not found."}, status=status.HTTP_404_NOT_FOUND)

                    # Fetch the Purchase Invoice
                    purchase_invoice = PurchaseInvoice.objects.filter(client_id=client_pk, id=invoice_pk )\
                        .select_related('client_Location', 'vendor') \
                        .prefetch_related('product_summaries__hsn', 'product_summaries__prod_description') \
                        .first()
                    if not purchase_invoice:
                        return Response({"message": "Credit Note not found."}, status=status.HTTP_404_NOT_FOUND)

                    debit_notes = CreditNote.objects.filter(purchase_invoice=purchase_invoice)
                    product_unit_sums = defaultdict(int)

                    for dn in debit_notes:
                        for product_summary in dn.product_summaries.all():
                            product_name = product_summary.product_name()
                            product_unit = product_summary.unit() or 0  # Agar unit None ho toh 0 lelo
                            product_unit_sums[product_name] += product_unit

                    remaining_units = {}
                    for product in purchase_invoice.product_summaries.all():
                        product_name = product.product_name()
                        si_unit = product.unit() or 0  # Sales Invoice ke unit
                        dn_unit_sum = product_unit_sums[product_name]  # Debit Notes ka unit sum
                        remaining_units[product_name] = si_unit - dn_unit_sum  # Remaining units

                    print('gggggggg',remaining_units)

                    
                    # Validations..........
                    if all(unit == 0 for unit in remaining_units.values()):
                        return Response(
                            {"message": "Cannot create a new Credit Note. All product remaining units are 0."},
                            status=status.HTTP_400_BAD_REQUEST
                        )
                    
                    for product_name, unit in remaining_units.items():
                        if unit is None:
                            return Response({"message" : f"Remaining units for product '{product_name}' is missing or undefined."}, status=status.HTTP_400_BAD_REQUEST)
                        if unit < 0:
                            return Response({"message" : f"Remaining units for prooooooooooooduct '{product_name}' cannot be negative. Current value: {unit}"}, status=status.HTTP_400_BAD_REQUEST)


                    # Extract rows dynamically
                    rows_data = defaultdict(dict)
                    for key, value in payload.items():
                        if key.startswith("rows["):  # Check if the key corresponds to rows
                            row_index = key.split('[')[1].split(']')[0]
                            field_name = key.split('[')[2].split(']')[0]
                            rows_data[int(row_index)][field_name] = value
                    rows = [rows_data[index] for index in sorted(rows_data.keys())]
                    print('rows', rows)

                    # Extract form data, vendor data, and invoice data
                    form_data = {
                        "offLocID": payload.get("formData[offLocID]"),
                        "location": payload.get("formData[location]"),
                        "contact": payload.get("formData[contact]"),
                        "address": payload.get("formData[address]"),
                        "city": payload.get("formData[city]"),
                        "state": payload.get("formData[state]"),
                        "country": payload.get("formData[country]"),
                        "branchID": payload.get("formData[branchID]"),
                    }
                    vendor_data = {
                        "name": payload.get("vendorData[name]"),
                        "gst_no": payload.get("vendorData[gst_no]"),
                        "pan": payload.get("vendorData[pan]"),
                        "customer_address": payload.get("vendorData[customer_address]"),
                        "email": payload.get("vendorData[email]"),
                        "contact": payload.get("vendorData[contact]"),
                        "customer": payload.get("vendorData[customer]", "").lower() == "true",
                        "vendor": payload.get("vendorData[vendor]", "").lower() == "true",
                    }
                    invoice_data = {
                        "invoice_no": payload.get("invoiceData[0][invoice_no]"),
                        "invoice_date": payload.get("invoiceData[0][invoice_date]"),
                        "month": payload.get("invoiceData[0][month]"),
                        "invoice_type": payload.get("invoiceData[0][invoice_type]"),
                        "entry_type": payload.get("invoiceData[0][entry_type]"),
                        "taxable_amount": payload.get("invoiceData[0][taxable_amount]"),
                        "totalall_gst": payload.get("invoiceData[0][totalall_gst]"),
                        "total_invoice_value": payload.get("invoiceData[0][total_invoice_value]"),
                        "tds_tcs_rate": payload.get("invoiceData[0][tds_tcs_rate]"),
                        "tcs": payload.get("invoiceData[0][tcs]"),
                        "tds": payload.get("invoiceData[0][tds]"),
                        "amount_receivable": payload.get("invoiceData[0][amount_receivable]"),
                        "utilise_edit": payload.get("invoiceData[0][utilise_edit]", "").lower() == "true", #nnnn
                        "utilise_month": payload.get("invoiceData[0][utilise_month]"), #nnnnnn
                        
                    }
                    if invoice_data["invoice_date"]:
                        try:
                            invoice_data["invoice_date"] = datetime.strptime(invoice_data["invoice_date"], "%d-%m-%Y").date()
                        except ValueError:
                            return Response({"error_message": "Invalid Invoice date format, expected DD-MM-YYYY."}, status=status.HTTP_400_BAD_REQUEST)
                    print("Final invoice date:", invoice_data["invoice_date"])

                    if invoice_data["month"]:
                        try:
                            invoice_data["month"] = datetime.strptime(invoice_data["month"], "%d-%m-%Y").date()
                        except ValueError:
                            return Response({"error_message": "Invalid Month date format, expected DD-MM-YYYY."},status=status.HTTP_400_BAD_REQUEST)

                    if invoice_data["utilise_month"]:
                        try:
                            invoice_data["utilise_month"] = datetime.strptime(invoice_data["utilise_month"], "%d-%m-%Y").date()
                        except ValueError:
                            return Response({"error_message": "Invalid Utilise Month date format, expected DD-MM-YYYY."},status=status.HTTP_400_BAD_REQUEST)

                    print("Final month:", invoice_data["month"])
                    attach_invoice = request.FILES.get("invoiceData[0][attach_invoice]")
                    attach_e_way_bill = request.FILES.get("invoiceData[0][attach_e_way_bill]")

                    # Handle Vendor creation or update
                    vendor_obj = None
                    if vendor_data.get("gst_no"):
                        existing_vendor = Customer.objects.filter(client=client, gst_no=vendor_data["gst_no"]).first()
                        if existing_vendor:
                            vendor_serializer = CustomerVendorSerializer(existing_vendor, data=vendor_data, partial=True)
                            if vendor_serializer.is_valid():
                                vendor_obj = vendor_serializer.save()
                            else:
                                return Response({"vendor_errors": vendor_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
                        else:
                            vendor_serializer = CustomerVendorSerializer(data=vendor_data)
                            if vendor_serializer.is_valid():
                                vendor_obj = vendor_serializer.save(client=client)
                            else:
                                return Response({"vendor_errors": vendor_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

                    if any(unit is None for unit in remaining_units.values()):
                        return Response({"message" : f"Remaining units for product {product} is missing."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
                    else:
                        credit_note_serializer = CreditNoteSerializer(data=data)

                        if credit_note_serializer.is_valid():
                            # Handle Office Location creation or selection
                            location_obj = None
                            if form_data["offLocID"]:
                                location_obj = OfficeLocation.objects.filter(id=form_data["offLocID"]).first()
                                if not location_obj:
                                    return Response({"message": "Office Location not found."}, status=status.HTTP_404_NOT_FOUND)
                            else:
                                branch_instance = Branch.objects.filter(id=form_data["branchID"], client=client).first()
                                if not branch_instance:
                                    return Response({"message": f"Branch with ID {form_data['branchID']} not found or doesn't belong to the client."},
                                                    status=status.HTTP_404_NOT_FOUND)
                                location_obj, _ = OfficeLocation.objects.get_or_create(
                                    location=form_data.get("location"),
                                    contact=form_data.get("contact"),
                                    address=form_data.get("address"),
                                    city=form_data.get("city"),
                                    state=form_data.get("state"),
                                    country=form_data.get("country"),
                                    branch=branch_instance
                                )
                            for row in rows:
                                product_name = row.get('product')
                                unit_value = safe_decimal(row.get('unit', '0'))
                                
                                # Skip rows with unit value as 0 or negative
                                if unit_value <= 0:
                                    continue  # Skip processing this row
                                
                                # Validate remaining units for the product
                                if product_name in remaining_units:
                                    if remaining_units[product_name] < unit_value:
                                        return Response(
                                            {
                                                "message": f"Not enough units remaining for the product '{product_name}'. "
                                                        f"Available: {remaining_units[product_name]}, Requested: {unit_value}."
                                            },
                                            status=status.HTTP_400_BAD_REQUEST
                                        )
                                    
                                    # Deduct the used units from the remaining_units
                                    remaining_units[product_name] -= unit_value
                                else:
                                    return Response(
                                        {"message": f"Product '{product_name}' not found in remaining units."},
                                        status=status.HTTP_400_BAD_REQUEST
                                    )

                            # Create the Debit Note
                            credit_note = CreditNote.objects.create(
                                client=client,
                                purchase_invoice=purchase_invoice,
                                client_Location=location_obj,
                                vendor=vendor_obj,
                                attach_invoice=attach_invoice,
                                attach_e_way_bill=attach_e_way_bill,
                                **invoice_data
                            )
                            product_summaries = []
                            for row in rows:
                                product_name = row.get('product')
                                unit_value = safe_decimal(row.get('unit', '0'))
                                hsn_code = row.get('hsnCode')
                                gst_rate = safe_decimal(row.get('gstRate', '0'))
                                product_id = row.get('product_id')
                                description_text = row.get('description', '')
                                rate_value = safe_decimal(row.get('rate', '0'))
                                amount = safe_decimal(row.get('product_amount', '0'))
                                cgst = safe_decimal(row.get('cgst', '0'))
                                sgst = safe_decimal(row.get('sgst', '0'))
                                igst = safe_decimal(row.get('igst', '0'))

                                hsn_code_obj, _ = HSNCode.objects.get_or_create(
                                    hsn_code=hsn_code, defaults={'gst_rate': gst_rate}
                                )

                                if product_id:
                                    product_obj = Product.objects.filter(id=product_id).first()
                                    if not product_obj:
                                        return Response({"message": f"Product with ID {product_id} not found."}, status=status.HTTP_404_NOT_FOUND)
                                else:
                                    product_obj, _ = Product.objects.get_or_create(
                                        product_name=product_name, hsn=hsn_code_obj, defaults={'unit_of_measure': unit_value}
                                    )
                                    product_description_obj = ProductDescription.objects.create(
                                        product=product_obj,
                                        description=description_text,
                                        unit=unit_value,
                                        rate=rate_value,
                                        product_amount=amount,
                                        cgst=cgst,
                                        sgst=sgst,
                                        igst=igst,
                                    )
                                    l = product_description_obj.unit
                                    product_summary = ProductSummaryCreditNote.objects.create(
                                        hsn=hsn_code_obj,
                                        product=product_obj,
                                        prod_description=product_description_obj
                                    )
                                    product_summaries.append(product_summary)
                                    credit_note.product_summaries.add(product_summary)
                                    
                                credit_note.save()
                            return Response({"message": "Credit Note created successfully."}, status=status.HTTP_200_OK)
                        else: 
                            return Response({"message": credit_note_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
                # else:
                #     raise Exception("Client data is invalid!")  # Force rollback
            except Exception as e:
                return Response({"error": str(e)}, status=400)

    except Exception as e:
        error_details = traceback.format_exc()
        # print({"Error in update_sales_invoice" : str(e)})
        return Response({"error_message":str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['DELETE'])
def delete_credit_note(request, client_pk, invoice_pk, credit_pk):
    """
    Deletes a SalesInvoice by its primary key (ID) along with its associated product summaries.
    """
    try:
        # Retrieve the Client instance
        client = Client.objects.filter(id=client_pk).first()

        if not client:
            return Response({"error_message": "Client not found."}, status=status.HTTP_404_NOT_FOUND)
        
        purchase = PurchaseInvoice.objects.filter(pk=invoice_pk, client=client).first()
        if not purchase:
            return Response({"error_message": "Purchase Invoice not found or does not belong to the client."},
                            status=status.HTTP_404_NOT_FOUND)

        # Retrieve the SalesInvoice instance
        credit_note = CreditNote.objects.filter(id=credit_pk, purchase_invoice=purchase ,client=client).first()

        if not credit_note:
            return Response({"error_message": "Credit Note not found."}, status=status.HTTP_404_NOT_FOUND)

        # Handle deletion of associated ProductSummary instances
        product_summaries = credit_note.product_summaries.all()
        for product_summary in product_summaries:
            product_summary.delete()

        # Delete the SalesInvoice instance
        credit_note.delete()

        return Response({"message": "Credit Note deleted successfully."}, status=status.HTTP_200_OK)

    # except Exception as e:
    #     return Response({"error_message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    except Exception as e:
        error_details = traceback.format_exc()
        # print({"Error in update_sales_invoice" : str(e)})
        return Response({"error_message":str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET', 'PATCH'])
def credit_note_detail_view(request, client_pk, invoice_pk, credit_pk):
    try:
        # Fetch the sales invoice object
        credit_note = CreditNote.objects.get(client=client_pk, purchase_invoice=invoice_pk, pk=credit_pk)
    except PurchaseInvoice.DoesNotExist:
        return Response({"error_message": "Purchase invoice not found."}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        # Serialize and return the sales invoice details
        serializer = CreditNoteSerializerList(credit_note)
        return Response(serializer.data, status=status.HTTP_200_OK)

    elif request.method == 'PATCH':
        # Update only specific fields of the sales invoice
        data = request.data

        # Update fields from the request payload
        credit_note.invoice_no = data.get('invoice_no', credit_note.invoice_no)
        credit_note.invoice_date = data.get('invoice_date', credit_note.invoice_date)
        credit_note.month = data.get('month', credit_note.month)
        credit_note.invoice_type = data.get('invoice_type', credit_note.invoice_type)
        credit_note.entry_type = data.get('entry_type', credit_note.entry_type)
        credit_note.taxable_amount = safe_decimal(data.get('taxable_amount', credit_note.taxable_amount))
        credit_note.totalall_gst = safe_decimal(data.get('totalall_gst', credit_note.totalall_gst))
        credit_note.total_invoice_value = safe_decimal(data.get('total_invoice_value', credit_note.total_invoice_value))
        credit_note.tds_tcs_rate = safe_decimal(data.get('tds_tcs_rate', credit_note.tds_tcs_rate))
        credit_note.tds = safe_decimal(data.get('tds', credit_note.tds))
        credit_note.tcs = safe_decimal(data.get('tcs', credit_note.tcs))
        credit_note.amount_receivable = safe_decimal(data.get('amount_receivable', credit_note.amount_receivable))
        credit_note.utilise_edit = data.get('utilise_edit', credit_note.utilise_edit)
        credit_note.utilise_month = data.get('utilise_month', credit_note.utilise_month)


        # Update product summaries if provided
        rows = data.get('rows', [])
        product_summaries = []
        for row in rows:
            hsn_code = row.get('hsnCode')
            gst_rate = safe_decimal(row.get('gstRate', '0'))
            product_name = row.get('product')
            description_text = row.get('description')
            unit_value = safe_decimal(row.get('unit', '0'))
            rate_value = safe_decimal(row.get('rate', '0'))
            amount = safe_decimal(row.get('product_amount', '0'))
            cgst = safe_decimal(row.get('cgst', '0'))
            sgst = safe_decimal(row.get('sgst', '0'))
            igst = safe_decimal(row.get('igst', '0'))

            # Handle HSNCode
            hsn_code_obj, created = HSNCode.objects.get_or_create(
                hsn_code=hsn_code,
                defaults={'gst_rate': gst_rate}
            )
            if not created and hsn_code_obj.gst_rate != gst_rate:
                hsn_code_obj.gst_rate = gst_rate
                hsn_code_obj.save()

            # Handle Product
            product_obj = Product.objects.filter(product_name=product_name, hsn=hsn_code_obj).first()
            if not product_obj:
                product_obj = Product.objects.create(
                    product_name=product_name,
                    hsn=hsn_code_obj,
                    unit_of_measure=unit_value
                )

            # Handle ProductDescription
            product_description_obj, created = ProductDescription.objects.get_or_create(
                product=product_obj,
                description=description_text,
                defaults={
                    'unit': unit_value,
                    'rate': rate_value,
                    'product_amount': amount,
                    'cgst': cgst,
                    'sgst': sgst,
                    'igst': igst
                }
            )
            if not created:
                # Update fields in ProductDescription
                product_description_obj.unit = unit_value
                product_description_obj.rate = rate_value
                product_description_obj.product_amount = amount
                product_description_obj.cgst = cgst
                product_description_obj.sgst = sgst
                product_description_obj.igst = igst
                product_description_obj.save()

            # Create ProductSummary
            product_summary = ProductSummaryCreditNote.objects.create(
                hsn=hsn_code_obj,
                product=product_obj,
                prod_description=product_description_obj
            )
            product_summaries.append(product_summary)

        # Save updates to the sales invoice
        if product_summaries:
            credit_note.product_summaries.set(product_summaries)
        credit_note.save()

        # Return the updated object
        updated_serializer = CreditNoteSerializer(credit_note)
        return Response(updated_serializer.data, status=status.HTTP_200_OK)

# ******************************************************Income****************************************************
   
@api_view(['GET','POST'])
def create_income_get(request, pk):
    try:
        client = Client.objects.get(id=pk)
    except Client.DoesNotExist:
        return Response({'error_message':'Client not Found.'}, status=404)

    if request.method == 'GET':
        context = {
            'serializer_customer':[],
            'serializer': [],
            'product_serializer':[],
            'message':None,
            'hsn':None,
            'location':None,
        }
        received_value = request.GET.get('newValue')
        product_Id = request.GET.get('productID')
        if product_Id:
            try:
                hsn_cc = Product.objects.get(id=product_Id).hsn
                context.update({
                    'message':'Product HSN found',
                    'hsn':HSNSerializer(hsn_cc).data
                })
            except (ValueError, Product.DoesNotExist):
                return Response({'error_message':'Invalid Product ID'}, status=400)

        if received_value:
            try:
                received_value = int(received_value)
                location = OfficeLocation.objects.get(id=received_value)
                branch_gst = location.branch.gst_no
                print('branch val', branch_gst)
                context.update({
                    "message":"Location Found",
                    "location" : OfficeLocationSerializer(location).data,
                    "branch_gst": branch_gst
                })
            except (ValueError, OfficeLocation.DoesNotExist):
                return Response({'error_message':'Invalid location ID.'}, status=400)

        if not product_Id and not received_value:
            off = OfficeLocation.objects.filter(branch__client=client)
            customer = Customer.objects.filter(client=client, customer=True)
            product = Product.objects.all()
            branch = Branch.objects.filter(client=client)

            serializer_customer = CustomerVendorSerializer(customer, many=True)
            serializers = OfficeLocationSerializer(off, many=True)
            product_serializer = ProductSerializer(product, many=True)
            branch_serializer = BranchSerailizer(branch, many= True)

            context.update({
                'serializer_customer' : serializer_customer.data,
                'serializer' : serializers.data,
                'product_serializer' : product_serializer.data,
                'branch_serializer' : branch_serializer.data
            })
        return Response(context)

@api_view(['POST'])
def create_income(request, pk):
    client = Client.objects.get(id=pk)

    # Check if 'attach_e_way_bill' is in the request files
    if 'attach_invoice' in request.FILES:
        # Iterate through each file in the 'attach_e_way_bill' field
        for e_way_bill in request.FILES.getlist('attach_invoice'):
            # Prepare data for each file
            income_data = {
                'attach_invoice': e_way_bill,  # The file being uploaded
                'client': client.id,  # Associate the file with the client
            }

            # Initialize the serializer for each file
            serializer = IncomeSerializer2(data=income_data)

            # Check if the serializer is valid
            if serializer.is_valid():
                # Save each file as a separate object
                serializer.save()
            else:
                return Response({'error_message': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

        # If all files are processed, return success response
        return Response({'message': 'Income E-way bill(s) uploaded successfully.'}, status=status.HTTP_200_OK)
    else:
        return Response({'error_message': 'No files uploaded in the request.'}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET', 'PUT'])
def update_income(request, client_pk, invoice_pk):

    try:
        # Handle GET request
        if request.method == 'GET':
            income = Income.objects.filter(client_id=client_pk, id=invoice_pk).first()
            if not income:
                return Response({"error_message": "Income not found."}, status=status.HTTP_404_NOT_FOUND)

            income_data = IncomeSerializer3(income).data
            
            print('income',income_data)

            product_summaries = income.product_summaries.all()
            product_summary_data = [
                {
                    "id": summary.id,
                    "hsnCode": summary.hsn.hsn_code,
                    "gstRate": summary.hsn.gst_rate,
                    "product": summary.product.product_name,
                    "description": summary.prod_description.description,
                    "unit": summary.prod_description.unit,
                    "rate": summary.prod_description.rate,
                    "product_amount": summary.prod_description.product_amount,
                    "cgst": summary.prod_description.cgst,
                    "sgst": summary.prod_description.sgst,
                    "igst": summary.prod_description.igst,
                }
                for summary in product_summaries
            ]

            if income.month:
                formatted_month = income.month.strftime("%d-%m-%Y")  # Directly format date
            else:
                formatted_month = None

            if income.invoice_date:
                formatted_invoice_date = income.invoice_date.strftime("%d-%m-%Y")  # Directly format date
            else:
                formatted_invoice_date = None

            response_data = {
                "income": {
                    **income_data,
                    "month": formatted_month,  # Ensure month is sent in dd/mm/yyyy format
                    "invoice_date" : formatted_invoice_date,},
                "product_summaries": product_summary_data,
                "client_location": {
                    "id": income.client_Location.id if income.client_Location else None,
                    "location": income.client_Location.location if income.client_Location else None,
                    "contact": income.client_Location.contact if income.client_Location else None,
                    "address": income.client_Location.address if income.client_Location else None,
                    "city": income.client_Location.city if income.client_Location else None,
                    "state": income.client_Location.state if income.client_Location else None,
                    "country": income.client_Location.country if income.client_Location else None,
                    "branchID": income.client_Location.branch.id if income.client_Location else None,
                },
                "customer": {
                    "id": income.customer.id if income.customer else None,
                    "name": income.customer.name if income.customer else None,
                    "gst_no": income.customer.gst_no if income.customer else None,
                    "pan": income.customer.pan if income.customer else None,
                    "customer_address": income.customer.address if income.customer else None,
                    "customer": income.customer.customer if income.customer else None,
                    "vendor": income.customer.vendor if income.customer else None,
                    "email": income.customer.email if income.customer else None,
                    "contact": income.customer.contact if income.customer else None,
                },
            }
            # print('bbbbbbbbbbbbbbbbbbbbbbb',response_data["income"])
            return Response(response_data, status=status.HTTP_200_OK)

        # Handle PUT request
        elif request.method == 'PUT':
            print('youuuuuuuuu',request.FILES)
            income = Income.objects.filter(client_id=client_pk, id=invoice_pk).first()
            if not income:
                return Response({"error_message": "Income not found."}, status=status.HTTP_404_NOT_FOUND)

            # Extract data from the request
            # form_data = request.data.get('formData', {})
            # vendor_data = request.data.get('vendorData', {})
            # rows = request.data.get('rows', [])
            payload = request.data

            # Container to hold rows
            rows_data = defaultdict(dict)

            # Iterate over keys dynamically
            for key, value in payload.items():
                if key.startswith("rows["):  # Check if the key corresponds to rows
                    # Extract the row index and field name
                    row_index = key.split('[')[1].split(']')[0]  # Get the index (e.g., '0', '1', '2')
                    field_name = key.split('[')[2].split(']')[0]  # Get the field name (e.g., 'product')

                    # Store the value in the corresponding row and field
                    rows_data[int(row_index)][field_name] = value

            # Convert defaultdict to a regular list for easier use
            rows = [rows_data[index] for index in sorted(rows_data.keys())]

            # Output the rows data
            print(rows)

            invoice_data = request.data.get('invoiceData', [{}])[0]  # Extract the first item
            invoice_file = request.data.get('invoice_file')
            print('invoice_file',invoice_data)
            form_data = {
                "offLocID": request.data.get("formData[offLocID]"),
                "location": request.data.get("formData[location]"),
                "contact": request.data.get("formData[contact]"),
                "address": request.data.get("formData[address]"),
                "city": request.data.get("formData[city]"),
                "state": request.data.get("formData[state]"),
                "country": request.data.get("formData[country]"),
                "branchID": request.data.get("formData[branchID]"),
            }
            vendor_data = {
                "name": request.data.get("vendorData[name]"),
                "gst_no": request.data.get("vendorData[gst_no]"),
                "pan": request.data.get("vendorData[pan]"),
                "customer_address": request.data.get("vendorData[customer_address]"),
                "email": request.data.get("vendorData[email]"),
                "contact": request.data.get("vendorData[contact]"),
                "customer": request.data.get("vendorData[customer]").lower() == "true" if request.data.get("vendorData[customer]") else None,
                "vendor": request.data.get("vendorData[vendor]").lower() == "true" if request.data.get("vendorData[vendor]") else None,
            }

            # Access invoice_data in a similar way
            invoice_data = {
                "invoice_no": request.data.get("invoiceData[0][invoice_no]"),
                "invoice_date": request.data.get("invoiceData[0][invoice_date]"),
                "month": request.data.get("invoiceData[0][month]"),
                "invoice_type": request.data.get("invoiceData[0][invoice_type]"),
                "entry_type": request.data.get("invoiceData[0][entry_type]"),
                "taxable_amount": request.data.get("invoiceData[0][taxable_amount]"),
                "totalall_gst": request.data.get("invoiceData[0][totalall_gst]"),
                "total_invoice_value": request.data.get("invoiceData[0][total_invoice_value]"),
                "tds_tcs_rate": request.data.get("invoiceData[0][tds_tcs_rate]"),
                "tcs": request.data.get("invoiceData[0][tcs]"),
                "tds": request.data.get("invoiceData[0][tds]"),
                "amount_receivable": request.data.get("invoiceData[0][amount_receivable]"),
                "attach_invoice": request.data.get("invoiceData[0][attach_invoice]"),
                "attach_e_way_bill": request.data.get("invoiceData[0][attach_e_way_bill]"),
            }

            if invoice_data["invoice_date"]:
                try:
                    invoice_data["invoice_date"] = datetime.strptime(invoice_data["invoice_date"], "%d-%m-%Y").date()
                except ValueError:
                    return Response({"error_message": "Invalid Invoice date format, expected DD-MM-YYYY."}, status=status.HTTP_400_BAD_REQUEST)
            print("Final invoice date:", invoice_data["invoice_date"])

            if invoice_data["month"]:
                try:
                    invoice_data["month"] = datetime.strptime(invoice_data["month"], "%d-%m-%Y").date()
                except ValueError:
                    return Response({"error_message": "Invalid Month date format, expected DD-MM-YYYY."},status=status.HTTP_400_BAD_REQUEST)


            attach_invoice = request.FILES.get("invoiceData[0][attach_invoice]")
            attach_e_way_bill = request.FILES.get("invoiceData[0][attach_e_way_bill]")
                        # Update the income instance fields
            if attach_invoice:
                income.attach_invoice = attach_invoice

            if attach_e_way_bill:
                income.attach_e_way_bill = attach_e_way_bill

            for field, value in invoice_data.items():
                if field not in ['attach_invoice', 'attach_e_way_bill']:  # Skip file fields
                    if hasattr(income, field):
                        setattr(
                            income,
                            field,
                            safe_decimal(value) if field in [
                                'taxable_amount', 'totalall_gst', 'total_invoice_value',
                                'tds_tcs_rate', 'tds', 'tcs', 'amount_receivable'
                            ] else value
                        )


            attach_invoice = invoice_data.get('attach_invoice')
            print('uoiuoiuiuiuiuio',attach_invoice)
             # Log the flattened data
            # print("Flattened form_data:", form_data)
            print("Flattened invoice_data:", invoice_data)

            print('request payload',request.data)
            # print('form_data',form_data)
            # Update invoice file
            if invoice_file:
                income.invoice_file = invoice_file
                income.save()
                return Response({"message": "Invoice file uploaded successfully."}, status=status.HTTP_200_OK)

            # Update or create vendor
            # Update or create vendor (Customer)
            if vendor_data:
                # Check for 'customer_address' in vendor data and map it to 'address'
                if 'customer_address' in vendor_data:
                    vendor_data['address'] = vendor_data.pop('customer_address')  # Replace 'customer_address' with 'address'

                vendor_id = request.data.get("vendorData[vendorID]")  # Retrieve vendorID if provided

                if vendor_id:
                    vendor_obj = Customer.objects.filter(id=vendor_id).first()
                    if vendor_obj:
                        if vendor_obj.gst_no == vendor_data.get("gst_no"):
                            vendor_serializer = CustomerVendorSerializer(vendor_obj, data=vendor_data, partial=True)
                            if vendor_serializer.is_valid():
                                vendor_serializer.save()
                            else:
                                return Response({"vendor_errors": vendor_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
                        else:
                            # Create a new vendor if gst_no is changed
                            vendor_serializer = CustomerVendorSerializer(data=vendor_data)
                            if vendor_serializer.is_valid():
                                vendor_obj = vendor_serializer.save(client=income.client)
                            else:
                                return Response({"vendor_errors": vendor_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
                    else:
                        return Response({"error_message": f"Vendor with ID {vendor_id} not found."}, status=status.HTTP_404_NOT_FOUND)
                else:
                    existing_vendor = Customer.objects.filter(client=income.client, gst_no=vendor_data.get("gst_no")).first()
                    if existing_vendor:
                        vendor_obj = existing_vendor
                        vendor_serializer = CustomerVendorSerializer(vendor_obj, data=vendor_data, partial=True)
                        if vendor_serializer.is_valid():
                            vendor_serializer.save()
                        else:
                            return Response({"vendor_errors": vendor_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
                    else:
                        vendor_serializer = CustomerVendorSerializer(data=vendor_data)
                        if vendor_serializer.is_valid():
                            vendor_obj = vendor_serializer.save(client=income.client)
                        else:
                            return Response({"vendor_errors": vendor_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
                
                # Assign vendor_obj to income.vendor if defined
                if vendor_obj:
                    income.customer = vendor_obj

            # Process rows for product summaries
            product_summaries = []
            # processed_product_names = set()
            for row in rows:
                hsn_code = row.get('hsnCode')
                gst_rate = safe_decimal(row.get('gstRate', '0'))
                product_name = row.get('product')
                description_text = row.get('description')
                unit_value = safe_decimal(row.get('unit', '0'))
                rate_value = safe_decimal(row.get('rate', '0'))
                amount = safe_decimal(row.get('product_amount', '0'))
                cgst = safe_decimal(row.get('cgst', '0'))
                sgst = safe_decimal(row.get('sgst', '0'))
                igst = safe_decimal(row.get('igst', '0'))

                # Update or create HSNCode
                hsn_code_obj, _ = HSNCode.objects.update_or_create(
                    hsn_code=hsn_code,
                    defaults={'gst_rate': gst_rate}
                )

                existing_product = Product.objects.filter(product_name=product_name).exclude(hsn=hsn_code_obj).first()
                if existing_product:
                    return Response(
                        {"error_message": f"Product name '{product_name}' already exists with a different HSN code."},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                # Update or create Product
                product_obj, _ = Product.objects.update_or_create(
                    product_name=product_name,
                    hsn=hsn_code_obj,
                    defaults={'unit_of_measure': unit_value}
                )

                # Update or create ProductDescription
                product_description_obj, _ = ProductDescription.objects.update_or_create(
                    product=product_obj,
                    description=description_text,
                    defaults={
                        'unit': unit_value,
                        'rate': rate_value,
                        'product_amount': amount,
                        'cgst': cgst,
                        'sgst': sgst,
                        'igst': igst
                    }
                )

                # Update or create ProductSummary
                product_summary, _ = ProductSummaryIncome.objects.update_or_create(
                    hsn=hsn_code_obj,
                    product=product_obj,
                    prod_description=product_description_obj
                )
                product_summaries.append(product_summary)

            income.product_summaries.set(product_summaries)
            # income.save()
            location_data = form_data.get('location')  # Location name entered by user
            location_id = form_data.get('offLocID')   # Existing location ID, if provided
            branch_id = form_data.get('branchID')     # Branch ID selected for new location

            if location_id:  # Update existing location
                # Fetch the existing location
                location_obj = OfficeLocation.objects.filter(id=location_id).first()
                if not location_obj:
                    return Response({"error_message": "Office Location not found."}, status=status.HTTP_404_NOT_FOUND)

                # Update the location details
                location_obj.location = location_data
                location_obj.contact = form_data.get('contact')
                location_obj.address = form_data.get('address')
                location_obj.city = form_data.get('city')
                location_obj.state = form_data.get('state')
                location_obj.country = form_data.get('country')
                location_obj.save()

            else:  # Create a new location
                # Validate branch selection
                if not branch_id:
                    return Response({"error_message": "Branch ID is required for creating a new location."}, status=status.HTTP_400_BAD_REQUEST)

                branch_instance = Branch.objects.filter(id=branch_id, client_id=income.client.id).first()
                if not branch_instance:
                    return Response({"error_message": f"Branch with ID {branch_id} not found or doesn't belong to the client."},
                                    status=status.HTTP_404_NOT_FOUND)

                # Create the new location
                location_obj, _  = OfficeLocation.objects.get_or_create(
                    location=location_data,
                    contact=form_data.get('contact'),
                    address=form_data.get('address'),
                    city=form_data.get('city'),
                    state=form_data.get('state'),
                    country=form_data.get('country'),
                    branch=branch_instance  # Associate with the selected branch
                )

            # Associate the updated or newly created location with the sales invoice
            income.client_Location = location_obj
            income.save()

            response_data = {
                'message': 'Income updated successfully.',
                'income_data': IncomeSerializer(income).data,
                'product_summaries': [{'id': summary.id, 'product_name': summary.product.product_name} for summary in product_summaries]
            }
            return Response(response_data, status=status.HTTP_200_OK)

    except Exception as e:
        error_details = traceback.format_exc()
        # print({"Error in update_sales_invoice" : str(e)})
        return Response({"error_message":str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
def create_income2(request, client_pk):
    try:
        with transaction.atomic():
            payload = request.data
            print('payload', payload)

            # Extract rows dynamically
            rows_data = defaultdict(dict)
            for key, value in payload.items():
                if key.startswith("rows["):  # Check if the key corresponds to rows
                    row_index = key.split('[')[1].split(']')[0]
                    field_name = key.split('[')[2].split(']')[0]
                    rows_data[int(row_index)][field_name] = value
            rows = [rows_data[index] for index in sorted(rows_data.keys())]
            print('jjjjjjjjjjjjjj', rows)

            # Extract form data, vendor data, and invoice data
            form_data = {
                "offLocID": payload.get("formData[offLocID]"),
                "location": payload.get("formData[location]"),
                "contact": payload.get("formData[contact]"),
                "address": payload.get("formData[address]"),
                "city": payload.get("formData[city]"),
                "state": payload.get("formData[state]"),
                "country": payload.get("formData[country]"),
                "branchID": payload.get("formData[branchID]"),                
            }
            vendor_data = {
                "name": payload.get("vendorData[name]"),
                "gst_no": payload.get("vendorData[gst_no]"),
                "pan": payload.get("vendorData[pan]"),
                "address": payload.get("vendorData[customer_address]"),
                "email": payload.get("vendorData[email]"),
                "contact": payload.get("vendorData[contact]"),
                "customer": payload.get("vendorData[customer]", "").lower() == "true",
                "vendor": payload.get("vendorData[vendor]", "").lower() == "true",
            }
            invoice_data = {
                "invoice_no": payload.get("invoiceData[0][invoice_no]"),
                "invoice_date": payload.get("invoiceData[0][invoice_date]"),
                "month": payload.get("invoiceData[0][month]"),
                "invoice_type": payload.get("invoiceData[0][invoice_type]"),
                "entry_type": payload.get("invoiceData[0][entry_type]"),
                "taxable_amount": payload.get("invoiceData[0][taxable_amount]"),
                "totalall_gst": payload.get("invoiceData[0][totalall_gst]"),
                "total_invoice_value": payload.get("invoiceData[0][total_invoice_value]"),
                "tds_tcs_rate": payload.get("invoiceData[0][tds_tcs_rate]"),
                "tcs": payload.get("invoiceData[0][tcs]"),
                "tds": payload.get("invoiceData[0][tds]"),
                "amount_receivable": payload.get("invoiceData[0][amount_receivable]"),
            }

            if invoice_data["invoice_date"]:
                try:
                    invoice_data["invoice_date"] = datetime.strptime(invoice_data["invoice_date"], "%d-%m-%Y").date()
                except ValueError:
                    return Response({"error_message": "Invalid date format, expected DD-MM-YYYY."}, status=status.HTTP_400_BAD_REQUEST)
            print("Final invoice date:", invoice_data["invoice_date"])

            if invoice_data["month"]:
                try:
                    invoice_data["month"] = datetime.strptime(invoice_data["month"], "%d-%m-%Y").date()
                except ValueError:
                    return Response({"error_message": "Invalid Month date format, expected DD-MM-YYYY."},status=status.HTTP_400_BAD_REQUEST)

            print("Final month:", invoice_data["month"])

            attach_invoice = request.FILES.get("invoiceData[0][attach_invoice]")
            attach_e_way_bill = request.FILES.get("invoiceData[0][attach_e_way_bill]")

            # Handle Vendor creation or update
            vendor_obj = None
            if vendor_data.get("gst_no"):
                existing_vendor = Customer.objects.filter(client_id=client_pk, gst_no=vendor_data["gst_no"]).first()
                if existing_vendor:
                    vendor_serializer = CustomerVendorSerializer(existing_vendor, data=vendor_data, partial=True)
                    if vendor_serializer.is_valid():
                        vendor_obj = vendor_serializer.save()
                    else:
                        return Response({"vendor_errors": vendor_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
                else:
                    vendor_serializer = CustomerVendorSerializer(data=vendor_data)
                    if vendor_serializer.is_valid():
                        vendor_obj = vendor_serializer.save(client_id=client_pk)
                    else:
                        return Response({"vendor_errors": vendor_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

            # Temporary list to store product summaries
            product_summaries_to_save = []
            for row in rows:
                hsn_code = row.get('hsnCode')
                gst_rate = safe_decimal(row.get('gstRate', '0'))
                product_name = row.get('product')
                product_id = row.get('product_id')  # Assuming the frontend sends this if selecting an existing product
                description_text = row.get('description', '')
                unit_value = safe_decimal(row.get('unit', '0'))
                rate_value = safe_decimal(row.get('rate', '0'))
                amount = safe_decimal(row.get('product_amount', '0'))
                cgst = safe_decimal(row.get('cgst', '0'))
                sgst = safe_decimal(row.get('sgst', '0'))
                igst = safe_decimal(row.get('igst', '0'))

                # Handle HSNCode
                hsn_code_obj, _ = HSNCode.objects.get_or_create(
                    hsn_code=hsn_code, defaults={'gst_rate': gst_rate}
                )

                existing_product = Product.objects.filter(product_name=product_name).exclude(hsn=hsn_code_obj).first()
                if existing_product:
                    return Response(
                        {"error_message": f"Product name '{product_name}' already exists with a different HSN code."},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                # Handle Product (existing or new)
                if product_id:
                    # Use existing product
                    product_obj = Product.objects.filter(id=product_id).first()
                    if not product_obj:
                        return Response({"error_message": f"Product with ID {product_id} not found."}, status=status.HTTP_404_NOT_FOUND)
                else:
                    # Create new product
                    product_obj, _ = Product.objects.get_or_create(
                        product_name=product_name, hsn=hsn_code_obj, defaults={'unit_of_measure': unit_value}
                    )

                # Handle ProductDescription
                product_description_obj, _ = ProductDescription.objects.get_or_create(
                    product=product_obj,
                    description=description_text,
                    defaults={
                        'unit': unit_value,
                        'rate': rate_value,
                        'product_amount': amount,
                        'cgst': cgst,
                        'sgst': sgst,
                        'igst': igst,
                    }
                )

                # Create ProductSummary but do not save it yet
                product_summary = ProductSummaryIncome(
                    hsn=hsn_code_obj,
                    product=product_obj,
                    prod_description=product_description_obj
                )
                product_summaries_to_save.append(product_summary)

            location_obj = None
            if form_data["offLocID"]:
                location_obj = OfficeLocation.objects.filter(id=form_data["offLocID"]).first()
                if not location_obj:
                    return Response({"error_message": "Office Location not found."}, status=status.HTTP_404_NOT_FOUND)
            else:
                branch_instance = Branch.objects.filter(id=form_data["branchID"], client_id=client_pk).first()
                if not branch_instance:
                    return Response({"error_message": f"Branch with ID {form_data['branchID']} not found or doesn't belong to the client."},
                                    status=status.HTTP_404_NOT_FOUND)
                location_obj, _ = OfficeLocation.objects.get_or_create(
                    location=form_data.get("location"),
                    contact=form_data.get("contact"),
                    address=form_data.get("address"),
                    city=form_data.get("city"),
                    state=form_data.get("state"),
                    country=form_data.get("country"),
                    branch=branch_instance
                )

            income = Income.objects.create(
                client_id=client_pk,
                customer=vendor_obj,
                attach_invoice=attach_invoice,
                attach_e_way_bill=attach_e_way_bill,
                client_Location=location_obj,
                **invoice_data
            )

            # Save product summaries at the end
            for product_summary in product_summaries_to_save:
                product_summary.save()
                income.product_summaries.add(product_summary)

            return Response({"message": "Income created successfully.", "invoice_id": income.id}, status=status.HTTP_201_CREATED)

    except Exception as e:
        error_details = traceback.format_exc()
        return Response({"error_message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['DELETE'])
def delete_income(request, client_pk, pk):
    """
    Deletes a SalesInvoice by its primary key (ID) along with its associated product summaries.
    """
    try:
        # Retrieve the Client instance
        client = Client.objects.filter(id=client_pk).first()

        if not client:
            return Response({"error_message": "Client not found."}, status=status.HTTP_404_NOT_FOUND)

        # Retrieve the SalesInvoice instance
        income = Income.objects.filter(id=pk, client=client).first()

        if not income:
            return Response({"error_message": "Income not found."}, status=status.HTTP_404_NOT_FOUND)

        # Handle deletion of associated ProductSummary instances
        product_summaries = income.product_summaries.all()
        for product_summary in product_summaries:
            product_summary.delete()

        # Delete the SalesInvoice instance
        income.delete()

        return Response({"message": "Income deleted successfully."}, status=status.HTTP_200_OK)

    # except Exception as e:
    #     return Response({"error_message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    except Exception as e:
        error_details = traceback.format_exc()
        # print({"Error in update_sales_invoice" : str(e)})
        return Response({"error_message":str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET', 'PATCH'])
def income_detail_view(request, client_pk, invoice_pk):
    try:
        # Fetch the sales invoice object
        income = Income.objects.get(client=client_pk, pk=invoice_pk)
    except Income.DoesNotExist:
        return Response({"error_message": "Income not found."}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        # Serialize and return the sales invoice details
        serializer = IncomeSerializerList(income)
        return Response(serializer.data, status=status.HTTP_200_OK)

    elif request.method == 'PATCH':
        # Update only specific fields of the sales invoice
        data = request.data

        # Update fields from the request payload
        income.invoice_no = data.get('invoice_no', income.invoice_no)
        income.invoice_date = data.get('invoice_date', income.invoice_date)
        income.month = data.get('month', income.month)
        income.invoice_type = data.get('invoice_type', income.invoice_type)
        income.entry_type = data.get('entry_type', income.entry_type)
        income.taxable_amount = safe_decimal(data.get('taxable_amount', income.taxable_amount))
        income.totalall_gst = safe_decimal(data.get('totalall_gst', income.total_gst))
        income.total_invoice_value = safe_decimal(data.get('totalall_invoice_value', income.totalall_invoice_value))
        income.tds_tcs_rate = safe_decimal(data.get('tds_tcs_rate', income.tds_tcs_rate))
        income.tds = safe_decimal(data.get('tds', income.tds))
        income.tcs = safe_decimal(data.get('tcs', income.tcs))
        income.amount_receivable = safe_decimal(data.get('amount_receivable', income.amount_receivable))

        # Update product summaries if provided
        rows = data.get('rows', [])
        product_summaries = []
        for row in rows:
            hsn_code = row.get('hsnCode')
            gst_rate = safe_decimal(row.get('gstRate', '0'))
            product_name = row.get('product')
            description_text = row.get('description')
            unit_value = safe_decimal(row.get('unit', '0'))
            rate_value = safe_decimal(row.get('rate', '0'))
            amount = safe_decimal(row.get('product_amount', '0'))
            cgst = safe_decimal(row.get('cgst', '0'))
            sgst = safe_decimal(row.get('sgst', '0'))
            igst = safe_decimal(row.get('igst', '0'))

            # Handle HSNCode
            hsn_code_obj, created = HSNCode.objects.get_or_create(
                hsn_code=hsn_code,
                defaults={'gst_rate': gst_rate}
            )
            if not created and hsn_code_obj.gst_rate != gst_rate:
                hsn_code_obj.gst_rate = gst_rate
                hsn_code_obj.save()

            # Handle Product
            product_obj = Product.objects.filter(product_name=product_name, hsn=hsn_code_obj).first()
            if not product_obj:
                product_obj = Product.objects.create(
                    product_name=product_name,
                    hsn=hsn_code_obj,
                    unit_of_measure=unit_value
                )

            # Handle ProductDescription
            product_description_obj, created = ProductDescription.objects.get_or_create(
                product=product_obj,
                description=description_text,
                defaults={
                    'unit': unit_value,
                    'rate': rate_value,
                    'product_amount': amount,
                    'cgst': cgst,
                    'sgst': sgst,
                    'igst': igst
                }
            )
            if not created:
                # Update fields in ProductDescription
                product_description_obj.unit = unit_value
                product_description_obj.rate = rate_value
                product_description_obj.product_amount = amount
                product_description_obj.cgst = cgst
                product_description_obj.sgst = sgst
                product_description_obj.igst = igst
                product_description_obj.save()

            # Create ProductSummary
            product_summary = ProductSummaryIncome.objects.create(
                hsn=hsn_code_obj,
                product=product_obj,
                prod_description=product_description_obj
            )
            product_summaries.append(product_summary)

        # Save updates to the sales invoice
        if product_summaries:
            income.product_summaries.set(product_summaries)
        income.save()

        # Return the updated object
        updated_serializer = IncomeSerializer(income)
        return Response(updated_serializer.data, status=status.HTTP_200_OK)

# ************************************************Expenses************************************************

@api_view(['GET','POST'])
def create_expenses_get(request, pk):
    try:
        client = Client.objects.get(id=pk)
    except Client.DoesNotExist:
        return Response({'error_message':'Client not Found.'}, status=404)

    if request.method == 'GET':
        context = {
            'serializer_customer':[],
            'serializer': [],
            'product_serializer':[],
            'message':None,
            'hsn':None,
            'location':None,
        }
        received_value = request.GET.get('newValue')
        product_Id = request.GET.get('productID')
        if product_Id:
            try:
                hsn_cc = Product.objects.get(id=product_Id).hsn
                context.update({
                    'message':'Product HSN found',
                    'hsn':HSNSerializer(hsn_cc).data
                })
            except (ValueError, Product.DoesNotExist):
                return Response({'error_message':'Invalid Product ID'}, status=400)

        if received_value:
            try:
                received_value = int(received_value)
                location = OfficeLocation.objects.get(id=received_value)
                branch_gst = location.branch.gst_no
                print('branch val', branch_gst)
                context.update({
                    "message":"Location Found",
                    "location" : OfficeLocationSerializer(location).data,
                    "branch_gst": branch_gst
                })
            except (ValueError, OfficeLocation.DoesNotExist):
                return Response({'error_message':'Invalid location ID.'}, status=400)

        if not product_Id and not received_value:
            off = OfficeLocation.objects.filter(branch__client=client)
            customer = Customer.objects.filter(client=client, vendor=True)
            product = Product.objects.all()
            branch = Branch.objects.filter(client=client)

            serializer_customer = CustomerVendorSerializer(customer, many=True)
            serializers = OfficeLocationSerializer(off, many=True)
            product_serializer = ProductSerializer(product, many=True)
            branch_serializer = BranchSerailizer(branch, many= True)

            context.update({
                'serializer_customer' : serializer_customer.data,
                'serializer' : serializers.data,
                'product_serializer' : product_serializer.data,
                'branch_serializer' : branch_serializer.data
            })
        return Response(context)
    
@api_view(['POST'])
def create_expenses(request, pk):
    client = Client.objects.get(id=pk)

    # Check if 'attach_e_way_bill' is in the request files
    if 'attach_invoice' in request.FILES:
        # Iterate through each file in the 'attach_e_way_bill' field
        for e_way_bill in request.FILES.getlist('attach_invoice'):
            # Prepare data for each file
            expenses_data = {
                'attach_invoice': e_way_bill,  # The file being uploaded
                'client': client.id,  # Associate the file with the client
            }

            # Initialize the serializer for each file
            serializer = ExpensesSerializer2(data=expenses_data)

            # Check if the serializer is valid
            if serializer.is_valid():
                # Save each file as a separate object
                serializer.save()
            else:
                return Response({'error_message': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

        # If all files are processed, return success response
        return Response({'message': 'Expenses E-way bill(s) uploaded successfully.'}, status=status.HTTP_201_CREATED)
    else:
        return Response({'error_message': 'No files uploaded in the request.'}, status=status.HTTP_400_BAD_REQUEST)
        
@api_view(['GET','PUT'])
def update_expenses(request, client_pk, invoice_pk):
    try:
        # print('my payload',request.data)
        if request.method == 'GET':
            expenses = Expenses.objects.filter(client_id=client_pk, id= invoice_pk).first()
            if not expenses:
                return Response({"error_message": "Expenses not found."}, status=status.HTTP_404_NOT_FOUND)
            expenses_data = ExpensesSerializer3(expenses).data
            product_summaries = expenses.product_summaries.all()
            product_summary_data =[
                {
                    "id": summary.id,
                    "hsnCode": summary.hsn.hsn_code,
                    "gstRate": summary.hsn.gst_rate,
                    "product": summary.product.product_name,
                    "description": summary.prod_description.description,
                    "unit": summary.prod_description.unit,
                    "rate": summary.prod_description.rate,
                    "product_amount": summary.prod_description.product_amount,
                    "cgst": summary.prod_description.cgst,
                    "sgst": summary.prod_description.sgst,
                    "igst": summary.prod_description.igst,
                }
                for summary in product_summaries
            ]

            if expenses.month:
                formatted_month = expenses.month.strftime("%d-%m-%Y")  # Directly format date
            else:
                formatted_month = None

            if expenses.invoice_date:
                formatted_invoice_date = expenses.invoice_date.strftime("%d-%m-%Y")  # Directly format date
            else:
                formatted_invoice_date = None

            if expenses.utilise_month:
                formatted_utilise = expenses.utilise_month.strftime("%d-%m-%Y")
            else:
                formatted_utilise = None
            
            response_data ={
                "expenses" : {
                    **expenses_data,
                    "month": formatted_month,  # Ensure month is sent in dd/mm/yyyy format
                    "invoice_date" : formatted_invoice_date,
                    "utilise_month" : formatted_utilise,
                    },
                "product_summaries" : product_summary_data,
                "client_location": {
                    "id": expenses.client_Location.id if expenses.client_Location else None,
                    "location": expenses.client_Location.location if expenses.client_Location else None,
                    "contact": expenses.client_Location.contact if expenses.client_Location else None,
                    "address": expenses.client_Location.address if expenses.client_Location else None,
                    "city": expenses.client_Location.city if expenses.client_Location else None,
                    "state": expenses.client_Location.state if expenses.client_Location else None,
                    "country": expenses.client_Location.country if expenses.client_Location else None,
                    "branchID": expenses.client_Location.branch.id if expenses.client_Location else None,
                },
                "vendor": {
                    "id": expenses.vendor.id if expenses.vendor else None,
                    "name": expenses.vendor.name if expenses.vendor else None,
                    "gst_no": expenses.vendor.gst_no if expenses.vendor else None,
                    "pan": expenses.vendor.pan if expenses.vendor else None,
                    "vendor_address": expenses.vendor.address if expenses.vendor else None,
                    "customer": expenses.vendor.customer if expenses.vendor else None,
                    "vendor": expenses.vendor.vendor if expenses.vendor else None,
                    "email": expenses.vendor.email if expenses.vendor else None,
                    "contact": expenses.vendor.contact if expenses.vendor else None,
                },
            }
            #  print('bbbbbbbbbbbbbbbbbbbbbbb',response_data["sales_invoice"])
            return Response(response_data, status=status.HTTP_200_OK)
        elif request.method == 'PUT':
            expenses = Expenses.objects.filter(client_id=client_pk, id= invoice_pk).first()
            if not expenses:
                    return Response({"error_message":"Expenses not found"}, status=status.HTTP_404_NOT_FOUND)
                
            payload = request.data
            rows_data = defaultdict(dict)
            for key, value in payload.items():
                if key.startswith("rows["):  # Check if the key corresponds to rows
                    # Extract the row index and field name
                    row_index = key.split('[')[1].split(']')[0]  # Get the index (e.g., '0', '1', '2')
                    field_name = key.split('[')[2].split(']')[0]  # Get the field name (e.g., 'product')

                    # Store the value in the corresponding row and field
                    rows_data[int(row_index)][field_name] = value
                    
            rows = [rows_data[index] for index in sorted(rows_data.keys())]
            print(rows)
            invoice_data = request.data.get('invoiceData', [{}])[0]  # Extract the first item
            invoice_file = request.data.get('invoice_file')
            print('invoice_file',invoice_data)
            form_data = {
                "offLocID": request.data.get("formData[offLocID]"),
                "location": request.data.get("formData[location]"),
                "contact": request.data.get("formData[contact]"),
                "address": request.data.get("formData[address]"),
                "city": request.data.get("formData[city]"),
                "state": request.data.get("formData[state]"),
                "country": request.data.get("formData[country]"),
                "branchID": request.data.get("formData[branchID]"),
            }
            vendor_data = {
                "name": request.data.get("vendorData[name]"),
                "gst_no": request.data.get("vendorData[gst_no]"),
                "pan": request.data.get("vendorData[pan]"),
                "vendor_address": request.data.get("vendorData[vendor_address]"),
                "email": request.data.get("vendorData[email]"),
                "contact": request.data.get("vendorData[contact]"),
                "customer": request.data.get("vendorData[customer]").lower() == "true" if request.data.get("vendorData[customer]") else None,
                "vendor": request.data.get("vendorData[vendor]").lower() == "true" if request.data.get("vendorData[vendor]") else None,
            }
            invoice_data = {
                "invoice_no": request.data.get("invoiceData[0][invoice_no]"),
                "invoice_date": request.data.get("invoiceData[0][invoice_date]"),
                "month": request.data.get("invoiceData[0][month]"),
                "invoice_type": request.data.get("invoiceData[0][invoice_type]"),
                "entry_type": request.data.get("invoiceData[0][entry_type]"),
                "taxable_amount": request.data.get("invoiceData[0][taxable_amount]"),
                "totalall_gst": request.data.get("invoiceData[0][totalall_gst]"),
                "total_invoice_value": request.data.get("invoiceData[0][total_invoice_value]"),
                "tds_tcs_rate": request.data.get("invoiceData[0][tds_tcs_rate]"),
                "tcs": request.data.get("invoiceData[0][tcs]"),
                "tds": request.data.get("invoiceData[0][tds]"),
                "amount_receivable": request.data.get("invoiceData[0][amount_receivable]"),
                "attach_invoice": request.data.get("invoiceData[0][attach_invoice]"),
                "attach_e_way_bill": request.data.get("invoiceData[0][attach_e_way_bill]"),
                "utilise_edit": payload.get("invoiceData[0][utilise_edit]", "").lower() == "true", #nnnn
                "utilise_month": payload.get("invoiceData[0][utilise_month]"), #nnnnnn

            }

            if invoice_data["invoice_date"]:
                try:
                    invoice_data["invoice_date"] = datetime.strptime(invoice_data["invoice_date"], "%d-%m-%Y").date()
                except ValueError:
                    return Response({"error_message": "Invalid Invoice date format, expected DD-MM-YYYY."}, status=status.HTTP_400_BAD_REQUEST)
            print("Final invoice date:", invoice_data["invoice_date"])

            if invoice_data["month"]:
                try:
                    invoice_data["month"] = datetime.strptime(invoice_data["month"], "%d-%m-%Y").date()
                except ValueError:
                    return Response({"error_message": "Invalid Month date format, expected DD-MM-YYYY."},status=status.HTTP_400_BAD_REQUEST)

            if invoice_data["utilise_month"]:
                try:
                    invoice_data["utilise_month"] = datetime.strptime(invoice_data["utilise_month"], "%d-%m-%Y").date()
                except ValueError:
                    return Response({"error_message": "Invalid Utilise Month date format, expected DD-MM-YYYY."},status=status.HTTP_400_BAD_REQUEST)

            print("Final month:", invoice_data["month"])

            attach_invoice = request.FILES.get("invoiceData[0][attach_invoice]")
            attach_e_way_bill = request.FILES.get("invoiceData[0][attach_e_way_bill]")
            if attach_invoice:
                expenses.attach_invoice = attach_invoice
                
            if attach_e_way_bill:
                expenses.attach_e_way_bill = attach_e_way_bill
            
            for field, value in invoice_data.items():
                if field not in ['attach_invoice','attach_e_way_bill']:
                    if hasattr(expenses, field):
                        setattr(
                            expenses,
                            field,
                            safe_decimal(value) if field in [
                                'taxable_amount', 'totalall_gst', 'total_invoice_value',
                                'tds_tcs_rate', 'tds', 'tcs', 'amount_receivable'
                            ] else value
                        )
            
            attach_invoice = invoice_data.get('attach_invoice')
            print('attach_invoice',attach_invoice)
             # Log the flattened data
            # print("Flattened form_data:", form_data)
            print("Flattened invoice_data:", invoice_data)
            print('request payload', request.data)
            if invoice_file:
                expenses.invoice_file = invoice_file
                expenses.save()
                return Response ({'message':'Invoice file uploaded successfully '},status=status.HTTP_400_BAD_REQUEST)
            
            if vendor_data:
                if 'vendor_address' in vendor_data:
                    vendor_data['address'] = vendor_data.pop('vendor_address')
                    
                vendor_id = request.data.get("vendorData[vendorID]")

                if vendor_id:
                    vendor_obj = Customer.objects.filter(id=vendor_id).first()
                    if vendor_obj:
                        if vendor_obj.gst_no == vendor_data.get("gst_no"):
                            vendor_serializer = CustomerVendorSerializer(vendor_obj, data=vendor_data, partial=True)
                            if vendor_serializer.is_valid():
                                vendor_serializer.save()
                            else:
                                return Response({"vendor_errors": vendor_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
                        else:
                            # Create a new vendor if gst_no is changed
                            vendor_serializer = CustomerVendorSerializer(data=vendor_data)
                            if vendor_serializer.is_valid():
                                vendor_obj = vendor_serializer.save(client=expenses.client)
                            else:
                                return Response({"vendor_errors": vendor_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
                    else:
                        return Response({"error_message": f"Vendor with ID {vendor_id} not found."}, status=status.HTTP_404_NOT_FOUND)
                else:
                    existing_vendor = Customer.objects.filter(client=expenses.client, gst_no=vendor_data.get("gst_no")).first()
                    if existing_vendor:
                        vendor_obj = existing_vendor
                        vendor_serializer = CustomerVendorSerializer(vendor_obj, data=vendor_data, partial=True)
                        if vendor_serializer.is_valid():
                            vendor_serializer.save()
                        else:
                            return Response({"vendor_errors": vendor_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
                    else:
                        vendor_serializer = CustomerVendorSerializer(data=vendor_data)
                        if vendor_serializer.is_valid():
                            vendor_obj = vendor_serializer.save(client=expenses.client)
                        else:
                            return Response({"vendor_errors": vendor_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
                
                # Assign vendor_obj to expenses.vendor if defined
                if vendor_obj:
                    expenses.vendor = vendor_obj
                        
            product_summaries = []
            # processed_product_names = set()
            for row in rows:
                hsn_code = row.get('hsnCode')
                gst_rate = safe_decimal(row.get('gstRate', '0'))
                product_name = row.get('product')
                description_text = row.get('description')
                unit_value = safe_decimal(row.get('unit', '0'))
                rate_value = safe_decimal(row.get('rate', '0'))
                amount = safe_decimal(row.get('product_amount', '0'))
                cgst = safe_decimal(row.get('cgst', '0'))
                sgst = safe_decimal(row.get('sgst', '0'))
                igst = safe_decimal(row.get('igst', '0'))

                # Update or create HSNCode
                hsn_code_obj, _ = HSNCode.objects.update_or_create(
                    hsn_code=hsn_code,
                    defaults={'gst_rate': gst_rate}
                )

                existing_product = Product.objects.filter(product_name=product_name).exclude(hsn=hsn_code_obj).first()
                if existing_product:
                    return Response(
                        {"error_message": f"Product name '{product_name}' already exists with a different HSN code."},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                # Update or create Product
                product_obj, _ = Product.objects.update_or_create(
                    product_name=product_name,
                    hsn=hsn_code_obj,
                    defaults={'unit_of_measure': unit_value}
                )

                # Update or create ProductDescription
                product_description_obj, _ = ProductDescription.objects.update_or_create(
                    product=product_obj,
                    description=description_text,
                    defaults={
                        'unit': unit_value,
                        'rate': rate_value,
                        'product_amount': amount,
                        'cgst': cgst,
                        'sgst': sgst,
                        'igst': igst
                    }
                )
                
                product_summary, _ = ProductSummaryExpenses.objects.update_or_create(
                    hsn=hsn_code_obj,
                    product=product_obj,
                    prod_description=product_description_obj
                )
                product_summaries.append(product_summary)
            expenses.product_summaries.set(product_summaries)
            # expenses.save()

            location_data = form_data.get('location')  # Location name entered by user
            location_id = form_data.get('offLocID')   # Existing location ID, if provided
            branch_id = form_data.get('branchID')     # Branch ID selected for new location
            
            if location_id:
                location_obj = OfficeLocation.objects.filter(id=location_id).first()
                if not location_obj:
                    return Response({'error_message':'Office Location not found. '}, status=status.HTTP_400_BAD_REQUEST)
                location_obj.location = location_data
                location_obj.contact = form_data.get('contact')
                location_obj.address = form_data.get('address')
                location_obj.city = form_data.get('city')
                location_obj.state = form_data.get('state')
                location_obj.country = form_data.get('country')
                location_obj.save()
            else:
                if not branch_id:
                    return Response({"error_message": "Branch ID is required for creating a new location."}, status=status.HTTP_400_BAD_REQUEST)
                branch_instance = Branch.objects.filter(id=branch_id, client_id=expenses.client.id).first()
                if not branch_instance:
                    return Response({"error_message": f"Branch with ID {branch_id} not found or doesn't belong to the client."},
                                    status=status.HTTP_404_NOT_FOUND)
                    
                location_obj, _ = OfficeLocation.objects.get_or_create(
                    location = location_data,
                    contact = form_data.get('contact'),
                    address = form_data.get('address'),
                    city = form_data.get('city'),
                    state = form_data.get('state'),
                    country = form_data.get('country'),
                    branch = branch_instance
                )
                
            expenses.client_Location = location_obj
            expenses.save()
            
            response_data = {
                'message' : 'Expenses Updated successfully. ',
                'expenses_data' : ExpensesSerializer(expenses).data,
                'product_summaries' : [{'id': summary.id, 'product_name': summary.product.product_name} for summary in product_summaries]
            }
            return Response(response_data, status=status.HTTP_200_OK)

    except Exception as e:
        error_details = traceback.format_exc()
        # print({"Error in update_sales_invoice" : str(e)})
        return Response({"error_message":str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
def create_expenses2(request, client_pk):
    try:
        with transaction.atomic():
            payload = request.data
            print('payload',payload)

            rows_data = defaultdict(dict)
            for key, value in payload.items():
                if key.startswith("rows["):
                    row_index = key.split('[')[1].split(']')[0]
                    field_name = key.split('[')[2].split(']')[0]
                    rows_data[int(row_index)][field_name] = value
            rows = [rows_data[index] for index in sorted(rows_data.keys())]
            form_data = {
                "offLocID" : payload.get("formData[offLocID]"),
                "location" : payload.get("formData[location]"),
                "contact" :  payload.get("formData[contact]"),
                "address" : payload.get("formData[address]"),
                "city" : payload.get("formData[city]"),
                "state" : payload.get("formData[state]"),
                "country" : payload.get("formData[country]"),
                "branchID" : payload.get("formData[branchID]"),
            }
            vendor_data = {
                "name" : payload.get("vendorData[name]"),
                "gst_no" : payload.get("vendorData[gst_no]"),
                "pan" : payload.get("vendorData[pan]"),
                "address" : payload.get("vendorData[vendor_address]"), #nnnnnnnnn
                "customer" : payload.get("vendorData[customer]", "").lower() == "true",
                "vendor" : payload.get("vendorData[vendor]", "").lower() == "true",
                "email" : payload.get("vendorData[email]"),
                "contact" : payload.get("vendorData[contact]"),
            }
            invoice_data = {
                "invoice_no" : payload.get("invoiceData[0][invoice_no]"),
                "invoice_date" : payload.get("invoiceData[0][invoice_date]"),
                "month" : payload.get("invoiceData[0][month]"),
                "invoice_type" : payload.get("invoiceData[0][invoice_type]"),
                "entry_type" : payload.get("invoiceData[0][entry_type]"),
                "taxable_amount" : payload.get("invoiceData[0][taxable_amount]"),
                "totalall_gst" : payload.get("invoiceData[0][totalall_gst]"),
                "total_invoice_value" : payload.get("invoiceData[0][total_invoice_value]"),
                "tds_tcs_rate": payload.get("invoiceData[0][tds_tcs_rate]"),
                "tcs": payload.get("invoiceData[0][tcs]"),
                "tds": payload.get("invoiceData[0][tds]"),
                "amount_receivable": payload.get("invoiceData[0][amount_receivable]"),
                # "utilise_edit": payload.get("invoiceData[0][utilise_edit]", "").lower() == "true", #nnnn
                # "utilise_month": payload.get("invoiceData[0][utilise_month]").lower(), #nnnnnn
                
            }

            if invoice_data["invoice_date"]:
                try:
                    invoice_data["invoice_date"] = datetime.strptime(invoice_data["invoice_date"], "%d-%m-%Y").date()
                except ValueError:
                    return Response({"error_message": "Invalid date format, expected DD-MM-YYYY."}, status=status.HTTP_400_BAD_REQUEST)
            print("Final invoice date:", invoice_data["invoice_date"])

            if invoice_data["month"]:
                try:
                    invoice_data["month"] = datetime.strptime(invoice_data["month"], "%d-%m-%Y").date()
                except ValueError:
                    return Response({"error_message": "Invalid Month date format, expected DD-MM-YYYY."},status=status.HTTP_400_BAD_REQUEST)

            print("Final month:", invoice_data["month"])


            attach_invoice = request.FILES.get("invoiceData[0][attach_invoice]")
            attach_e_way_bill = request.FILES.get("invoiceData[0][attach_e_way_bill]")

            vendor_obj = None
            if vendor_data.get("gst_no"):
                existing_vendor = Customer.objects.filter(client_id=client_pk, gst_no=vendor_data["gst_no"]).first()
                if existing_vendor:
                    vendor_serializer = CustomerVendorSerializer(existing_vendor, data=vendor_data, partial=True)
                    if vendor_serializer.is_valid():
                        vendor_obj = vendor_serializer.save(status=status.HTTP_200_OK)
                    else:
                        return Response({"vendor_errors": vendor_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
                else:
                    vendor_serializer = CustomerVendorSerializer(data=vendor_data)
                    if vendor_serializer.is_valid():
                        vendor_obj = vendor_serializer.save(client_id=client_pk)
                    else:
                        return Response({"vendor_errors": vendor_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
            
            product_summaries = []  # To store created product summaries
            # processed_product_names = set()
            for row in rows:
                hsn_code = row.get('hsnCode')
                gst_rate = safe_decimal(row.get('gstRate', '0'))
                product_name = row.get('product')
                product_id = row.get('product_id')  # Assuming the frontend sends this if selecting an existing product
                description_text = row.get('description', '')
                unit_value = safe_decimal(row.get('unit', '0'))
                rate_value = safe_decimal(row.get('rate', '0'))
                amount = safe_decimal(row.get('product_amount', '0'))
                cgst = safe_decimal(row.get('cgst', '0'))
                sgst = safe_decimal(row.get('sgst', '0'))
                igst = safe_decimal(row.get('igst', '0'))

                # Handle HSNCode
                hsn_code_obj, _ = HSNCode.objects.get_or_create(
                    hsn_code=hsn_code, defaults={'gst_rate': gst_rate}
                )

                existing_product = Product.objects.filter(product_name=product_name).exclude(hsn=hsn_code_obj).first()
                if existing_product:
                    return Response(
                        {"error_message": f"Product name '{product_name}' already exists with a different HSN code."},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                # Handle Product (existing or new)
                if product_id:
                    # Use existing product
                    product_obj = Product.objects.filter(id=product_id).first()
                    if not product_obj:
                        return Response({"error_message": f"Product with ID {product_id} not found."}, status=status.HTTP_404_NOT_FOUND)
                else:
                    # Create new product
                    product_obj, _ = Product.objects.get_or_create(
                        product_name=product_name, hsn=hsn_code_obj, defaults={'unit_of_measure': unit_value}
                    )

                # Handle ProductDescription
                product_description_obj, _ = ProductDescription.objects.get_or_create(
                    product=product_obj,
                    description=description_text,
                    defaults={
                        'unit': unit_value,
                        'rate': rate_value,
                        'product_amount': amount,
                        'cgst': cgst,
                        'sgst': sgst,
                        'igst': igst,
                    }
                )

                # Create ProductSummary
                product_summary = ProductSummaryExpenses(
                    hsn=hsn_code_obj,
                    product=product_obj,
                    prod_description=product_description_obj
                )
                product_summaries.append(product_summary)

                # Link ProductSummary to the SalesInvoice
                # expenses.product_summaries.add(product_summary)  # Add the product summary to the invoice
            location_obj = None
            if form_data["offLocID"] :
                location_obj = OfficeLocation.objects.filter(id=form_data["offLocID"]).first()
                if not location_obj:
                    return Response({"error_message":"Office Location not found"}, status=status.HTTP_400_BAD_REQUEST)
            else:
                branch_instance = Branch.objects.filter(id=form_data["branchID"], client_id=client_pk).first()
                if not branch_instance:
                    return Response({"error_message": f"Branch with ID {form_data['branchID']} not found or doesn't belong to the client."},
                                    status=status.HTTP_404_NOT_FOUND)
                location_obj, _ = OfficeLocation.objects.get_or_create(
                    location = form_data.get("location"),
                    contact = form_data.get("contact"),
                    address = form_data.get("address"),
                    city = form_data.get("city"),
                    state = form_data.get("state"),
                    country = form_data.get("country"),
                    branch = branch_instance

                )
            
            expenses = Expenses.objects.create(
                client_id = client_pk,
                vendor = vendor_obj,
                attach_invoice = attach_invoice,
                attach_e_way_bill = attach_e_way_bill,
                client_Location = location_obj,
                **invoice_data
            )

            for product_summary in product_summaries:
                product_summary.save()
                expenses.product_summaries.add(product_summary)
        
            print('eeeeee',payload)
            return Response({"message": "Expenses created successfully.", "invoice_id": expenses.id}, status=status.HTTP_200_OK)
        
    except Exception as e:
        error_details = traceback.format_exc()
        # print({"Error in update_sales_invoice" : str(e)})
        return Response({"error_message":str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET','PATCH'])
def expenses_detail_view(request, client_pk, invoice_pk):
    try:
        expenses = Expenses.objects.get(client=client_pk, pk=invoice_pk)
    except Expenses.DoesNotExist:
        return Response({'error_message':'Expenses not found'}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        serializers = ExpensesSerializerList(expenses)
        return Response(serializers.data, status=status.HTTP_200_OK)

    elif request.method == 'PATCH':
        data = request.data
        expenses.invoice_no = data.get('invoice_no',expenses.invoice_no)
        expenses.invoice_date = data.get('invoice_date', expenses.invoice_date)
        expenses.month = data.get('invoice_month',expenses.month)
        expenses.invoice_type = data.get('invoice_type', expenses.invoice_type)
        expenses.entry_type= data.get('entry_type',expenses.entry_type)
        expenses.taxable_amount = safe_decimal(data.get('taxable_amount', expenses.taxable_amount))
        expenses.totalall_gst = safe_decimal(data.get('totalall_gst', expenses.totalall_gst))
        expenses.total_invoice_value = safe_decimal(data.get('total_invoice_value', expenses.total_invoice_value))
        expenses.tds_tcs_rate = safe_decimal(data.get('tds_tcs_rate', expenses.tds_tcs_rate))
        expenses.tds = safe_decimal(data.get('tds',expenses.tds))
        expenses.tcs = safe_decimal(data.get('tcs', expenses.tcs))
        expenses.amount_receivable = safe_decimal(data.get('amount_receivable', expenses.amount_receivable))
        expenses.utilise_edit = data.get('utilise_edit', expenses.utilise_edit)
        expenses.utilise_month = data.get('utilise_month', expenses.utilise_month)

        rows = data.get('rows',[])
        product_summaries = []
        for row in rows:
            hsn_code = row.get('hsnCode')
            gst_rate = safe_decimal(row.get('gstRate','0'))
            product_name = row.get('product')
            description_text = row.get('description')
            unit_value = safe_decimal(row.get('unit','0'))
            rate_value = safe_decimal(row.get('rate','0'))
            amount = safe_decimal(row.get('product_amount','0'))
            cgst = safe_decimal(row.get('cgst','0'))
            sgst = safe_decimal(row.get('sgst','0'))
            igst = safe_decimal(row.get('igst','0'))

            hsn_code_obj, created = HSNCode.objects.get_or_create(
                hsn_code = hsn_code,
                defaults={'gst_rate': gst_rate}
            )
            if not created and hsn_code_obj.gst_rate != gst_rate:
                hsn_code_obj.gst_rate = gst_rate
                hsn_code_obj.save()

            product_obj = Product.objects.filter(product_name=product_name, hsn=hsn_code_obj).first()
            if not product_obj:
                product_obj = Product.objects.create(
                    product_name = product_name,
                    hsn = hsn_code_obj,
                    unit_of_measure = unit_value
                )

            product_description_obj, created = ProductDescription.objects.get_or_create(
                product = product_obj,
                description = description_text,
                defaults={
                    'unit': unit_value,
                    'rate': rate_value,
                    'product_amount' : amount,
                    'cgst' : cgst,
                    'sgst' : sgst,
                    'igst' : igst,
                }
            )
            if not created:
                product_description_obj.unit = unit_value
                product_description_obj.rate = rate_value
                product_description_obj.product_amount = amount
                product_description_obj.cgst = cgst
                product_description_obj.sgst = sgst
                product_description_obj.igst = igst
                product_description_obj.save()

            product_summary = ProductSummaryExpenses.objects.create(
                hsn= hsn_code_obj,
                product = product_obj,
                prod_description = product_description_obj
            )
            product_summaries.append(product_summary)

        if product_summaries:
            expenses.product_summaries.set(product_summaries)
        expenses.save()

        updated_serializer = ExpensesSerializer(expenses)
        return Response(updated_serializer.data, status=status.HTTP_200_OK)
    
@api_view(['DELETE'])
def delete_expenses(request, client_pk, pk):
    """
    Deletes a PurchaseInvoice by its primary key (ID) along with its associated product summaries.
    """
    try:
        # Retrieve the Client instance
        client = Client.objects.filter(id=client_pk).first()

        if not client:
            return Response({"error_message": "Client not found."}, status=status.HTTP_404_NOT_FOUND)

        # Retrieve the SalesInvoice instance
        expenses = Expenses.objects.filter(id=pk, client=client).first()

        if not expenses:
            return Response({"error_message": "Expenses not found."}, status=status.HTTP_404_NOT_FOUND)

        # Handle deletion of associated ProductSummary instances
        product_summaries = expenses.product_summaries.all()
        for product_summary in product_summaries:
            product_summary.delete()

        # Delete the SalesInvoice instance
        expenses.delete()

        return Response({"message": "Expenses deleted successfully."}, status=status.HTTP_200_OK)
    
    except Exception as e:
        error_details = traceback.format_exc()
        # print({"Error in update_sales_invoice" : str(e)})
        return Response({"error_message":str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


    # except Exception as e:
    #     return Response({"error_message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
# ************************************************Zip Upload*********************************************************

# @api_view(['POST'])
# def create_zipupload(request, pk):
#     client = Client.objects.get(id=pk)

#     # Check if 'attach_e_way_bill' is in the request files
#     if 'files' in request.FILES:
#         # Iterate through each file in the 'attach_e_way_bill' field
#         for file in request.FILES.getlist('files'):
#             # Prepare data for each file
#             zipupload_data = {
#                 'files': file,  # The file being uploaded
#                 'client': client.id,  # Associate the file with the client
#             }

#             # Initialize the serializer for each file
#             serializer = ZipUploadSerializer2(data=zipupload_data)

#             # Check if the serializer is valid
#             if serializer.is_valid():
#                 # Save each file as a separate object
#                 serializer.save()
#             else:
#                 return Response({'error_message': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

#         # If all files are processed, return success response
#         return Response({'message': 'Zip Files uploaded successfully.', 'data': serializer.data}, status=status.HTTP_201_CREATED)
#     else:
#         return Response({'error_message': 'No files uploaded in the request.'}, status=status.HTTP_400_BAD_REQUEST)
    

@api_view(['POST'])
def create_zipupload(request, pk):
    try:
        client = Client.objects.get(id=pk)
    except Client.DoesNotExist:
        return Response({'error': 'Client not found'}, status=status.HTTP_404_NOT_FOUND)

    data = {key: value for key, value in request.data.items()}
    serializer = ZipUploadSerializer2(data=data)

    if serializer.is_valid(raise_exception=True):
        zipupload_instance = serializer.save(client=client)

        # Handle file uploads
        files = request.FILES.getlist('files')
        if files:
            for file in files:
                file_data = {
                    'zipupload': zipupload_instance.pk,
                    'files': file,
                }
                file_serializer = FilesSerializer(data=file_data)
                file_serializer.is_valid(raise_exception=True)
                file_serializer.save()

            return Response({
                'message': 'Zip Upload created successfully.',
                'data': serializer.data
            }, status=status.HTTP_201_CREATED)
        else:
            return Response({
                'message': 'No files provided.'
            }, status=status.HTTP_400_BAD_REQUEST)
 

@api_view(['DELETE'])
def delete_zipupload(request, client_pk, pk):
    """
    Deletes a PurchaseInvoice by its primary key (ID) along with its associated product summaries.
    """
    try:
        # Retrieve the Client instance
        client = Client.objects.filter(id=client_pk).first()

        if not client:
            return Response({"error_message": "Client not found."}, status=status.HTTP_404_NOT_FOUND)

        # Retrieve the SalesInvoice instance
        zipupload = ZipUpload.objects.filter(id=pk, client=client).first()

        if not zipupload:
            return Response({"error_message": "Zip Uploaded not found."}, status=status.HTTP_404_NOT_FOUND)

        # Handle deletion of associated ProductSummary instances
        # product_summaries = expenses.product_summaries.all()
        # for product_summary in product_summaries:
        #     product_summary.delete()

        # Delete the SalesInvoice instance
        zipupload.delete()

        return Response({"message": "File deleted successfully."}, status=status.HTTP_200_OK)

    # except Exception as e:
    #     return Response({"error_message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    except Exception as e:
        error_details = traceback.format_exc()
        # print({"Error in update_sales_invoice" : str(e)})
        return Response({"error_message":str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
# ******************************************************Income Debit Note******************************************************
@api_view(['GET', 'POST'])
def create_income_debit_note_get(request, pk):
    try:
        client = Client.objects.get(id=pk)

    except Client.DoesNotExist:
        return Response({"error_message": "Client not found."}, status=404)

    if request.method == 'GET':
        # Initialize response context with default values
        context = {
            'serializer_customer': [],
            'serializer': [],
            'product_serializer': [],
            'message': None,
            'hsn': None,
            'location': None
        }

        # Retrieve query parameters
        received_value = request.GET.get('newValue')
        product_id = request.GET.get('productID')
        print('daadasdasdsa',received_value)
        # Process productID if it exists
        if product_id:
            try:
                hsn_cc = Product.objects.get(id=product_id).hsn
                context.update({
                    "message": "Product HSN found",
                    "hsn": HSNSerializer(hsn_cc).data
                })
            except (ValueError, Product.DoesNotExist):
                return Response({"error_message": "Invalid Product ID."}, status=400)

        # Process received_value if it exists
        if received_value:
            try:
                received_value = int(received_value)
                location = OfficeLocation.objects.get(id=received_value)
                branch_gst = location.branch.gst_no
                print("branch val",branch_gst)
                context.update({
                    "message": "Location found",
                    "location": OfficeLocationSerializer(location).data,
                    "branch_gst": branch_gst
                })
            except (ValueError, OfficeLocation.DoesNotExist):
                return Response({"error_message": "Invalid location ID."}, status=400)

        # Add data only if productID and received_value were not provided
        if not product_id and not received_value:
            off = OfficeLocation.objects.filter(branch__client=client)
            customer = Customer.objects.filter(client=client, customer=True)
            product = Product.objects.all()
            branch = Branch.objects.filter(client=client)

            serializer_customer = CustomerVendorSerializer(customer, many=True)
            serializer = OfficeLocationSerializer(off, many=True)
            product_serializer = ProductSerializer(product, many=True)
            branch_serializer = BranchSerailizer(branch, many=True)

            context.update({
                'serializer_customer': serializer_customer.data,
                'serializer': serializer.data,
                'product_serializer': product_serializer.data,
                'branch_serializer': branch_serializer.data
            })

        return Response(context)

@api_view(['POST'])
def create_income_debit_note(request, client_pk, income_pk):
    client = Client.objects.get(id=client_pk)
    income = Income.objects.get(id=income_pk, client=client)

    # Check if 'attach_e_way_bill' is in the request files
    if 'attach_invoice' in request.FILES:
        # Iterate through each file in the 'attach_e_way_bill' field
        for e_way_bill in request.FILES.getlist('attach_invoice'):
            # Prepare data for each file
            debitnote_data = {
                'attach_invoice': e_way_bill,  # The file being uploaded
                'client': client.id,  # Associate the file with the client
                'income' : income.id,
            }

            # Initialize the serializer for each file
            serializer = IncomeDebitNoteSerializer2(data=debitnote_data)

            # Check if the serializer is valid
            if serializer.is_valid():
                # Save each file as a separate object
                serializer.save()
            else:
                return Response({'error_message': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

        # If all files are processed, return success response
        return Response({'message': 'Debit Note E-way bill(s) uploaded successfully.'}, status=status.HTTP_200_OK)
    else:
        return Response({'error_message': 'No files uploaded in the request.'}, status=status.HTTP_400_BAD_REQUEST)
    
@api_view(['GET'])
def get_income_debit_note_data(request, client_pk, income_pk, debit_pk):
    """
    GET request to fetch sales invoice data along with related product summaries and client info.
    """
    # Fetch the sales invoice with related client information
    debit_note = IncomeDebitNote.objects.filter(client_id=client_pk, sales_invoice_id=income_pk, id=debit_pk).first()

    if not debit_note:
        return Response({"error_message": "Debit Note not found."}, status=status.HTTP_404_NOT_FOUND)

    # Serialize the sales invoice
    debit_note_data = IncomeDebitNoteSerializer3(debit_note).data

    # Prepare product summaries data
    product_summaries = debit_note.product_summaries.all()
    product_summary_data = [
        {
            "id": summary.id,
            "hsnCode": summary.hsn.hsn_code,
            "gstRate": summary.hsn.gst_rate,
            "product": summary.product.product_name,
            "description": summary.prod_description.description,
            "unit": summary.prod_description.unit,
            "rate": summary.prod_description.rate,
            "product_amount": summary.prod_description.product_amount,
            "cgst": summary.prod_description.cgst,
            "sgst": summary.prod_description.sgst,
            "igst": summary.prod_description.igst,
        }
        for summary in product_summaries
    ]

    # Helper function to safely get attributes from related objects
    def get_safe_attr(obj, attr, default=None):
        return getattr(obj, attr, default) if obj else default

    # Prepare client location data (with safe attribute access)
    client_location = debit_note.client_Location
    client_location_data = {
        "id": get_safe_attr(client_location, 'id'),
        "location": get_safe_attr(client_location, 'location'),
        "contact": get_safe_attr(client_location, 'contact'),
        "address": get_safe_attr(client_location, 'address'),
        "city": get_safe_attr(client_location, 'city'),
        "state": get_safe_attr(client_location, 'state'),
        "country": get_safe_attr(client_location, 'country'),
    }

    # Prepare customer data (with safe attribute access)
    customer = debit_note.customer
    customer_data = {
        "id": get_safe_attr(customer, 'id'),
        "name": get_safe_attr(customer, 'name'),
        "gst_no": get_safe_attr(customer, 'gst_no'),
        "pan": get_safe_attr(customer, 'pan'),
        "customer_address": get_safe_attr(customer, 'address'),
        "customer": get_safe_attr(customer, 'customer'),
        "vendor": get_safe_attr(customer, 'vendor'),
    }

    # Prepare final response data
    response_data = {
        "debit_note": debit_note_data,
        "product_summaries": product_summary_data,
        "client_location": client_location_data,
        "customer": customer_data,
    }

    return Response(response_data, status=status.HTTP_200_OK)

@api_view(['GET', 'POST'])
def update_income_debit_note(request, client_pk, income_pk):
    try:
        # Handle GET request
        if request.method == 'GET':
            debit_note = Income.objects.filter(client_id=client_pk, id=income_pk) \
                .select_related('client_Location', 'customer') \
                .prefetch_related('product_summaries__hsn', 'product_summaries__prod_description') \
                .first()

            if not debit_note:
                return Response({"error_message": "Debit Note not found."}, status=status.HTTP_404_NOT_FOUND)

            debit_notes = IncomeDebitNote.objects.filter(income=debit_note)
            product_unit_sums = defaultdict(int)

            for dn in debit_notes:
                for product_summary in dn.product_summaries.all():
                    product_name = product_summary.product_name()
                    product_unit = product_summary.unit() or 0  # Agar unit None ho toh 0 lelo
                    product_unit_sums[product_name] += product_unit

            remaining_units = {}
            for product in debit_note.product_summaries.all():
                product_name = product.product_name()
                si_unit = product.unit() or 0  # Sales Invoice ke unit
                dn_unit_sum = product_unit_sums[product_name]  # Debit Notes ka unit sum
                remaining_units[product_name] = si_unit - dn_unit_sum  # Remaining units

            print('ooooooooo',remaining_units)

            debit_notes_details = []
            for debit_note in debit_notes:
                product_summaries = debit_note.product_summaries.all()
                print(f"Found {len(product_summaries)} ProductSummaries for DebitNote {debit_note.id}")

                for product_summary in product_summaries:
                    product_details = {
                        'debit_note_id': debit_note.id,
                        'product_name': product_summary.product_name(),
                        'unit': product_summary.unit(),
                    }
                    debit_notes_details.append(product_details)

            dd = debit_notes_details
            print('Debit Notes',dd)

            debit_note_data = IncomeSerializer3(debit_note).data

            product_summaries = debit_note.product_summaries.all()
            product_summary_data = [
                {
                    "id": summary.id,
                    "hsnCode": summary.hsn.hsn_code,
                    "gstRate": summary.hsn.gst_rate,
                    "product": summary.product.product_name,
                    "description": summary.prod_description.description,
                    "rate": summary.prod_description.rate,
                    "product_amount": summary.prod_description.product_amount,
                    "unit": remaining_units.get(summary.product_name(), summary.prod_description.unit),
                    "cgst": summary.prod_description.cgst,
                    "sgst": summary.prod_description.sgst,
                    "igst": summary.prod_description.igst,
                }
                for summary in product_summaries
            ]

            # if debit_notes.month:
            #     formatted_month = debit_notes.month.strftime("%d-%m-%Y")  # Directly format date
            # else:
            #     formatted_month = None

            # if debit_notes.invoice_date:
            #     formatted_invoice_date = debit_notes.invoice_date.strftime("%d-%m-%Y")  # Directly format date
            # else:
            #     formatted_invoice_date = None


            response_data = {
                # "debit_note": {
                # **debit_note_data,
                # "month": formatted_month,  # Ensure month is sent in dd/mm/yyyy format
                # "invoice_date" : formatted_invoice_date,},
                # "product_summaries": product_summary_data,
                # "remaining_units": remaining_units,
                "debit_note": debit_note_data,
                "product_summaries": product_summary_data,
                "remaining_units": remaining_units,
                "client_location": {
                    "id": debit_note.client_Location.id if debit_note.client_Location else None,
                    "location": debit_note.client_Location.location if debit_note.client_Location else None,
                    "contact": debit_note.client_Location.contact if debit_note.client_Location else None,
                    "address": debit_note.client_Location.address if debit_note.client_Location else None,
                    "city": debit_note.client_Location.city if debit_note.client_Location else None,
                    "state": debit_note.client_Location.state if debit_note.client_Location else None,
                    "country": debit_note.client_Location.country if debit_note.client_Location else None,
                    "branchID": debit_note.client_Location.branch.id if debit_note.client_Location else None,
                },
                "customer": {
                    "id": debit_note.customer.id if debit_note.customer else None,
                    "name": debit_note.customer.name if debit_note.customer else None,
                    "gst_no": debit_note.customer.gst_no if debit_note.customer else None,
                    "pan": debit_note.customer.pan if debit_note.customer else None,
                    "email": debit_note.customer.email if debit_note.customer else None,
                    "contact": debit_note.customer.contact if debit_note.customer else None,
                    "customer_address": debit_note.customer.address if debit_note.customer else None,
                    "customer": debit_note.customer.customer if debit_note.customer else None,
                    "vendor": debit_note.customer.vendor if debit_note.customer else None,
                },
            }
            return Response(response_data, status=status.HTTP_200_OK)

        # Handle PUT request
        elif request.method == 'POST':
            try:
                with transaction.atomic(): 
                    payload = request.data
                    data = request.data

                    # Fetch the Client
                    client = Client.objects.filter(pk=client_pk).first()
                    if not client:
                        return Response({"error_message": "Client not found."}, status=status.HTTP_404_NOT_FOUND)

                    # Fetch the Sales Invoice
                    income = Income.objects.filter(client_id=client_pk, id=income_pk) \
                        .select_related('client_Location', 'customer') \
                        .prefetch_related('product_summaries__hsn', 'product_summaries__prod_description') \
                        .first()

                    if not income:
                        return Response({"error_message": "Income not found or does not belong to the client."},
                                        status=status.HTTP_404_NOT_FOUND)

                    # Calculate remaining units
                    debit_notes = IncomeDebitNote.objects.filter(income=income)
                    product_unit_sums = defaultdict(int)

                    for dn in debit_notes:
                        for product_summary in dn.product_summaries.all():
                            product_name = product_summary.product_name()
                            product_unit = product_summary.unit() or 0
                            product_unit_sums[product_name] += product_unit

                    remaining_units = {}
                    for product in income.product_summaries.all():
                        product_name = product.product_name()
                        si_unit = product.unit() or 0
                        dn_unit_sum = product_unit_sums[product_name]
                        remaining_units[product_name] = si_unit - dn_unit_sum

                    print('gggggggggg',remaining_units)
                    
                    # Validations..........
                    if all(unit == 0 for unit in remaining_units.values()):
                        return Response(
                            {"message": "Cannot create a new Income Debit Note. All product remaining units are 0."},
                            status=status.HTTP_400_BAD_REQUEST
                        )
                    
                    for product_name, unit in remaining_units.items():
                        if unit is None:
                            return Response({"message" : f"Remaining units for product '{product_name}' is missing or undefined."}, status=status.HTTP_400_BAD_REQUEST)
                        if unit < 0:
                            return Response({"message" : f"Remaining units for prooooooooooooduct '{product_name}' cannot be negative. Current value: {unit}"}, status=status.HTTP_400_BAD_REQUEST)

                    # Extract rows dynamically
                    rows_data = defaultdict(dict)
                    for key, value in payload.items():
                        if key.startswith("rows["):
                            row_index = key.split('[')[1].split(']')[0]
                            field_name = key.split('[')[2].split(']')[0]
                            rows_data[int(row_index)][field_name] = value
                    rows = [rows_data[index] for index in sorted(rows_data.keys())]

                    # Extract form data, vendor data, and invoice data
                    form_data = {
                        "offLocID": payload.get("formData[offLocID]"),
                        "location": payload.get("formData[location]"),
                        "contact": payload.get("formData[contact]"),
                        "address": payload.get("formData[address]"),
                        "city": payload.get("formData[city]"),
                        "state": payload.get("formData[state]"),
                        "country": payload.get("formData[country]"),
                        "branchID": payload.get("formData[branchID]"),
                    }
                    vendor_data = {
                        "name": payload.get("vendorData[name]"),
                        "gst_no": payload.get("vendorData[gst_no]"),
                        "pan": payload.get("vendorData[pan]"),
                        "customer_address": payload.get("vendorData[customer_address]"),
                        "email": payload.get("vendorData[email]"),
                        "contact": payload.get("vendorData[contact]"),
                        "customer": payload.get("vendorData[customer]", "").lower() == "true",
                        "vendor": payload.get("vendorData[vendor]", "").lower() == "true",
                    }
                    invoice_data = {
                        "invoice_no": payload.get("invoiceData[0][invoice_no]"),
                        "invoice_date": payload.get("invoiceData[0][invoice_date]"),
                        "month": payload.get("invoiceData[0][month]"),
                        "invoice_type": payload.get("invoiceData[0][invoice_type]"),
                        "entry_type": payload.get("invoiceData[0][entry_type]"),
                        "taxable_amount": payload.get("invoiceData[0][taxable_amount]"),
                        "totalall_gst": payload.get("invoiceData[0][totalall_gst]"),
                        "total_invoice_value": payload.get("invoiceData[0][total_invoice_value]"),
                        "tds_tcs_rate": payload.get("invoiceData[0][tds_tcs_rate]"),
                        "tcs": payload.get("invoiceData[0][tcs]"),
                        "tds": payload.get("invoiceData[0][tds]"),
                        "amount_receivable": payload.get("invoiceData[0][amount_receivable]"),
                    }
                    if invoice_data["invoice_date"]:
                        try:
                            invoice_data["invoice_date"] = datetime.strptime(invoice_data["invoice_date"], "%d-%m-%Y").date()
                        except ValueError:
                            return Response({"error_message": "Invalid Invoice date format, expected DD-MM-YYYY."}, status=status.HTTP_400_BAD_REQUEST)
                    print("Final invoice date:", invoice_data["invoice_date"])

                    if invoice_data["month"]:
                        try:
                            invoice_data["month"] = datetime.strptime(invoice_data["month"], "%d-%m-%Y").date()
                        except ValueError:
                            return Response({"error_message": "Invalid Month date format, expected DD-MM-YYYY."},status=status.HTTP_400_BAD_REQUEST)

                    # if invoice_data["utilise_month"]:
                    #     try:
                    #         invoice_data["utilise_month"] = datetime.strptime(invoice_data["utilise_month"], "%d-%m-%Y").date()
                    #     except ValueError:
                    #         return Response({"error_message": "Invalid Utilise Month date format, expected DD-MM-YYYY."},status=status.HTTP_400_BAD_REQUEST)

                    # print("Final month:", invoice_data["month"])


                    attach_invoice = request.FILES.get("invoiceData[0][attach_invoice]")
                    attach_e_way_bill = request.FILES.get("invoiceData[0][attach_e_way_bill]")

                    # Handle Office Location creation or selection
                    # location_obj = None
                    # if form_data["offLocID"]:
                    #     location_obj = OfficeLocation.objects.filter(id=form_data["offLocID"]).first()
                    #     if not location_obj:
                    #         return Response({"error_message": "Office Location not found."}, status=status.HTTP_404_NOT_FOUND)
                    # else:
                    #     branch_instance = Branch.objects.filter(id=form_data["branchID"], client=client).first()
                    #     if not branch_instance:
                    #         return Response({"error_message": f"Branch with ID {form_data['branchID']} not found or doesn't belong to the client."},
                    #                         status=status.HTTP_404_NOT_FOUND)
                    #     location_obj = OfficeLocation.objects.create(
                    #         location=form_data.get("location"),
                    #         contact=form_data.get("contact"),
                    #         address=form_data.get("address"),
                    #         city=form_data.get("city"),
                    #         state=form_data.get("state"),
                    #         country=form_data.get("country"),
                    #         branch=branch_instance
                    #     )

                    # Handle Vendor creation or update
                    vendor_obj = None
                    if vendor_data.get("gst_no"):
                        existing_vendor = Customer.objects.filter(client=client, gst_no=vendor_data["gst_no"]).first()
                        if existing_vendor:
                            vendor_serializer = CustomerVendorSerializer(existing_vendor, data=vendor_data, partial=True)
                            if vendor_serializer.is_valid():
                                vendor_obj = vendor_serializer.save()
                            else:
                                return Response({"vendor_errors": vendor_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
                        else:
                            vendor_serializer = CustomerVendorSerializer(data=vendor_data)
                            if vendor_serializer.is_valid():
                                vendor_obj = vendor_serializer.save(client=client)
                            else:
                                return Response({"vendor_errors": vendor_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

                    if any(unit is None for unit in remaining_units.values()):
                        return Response({"message" : f"Remaining units for product {product} is missing."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
                    else:
                        debit_note_serializer = IncomeDebitNoteSerializer(data=data)

                        if debit_note_serializer.is_valid():
                            location_obj = None
                            if form_data["offLocID"]:
                                location_obj = OfficeLocation.objects.filter(id=form_data["offLocID"]).first()
                                if not location_obj:
                                    return Response({"error_message": "Office Location not found."}, status=status.HTTP_404_NOT_FOUND)
                            else:
                                branch_instance = Branch.objects.filter(id=form_data["branchID"], client=client).first()
                                if not branch_instance:
                                    return Response({"error_message": f"Branch with ID {form_data['branchID']} not found or doesn't belong to the client."},
                                                    status=status.HTTP_404_NOT_FOUND)
                                location_obj, _ = OfficeLocation.objects.get_or_create(
                                    location=form_data.get("location"),
                                    contact=form_data.get("contact"),
                                    address=form_data.get("address"),
                                    city=form_data.get("city"),
                                    state=form_data.get("state"),
                                    country=form_data.get("country"),
                                    branch=branch_instance
                                )

                            for row in rows:
                                product_name = row.get('product')
                                unit_value = safe_decimal(row.get('unit', '0'))
                                
                                # Skip rows with unit value as 0 or negative
                                if unit_value <= 0:
                                    continue  # Skip processing this row
                                
                                # Validate remaining units for the product
                                if product_name in remaining_units:
                                    if remaining_units[product_name] < unit_value:
                                        return Response(
                                            {
                                                "message": f"Not enough units remaining for the product '{product_name}'. "
                                                        f"Available: {remaining_units[product_name]}, Requested: {unit_value}."
                                            },
                                            status=status.HTTP_400_BAD_REQUEST
                                        )
                                    
                                    # Deduct the used units from the remaining_units
                                    remaining_units[product_name] -= unit_value
                                else:
                                    return Response(
                                        {"message": f"Product '{product_name}' not found in remaining units."},
                                        status=status.HTTP_400_BAD_REQUEST
                                    )
                
                        # if not debit_note_serializer.is_valid():
                            debit_note = IncomeDebitNote.objects.create(
                            client=client,
                            # sales_invoice=sales_invoice,
                            income = income,
                            client_Location=location_obj,
                            customer=vendor_obj,
                            attach_invoice=attach_invoice,
                            attach_e_way_bill=attach_e_way_bill,
                            **invoice_data
                            )
                
                            product_summaries = []
                            for row in rows:
                                product_name = row.get('product')
                                unit_value = safe_decimal(row.get('unit', '0'))
                                hsn_code = row.get('hsnCode')
                                gst_rate = safe_decimal(row.get('gstRate', '0'))
                                product_id = row.get('product_id')
                                description_text = row.get('description', '')
                                rate_value = safe_decimal(row.get('rate', '0'))
                                amount = safe_decimal(row.get('product_amount', '0'))
                                cgst = safe_decimal(row.get('cgst', '0'))
                                sgst = safe_decimal(row.get('sgst', '0'))
                                igst = safe_decimal(row.get('igst', '0'))

                                hsn_code_obj, _ = HSNCode.objects.get_or_create(
                                    hsn_code=hsn_code, defaults={'gst_rate': gst_rate}
                                )

                                if product_id:
                                    product_obj = Product.objects.filter(id=product_id).first()
                                    if not product_obj:
                                        return Response({"message": f"Product with ID {product_id} not found."}, status=status.HTTP_404_NOT_FOUND)
                                else:
                                    product_obj, _ = Product.objects.get_or_create(
                                        product_name=product_name, hsn=hsn_code_obj, defaults={'unit_of_measure': unit_value}
                                    )
                                    product_description_obj = ProductDescription.objects.create(
                                        product=product_obj,
                                        description=description_text,
                                        unit=unit_value,
                                        rate=rate_value,
                                        product_amount=amount,
                                        cgst=cgst,
                                        sgst=sgst,
                                        igst=igst,
                                    )
                                    l = product_description_obj.unit
                                    product_summary = ProductSummaryIncomeDebitNote.objects.create(
                                        hsn=hsn_code_obj,
                                        product=product_obj,
                                        prod_description=product_description_obj
                                    )
                                    product_summaries.append(product_summary)
                                    debit_note.product_summaries.add(product_summary)
                                    
                                debit_note.save()
                            return Response({"message": "Income Debit Note created successfully."}, status=status.HTTP_200_OK)
            except Exception as e:
                return Response({"error": str(e)}, status=400)

    # except Exception as e:
    #     return Response({"error_message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    except Exception as e:
        error_details = traceback.format_exc()
        # print({"Error in update_sales_invoice" : str(e)})
        return Response({"error_message":str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET', 'PATCH'])
def income_debit_note_detail_view(request, client_pk, income_pk, debit_pk):
    try:
        # Fetch the sales invoice object
        debit_note = IncomeDebitNote.objects.get(client=client_pk, income_id=income_pk, pk=debit_pk)
    except Income.DoesNotExist:
        return Response({"error_message": "Income not found."}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        # Serialize and return the sales invoice details
        serializer = IncomeDebitNoteSerializerList(debit_note)
        return Response(serializer.data, status=status.HTTP_200_OK)

    elif request.method == 'PATCH':
        # Update only specific fields of the sales invoice
        data = request.data

        # Update fields from the request payload
        debit_note.invoice_no = data.get('invoice_no', debit_note.invoice_no)
        debit_note.invoice_date = data.get('invoice_date', debit_note.invoice_date)
        debit_note.month = data.get('month', debit_note.month)
        debit_note.invoice_type = data.get('invoice_type', debit_note.invoice_type)
        debit_note.entry_type = data.get('entry_type', debit_note.entry_type)
        debit_note.taxable_amount = safe_decimal(data.get('taxable_amount', debit_note.taxable_amount))
        debit_note.totalall_gst = safe_decimal(data.get('totalall_gst', debit_note.totalall_gst))
        debit_note.total_invoice_value = safe_decimal(data.get('total_invoice_value', debit_note.total_invoice_value))
        debit_note.tds_tcs_rate = safe_decimal(data.get('tds_tcs_rate', debit_note.tds_tcs_rate))
        debit_note.tds = safe_decimal(data.get('tds', debit_note.tds))
        debit_note.tcs = safe_decimal(data.get('tcs', debit_note.tcs))
        debit_note.amount_receivable = safe_decimal(data.get('amount_receivable', debit_note.amount_receivable))

        # Update product summaries if provided
        rows = data.get('rows', [])
        product_summaries = []
        for row in rows:
            hsn_code = row.get('hsnCode')
            gst_rate = safe_decimal(row.get('gstRate', '0'))
            product_name = row.get('product')
            description_text = row.get('description')
            unit_value = safe_decimal(row.get('unit', '0'))
            rate_value = safe_decimal(row.get('rate', '0'))
            amount = safe_decimal(row.get('product_amount', '0'))
            cgst = safe_decimal(row.get('cgst', '0'))
            sgst = safe_decimal(row.get('sgst', '0'))
            igst = safe_decimal(row.get('igst', '0'))

            # Handle HSNCode
            hsn_code_obj, created = HSNCode.objects.get_or_create(
                hsn_code=hsn_code,
                defaults={'gst_rate': gst_rate}
            )
            if not created and hsn_code_obj.gst_rate != gst_rate:
                hsn_code_obj.gst_rate = gst_rate
                hsn_code_obj.save()

            # Handle Product
            product_obj = Product.objects.filter(product_name=product_name, hsn=hsn_code_obj).first()
            if not product_obj:
                product_obj = Product.objects.create(
                    product_name=product_name,
                    hsn=hsn_code_obj,
                    unit_of_measure=unit_value
                )

            # Handle ProductDescription
            product_description_obj, created = ProductDescription.objects.get_or_create(
                product=product_obj,
                description=description_text,
                defaults={
                    'unit': unit_value,
                    'rate': rate_value,
                    'product_amount': amount,
                    'cgst': cgst,
                    'sgst': sgst,
                    'igst': igst
                }
            )
            if not created:
                # Update fields in ProductDescription
                product_description_obj.unit = unit_value
                product_description_obj.rate = rate_value
                product_description_obj.product_amount = amount
                product_description_obj.cgst = cgst
                product_description_obj.sgst = sgst
                product_description_obj.igst = igst
                product_description_obj.save()

            # Create ProductSummary
            product_summary = ProductSummaryIncomeDebitNote.objects.create(
                hsn=hsn_code_obj,
                product=product_obj,
                prod_description=product_description_obj
            )
            product_summaries.append(product_summary)

        # Save updates to the sales invoice
        if product_summaries:
            debit_note.product_summaries.set(product_summaries)
        debit_note.save()

        # Return the updated object
        updated_serializer = IncomeDebitNoteSerializer(debit_note)
        return Response(updated_serializer.data, status=status.HTTP_200_OK)
        
@api_view(['DELETE'])
def delete_income_debit_note(request, client_pk, income_pk, pk):
    """
    Deletes a SalesInvoice by its primary key (ID) along with its associated product summaries.
    """
    try:
        # Retrieve the Client instance
        client = Client.objects.filter(id=client_pk).first()

        if not client:
            return Response({"error_message": "Client not found."}, status=status.HTTP_404_NOT_FOUND)
        
        income = Income.objects.filter(pk=income_pk, client=client).first()
        if not income:
            return Response({"error_message": "Income not found or does not belong to the client."},
                            status=status.HTTP_404_NOT_FOUND)

        # Retrieve the SalesInvoice instance
        debit_note = IncomeDebitNote.objects.filter(id=pk, income=income ,client=client).first()

        if not debit_note:
            return Response({"error_message": "Debit Note not found."}, status=status.HTTP_404_NOT_FOUND)

        # Handle deletion of associated ProductSummary instances
        product_summaries = debit_note.product_summaries.all()
        for product_summary in product_summaries:
            product_summary.delete()

        # Delete the SalesInvoice instance
        debit_note.delete()

        return Response({"message": "Debit Note deleted successfully."}, status=status.HTTP_200_OK)

    except Exception as e:
        return Response({"error_message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
@api_view(['POST'])
def create_income_debit_note2(request, client_pk, income_pk):
    try:
        payload = request.data
        print('payload', payload)

        # Fetch the Client
        client = Client.objects.filter(pk=client_pk).first()
        if not client:
            return Response({"error_message": "Client not found."}, status=status.HTTP_404_NOT_FOUND)

        # Fetch the Sales Invoice
        income = Income.objects.filter(pk=income_pk, client=client).first()
        if not income:
            return Response({"error_message": "Income not found or does not belong to the client."},
                            status=status.HTTP_404_NOT_FOUND)

        # Extract rows dynamically
        rows_data = defaultdict(dict)
        for key, value in payload.items():
            if key.startswith("rows["):  # Check if the key corresponds to rows
                row_index = key.split('[')[1].split(']')[0]
                field_name = key.split('[')[2].split(']')[0]
                rows_data[int(row_index)][field_name] = value
        rows = [rows_data[index] for index in sorted(rows_data.keys())]
        print('rows', rows)

        # Extract form data, vendor data, and invoice data
        form_data = {
            "offLocID": payload.get("formData[offLocID]"),
            "location": payload.get("formData[location]"),
            "contact": payload.get("formData[contact]"),
            "address": payload.get("formData[address]"),
            "city": payload.get("formData[city]"),
            "state": payload.get("formData[state]"),
            "country": payload.get("formData[country]"),
            "branchID": payload.get("formData[branchID]"),
        }
        vendor_data = {
            "name": payload.get("vendorData[name]"),
            "gst_no": payload.get("vendorData[gst_no]"),
            "pan": payload.get("vendorData[pan]"),
            "customer_address": payload.get("vendorData[customer_address]"),
            "customer": payload.get("vendorData[customer]", "").lower() == "true",
            "vendor": payload.get("vendorData[vendor]", "").lower() == "true",
        }
        invoice_data = {
            "invoice_no": payload.get("invoiceData[0][invoice_no]"),
            "invoice_date": payload.get("invoiceData[0][invoice_date]"),
            "month": payload.get("invoiceData[0][month]"),
            "invoice_type": payload.get("invoiceData[0][invoice_type]"),
            "entry_type": payload.get("invoiceData[0][entry_type]"),
            "taxable_amount": payload.get("invoiceData[0][taxable_amount]"),
            "totalall_gst": payload.get("invoiceData[0][totalall_gst]"),
            "total_invoice_value": payload.get("invoiceData[0][total_invoice_value]"),
            "tds_tcs_rate": payload.get("invoiceData[0][tds_tcs_rate]"),
            "tcs": payload.get("invoiceData[0][tcs]"),
            "tds": payload.get("invoiceData[0][tds]"),
            "amount_receivable": payload.get("invoiceData[0][amount_receivable]"),
        }
        attach_invoice = request.FILES.get("invoiceData[0][attach_invoice]")
        attach_e_way_bill = request.FILES.get("invoiceData[0][attach_e_way_bill]")

        # Handle Office Location creation or selection
        location_obj = None
        if form_data["offLocID"]:
            location_obj = OfficeLocation.objects.filter(id=form_data["offLocID"]).first()
            if not location_obj:
                return Response({"error_message": "Office Location not found."}, status=status.HTTP_404_NOT_FOUND)
        else:
            branch_instance = Branch.objects.filter(id=form_data["branchID"], client=client).first()
            if not branch_instance:
                return Response({"error_message": f"Branch with ID {form_data['branchID']} not found or doesn't belong to the client."},
                                status=status.HTTP_404_NOT_FOUND)
            location_obj = OfficeLocation.objects.create(
                location=form_data.get("location"),
                contact=form_data.get("contact"),
                address=form_data.get("address"),
                city=form_data.get("city"),
                state=form_data.get("state"),
                country=form_data.get("country"),
                branch=branch_instance
            )

        # Handle Vendor creation or update
        vendor_obj = None
        if vendor_data.get("gst_no"):
            existing_vendor = Customer.objects.filter(client=client, gst_no=vendor_data["gst_no"]).first()
            if existing_vendor:
                vendor_serializer = CustomerVendorSerializer(existing_vendor, data=vendor_data, partial=True)
                if vendor_serializer.is_valid():
                    vendor_obj = vendor_serializer.save()
                else:
                    return Response({"vendor_errors": vendor_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
            else:
                vendor_serializer = CustomerVendorSerializer(data=vendor_data)
                if vendor_serializer.is_valid():
                    vendor_obj = vendor_serializer.save(client=client)
                else:
                    return Response({"vendor_errors": vendor_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

        # Create the Debit Note
        debit_note = IncomeDebitNote.objects.create(
            client=client,
            income = income,
            client_Location=location_obj,
            customer=vendor_obj,
            attach_invoice=attach_invoice,
            attach_e_way_bill=attach_e_way_bill,
            **invoice_data
        )

        # Handle Product Summaries (same logic as before)
        product_summaries = []  # To store created product summaries
        for row in rows:
            hsn_code = row.get('hsnCode')
            gst_rate = safe_decimal(row.get('gstRate', '0'))
            product_name = row.get('product')
            product_id = row.get('product_id')  # Assuming the frontend sends this if selecting an existing product
            description_text = row.get('description', '')
            unit_value = safe_decimal(row.get('unit', '0'))
            rate_value = safe_decimal(row.get('rate', '0'))
            amount = safe_decimal(row.get('product_amount', '0'))
            cgst = safe_decimal(row.get('cgst', '0'))
            sgst = safe_decimal(row.get('sgst', '0'))
            igst = safe_decimal(row.get('igst', '0'))

            # Handle HSNCode
            hsn_code_obj, _ = HSNCode.objects.get_or_create(
                hsn_code=hsn_code, defaults={'gst_rate': gst_rate}
            )

            # Handle Product (existing or new)
            if product_id:
                # Use existing product
                product_obj = Product.objects.filter(id=product_id).first()
                if not product_obj:
                    return Response({"error_message": f"Product with ID {product_id} not found."}, status=status.HTTP_404_NOT_FOUND)
            else:
                # Create new product
                product_obj, _ = Product.objects.get_or_create(
                    product_name=product_name, hsn=hsn_code_obj, defaults={'unit_of_measure': unit_value}
                )

            # Handle ProductDescription
            product_description_obj, _ = ProductDescription.objects.get_or_create(
                product=product_obj,
                description=description_text,
                defaults={
                    'unit': unit_value,
                    'rate': rate_value,
                    'product_amount': amount,
                    'cgst': cgst,
                    'sgst': sgst,
                    'igst': igst,
                }
            )

            # Create ProductSummary
            product_summary = ProductSummaryIncomeDebitNote.objects.create(
                hsn=hsn_code_obj,
                product=product_obj,
                prod_description=product_description_obj
            )
            product_summaries.append(product_summary)

            # Link ProductSummary to the SalesInvoice
            debit_note.product_summaries.add(product_summary)  # Add the product summary to the invoice
            print('cccccccc', request.FILES)

        return Response({"message": "Debit Note created successfully.", "invoice_id": debit_note.id}, status=status.HTTP_200_OK)

    # except Exception as e:
    #     return Response({"error_message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    except Exception as e:
        error_details = traceback.format_exc()
        # print({"Error in update_sales_invoice" : str(e)})
        return Response({"error_message":str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# ******************************************************Credit Note****************************************************   

@api_view(['GET', 'POST'])
def create_expenses_credit_note_get(request, pk):
    try:
        client = Client.objects.get(id=pk)

    except Client.DoesNotExist:
        return Response({"error_message": "Client not found."}, status=404)

    if request.method == 'GET':
        # Initialize response context with default values
        context = {
            'serializer_customer': [],
            'serializer': [],
            'product_serializer': [],
            'message': None,
            'hsn': None,
            'location': None
        }

        # Retrieve query parameters
        received_value = request.GET.get('newValue')
        product_id = request.GET.get('productID')
        print('daadasdasdsa',received_value)
        # Process productID if it exists
        if product_id:
            try:
                hsn_cc = Product.objects.get(id=product_id).hsn
                context.update({
                    "message": "Product HSN found",
                    "hsn": HSNSerializer(hsn_cc).data
                })
            except (ValueError, Product.DoesNotExist):
                return Response({"error_message": "Invalid Product ID."}, status=400)

        # Process received_value if it exists
        if received_value:
            try:
                received_value = int(received_value)
                location = OfficeLocation.objects.get(id=received_value)
                branch_gst = location.branch.gst_no
                print("branch val",branch_gst)
                context.update({
                    "message": "Location found",
                    "location": OfficeLocationSerializer(location).data,
                    "branch_gst": branch_gst
                })
            except (ValueError, OfficeLocation.DoesNotExist):
                return Response({"error_message": "Invalid location ID."}, status=400)

        # Add data only if productID and received_value were not provided
        if not product_id and not received_value:
            off = OfficeLocation.objects.filter(branch__client=client)
            customer = Customer.objects.filter(client=client, customer=True)
            product = Product.objects.all()
            branch = Branch.objects.filter(client=client)

            serializer_customer = CustomerVendorSerializer(customer, many=True)
            serializer = OfficeLocationSerializer(off, many=True)
            product_serializer = ProductSerializer(product, many=True)
            branch_serializer = BranchSerailizer(branch, many=True)

            context.update({
                'serializer_customer': serializer_customer.data,
                'serializer': serializer.data,
                'product_serializer': product_serializer.data,
                'branch_serializer': branch_serializer.data
            })

        return Response(context)
    
@api_view(['POST'])
def create_expenses_credit_note2(request, client_pk, expenses_pk):
    try:
        payload = request.data
        print('payload', payload)

        # Fetch the Client
        client = Client.objects.filter(pk=client_pk).first()
        if not client:
            return Response({"error_message": "Client not found."}, status=status.HTTP_404_NOT_FOUND)

        # Fetch the Purchase Invoice
        expenses = Expenses.objects.filter(pk=expenses_pk, client=client).first()
        if not expenses:
            return Response({"error_message": "Expenses not found or does not belong to the client."},
                            status=status.HTTP_404_NOT_FOUND)

        # Extract rows dynamically
        rows_data = defaultdict(dict)
        for key, value in payload.items():
            if key.startswith("rows["):  # Check if the key corresponds to rows
                row_index = key.split('[')[1].split(']')[0]
                field_name = key.split('[')[2].split(']')[0]
                rows_data[int(row_index)][field_name] = value
        rows = [rows_data[index] for index in sorted(rows_data.keys())]
        print('rows', rows)

        # Extract form data, vendor data, and invoice data
        form_data = {
            "offLocID": payload.get("formData[offLocID]"),
            "location": payload.get("formData[location]"),
            "contact": payload.get("formData[contact]"),
            "address": payload.get("formData[address]"),
            "city": payload.get("formData[city]"),
            "state": payload.get("formData[state]"),
            "country": payload.get("formData[country]"),
            "branchID": payload.get("formData[branchID]"),
        }
        vendor_data = {
            "name": payload.get("vendorData[name]"),
            "gst_no": payload.get("vendorData[gst_no]"),
            "pan": payload.get("vendorData[pan]"),
            "customer_address": payload.get("vendorData[customer_address]"),
            "customer": payload.get("vendorData[customer]", "").lower() == "true",
            "vendor": payload.get("vendorData[vendor]", "").lower() == "true",
        }
        invoice_data = {
            "invoice_no": payload.get("invoiceData[0][invoice_no]"),
            "invoice_date": payload.get("invoiceData[0][invoice_date]"),
            "month": payload.get("invoiceData[0][month]"),
            "invoice_type": payload.get("invoiceData[0][invoice_type]"),
            "entry_type": payload.get("invoiceData[0][entry_type]"),
            "taxable_amount": payload.get("invoiceData[0][taxable_amount]"),
            "totalall_gst": payload.get("invoiceData[0][totalall_gst]"),
            "total_invoice_value": payload.get("invoiceData[0][total_invoice_value]"),
            "tds_tcs_rate": payload.get("invoiceData[0][tds_tcs_rate]"),
            "tcs": payload.get("invoiceData[0][tcs]"),
            "tds": payload.get("invoiceData[0][tds]"),
            "amount_receivable": payload.get("invoiceData[0][amount_receivable]"),
            "utilise_edit": payload.get("invoiceData[0][utilise_edit]", "").lower() == "true", #nnnn
            "utilise_month": payload.get("invoiceData[0][utilise_month]").lower(), #nnnnnn
            
        }
        attach_invoice = request.FILES.get("invoiceData[0][attach_invoice]")
        attach_e_way_bill = request.FILES.get("invoiceData[0][attach_e_way_bill]")

        # Handle Office Location creation or selection
        location_obj = None
        if form_data["offLocID"]:
            location_obj = OfficeLocation.objects.filter(id=form_data["offLocID"]).first()
            if not location_obj:
                return Response({"error_message": "Office Location not found."}, status=status.HTTP_404_NOT_FOUND)
        else:
            branch_instance = Branch.objects.filter(id=form_data["branchID"], client=client).first()
            if not branch_instance:
                return Response({"error_message": f"Branch with ID {form_data['branchID']} not found or doesn't belong to the client."},
                                status=status.HTTP_404_NOT_FOUND)
            location_obj = OfficeLocation.objects.create(
                location=form_data.get("location"),
                contact=form_data.get("contact"),
                address=form_data.get("address"),
                city=form_data.get("city"),
                state=form_data.get("state"),
                country=form_data.get("country"),
                branch=branch_instance
            )

        # Handle Vendor creation or update
        vendor_obj = None
        if vendor_data.get("gst_no"):
            existing_vendor = Customer.objects.filter(client=client, gst_no=vendor_data["gst_no"]).first()
            if existing_vendor:
                vendor_serializer = CustomerVendorSerializer(existing_vendor, data=vendor_data, partial=True)
                if vendor_serializer.is_valid():
                    vendor_obj = vendor_serializer.save()
                else:
                    return Response({"vendor_errors": vendor_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
            else:
                vendor_serializer = CustomerVendorSerializer(data=vendor_data)
                if vendor_serializer.is_valid():
                    vendor_obj = vendor_serializer.save(client=client)
                else:
                    return Response({"vendor_errors": vendor_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

        # Create the Debit Note
        credit_note = ExpensesCreditNote.objects.create(
            client=client,
            expenses=expenses,
            client_Location=location_obj,
            vendor=vendor_obj,
            attach_invoice=attach_invoice,
            attach_e_way_bill=attach_e_way_bill,
            **invoice_data
        )
        product_summaries = []  # To store created product summaries
        for row in rows:
            hsn_code = row.get('hsnCode')
            gst_rate = safe_decimal(row.get('gstRate', '0'))
            product_name = row.get('product')
            product_id = row.get('product_id')  # Assuming the frontend sends this if selecting an existing product
            description_text = row.get('description', '')
            unit_value = safe_decimal(row.get('unit', '0'))
            rate_value = safe_decimal(row.get('rate', '0'))
            amount = safe_decimal(row.get('product_amount', '0'))
            cgst = safe_decimal(row.get('cgst', '0'))
            sgst = safe_decimal(row.get('sgst', '0'))
            igst = safe_decimal(row.get('igst', '0'))

            # Handle HSNCode
            hsn_code_obj, _ = HSNCode.objects.get_or_create(
                hsn_code=hsn_code, defaults={'gst_rate': gst_rate}
            )

            # Handle Product (existing or new)
            if product_id:
                # Use existing product
                product_obj = Product.objects.filter(id=product_id).first()
                if not product_obj:
                    return Response({"error_message": f"Product with ID {product_id} not found."}, status=status.HTTP_404_NOT_FOUND)
            else:
                # Create new product
                product_obj, _ = Product.objects.get_or_create(
                    product_name=product_name, hsn=hsn_code_obj, defaults={'unit_of_measure': unit_value}
                )

            # Handle ProductDescription
            product_description_obj, _ = ProductDescription.objects.get_or_create(
                product=product_obj,
                description=description_text,
                defaults={
                    'unit': unit_value,
                    'rate': rate_value,
                    'product_amount': amount,
                    'cgst': cgst,
                    'sgst': sgst,
                    'igst': igst,
                }
            )

            # Create ProductSummary
            product_summary = ProductSummaryExpensesCreditNote.objects.create(
                hsn=hsn_code_obj,
                product=product_obj,
                prod_description=product_description_obj
            )
            product_summaries.append(product_summary)

            # Link ProductSummary to the SalesInvoice
            credit_note.product_summaries.add(product_summary)  # Add the product summary to the invoice


        # Handle Product Summaries (same logic as before)

        return Response({"message": "Credit Note created successfully.", "invoice_id": credit_note.id}, status=status.HTTP_200_OK)

    # except Exception as e:
    #     return Response({"error_message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    except Exception as e:
        error_details = traceback.format_exc()
        # print({"Error in update_sales_invoice" : str(e)})
        return Response({"error_message":str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
@api_view(['POST'])
def create_expenses_credit_note(request, client_pk, expenses_pk):
    client = Client.objects.get(id=client_pk)
    expenses = Expenses.objects.get(id=expenses_pk, client=client)

    # Check if 'attach_e_way_bill' is in the request files
    if 'attach_invoice' in request.FILES:
        # Iterate through each file in the 'attach_e_way_bill' field
        for e_way_bill in request.FILES.getlist('attach_invoice'):
            # Prepare data for each file
            creditnote_data = {
                'attach_invoice': e_way_bill,  # The file being uploaded
                'client': client.id,  # Associate the file with the client
                'expenses' : expenses.id,
            }

            # Initialize the serializer for each file
            serializer = ExpensesCreditNoteSerializer2(data=creditnote_data)

            # Check if the serializer is valid
            if serializer.is_valid():
                # Save each file as a separate object
                serializer.save()
            else:
                return Response({'error_message': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

        # If all files are processed, return success response
        return Response({'message': 'Credit Note E-way bill(s) uploa ded successfully.'}, status=status.HTTP_200_OK)
    else:
        return Response({'error_message': 'No files uploaded in the request.'}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET', 'POST'])
def update_expenses_credit_note(request, client_pk, expenses_pk):

    try:
        # Handle GET request
        if request.method == 'GET':
            credit_note = Expenses.objects.filter(client_id=client_pk, id=expenses_pk )\
                .select_related('client_Location', 'vendor') \
                .prefetch_related('product_summaries__hsn', 'product_summaries__prod_description') \
                .first()
            if not credit_note:
                return Response({"error_message": "Expenses not found."}, status=status.HTTP_404_NOT_FOUND)

            credit_notes = ExpensesCreditNote.objects.filter(expenses=credit_note)
            product_unit_sums = defaultdict(int)

            for dn in credit_notes:
                for product_summary in dn.product_summaries.all():
                    product_name = product_summary.product_name()
                    product_unit = product_summary.unit() or 0  # Agar unit None ho toh 0 lelo
                    product_unit_sums[product_name] += product_unit

            remaining_units = {}
            for product in credit_note.product_summaries.all():
                product_name = product.product_name()
                si_unit = product.unit() or 0  # Sales Invoice ke unit
                dn_unit_sum = product_unit_sums[product_name]  # Debit Notes ka unit sum
                remaining_units[product_name] = si_unit - dn_unit_sum  # Remaining units

            print('ooooooooo',remaining_units)

            credit_notes_details = []
            for credit_note in credit_notes:
                product_summaries = credit_note.product_summaries.all()
                print(f"Found {len(product_summaries)} ProductSummaries for DebitNote {credit_note.id}")

                for product_summary in product_summaries:
                    product_details = {
                        'credit_note_id': credit_note.id,
                        'product_name': product_summary.product_name(),
                        'unit': product_summary.unit(),
                    }
                    credit_notes_details.append(product_details)

            dd = credit_notes_details
            print('Credit Notes',dd)


            credit_note_data = ExpensesSerializer3(credit_note).data
            
            print('creditnote',credit_note_data)

            product_summaries = credit_note.product_summaries.all()
            product_summary_data = [
                {
                    "id": summary.id,
                    "hsnCode": summary.hsn.hsn_code,
                    "gstRate": summary.hsn.gst_rate,
                    "product": summary.product.product_name,
                    "description": summary.prod_description.description,
                    # "unit": summary.prod_description.unit,
                    "unit": remaining_units.get(summary.product_name(), summary.prod_description.unit),
                    "rate": summary.prod_description.rate,
                    "product_amount": summary.prod_description.product_amount,
                    "cgst": summary.prod_description.cgst,
                    "sgst": summary.prod_description.sgst,
                    "igst": summary.prod_description.igst,
                }
                for summary in product_summaries
            ]

            response_data = {
                "credit_note": credit_note_data,
                "product_summaries": product_summary_data,
                "client_location": {
                    "id": credit_note.client_Location.id if credit_note.client_Location else None,
                    "location": credit_note.client_Location.location if credit_note.client_Location else None,
                    "contact": credit_note.client_Location.contact if credit_note.client_Location else None,
                    "address": credit_note.client_Location.address if credit_note.client_Location else None,
                    "city": credit_note.client_Location.city if credit_note.client_Location else None,
                    "state": credit_note.client_Location.state if credit_note.client_Location else None,
                    "country": credit_note.client_Location.country if credit_note.client_Location else None,
                    "branchID": credit_note.client_Location.branch.id if credit_note.client_Location else None,
                },
                "vendor": {
                    "id": credit_note.vendor.id if credit_note.vendor else None,
                    "name": credit_note.vendor.name if credit_note.vendor else None,
                    "gst_no": credit_note.vendor.gst_no if credit_note.vendor else None,
                    "pan": credit_note.vendor.pan if credit_note.vendor else None,
                    "email": credit_note.vendor.email if credit_note.vendor else None,
                    "contact": credit_note.vendor.contact if credit_note.vendor else None,
                    "customer_address": credit_note.vendor.address if credit_note.vendor else None,
                    "customer": credit_note.vendor.customer if credit_note.vendor else None,
                    "vendor": credit_note.vendor.vendor if credit_note.vendor else None,
                },
            }
            print('bbbbbbbbbbbbbbbbbbbbbbb',response_data["credit_note"])
            return Response(response_data, status=status.HTTP_200_OK)

        # Handle PUT request
        elif request.method == 'POST':
            try:
                with transaction.atomic():
                    payload = request.data
                    data = request.data
                    # print('payload', payload)

                    # Fetch the Client
                    client = Client.objects.filter(pk=client_pk).first()
                    if not client:
                        return Response({"message": "Client not found."}, status=status.HTTP_404_NOT_FOUND)

                    # Fetch the Purchase Invoice
                    expenses = Expenses.objects.filter(client_id=client_pk, id=expenses_pk )\
                        .select_related('client_Location', 'vendor') \
                        .prefetch_related('product_summaries__hsn', 'product_summaries__prod_description') \
                        .first()
                    if not expenses:
                        return Response({"message": "Credit Note not found."}, status=status.HTTP_404_NOT_FOUND)

                    debit_notes = ExpensesCreditNote.objects.filter(expenses=expenses)
                    product_unit_sums = defaultdict(int)

                    for dn in debit_notes:
                        for product_summary in dn.product_summaries.all():
                            product_name = product_summary.product_name()
                            product_unit = product_summary.unit() or 0  # Agar unit None ho toh 0 lelo
                            product_unit_sums[product_name] += product_unit

                    remaining_units = {}
                    for product in expenses.product_summaries.all():
                        product_name = product.product_name()
                        si_unit = product.unit() or 0  # Sales Invoice ke unit
                        dn_unit_sum = product_unit_sums[product_name]  # Debit Notes ka unit sum
                        remaining_units[product_name] = si_unit - dn_unit_sum  # Remaining units

                    print('gggggggg',remaining_units)

                    
                    # Validations..........
                    if all(unit == 0 for unit in remaining_units.values()):
                        return Response(
                            {"message": "Cannot create a new Expenses Credit Note. All product remaining units are 0."},
                            status=status.HTTP_400_BAD_REQUEST
                        )
                    
                    for product_name, unit in remaining_units.items():
                        if unit is None:
                            return Response({"message" : f"Remaining units for product '{product_name}' is missing or undefined."}, status=status.HTTP_400_BAD_REQUEST)
                        if unit < 0:
                            return Response({"message" : f"Remaining units for prooooooooooooduct '{product_name}' cannot be negative. Current value: {unit}"}, status=status.HTTP_400_BAD_REQUEST)


                    # Extract rows dynamically
                    rows_data = defaultdict(dict)
                    for key, value in payload.items():
                        if key.startswith("rows["):  # Check if the key corresponds to rows
                            row_index = key.split('[')[1].split(']')[0]
                            field_name = key.split('[')[2].split(']')[0]
                            rows_data[int(row_index)][field_name] = value
                    rows = [rows_data[index] for index in sorted(rows_data.keys())]
                    print('rows', rows)

                    # Extract form data, vendor data, and invoice data
                    form_data = {
                        "offLocID": payload.get("formData[offLocID]"),
                        "location": payload.get("formData[location]"),
                        "contact": payload.get("formData[contact]"),
                        "address": payload.get("formData[address]"),
                        "city": payload.get("formData[city]"),
                        "state": payload.get("formData[state]"),
                        "country": payload.get("formData[country]"),
                        "branchID": payload.get("formData[branchID]"),
                    }
                    vendor_data = {
                        "name": payload.get("vendorData[name]"),
                        "gst_no": payload.get("vendorData[gst_no]"),
                        "pan": payload.get("vendorData[pan]"),
                        "email": payload.get("vendorData[email]"),
                        "contact": payload.get("vendorData[contact]"),
                        "customer_address": payload.get("vendorData[customer_address]"),
                        "customer": payload.get("vendorData[customer]", "").lower() == "true",
                        "vendor": payload.get("vendorData[vendor]", "").lower() == "true",
                    }
                    invoice_data = {
                        "invoice_no": payload.get("invoiceData[0][invoice_no]"),
                        "invoice_date": payload.get("invoiceData[0][invoice_date]"),
                        "month": payload.get("invoiceData[0][month]"),
                        "invoice_type": payload.get("invoiceData[0][invoice_type]"),
                        "entry_type": payload.get("invoiceData[0][entry_type]"),
                        "taxable_amount": payload.get("invoiceData[0][taxable_amount]"),
                        "totalall_gst": payload.get("invoiceData[0][totalall_gst]"),
                        "total_invoice_value": payload.get("invoiceData[0][total_invoice_value]"),
                        "tds_tcs_rate": payload.get("invoiceData[0][tds_tcs_rate]"),
                        "tcs": payload.get("invoiceData[0][tcs]"),
                        "tds": payload.get("invoiceData[0][tds]"),
                        "amount_receivable": payload.get("invoiceData[0][amount_receivable]"),
                        "utilise_edit": payload.get("invoiceData[0][utilise_edit]", "").lower() == "true", #nnnn
                        "utilise_month": payload.get("invoiceData[0][utilise_month]"), #nnnnnn
                        
                    }

                    if invoice_data["invoice_date"]:
                        try:
                            invoice_data["invoice_date"] = datetime.strptime(invoice_data["invoice_date"], "%d-%m-%Y").date()
                        except ValueError:
                            return Response({"error_message": "Invalid Invoice date format, expected DD-MM-YYYY."}, status=status.HTTP_400_BAD_REQUEST)
                    print("Final invoice date:", invoice_data["invoice_date"])

                    if invoice_data["month"]:
                        try:
                            invoice_data["month"] = datetime.strptime(invoice_data["month"], "%d-%m-%Y").date()
                        except ValueError:
                            return Response({"error_message": "Invalid Month date format, expected DD-MM-YYYY."},status=status.HTTP_400_BAD_REQUEST)

                    if invoice_data["utilise_month"]:
                        try:
                            invoice_data["utilise_month"] = datetime.strptime(invoice_data["utilise_month"], "%d-%m-%Y").date()
                        except ValueError:
                            return Response({"error_message": "Invalid Utilise Month date format, expected DD-MM-YYYY."},status=status.HTTP_400_BAD_REQUEST)

                    print("Final month:", invoice_data["month"])


                    attach_invoice = request.FILES.get("invoiceData[0][attach_invoice]")
                    attach_e_way_bill = request.FILES.get("invoiceData[0][attach_e_way_bill]")

                    # Handle Office Location creation or selection
                    # location_obj = None
                    # if form_data["offLocID"]:
                    #     location_obj = OfficeLocation.objects.filter(id=form_data["offLocID"]).first()
                    #     if not location_obj:
                    #         return Response({"error_message": "Office Location not found."}, status=status.HTTP_404_NOT_FOUND)
                    # else:
                    #     branch_instance = Branch.objects.filter(id=form_data["branchID"], client=client).first()
                    #     if not branch_instance:
                    #         return Response({"error_message": f"Branch with ID {form_data['branchID']} not found or doesn't belong to the client."},
                    #                         status=status.HTTP_404_NOT_FOUND)
                    #     location_obj = OfficeLocation.objects.create(
                    #         location=form_data.get("location"),
                    #         contact=form_data.get("contact"),
                    #         address=form_data.get("address"),
                    #         city=form_data.get("city"),
                    #         state=form_data.get("state"),
                    #         country=form_data.get("country"),
                    #         branch=branch_instance
                    #     )

                    # Handle Vendor creation or update
                    vendor_obj = None
                    if vendor_data.get("gst_no"):
                        existing_vendor = Customer.objects.filter(client=client, gst_no=vendor_data["gst_no"]).first()
                        if existing_vendor:
                            vendor_serializer = CustomerVendorSerializer(existing_vendor, data=vendor_data, partial=True)
                            if vendor_serializer.is_valid():
                                vendor_obj = vendor_serializer.save()
                            else:
                                return Response({"vendor_errors": vendor_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
                        else:
                            vendor_serializer = CustomerVendorSerializer(data=vendor_data)
                            if vendor_serializer.is_valid():
                                vendor_obj = vendor_serializer.save(client=client)
                            else:
                                return Response({"vendor_errors": vendor_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

                    if any(unit is None for unit in remaining_units.values()):
                        return Response({"message" : f"Remaining units for product {product} is missing."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
                    else:
                        credit_note_serializer = CreditNoteSerializer(data=data)

                        if credit_note_serializer.is_valid():
                            # Handle Office Location creation or selection
                            location_obj = None
                            if form_data["offLocID"]:
                                location_obj = OfficeLocation.objects.filter(id=form_data["offLocID"]).first()
                                if not location_obj:
                                    return Response({"message": "Office Location not found."}, status=status.HTTP_404_NOT_FOUND)
                            else:
                                branch_instance = Branch.objects.filter(id=form_data["branchID"], client=client).first()
                                if not branch_instance:
                                    return Response({"message": f"Branch with ID {form_data['branchID']} not found or doesn't belong to the client."},
                                                    status=status.HTTP_404_NOT_FOUND)
                                location_obj, _ = OfficeLocation.objects.get_or_create(
                                    location=form_data.get("location"),
                                    contact=form_data.get("contact"),
                                    address=form_data.get("address"),
                                    city=form_data.get("city"),
                                    state=form_data.get("state"),
                                    country=form_data.get("country"),
                                    branch=branch_instance
                                )
                            for row in rows:
                                product_name = row.get('product')
                                unit_value = safe_decimal(row.get('unit', '0'))
                                
                                # Skip rows with unit value as 0 or negative
                                if unit_value <= 0:
                                    continue  # Skip processing this row
                                
                                # Validate remaining units for the product
                                if product_name in remaining_units:
                                    if remaining_units[product_name] < unit_value:
                                        return Response(
                                            {
                                                "message": f"Not enough units remaining for the product '{product_name}'. "
                                                        f"Available: {remaining_units[product_name]}, Requested: {unit_value}."
                                            },
                                            status=status.HTTP_400_BAD_REQUEST
                                        )
                                    
                                    # Deduct the used units from the remaining_units
                                    remaining_units[product_name] -= unit_value
                                else:
                                    return Response(
                                        {"message": f"Product '{product_name}' not found in remaining units."},
                                        status=status.HTTP_400_BAD_REQUEST
                                    )

                            # Create the Debit Note
                            credit_note = ExpensesCreditNote.objects.create(
                                client=client,
                                # purchase_invoice=purchase_invoice,
                                expenses = expenses,
                                client_Location=location_obj,
                                vendor=vendor_obj,
                                attach_invoice=attach_invoice,
                                attach_e_way_bill=attach_e_way_bill,
                                **invoice_data
                            )
                            product_summaries = []
                            for row in rows:
                                product_name = row.get('product')
                                unit_value = safe_decimal(row.get('unit', '0'))
                                hsn_code = row.get('hsnCode')
                                gst_rate = safe_decimal(row.get('gstRate', '0'))
                                product_id = row.get('product_id')
                                description_text = row.get('description', '')
                                rate_value = safe_decimal(row.get('rate', '0'))
                                amount = safe_decimal(row.get('product_amount', '0'))
                                cgst = safe_decimal(row.get('cgst', '0'))
                                sgst = safe_decimal(row.get('sgst', '0'))
                                igst = safe_decimal(row.get('igst', '0'))

                                hsn_code_obj, _ = HSNCode.objects.get_or_create(
                                    hsn_code=hsn_code, defaults={'gst_rate': gst_rate}
                                )

                                if product_id:
                                    product_obj = Product.objects.filter(id=product_id).first()
                                    if not product_obj:
                                        return Response({"message": f"Product with ID {product_id} not found."}, status=status.HTTP_404_NOT_FOUND)
                                else:
                                    product_obj, _ = Product.objects.get_or_create(
                                        product_name=product_name, hsn=hsn_code_obj, defaults={'unit_of_measure': unit_value}
                                    )
                                    product_description_obj = ProductDescription.objects.create(
                                        product=product_obj,
                                        description=description_text,
                                        unit=unit_value,
                                        rate=rate_value,
                                        product_amount=amount,
                                        cgst=cgst,
                                        sgst=sgst,
                                        igst=igst,
                                    )
                                    l = product_description_obj.unit
                                    product_summary = ProductSummaryExpensesCreditNote.objects.create(
                                        hsn=hsn_code_obj,
                                        product=product_obj,
                                        prod_description=product_description_obj
                                    )
                                    product_summaries.append(product_summary)
                                    credit_note.product_summaries.add(product_summary)
                                    
                                credit_note.save()
                            return Response({"message": "Expenses Credit Note created successfully."}, status=status.HTTP_200_OK)
            except Exception as e:
                return Response({"error": str(e)}, status=400)

    # except Exception as e:
    #     print("Error in update_credit_note:", str(e))
    #     return Response({"error_message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    except Exception as e:
        error_details = traceback.format_exc()
        # print({"Error in update_sales_invoice" : str(e)})
        return Response({"error_message":str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
@api_view(['DELETE'])
def delete_expenses_credit_note(request, client_pk, expenses_pk, credit_pk):
    """
    Deletes a SalesInvoice by its primary key (ID) along with its associated product summaries.
    """
    try:
        # Retrieve the Client instance
        client = Client.objects.filter(id=client_pk).first()

        if not client:
            return Response({"error_message": "Client not found."}, status=status.HTTP_404_NOT_FOUND)
        
        expenses = Expenses.objects.filter(pk=expenses_pk, client=client).first()
        if not expenses:
            return Response({"error_message": "Expenses not found or does not belong to the client."},
                            status=status.HTTP_404_NOT_FOUND)

        # Retrieve the SalesInvoice instance
        credit_note = ExpensesCreditNote.objects.filter(id=credit_pk, expenses=expenses ,client=client).first()

        if not credit_note:
            return Response({"error_message": "Credit Note not found."}, status=status.HTTP_404_NOT_FOUND)

        # Handle deletion of associated ProductSummary instances
        product_summaries = credit_note.product_summaries.all()
        for product_summary in product_summaries:
            product_summary.delete()

        # Delete the SalesInvoice instance
        credit_note.delete()

        return Response({"message": "Credit Note deleted successfully."}, status=status.HTTP_200_OK)

    # except Exception as e:
    #     return Response({"error_message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    except Exception as e:
        error_details = traceback.format_exc()
        # print({"Error in update_sales_invoice" : str(e)})
        return Response({"error_message":str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET', 'PATCH'])
def expenses_credit_note_detail_view(request, client_pk, expenses_pk, credit_pk):
    try:
        # Fetch the sales invoice object
        credit_note = ExpensesCreditNote.objects.get(client=client_pk, expenses=expenses_pk, pk=credit_pk)
    except Expenses.DoesNotExist:
        return Response({"error_message": "Expenses not found."}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        # Serialize and return the sales invoice details
        serializer = ExpensesCreditNoteSerializerList(credit_note)
        return Response(serializer.data, status=status.HTTP_200_OK)

    elif request.method == 'PATCH':
        # Update only specific fields of the sales invoice
        data = request.data

        # Update fields from the request payload
        credit_note.invoice_no = data.get('invoice_no', credit_note.invoice_no)
        credit_note.invoice_date = data.get('invoice_date', credit_note.invoice_date)
        credit_note.month = data.get('month', credit_note.month)
        credit_note.invoice_type = data.get('invoice_type', credit_note.invoice_type)
        credit_note.entry_type = data.get('entry_type', credit_note.entry_type)
        credit_note.taxable_amount = safe_decimal(data.get('taxable_amount', credit_note.taxable_amount))
        credit_note.totalall_gst = safe_decimal(data.get('totalall_gst', credit_note.totalall_gst))
        credit_note.total_invoice_value = safe_decimal(data.get('total_invoice_value', credit_note.total_invoice_value))
        credit_note.tds_tcs_rate = safe_decimal(data.get('tds_tcs_rate', credit_note.tds_tcs_rate))
        credit_note.tds = safe_decimal(data.get('tds', credit_note.tds))
        credit_note.tcs = safe_decimal(data.get('tcs', credit_note.tcs))
        credit_note.amount_receivable = safe_decimal(data.get('amount_receivable', credit_note.amount_receivable))
        credit_note.utilise_edit = data.get('utilise_edit', credit_note.utilise_edit)
        credit_note.utilise_month = data.get('utilise_month', credit_note.utilise_month)


        # Update product summaries if provided
        rows = data.get('rows', [])
        product_summaries = []
        for row in rows:
            hsn_code = row.get('hsnCode')
            gst_rate = safe_decimal(row.get('gstRate', '0'))
            product_name = row.get('product')
            description_text = row.get('description')
            unit_value = safe_decimal(row.get('unit', '0'))
            rate_value = safe_decimal(row.get('rate', '0'))
            amount = safe_decimal(row.get('product_amount', '0'))
            cgst = safe_decimal(row.get('cgst', '0'))
            sgst = safe_decimal(row.get('sgst', '0'))
            igst = safe_decimal(row.get('igst', '0'))

            # Handle HSNCode
            hsn_code_obj, created = HSNCode.objects.get_or_create(
                hsn_code=hsn_code,
                defaults={'gst_rate': gst_rate}
            )
            if not created and hsn_code_obj.gst_rate != gst_rate:
                hsn_code_obj.gst_rate = gst_rate
                hsn_code_obj.save()

            # Handle Product
            product_obj = Product.objects.filter(product_name=product_name, hsn=hsn_code_obj).first()
            if not product_obj:
                product_obj = Product.objects.create(
                    product_name=product_name,
                    hsn=hsn_code_obj,
                    unit_of_measure=unit_value
                )

            # Handle ProductDescription
            product_description_obj, created = ProductDescription.objects.get_or_create(
                product=product_obj,
                description=description_text,
                defaults={
                    'unit': unit_value,
                    'rate': rate_value,
                    'product_amount': amount,
                    'cgst': cgst,
                    'sgst': sgst,
                    'igst': igst
                }
            )
            if not created:
                # Update fields in ProductDescription
                product_description_obj.unit = unit_value
                product_description_obj.rate = rate_value
                product_description_obj.product_amount = amount
                product_description_obj.cgst = cgst
                product_description_obj.sgst = sgst
                product_description_obj.igst = igst
                product_description_obj.save()

            # Create ProductSummary
            product_summary = ProductSummaryExpensesCreditNote.objects.create(
                hsn=hsn_code_obj,
                product=product_obj,
                prod_description=product_description_obj
            )
            product_summaries.append(product_summary)

        # Save updates to the sales invoice
        if product_summaries:
            credit_note.product_summaries.set(product_summaries)
        credit_note.save()

        # Return the updated object
        updated_serializer = ExpensesCreditNoteSerializer(credit_note)
        return Response(updated_serializer.data, status=status.HTTP_200_OK)

# ************************************************Excel File**************************************************

@api_view(['POST'])
def create_excel_file(request):
    if request.method == 'POST':
        ser = ExcelFileSerializer(data=request.data)
        if ser.is_valid():
            ser.save()
            return Response({
                'message':  'Excel Upload',
                'data' : ser.data,
            }, status=status.HTTP_201_CREATED)
        else: 
            return Response({
                'message' : 'Fail to upload Excel',
                'error_message' : ser.errors,
            }, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
def get_excel(request):
    excel = ExcelFile.objects.all()
    if request.method == 'GET':
        ser = ExcelFileSerializer(excel, many=True)
        return Response(ser.data)
    else:
        return Response(ser.errors)

@api_view(['DELETE'])
def delete_excel(request, excel_pk):
    excel = ExcelFile.objects.get(id=excel_pk)
    if request.method == 'DELETE':
        excel.delete()
        return Response({'message': 'Excel Deleted'},status=status.HTTP_200_OK)
    else :
        return Response({'error_message':'Fail to Delete Excel'},status=status.HTTP_400_BAD_REQUEST)

# ************************************************Acknowledgement**************************************************

@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def create_acknowledgement(request,pk):
    client = Client.objects.get(id=pk)
    instance_data = request.data
    data = {key: value for key, value in instance_data.items()}

    serializer = AcknowledgementSerializer(data=data)
    if serializer.is_valid(raise_exception=True):
        print('create data qqq',request.data)
        ack_instance = serializer.save(client=client)
        print('create data',request.data)

        if request.FILES:
            # files = dict((request.FILES).lists()).get('return_file','computation_file',None)
            # filess = dict((request.FILES).lists()).get('files',None)
            # files = request.FILES.getlist('files')
            files_return = dict(request.FILES.lists()).get('return_file', [])
            files_computation = dict(request.FILES.lists()).get('computation_file', [])

            files = files_return 
            filess = files_computation
            if files :
                for file in files:
                    file_data = {
                        'ack': ack_instance.pk,
                        'return_file': file if file in files_return else None,
                        # 'computation_file': file if file in files_computation else None,
                    }
                    file_serializer= FilesSerializer(data=file_data)
                    if file_serializer.is_valid(raise_exception=True):
                        file_serializer.save()

            if filess:
                for file in filess:
                    file_data = {
                        'ack': ack_instance.pk,
                        'computation_file': file if file in files_computation else None,
                    }
                    file_serializer= FilesSerializer(data=file_data)
                    if file_serializer.is_valid(raise_exception=True):
                        file_serializer.save()

            

            return Response({'message':'Bank created successfully', 'data' : serializer.data},status=status.HTTP_201_CREATED)
    return Response({
        'message':'Fail to create bank', 
        'error_message' : serializer.errors
        },status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST', 'GET'])
@parser_classes([MultiPartParser, FormParser])
def update_acknowledgement(request, pk, ack_pk):
    try:
        client = Client.objects.get(id=pk)
        ack = Acknowledgement.objects.get(id=ack_pk)
    except Client.DoesNotExist:
        return Response({"message": "Client not found"}, status=status.HTTP_404_NOT_FOUND)
    except Acknowledgement.DoesNotExist:
        return Response({"message": "Acknowledgement not found"}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'POST':
        # Update Acknowledgement fields
        ser = AcknowledgementSerializer(instance=ack, data=request.data, partial=True)
        if ser.is_valid():
            print('Updated data before save:', request.data)
            ser.save(client=client)
            print('Updated data after save:', request.data)

            computation_files = request.FILES.getlist("computation_file")
            return_files = request.FILES.getlist("return_file")

            print("Received Computation Files:", computation_files)  # Debugging
            print("Received Return Files:", return_files)

            if return_files:
                Files.objects.filter(ack=ack).exclude(return_file="").delete()
                for file in return_files:
                    file_serializer = FilesSerializer(data={'ack': ack.pk, 'return_file': file})
                    if file_serializer.is_valid():
                        file_serializer.save()
                    else:
                        return Response(file_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

            if computation_files:
                Files.objects.filter(ack=ack).exclude(computation_file="").delete()
                for file in computation_files:
                    file_serializer = FilesSerializer(data={'ack': ack.pk, 'computation_file': file})
                    if file_serializer.is_valid():
                        file_serializer.save()
                    else:
                        return Response(file_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

            return Response({'message': 'Acknowledgement updated successfully', 'data': ser.data}, status=status.HTTP_200_OK)
        else:
            return Response({'message': 'Failed to update acknowledgement', 'error_message': ser.errors}, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'GET':
        ser = AcknowledgementSerializer(ack)
        return Response(ser.data)

@api_view(['DELETE'])
def delete_acknowledgement(request, pk, ack_pk):
    client = Client.objects.get(id=pk)
    ack = Acknowledgement.objects.get(id=ack_pk)
    if request.method == 'DELETE':
        ack.delete()
        return Response({'message': 'Acknowledgement Deleted'},status=status.HTTP_200_OK)
    else :
        return Response({'error_message':'Fail to Delete Acknowledgement'},status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
def single_acknowledgement(request, pk, ack_pk):
    client = Client.objects.get(id=pk)
    ack = Acknowledgement.objects.get(id=ack_pk)
    if request.method == 'GET':
        ser = AcknowledgementSerializer(ack)
        print(ser)
        return Response(ser.data)

@api_view(['GET'])
def get_acknowledgement(request, pk):
    client = Client.objects.get(id=pk)
    ack = Acknowledgement.objects.filter(client=client)
    if request.method == 'GET':
        ser = AcknowledgementSerializer(ack, many=True)
        return Response(ser.data)
    else:
        return Response(ser.errors)

@api_view(['GET'])
def download_computation_file(request, pk, ack_pk):
    client = Client.objects.get(id=pk)
    ack = Acknowledgement.objects.get(id=ack_pk, client = client)
    if request.method == 'GET':
        ser = AcknowledgementSerializer(ack)
        try:
            ack = Acknowledgement.objects.get(id=ack_pk, client=client)
        except Acknowledgement.DoesNotExist:
            return Response({"error": "Acknowledgement not found."}, status=status.HTTP_404_NOT_FOUND)

        files = Files.objects.filter(ack=ack).exclude(computation_file='')

        if not files.exists():
            return Response([], status=status.HTTP_200_OK)

        # Return only the 'file' key with the computation_file path
        file_data = [{"file": file.computation_file.url} for file in files]

        return Response(file_data, status=status.HTTP_200_OK)

# @api_view(['GET'])
# def download_computation_file(request, pk, ack_pk):
#     try:
#         client = Client.objects.get(id=pk)
#         ack = Acknowledgement.objects.get(id=ack_pk, client=client)
#     except (Client.DoesNotExist, Acknowledgement.DoesNotExist):
#         return Response({"error": "Client or Acknowledgement not found."}, status=status.HTTP_404_NOT_FOUND)

#     files = Files.objects.filter(ack=ack).exclude(computation_file='')

#     if not files.exists():
#         return Response([], status=status.HTTP_200_OK)

#     file_paths = [file.computation_file.path for file in files]

#     # If you're returning one file at a time
#     file_path = file_paths[0]
#     file_name = os.path.basename(file_path)

#     try:
#         return FileResponse(open(file_path, 'rb'), as_attachment=True, filename=file_name)
#     except FileNotFoundError:
#         raise Http404("File not found.")


@api_view(['GET'])
def download_return_file(request, pk, ack_pk):
    client = Client.objects.get(id=pk)
    ack = Acknowledgement.objects.get(id=ack_pk, client = client)
    if request.method == 'GET':
        ser = AcknowledgementSerializer(ack)
        try:
            ack = Acknowledgement.objects.get(id=ack_pk, client=client)
        except Acknowledgement.DoesNotExist:
            return Response({"error": "Acknowledgement not found."}, status=status.HTTP_404_NOT_FOUND)

        files = Files.objects.filter(ack=ack).exclude(return_file='')

        if not files.exists():
            return Response([], status=status.HTTP_200_OK)

        # Return only the 'file' key with the return_file path
        file_data = [{"file": file.return_file.url} for file in files]

        return Response(file_data, status=status.HTTP_200_OK)


# @api_view(['GET'])
# def serve_computation_file(request, file_id):
#     try:
#         file_obj = Files.objects.get(id=file_id)
#         file_path = file_obj.computation_file.path  # absolute path

#         if not os.path.exists(file_path):
#             raise Http404

#         response = FileResponse(open(file_path, 'rb'))
#         filename = os.path.basename(file_path)
#         response['Content-Disposition'] = f'attachment; filename="{filename}"'
#         return response

#     except Files.DoesNotExist:
#         raise Http404("File not found")

@api_view(['GET'])
def serve_computation_file(request, file_id):
    try:
        file_obj = Files.objects.get(id=file_id)
        file_path = file_obj.computation_file.path

        if not os.path.exists(file_path):
            raise Http404("File not found.")

        with open(file_path, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/octet-stream')  # Force download
            response['Content-Disposition'] = f'attachment; filename="{os.path.basename(file_path)}"'
            return response

    except Files.DoesNotExist:
        raise Http404("File not found.")


# ***********************************************Detail page API's*********************************************

import logging
logger = logging.getLogger(__name__)

@api_view(['GET'])
# @permission_classes([IsAuthenticated])
# @permission_classes([IsSuperAdminOrOwnClient])
@permission_classes([IsAuthenticated, IsSuperAdminOrOwnClient])
def detail_client(request, pk):
    client = Client.objects.get(id=pk)
    view_bank = Bank.objects.filter(client=client)
    view_owner = Owner.objects.filter(client=client)
    view_clientuser = CommonUser.objects.filter(client=client, role='clientuser')
    view_customeruser = CommonUser.objects.filter(client=client, role='customeruser')
    view_companydoc = FileInfo.objects.filter(client=client)
    view_branch = Branch.objects.filter(client=client)
    view_customer = Customer.objects.filter(client=client)
    view_income = IncomeTaxDocument.objects.filter(client=client)
    view_pf = PF.objects.filter(client=client)
    view_taxaudit = TaxAudit.objects.filter(client=client)
    view_air = AIR.objects.filter(client=client)
    view_sft = SFT.objects.filter(client=client)
    view_others = Others.objects.filter(client=client)
    view_tdspayment = TDSPayment.objects.filter(client=client)
    view_tds = TDSReturn.objects.filter(client=client)
    view_tdssection = TDSSection.objects.filter()
    view_sales = SalesInvoice.objects.filter(client=client)
    view_purchase = PurchaseInvoice.objects.filter(client=client)
    view_income = Income.objects.filter(client=client)
    view_expenses = Expenses.objects.filter(client=client)
    view_zipupload  = ZipUpload.objects.filter(client=client)
    view_acknowledgement = Acknowledgement.objects.filter(client=client)
    view_ack = Acknowledgement.objects.filter(client=client)
    
    # Try-except blocks for each serializer
    try:
        client_serializer = ClientSerializer(client)
    except Exception as e:
        logger.error(f"Error serializing Client: {e}")
        raise

    try:
        bank_serializer = BankSerializer(view_bank, many=True)
    except Exception as e:
        logger.error(f"Error serializing Bank: {e}")
        raise

    try:
        owner_serializer = OwnerSerializer(view_owner, many=True)
    except Exception as e:
        print(Owner.objects.all().query)

        logger.error(f"Error serializing Owner: {e}")
        raise

    try:
        clientuser = ClientuserSerializerWithToken(view_clientuser, many=True)
    except Exception as e:
        logger.error(f"Error serializing ClientUser: {e}")
        raise
    try:
        customeruser = CustomeruserSerializerWithToken(view_customeruser, many=True)
    except Exception as e:
        logger.error(f"Error serializing ClientUser: {e}")
        raise

    try:
        companydoc = FileInfoSerializer(view_companydoc, many=True)
    except Exception as e:
        logger.error(f"Error serializing Company_Document: {e}")
        raise

    try:
        branch_serializer = BranchSerailizer(view_branch, many=True)
    except Exception as e:
        logger.error(f"Error serializing Branch: {e}")
        raise

    try:
        customer_serializer = CustomerVendorSerializer(view_customer, many=True)
    except Exception as e:
        logger.error(f"Error serializing Customer_or_Vendor: {e}")
        raise

    try:
        income_serializer = IncomeTaxDocumentSerializer(view_income, many=True)
    except Exception as e:
        logger.error(f"Error serializing Income_Tax_Document: {e}")
        raise

    try:
        pf_serializer = PfSerializer(view_pf, many=True)
    except Exception as e:
        logger.error(f"Error serializing PF: {e}")
        raise

    try:
        taxaudit_serializer = TaxAuditSerializer(view_taxaudit, many=True)
    except Exception as e:
        logger.error(f"Error serializing Tax_Audit: {e}")
        raise

    try:
        air_serializer = AIRSerializer(view_air, many=True)
    except Exception as e:
        logger.error(f"Error serializing AIR: {e}")
        raise

    try:
        sft_serializer = SFTSerializer(view_sft, many=True)
    except Exception as e:
        logger.error(f"Error serializing SFT: {e}")
        raise

    try:
        others_serializer = OthersSerializer(view_others, many=True)
    except Exception as e:
        logger.error(f"Error serializing Others: {e}")
        raise

    try:
        tdspayment_serializer = TDSPaymentSerializer(view_tdspayment, many=True)
    except Exception as e:
        logger.error(f"Error serializing TDS_Payment: {e}")
        raise

    try:
        tdssection_serializer = TDSSectionSerializer(view_tdssection, many=True)
    except Exception as e:
        logger.error(f"Error serializing TDS_Section: {e}")
        raise

    try:
        tds_serializer = TDSReturnSerializer(view_tds, many=True)
    except Exception as e:
        logger.error(f"Error serializing TDS_Return: {e}")
        raise

    try:
        sales_serializer = SalesSerializerList(view_sales, many=True)
        print('ggggggggg')
    except Exception as e:
        print('sales',sales_serializer.error)
        logger.error(f"Error serializing Sales: {e}")
        raise

    try:
        purchase_serializer = PurchaseSerializerList(view_purchase, many=True)
    except Exception as e:
        logger.error(f"Error serializing Purchase: {e}")
        raise

    try:
        income_serializer = IncomeSerializerList(view_income, many=True)
    except Exception as e:
        logger.error(f"Error serializing Income: {e}")
        raise

    try:
        expenses_serializer = ExpensesSerializerList(view_expenses, many=True)
    except Exception as e:
        logger.error(f"Error serializing Expenses: {e}")
        raise

    try:
        zipupload_serializer = ZipUploadSerializer(view_zipupload, many=True)
    except Exception as e:
        logger.error(f"Error serializing ZipUpload: {e}")
        raise
    
    try:
        acknowledgement_serializer = AcknowledgementSerializer(view_acknowledgement, many=True)
    except Exception as e:
        logger.error(f"Error serializing Acknowledgement: {e}")
        raise

    try:
        ack_serializer = AcknowledgementSerializer(view_ack, many=True)
    except Exception as e:
        logger.error(f"Error serializing Acknowledgement: {e}")
        raise

    # Building the final response data
    data = {
        'Client': client_serializer.data,
        'Bank': bank_serializer.data,
        'Owner': owner_serializer.data,
        'ClientUser': clientuser.data,
        'CustomerUser' : customeruser.data,
        'Company_Document': companydoc.data,
        'Branch': branch_serializer.data,
        'Customer_or_Vendor': customer_serializer.data,
        'Income_Tax_Document': income_serializer.data,
        'PF': pf_serializer.data,
        'Tax_Audit': taxaudit_serializer.data,
        'AIR': air_serializer.data,
        'SFT': sft_serializer.data,
        'Others': others_serializer.data,
        'TDS_Payment': tdspayment_serializer.data,
        'TDS_Return': tds_serializer.data,
        'TDS_Section': tdssection_serializer.data,
        'sales_invoice': sales_serializer.data,
        'purchase_invoice': purchase_serializer.data,
        'income': income_serializer.data,
        'expenses': expenses_serializer.data,
        'zipupload': zipupload_serializer.data,
        'acknowledgement': acknowledgement_serializer.data,
        'ack': ack_serializer.data,
    }

    return Response(data)

@api_view(['GET'])
def detail_branch(request, pk, branch_pk):
    client = Client.objects.get(id = pk)
    branch = Branch.objects.get(id = branch_pk, client=client)
    officeloaction = OfficeLocation.objects.filter(branch = branch)
    view_branchdoc = BranchDocument.objects.filter(branch=branch)

    branch_serializer = BranchSerailizer(branch)
    officeloaction_serializer = OfficeLocationSerializer(officeloaction, many=True)
    branchdoc_serializer = BranchDocSerailizer(view_branchdoc, many=True)

    data = {
        'Client_Name' :client.client_name, # to only retrive client name
        'Branch' : branch_serializer.data,
        'Branch_Document' : branchdoc_serializer.data,
        'Office_Location' : officeloaction_serializer.data,
    }
    return Response(data)

@api_view(['GET'])
def detail_acknowledgement(request, pk):
    client = Client.objects.get(id=pk)
    ack = Acknowledgement.objects.filter(client=client)

    client_serializer = ClientSerializer(client)
    ack_serializer = AcknowledgementSerializer(ack, many=True)

    data = {
        'Client_Name' :client.client_name,
        'Client': client_serializer.data,  # to only retrieve client name
        'Acknowledgement': ack_serializer.data,
    }   
    return Response(data)

@api_view(['POST'])
def import_hsn_excel(request):
    if request.method == 'POST':
        # Check if an Excel file is provided
        if 'file' not in request.FILES:
            return Response({'error_message': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)

        # Get the uploaded Excel file
        excel_file = request.FILES['file']

        try:
            # Read the Excel file using pandas
            df = pd.read_excel(excel_file)

            # List to hold successful imports and errors
            successful_imports = []
            errors = []

            # Ensure there are at least two columns
            if df.shape[1] < 2:
                return Response({'error_message': 'The Excel file must contain at least two columns'}, status=status.HTTP_400_BAD_REQUEST)

            # Iterate through each row and create a new record
            for index, row in df.iterrows():
                data = {}

                # Assign the first column to hsn_code and the second to gst_rate
                try:
                    data['hsn_code'] = int(row[0])  # Convert to int
                    data['gst_rate'] = float(row[1])  # Convert to float
                except ValueError as ve:
                    errors.append({'row': index + 1, 'error_message': str(ve)})
                    continue  # Skip this row if there's a conversion error

                # Use the serializer to validate and save data
                serializer = HSNSerializer(data=data)
                if serializer.is_valid():
                    serializer.save()
                    successful_imports.append(data)
                else:
                    errors.append({'row': index + 1, 'errors': serializer.errors})

            # Prepare response
            response_message = {'message': 'HSN records imported successfully', 'Imported': successful_imports}
            if errors:
                response_message['Errors'] = errors

            return Response(response_message, status=status.HTTP_201_CREATED)

        # except Exception as e:
        #     return Response({'Error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            error_details = traceback.format_exc()
            # print({"Error in update_sales_invoice" : str(e)})
            return Response({"error_message":str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST','GET'])
def create_hsn(request):
    if request.method == 'POST':
        serializer = HSNSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response ({'message':'TDS Payment created', 'Data': serializer.data})
        return Response({'error_message':serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET', 'POST'])
def edit_hsn(request, pk):
    hsn = HSNCode.objects.get(id=pk)
    serializer = HSNSerializer(instance=hsn, data=request.data)
    if request.method == 'GET':
        hsn_serializer = HSNSerializer(hsn)
        # print(tds_serializer.data)
        return Response (hsn_serializer.data)
    elif request.method == 'POST':
        if serializer.is_valid():
            serializer.save()
            return Response ({'message':'hsn Code Updated'})
        return Response(serializer.errors,status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
def list_hsn(request):
    if request.method == 'GET':
        hsn_code = HSNCode.objects.all()
        serializer = HSNSerializer(hsn_code, many=True)
        return Response(serializer.data)

@api_view(['DELETE'])
def delete_hsn(request, pk):
    hsn = HSNCode.objects.get(id = pk)
    if request.method == 'DELETE':
        hsn.delete()
        return Response({'message':'HSN Return Delete'})
    return Response({'message':'Fail to delete HSN Return'} ,status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST','GET'])
def create_product(request):
    if request.method=="GET":
        hsn_list = HSNCode.objects.all()
        serializer =HSNSerializer(hsn_list,many=True)
        return Response(serializer.data)
    if request.method == 'POST':
        serializer = ProductSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response ({'message':'Product Payment created', 'Data': serializer.data})
        return Response({'error_message':serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET', 'POST'])
def edit_product(request, pk):
    product = Product.objects.get(id=pk)
    serializer = ProductSerializer(instance=product, data=request.data)
    if request.method == 'GET':
        product_serializer = ProductSerializer(product)
        # print(tds_serializer.data)
        return Response (product_serializer.data)
    elif request.method == 'POST':
        if serializer.is_valid():
            serializer.save()
            return Response ({'message':'Product Updated'})
        return Response(serializer.errors,status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
def list_product(request):
    if request.method == 'GET':
        product = Product.objects.all()
        serializer = ProductSerializer(product, many=True)
        print(serializer.data,'lklklkk')
        return Response(serializer.data)

@api_view(['DELETE'])
def delete_product(request, pk):
    product = Product.objects.get(id = pk)
    if request.method == 'DELETE':
        product.delete()
        return Response({'message':'Product Return Delete'})
    return Response({'message':'Fail to delete HSN Return'} ,status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST','GET'])
def create_product_description(request):
    if request.method=="GET":
        product = Product.objects.all()
        serializer =ProductSerializer(product,many=True)
        return Response(serializer.data)
    if request.method == 'POST':
        serializer = ProductDescriptionSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response ({'message':'Product Description created', 'Data': serializer.data})
        return Response({'error_message':serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
def list_product_description(request):
    if request.method == 'GET':
        product_description = ProductDescription.objects.all()
        serializer = ProductDescriptionSerializer(product_description, many=True)
        print(serializer.data,'lklklkk')
        return Response(serializer.data)

@api_view(['GET', 'POST'])
def edit_product_description(request, pk):
    product_description = ProductDescription.objects.get(id=pk)
    serializer = ProductDescriptionSerializer(instance=product_description, data=request.data)
    if request.method == 'GET':
        product_serializer = ProductDescriptionSerializer(product_description)
        # print(tds_serializer.data)
        return Response (product_serializer.data)
    elif request.method == 'POST':
        if serializer.is_valid():
            serializer.save()
            return Response ({'message':'Product Description Updated'})
        return Response(serializer.errors,status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
def delete_product_description(request, pk):
    product_description = ProductDescription.objects.get(id = pk)
    if request.method == 'DELETE':
        product_description.delete()
        return Response({'message':'Product Description Return Delete'})
    return Response({'message':'Fail to delete Product Description Return'} ,status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
def sales_invoice_list(request):
    """
    View to list all sales invoices with their related product summaries.
    """
    # Fetch all sales invoices with related data (optimized query with prefetch_related)
    sales_invoices = SalesInvoice.objects.prefetch_related(
        'product_summaries', 'product_summaries__product', 'product_summaries__hsn',
        'product_summaries__prod_description', 'customer', 'client_location'
    ).all()

    # Serialize the data
    serializer = SalesSerializer(sales_invoices, many=True)

    # Return the serialized data
    return Response(serializer.data, status=status.HTTP_200_OK)

@api_view(['GET'])
def purchase_invoice_list(request):
    """
    View to list all sales invoices with their related product summaries.
    """
    # Fetch all sales invoices with related data (optimized query with prefetch_related)
    purchase_invoices = PurchaseInvoice.objects.prefetch_related(
        'product_summaries', 'product_summaries__product', 'product_summaries__hsn',
        'product_summaries__prod_description', 'customer', 'client_location'
    ).all()

    # Serialize the data
    serializer = SalesSerializer(purchase_invoices, many=True)

    # Return the serialized data
    return Response(serializer.data, status=status.HTTP_200_OK)

@api_view(['GET'])
def debit_list(request,client_pk, invoice_pk):
    client = Client.objects.get(id=client_pk)
    invoice= SalesInvoice.objects.get(client=client, id=invoice_pk)
    debit_note= DebitNote.objects.filter(client=client, sales_invoice=invoice)
    debit_list = DebitNoteSerializerList(debit_note,many=True)
    # print('hghhg',debit_list.data)
    return Response(debit_list.data)

@api_view(['GET'])
def credit_list(request,client_pk, invoice_pk):
    client = Client.objects.get(id=client_pk)
    invoice= PurchaseInvoice.objects.get(client=client, id=invoice_pk)
    credit_note= CreditNote.objects.filter(client=client, purchase_invoice=invoice)
    credit_list = CreditNoteSerializerList(credit_note,many=True)
    print('hghhg',credit_list.data)
    return Response(credit_list.data)

@api_view(['GET'])
def income_debit_list(request,client_pk, income_pk):
    client = Client.objects.get(id=client_pk)
    income= Income.objects.get(client=client, id=income_pk)
    debit_note= IncomeDebitNote.objects.filter(client=client, income=income)
    debit_list = IncomeDebitNoteSerializerList(debit_note,many=True)
    print('hghhg',debit_list.data)
    return Response(debit_list.data)

@api_view(['GET'])
def expenses_credit_list(request,client_pk, expenses_pk):
    client = Client.objects.get(id=client_pk)
    expenses= Expenses.objects.get(client=client, id=expenses_pk)
    credit_note= ExpensesCreditNote.objects.filter(client=client, expenses=expenses)
    credit_list = ExpensesCreditNoteSerializerList(credit_note,many=True)
    print('hghhg',credit_list.data)
    return Response(credit_list.data)
   
